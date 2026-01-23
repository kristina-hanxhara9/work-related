#!/usr/bin/env python3
"""
=============================================================================
MOBILE TRUCK TYRE SERVICES SCRAPER - UK
=============================================================================
Specifically targets MOBILE-ONLY truck tyre fitting services:
- 24/7 emergency callout
- Roadside truck tyre fitting
- Mobile HGV tyre services
- Fleet mobile services

Sources:
1. Companies House API - search for mobile tyre companies
2. Google Search - find mobile services
3. Yell.com - business directory
4. FreeIndex - local services

Date: January 2026
=============================================================================
"""

import requests
from bs4 import BeautifulSoup
import json
import time
import re
import csv
from datetime import datetime
from urllib.parse import urlparse, quote
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Companies House API
CH_API_KEY = '48d17266-ff2e-425f-9b20-7dcc9b25bb79'
CH_BASE_URL = 'https://api.company-information.service.gov.uk'

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'en-GB,en;q=0.9',
}

OUTPUT_FILE = 'uk_mobile_truck_tyres'


def search_companies_house(query, items_per_page=100):
    """Search Companies House for mobile tyre companies"""
    companies = []

    try:
        url = f"{CH_BASE_URL}/search/companies"
        params = {
            'q': query,
            'items_per_page': items_per_page
        }

        response = requests.get(url, params=params, auth=(CH_API_KEY, ''), timeout=15)

        if response.status_code == 200:
            data = response.json()
            items = data.get('items', [])

            for item in items:
                name = item.get('title', '')
                status = item.get('company_status', '')

                # Only active companies
                if status.lower() not in ['active', 'open']:
                    continue

                # Must be tyre related
                name_lower = name.lower()
                if not any(word in name_lower for word in ['tyre', 'tire', 'tyres', 'tires']):
                    continue

                companies.append({
                    'name': name,
                    'company_number': item.get('company_number', ''),
                    'company_status': status,
                    'address': item.get('address_snippet', ''),
                    'date_created': item.get('date_of_creation', ''),
                    'company_type': item.get('company_type', ''),
                    'source': f'Companies House ({query})'
                })

        return companies
    except Exception as e:
        print(f"    Error searching CH for '{query}': {e}")
        return []


def get_company_details(company_number):
    """Get full details from Companies House"""
    try:
        url = f"{CH_BASE_URL}/company/{company_number}"
        response = requests.get(url, auth=(CH_API_KEY, ''), timeout=10)

        if response.status_code == 200:
            data = response.json()

            # Get registered office address
            address_parts = []
            reg_office = data.get('registered_office_address', {})
            for field in ['premises', 'address_line_1', 'address_line_2', 'locality', 'region', 'postal_code']:
                val = reg_office.get(field)
                if val:
                    address_parts.append(val)

            return {
                'full_address': ', '.join(address_parts),
                'postcode': reg_office.get('postal_code', ''),
                'locality': reg_office.get('locality', ''),
                'region': reg_office.get('region', ''),
                'sic_codes': data.get('sic_codes', []),
                'type': data.get('type', '')
            }
        return None
    except:
        return None


def scrape_yell(search_term, location="UK"):
    """Scrape Yell.com for mobile tyre services"""
    companies = []
    base_url = f"https://www.yell.com/ucs/UcsSearchAction.do"

    try:
        params = {
            'scrambleSeed': '',
            'keywords': search_term,
            'location': location
        }

        response = requests.get(base_url, params=params, headers=HEADERS, timeout=15)

        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')

            # Find business listings
            listings = soup.find_all('div', class_='businessCapsule--mainRow')

            for listing in listings[:50]:  # Limit per search
                try:
                    name_elem = listing.find('h2', class_='businessCapsule--name')
                    name = name_elem.get_text(strip=True) if name_elem else None

                    if not name:
                        continue

                    # Get address
                    addr_elem = listing.find('span', class_='businessCapsule--address')
                    address = addr_elem.get_text(strip=True) if addr_elem else ''

                    # Get phone
                    phone_elem = listing.find('span', class_='business--telephoneNumber')
                    phone = phone_elem.get_text(strip=True) if phone_elem else ''

                    # Get website
                    website = ''
                    web_link = listing.find('a', class_='businessCapsule--ctaItem', href=True)
                    if web_link and 'website' in str(web_link).lower():
                        website = web_link.get('href', '')

                    companies.append({
                        'name': name,
                        'address': address,
                        'phone': phone,
                        'website': website,
                        'source': f'Yell.com ({search_term})'
                    })
                except:
                    continue

    except Exception as e:
        print(f"    Yell.com error for '{search_term}': {e}")

    return companies


def scrape_google_places_text(search_term):
    """Search using Google (text-based, no API needed)"""
    companies = []

    # We'll use a different approach - search for specific terms
    # and extract what we can from search results

    search_url = f"https://www.google.com/search?q={quote(search_term)}+UK"

    try:
        response = requests.get(search_url, headers=HEADERS, timeout=10)

        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')

            # Extract business names from search results
            # This is limited but can find some businesses

            for div in soup.find_all(['div', 'span']):
                text = div.get_text(strip=True)

                # Look for business-like names with tyre keywords
                if len(text) > 5 and len(text) < 80:
                    text_lower = text.lower()
                    if 'mobile' in text_lower and ('tyre' in text_lower or 'tire' in text_lower):
                        if any(word in text_lower for word in ['ltd', 'limited', 'services', 'uk', 'tyres']):
                            # Avoid duplicates and non-company text
                            if not any(c['name'].lower() == text.lower() for c in companies):
                                companies.append({
                                    'name': text,
                                    'source': 'Google Search',
                                    'search_term': search_term
                                })
    except Exception as e:
        print(f"    Google search error: {e}")

    return companies[:20]  # Limit


def scrape_freeindex(search_term):
    """Scrape FreeIndex.co.uk"""
    companies = []

    url = f"https://www.freeindex.co.uk/search/?q={quote(search_term)}"

    try:
        response = requests.get(url, headers=HEADERS, timeout=15)

        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')

            listings = soup.find_all('div', class_='listing')

            for listing in listings[:30]:
                try:
                    name_elem = listing.find('h2') or listing.find('a', class_='listing_title')
                    name = name_elem.get_text(strip=True) if name_elem else None

                    if not name:
                        continue

                    # Get phone
                    phone = ''
                    phone_elem = listing.find('span', class_='telephone') or listing.find(text=re.compile(r'0\d{2,4}\s?\d{3,4}\s?\d{3,4}'))
                    if phone_elem:
                        phone = phone_elem.get_text(strip=True) if hasattr(phone_elem, 'get_text') else str(phone_elem)

                    companies.append({
                        'name': name,
                        'phone': phone,
                        'source': f'FreeIndex ({search_term})'
                    })
                except:
                    continue

    except Exception as e:
        print(f"    FreeIndex error: {e}")

    return companies


def verify_website(url):
    """Check if website works and extract contact info"""
    if not url or not url.startswith('http'):
        return None

    try:
        response = requests.get(url, headers=HEADERS, timeout=12, allow_redirects=True)

        if response.status_code == 200:
            text = response.text
            text_lower = text.lower()

            result = {
                'works': True,
                'final_url': response.url,
                'phone': None,
                'email': None,
                'is_mobile_service': False,
                'is_truck_tyre': False
            }

            # Check for mobile service indicators
            mobile_words = ['mobile', '24/7', '24 hour', 'callout', 'call out', 'roadside',
                          'emergency', 'breakdown', 'on-site', 'onsite', 'come to you']
            truck_words = ['truck', 'hgv', 'commercial', 'fleet', 'lorry', 'trailer', 'heavy']
            tyre_words = ['tyre', 'tire', 'fitting', 'fitter']

            mobile_count = sum(1 for w in mobile_words if w in text_lower)
            truck_count = sum(1 for w in truck_words if w in text_lower)
            tyre_count = sum(1 for w in tyre_words if w in text_lower)

            result['is_mobile_service'] = mobile_count >= 2
            result['is_truck_tyre'] = truck_count >= 1 and tyre_count >= 1

            # Extract phone
            phone_patterns = [
                r'0800[\s\-]?\d{3}[\s\-]?\d{3,4}',
                r'0808[\s\-]?\d{3}[\s\-]?\d{3,4}',
                r'01\d{2,3}[\s\-]?\d{3}[\s\-]?\d{3,4}',
                r'02\d[\s\-]?\d{4}[\s\-]?\d{4}',
                r'03\d{2}[\s\-]?\d{3}[\s\-]?\d{4}',
                r'07\d{3}[\s\-]?\d{3}[\s\-]?\d{3}',
            ]
            for pattern in phone_patterns:
                matches = re.findall(pattern, text)
                for m in matches:
                    cleaned = re.sub(r'[\s\-]', '', m)
                    if 10 <= len(cleaned) <= 12:
                        result['phone'] = m.strip()
                        break
                if result['phone']:
                    break

            # Extract email
            emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', text)
            for email in emails:
                if not any(x in email.lower() for x in ['.png', '.jpg', '.css', '.js', 'example.', 'wix.']):
                    result['email'] = email
                    break

            return result
        return None
    except:
        return None


def is_mobile_service(name, business_type=''):
    """Check if company name suggests mobile service"""
    text = f"{name} {business_type}".lower()

    mobile_indicators = ['mobile', '24/7', '24hr', '24 hour', 'callout', 'call out',
                        'roadside', 'emergency', 'breakdown', 'rapid', 'express',
                        'on-site', 'onsite', 'rescue', 'assist', 'response']

    return any(ind in text for ind in mobile_indicators)


def is_truck_related(name):
    """Check if company is truck/commercial vehicle related"""
    text = name.lower()

    truck_indicators = ['truck', 'hgv', 'commercial', 'fleet', 'lorry', 'trailer',
                       'heavy', 'lgv', 'van', 'coach', 'bus']

    return any(ind in text for ind in truck_indicators)


def extract_region(address):
    """Extract region from address"""
    if not address:
        return ''

    regions = {
        'Scotland': ['scotland', 'edinburgh', 'glasgow', 'aberdeen', 'dundee', 'inverness'],
        'Wales': ['wales', 'cardiff', 'swansea', 'newport', 'wrexham'],
        'Northern Ireland': ['northern ireland', 'belfast', 'derry', 'lisburn'],
        'North East': ['newcastle', 'sunderland', 'durham', 'middlesbrough', 'tyne'],
        'North West': ['manchester', 'liverpool', 'preston', 'blackpool', 'bolton', 'wigan'],
        'Yorkshire': ['leeds', 'sheffield', 'bradford', 'hull', 'york', 'doncaster'],
        'East Midlands': ['nottingham', 'derby', 'leicester', 'lincoln', 'northampton'],
        'West Midlands': ['birmingham', 'coventry', 'wolverhampton', 'stoke', 'walsall'],
        'East of England': ['norwich', 'cambridge', 'ipswich', 'peterborough', 'colchester'],
        'London': ['london', 'croydon', 'bromley', 'barnet', 'enfield'],
        'South East': ['brighton', 'southampton', 'portsmouth', 'reading', 'oxford', 'kent', 'surrey'],
        'South West': ['bristol', 'exeter', 'plymouth', 'bournemouth', 'bath', 'cornwall']
    }

    addr_lower = address.lower()
    for region, keywords in regions.items():
        if any(kw in addr_lower for kw in keywords):
            return region
    return ''


def main():
    print("=" * 80)
    print("MOBILE TRUCK TYRE SERVICES SCRAPER - UK")
    print("=" * 80)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    all_companies = {}  # Use dict to dedupe by name

    # =========================================================================
    # STEP 1: Companies House API - Multiple Search Terms
    # =========================================================================
    print("[1] SEARCHING COMPANIES HOUSE API")
    print("-" * 60)

    ch_search_terms = [
        'mobile tyre',
        'mobile tyres',
        'mobile truck tyre',
        'mobile hgv tyre',
        'mobile commercial tyre',
        '24 hour tyre',
        '24/7 tyre',
        'emergency tyre',
        'roadside tyre',
        'truck tyre mobile',
        'fleet mobile tyre',
        'callout tyre',
        'breakdown tyre',
        'mobile tyre fitting',
        'mobile tyre service',
        'rapid tyre',
        'express tyre mobile'
    ]

    ch_count = 0
    for term in ch_search_terms:
        print(f"    Searching: '{term}'...")
        results = search_companies_house(term)

        for c in results:
            name_key = c['name'].lower().strip()
            if name_key not in all_companies:
                # Get more details
                details = get_company_details(c['company_number'])
                if details:
                    c['full_address'] = details.get('full_address', c.get('address', ''))
                    c['postcode'] = details.get('postcode', '')
                    c['region'] = details.get('region', '') or extract_region(c.get('address', ''))
                    c['sic_codes'] = details.get('sic_codes', [])

                c['is_mobile'] = is_mobile_service(c['name'])
                c['is_truck'] = is_truck_related(c['name'])
                c['verified'] = True
                c['verification_method'] = 'Companies House API'

                all_companies[name_key] = c
                ch_count += 1
                print(f"      ✓ {c['name']}")

        time.sleep(0.5)  # Rate limiting

    print(f"\n    Found {ch_count} companies from Companies House")

    # =========================================================================
    # STEP 2: Yell.com Scraping
    # =========================================================================
    print("\n[2] SCRAPING YELL.COM")
    print("-" * 60)

    yell_searches = [
        'mobile truck tyre fitting',
        'mobile hgv tyre service',
        '24 hour truck tyres',
        'emergency truck tyre',
        'mobile commercial tyre',
        'roadside truck tyre'
    ]

    yell_count = 0
    for term in yell_searches:
        print(f"    Searching: '{term}'...")
        results = scrape_yell(term)

        for c in results:
            name_key = c['name'].lower().strip()
            if name_key not in all_companies:
                c['is_mobile'] = is_mobile_service(c['name'])
                c['is_truck'] = is_truck_related(c['name'])
                c['verified'] = False
                c['verification_method'] = 'Yell.com'
                c['region'] = extract_region(c.get('address', ''))

                all_companies[name_key] = c
                yell_count += 1
                print(f"      + {c['name']}")

        time.sleep(1)

    print(f"\n    Found {yell_count} new companies from Yell.com")

    # =========================================================================
    # STEP 3: FreeIndex Scraping
    # =========================================================================
    print("\n[3] SCRAPING FREEINDEX")
    print("-" * 60)

    freeindex_searches = [
        'mobile truck tyre',
        'mobile hgv tyre',
        '24 hour commercial tyre',
        'emergency tyre fitting'
    ]

    fi_count = 0
    for term in freeindex_searches:
        print(f"    Searching: '{term}'...")
        results = scrape_freeindex(term)

        for c in results:
            name_key = c['name'].lower().strip()
            if name_key not in all_companies:
                c['is_mobile'] = is_mobile_service(c['name'])
                c['is_truck'] = is_truck_related(c['name'])
                c['verified'] = False
                c['verification_method'] = 'FreeIndex'

                all_companies[name_key] = c
                fi_count += 1
                print(f"      + {c['name']}")

        time.sleep(1)

    print(f"\n    Found {fi_count} new companies from FreeIndex")

    # =========================================================================
    # STEP 4: Verify non-CH companies via Companies House search
    # =========================================================================
    print("\n[4] VERIFYING COMPANIES VIA COMPANIES HOUSE")
    print("-" * 60)

    verified_count = 0
    for name_key, c in all_companies.items():
        if not c.get('company_number'):
            # Search Companies House by name
            search_name = c['name'].replace('Ltd', '').replace('Limited', '').strip()

            try:
                url = f"{CH_BASE_URL}/search/companies"
                params = {'q': search_name, 'items_per_page': 5}
                response = requests.get(url, params=params, auth=(CH_API_KEY, ''), timeout=10)

                if response.status_code == 200:
                    data = response.json()
                    items = data.get('items', [])

                    for item in items:
                        ch_name = item.get('title', '').lower()
                        if search_name.lower()[:20] in ch_name or ch_name[:20] in search_name.lower():
                            c['company_number'] = item.get('company_number', '')
                            c['company_status'] = item.get('company_status', '')
                            c['verified'] = True
                            c['verification_method'] = 'Companies House Search'
                            verified_count += 1
                            print(f"    ✓ {c['name']} -> {item.get('company_number')}")
                            break

                time.sleep(0.3)
            except:
                pass

    print(f"\n    Verified {verified_count} additional companies")

    # =========================================================================
    # STEP 5: Check websites for unverified companies
    # =========================================================================
    print("\n[5] VERIFYING WEBSITES")
    print("-" * 60)

    website_verified = 0
    for name_key, c in all_companies.items():
        if not c.get('verified') and c.get('website'):
            result = verify_website(c['website'])
            if result and result.get('works'):
                c['website_works'] = True
                c['phone'] = c.get('phone') or result.get('phone', '')
                c['email'] = result.get('email', '')
                c['is_mobile'] = c.get('is_mobile') or result.get('is_mobile_service', False)
                c['is_truck'] = c.get('is_truck') or result.get('is_truck_tyre', False)

                if result.get('is_mobile_service') and result.get('is_truck_tyre'):
                    c['verified'] = True
                    c['verification_method'] = 'Website Content'
                    website_verified += 1
                    print(f"    ✓ {c['name']} - website verified as mobile truck tyre")

            time.sleep(0.5)

    print(f"\n    Website verified: {website_verified}")

    # =========================================================================
    # STEP 6: Filter to only mobile truck tyre services
    # =========================================================================
    print("\n[6] FILTERING TO MOBILE TRUCK TYRE SERVICES")
    print("-" * 60)

    # Convert to list and filter
    companies_list = list(all_companies.values())

    # Filter for mobile services
    mobile_companies = []
    for c in companies_list:
        name_lower = c['name'].lower()

        # Must have mobile/24hr/emergency indicator
        has_mobile = c.get('is_mobile', False) or any(word in name_lower for word in
            ['mobile', '24', 'emergency', 'roadside', 'callout', 'rapid', 'express', 'rescue', 'breakdown'])

        # Should be tyre related
        has_tyre = 'tyre' in name_lower or 'tire' in name_lower

        # Preferably truck/commercial
        has_truck = c.get('is_truck', False) or any(word in name_lower for word in
            ['truck', 'hgv', 'commercial', 'fleet', 'lorry', 'heavy'])

        if has_mobile and has_tyre:
            c['service_type'] = 'Mobile Truck Tyre' if has_truck else 'Mobile Tyre (General)'
            mobile_companies.append(c)

    # Sort by verification status and truck relevance
    mobile_companies.sort(key=lambda x: (
        not x.get('verified', False),
        not x.get('is_truck', False),
        x.get('name', '')
    ))

    print(f"    Total mobile tyre companies: {len(mobile_companies)}")
    print(f"    Verified: {len([c for c in mobile_companies if c.get('verified')])}")
    print(f"    Truck/HGV specific: {len([c for c in mobile_companies if c.get('is_truck')])}")

    # =========================================================================
    # STEP 7: Export
    # =========================================================================
    print("\n[7] EXPORTING DATA")
    print("-" * 60)

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Mobile Truck Tyre Services"

    headers = ['Company Name', 'Verified', 'Verification Method', 'Service Type',
               'Is Truck/HGV', 'Company Number', 'Status', 'Phone', 'Email',
               'Website', 'Address', 'Region', 'Source']

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    verified_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    truck_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row, c in enumerate(mobile_companies, 2):
        ws.cell(row=row, column=1, value=c.get('name', ''))

        verified_cell = ws.cell(row=row, column=2, value='Yes' if c.get('verified') else 'No')
        if c.get('verified'):
            verified_cell.fill = verified_fill

        ws.cell(row=row, column=3, value=c.get('verification_method', ''))
        ws.cell(row=row, column=4, value=c.get('service_type', ''))

        truck_cell = ws.cell(row=row, column=5, value='Yes' if c.get('is_truck') else 'No')
        if c.get('is_truck'):
            truck_cell.fill = truck_fill

        ws.cell(row=row, column=6, value=c.get('company_number', ''))
        ws.cell(row=row, column=7, value=c.get('company_status', ''))
        ws.cell(row=row, column=8, value=c.get('phone', ''))
        ws.cell(row=row, column=9, value=c.get('email', ''))
        ws.cell(row=row, column=10, value=c.get('website', ''))
        ws.cell(row=row, column=11, value=c.get('full_address', '') or c.get('address', ''))
        ws.cell(row=row, column=12, value=c.get('region', ''))
        ws.cell(row=row, column=13, value=c.get('source', ''))

    widths = [45, 10, 25, 20, 12, 12, 10, 18, 30, 45, 50, 15, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:M{len(mobile_companies) + 1}"

    wb.save(f"{OUTPUT_FILE}.xlsx")
    print(f"    Saved: {OUTPUT_FILE}.xlsx")

    # JSON
    with open(f"{OUTPUT_FILE}.json", 'w', encoding='utf-8') as f:
        json.dump(mobile_companies, f, indent=2, ensure_ascii=False, default=str)
    print(f"    Saved: {OUTPUT_FILE}.json")

    # CSV
    fieldnames = ['name', 'verified', 'verification_method', 'service_type', 'is_truck',
                  'company_number', 'company_status', 'phone', 'email', 'website',
                  'address', 'region', 'source']

    with open(f"{OUTPUT_FILE}.csv", 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        for c in mobile_companies:
            row = {
                'name': c.get('name', ''),
                'verified': 'Yes' if c.get('verified') else 'No',
                'verification_method': c.get('verification_method', ''),
                'service_type': c.get('service_type', ''),
                'is_truck': 'Yes' if c.get('is_truck') else 'No',
                'company_number': c.get('company_number', ''),
                'company_status': c.get('company_status', ''),
                'phone': c.get('phone', ''),
                'email': c.get('email', ''),
                'website': c.get('website', ''),
                'address': c.get('full_address', '') or c.get('address', ''),
                'region': c.get('region', ''),
                'source': c.get('source', '')
            }
            writer.writerow(row)
    print(f"    Saved: {OUTPUT_FILE}.csv")

    # =========================================================================
    # SUMMARY
    # =========================================================================
    print("\n" + "=" * 80)
    print("MOBILE TRUCK TYRE SCRAPER COMPLETE")
    print("=" * 80)

    total = len(mobile_companies)
    verified = len([c for c in mobile_companies if c.get('verified')])
    truck_specific = len([c for c in mobile_companies if c.get('is_truck')])
    with_phone = len([c for c in mobile_companies if c.get('phone')])
    with_ch = len([c for c in mobile_companies if c.get('company_number')])

    print(f"\nTotal Mobile Tyre Services: {total}")
    print(f"  ✓ Verified (CH or Website): {verified} ({verified/total*100:.1f}%)" if total > 0 else "")
    print(f"  ★ Truck/HGV Specific: {truck_specific}")
    print(f"  📞 With Phone Number: {with_phone}")
    print(f"  🏢 With CH Number: {with_ch}")

    print("\n\nTOP 20 VERIFIED MOBILE TRUCK TYRE SERVICES:")
    print("-" * 60)

    verified_list = [c for c in mobile_companies if c.get('verified') and c.get('is_truck')]
    for i, c in enumerate(verified_list[:20], 1):
        print(f"\n{i}. {c['name']}")
        if c.get('company_number'):
            print(f"   CH: {c['company_number']}")
        if c.get('phone'):
            print(f"   Phone: {c['phone']}")
        if c.get('region'):
            print(f"   Region: {c['region']}")


if __name__ == "__main__":
    main()
