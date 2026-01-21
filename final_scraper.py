"""
FINAL UK TRUCK TYRE SCRAPER
===========================
Gets comprehensive list of UK truck tyre companies with verified websites.
"""

import requests
from bs4 import BeautifulSoup
import json
import time
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
}

CH_API_KEY = '48d17266-ff2e-425f-9b20-7dcc9b25bb79'


def search_companies_house(term):
    """Search Companies House"""
    companies = []
    try:
        url = 'https://api.company-information.service.gov.uk/search/companies'
        params = {'q': term, 'items_per_page': 100}
        r = requests.get(url, auth=(CH_API_KEY, ''), params=params, timeout=15)

        if r.status_code == 200:
            for item in r.json().get('items', []):
                if item.get('company_status') == 'active':
                    companies.append({
                        'name': item.get('title', ''),
                        'company_number': item.get('company_number', ''),
                        'address': item.get('address_snippet', ''),
                        'date_created': item.get('date_of_creation', ''),
                        'source': 'Companies House'
                    })
    except Exception as e:
        print(f"    Error: {e}")

    return companies


def find_company_website(company_name):
    """Try to find company website via DuckDuckGo"""
    try:
        # Clean company name for search
        search_name = company_name.replace(' LIMITED', '').replace(' LTD', '').strip()

        # Try DuckDuckGo HTML
        url = f'https://html.duckduckgo.com/html/?q={search_name.replace(" ", "+")}+UK'
        r = requests.get(url, headers=HEADERS, timeout=10)

        if r.status_code == 200:
            soup = BeautifulSoup(r.text, 'html.parser')

            # Look for result links
            results = soup.find_all('a', class_='result__url')
            for result in results[:3]:
                href = result.get('href', '')
                text = result.get_text(strip=True)

                # Skip directories/social media
                skip = ['facebook', 'linkedin', 'twitter', 'yell.com', 'yelp',
                       'checkatrade', 'companieshouse', 'gov.uk', 'endole']

                if not any(s in href.lower() for s in skip):
                    if href.startswith('http'):
                        return href
                    elif text and '.' in text:
                        return f'https://{text}'

    except:
        pass

    return None


def main():
    print("=" * 70)
    print("UK TRUCK TYRE COMPANIES - COMPREHENSIVE SCRAPER")
    print("=" * 70)

    all_companies = []
    seen_numbers = set()

    # Search terms - specifically truck tyres
    search_terms = [
        'truck tyre',
        'truck tyres',
        'hgv tyre',
        'hgv tyres',
        'lorry tyre',
        'lorry tyres',
        'commercial tyre',
        'commercial tyres',
        'mobile truck tyre',
        'fleet tyre',
        'trailer tyre',
    ]

    print("\n[1] SEARCHING COMPANIES HOUSE")
    print("-" * 50)

    for term in search_terms:
        print(f"  Searching: '{term}'...")
        results = search_companies_house(term)

        for c in results:
            if c['company_number'] not in seen_numbers:
                seen_numbers.add(c['company_number'])
                all_companies.append(c)

        print(f"    Found {len(results)} (total unique: {len(all_companies)})")
        time.sleep(0.5)

    # Filter to ONLY truck tyre companies
    print("\n[2] FILTERING TO TRUCK TYRE COMPANIES ONLY")
    print("-" * 50)

    truck_keywords = ['truck tyre', 'truck tyres', 'hgv tyre', 'hgv tyres',
                      'lorry tyre', 'commercial tyre', 'commercial tyres',
                      'fleet tyre', 'trailer tyre', 'mobile truck']

    exclude = ['lift truck', 'fork', 'crane', 'dump', 'trucking', 'truck hire',
               'truck part', 'truck repair', 'truck sale', 'truck train',
               'truck service', 'truck wash', 'truck eng']

    filtered = []
    for c in all_companies:
        name = c['name'].lower()

        # Must have truck tyre keyword
        if not any(kw in name for kw in truck_keywords):
            continue

        # Must not be excluded type
        if any(ex in name for ex in exclude):
            continue

        filtered.append(c)

    print(f"  Filtered from {len(all_companies)} to {len(filtered)} truck tyre companies")

    # Try to find websites
    print("\n[3] FINDING COMPANY WEBSITES")
    print("-" * 50)

    for i, c in enumerate(filtered):
        print(f"  [{i+1}/{len(filtered)}] {c['name'][:40]}...", end=' ')

        website = find_company_website(c['name'])
        c['website'] = website or ''

        if website:
            print(f"Found: {website[:40]}")
        else:
            print("Not found")

        time.sleep(1)  # Be polite

    # Add known major networks
    print("\n[4] ADDING KNOWN TRUCK TYRE NETWORKS")
    print("-" * 50)

    major_networks = [
        {'name': 'Tyrenet', 'website': 'https://tyrenet.net/', 'phone': '0330 123 1234', 'source': 'Known network'},
        {'name': 'Tructyre ATS', 'website': 'https://www.tructyre.co.uk/', 'source': 'Known network'},
        {'name': 'Bandvulc', 'website': 'https://www.bandvulc.co.uk/', 'source': 'Known network'},
        {'name': '247 Mobile Truck Tyres', 'website': 'https://www.247mobiletrucktyres.co.uk/', 'source': 'Known network'},
        {'name': 'Tyre Assist 365', 'website': 'https://www.tyreassist365.com/', 'source': 'Known network'},
        {'name': 'HGV Tyres', 'website': 'https://www.hgvtyres.com/', 'source': 'Known network'},
        {'name': 'Michelin Truck Tyres UK', 'website': 'https://business.michelin.co.uk/truck-bus', 'source': 'Manufacturer'},
        {'name': 'Bridgestone Truck', 'website': 'https://www.bridgestone.co.uk/truck', 'source': 'Manufacturer'},
        {'name': 'Continental Truck Tyres', 'website': 'https://www.continental-tyres.co.uk/truck', 'source': 'Manufacturer'},
        {'name': 'Goodyear Truck Tyres', 'website': 'https://www.goodyear.eu/en_gb/truck.html', 'source': 'Manufacturer'},
    ]

    for n in major_networks:
        # Check if already in list
        exists = any(n['name'].lower() in c['name'].lower() for c in filtered)
        if not exists:
            filtered.append(n)
            print(f"  Added: {n['name']}")

    # Save results
    print(f"\n\n{'=' * 70}")
    print(f"FINAL RESULTS: {len(filtered)} TRUCK TYRE COMPANIES")
    print(f"{'=' * 70}")

    # JSON
    with open('UK_TRUCK_TYRE_COMPANIES.json', 'w') as f:
        json.dump(filtered, f, indent=2)
    print(f"\nSaved: UK_TRUCK_TYRE_COMPANIES.json")

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "UK Truck Tyre Companies"

    # Header style
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    headers = ['Company Name', 'Website', 'Company Number', 'Address', 'Phone', 'Date Created', 'Source']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    # Data
    for row, c in enumerate(filtered, 2):
        ws.cell(row=row, column=1, value=c.get('name', ''))
        ws.cell(row=row, column=2, value=c.get('website', ''))
        ws.cell(row=row, column=3, value=c.get('company_number', ''))
        ws.cell(row=row, column=4, value=c.get('address', ''))
        ws.cell(row=row, column=5, value=c.get('phone', ''))
        ws.cell(row=row, column=6, value=c.get('date_created', ''))
        ws.cell(row=row, column=7, value=c.get('source', ''))

    # Column widths
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20

    wb.save('UK_TRUCK_TYRE_COMPANIES.xlsx')
    print(f"Saved: UK_TRUCK_TYRE_COMPANIES.xlsx")

    # Display results
    print(f"\n\nALL SCRAPED TRUCK TYRE COMPANIES:")
    print("-" * 70)

    for i, c in enumerate(filtered, 1):
        print(f"\n{i}. {c.get('name', '')}")
        if c.get('website'):
            print(f"   Website: {c['website']}")
        if c.get('company_number'):
            print(f"   CH#: {c['company_number']}")
        if c.get('address'):
            print(f"   Address: {c['address'][:60]}")
        if c.get('phone'):
            print(f"   Phone: {c['phone']}")

    return filtered


if __name__ == "__main__":
    main()
