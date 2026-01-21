"""
UK Truck Tyre Companies - DETAILED Scraper
Pulls comprehensive data from Companies House API including:
- Full company profile
- Officers/Directors
- Filing history
- Charges (mortgages/loans)
- Persons with significant control (PSC)
- Accounts data
"""

import requests
import json
import csv
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Configuration
API_KEY = '48d17266-ff2e-425f-9b20-7dcc9b25bb79'
BASE_URL = 'https://api.company-information.service.gov.uk'
DELAY = 0.6  # Rate limiting

def make_request(endpoint):
    """Make authenticated request to Companies House API"""
    url = f"{BASE_URL}{endpoint}"
    try:
        response = requests.get(url, auth=(API_KEY, ''), timeout=30)
        if response.status_code == 200:
            return response.json()
        elif response.status_code == 404:
            return None
        else:
            print(f"  API Error {response.status_code}: {endpoint}")
            return None
    except Exception as e:
        print(f"  Request error: {e}")
        return None

def get_company_profile(company_number):
    """Get full company profile"""
    return make_request(f"/company/{company_number}")

def get_officers(company_number):
    """Get company officers/directors"""
    return make_request(f"/company/{company_number}/officers")

def get_filing_history(company_number, items_per_page=10):
    """Get recent filing history"""
    return make_request(f"/company/{company_number}/filing-history?items_per_page={items_per_page}")

def get_charges(company_number):
    """Get company charges (mortgages/loans)"""
    return make_request(f"/company/{company_number}/charges")

def get_psc(company_number):
    """Get persons with significant control"""
    return make_request(f"/company/{company_number}/persons-with-significant-control")

def get_registered_office(company_number):
    """Get registered office address"""
    return make_request(f"/company/{company_number}/registered-office-address")

def search_companies(query, items_per_page=100):
    """Search for companies"""
    return make_request(f"/search/companies?q={query}&items_per_page={items_per_page}")

def format_address(address_dict):
    """Format address dictionary into string"""
    if not address_dict:
        return ""
    parts = []
    for key in ['premises', 'address_line_1', 'address_line_2', 'locality', 'region', 'postal_code', 'country']:
        if key in address_dict and address_dict[key]:
            parts.append(address_dict[key])
    return ', '.join(parts)

def format_date(date_str):
    """Format date string"""
    if not date_str:
        return ""
    try:
        if isinstance(date_str, dict):
            # Handle date objects like {"year": 2020, "month": 1, "day": 15}
            return f"{date_str.get('year', '')}-{date_str.get('month', ''):02d}-{date_str.get('day', ''):02d}"
        return date_str
    except:
        return str(date_str)

# Truck tyre related search terms
SEARCH_TERMS = [
    "truck tyre",
    "truck tire",
    "lorry tyre",
    "hgv tyre",
    "commercial tyre",
    "fleet tyre",
    "tyre wholesale",
    "tyre retread",
    "truck wheel",
    "commercial vehicle tyre"
]

# Filter terms - must have BOTH truck AND tyre related
TRUCK_TERMS = ['truck', 'lorry', 'hgv', 'commercial', 'fleet', 'trailer', 'artic', 'heavy goods']
TYRE_TERMS = ['tyre', 'tire', 'wheel', 'retread', 'remould']

# Exclusion terms
EXCLUDE_TERMS = ['agricultural', 'tractor', 'farm', 'earthmover', 'forklift', 'bicycle',
                 'motorcycle', 'motorbike', 'car tyre centre', 'domestic']

def is_truck_tyre_company(name, sic_codes=None):
    """Check if company is specifically a truck tyre company"""
    name_lower = name.lower()

    # Check exclusions first
    for term in EXCLUDE_TERMS:
        if term in name_lower:
            return False

    # Must have truck-related term
    has_truck = any(term in name_lower for term in TRUCK_TERMS)

    # Must have tyre-related term
    has_tyre = any(term in name_lower for term in TYRE_TERMS)

    # If has both, it's a match
    if has_truck and has_tyre:
        return True

    # Also accept if just "tyre" with wholesale/retread/commercial
    if has_tyre and any(term in name_lower for term in ['wholesale', 'retread', 'fleet', 'commercial']):
        return True

    return False

def get_detailed_company_data(company_number, company_name):
    """Get all available data for a company"""
    print(f"  Fetching detailed data for: {company_name}")

    data = {
        'company_number': company_number,
        'company_name': company_name,
        'profile': {},
        'officers': [],
        'filings': [],
        'charges': [],
        'psc': []
    }

    # Get company profile
    time.sleep(DELAY)
    profile = get_company_profile(company_number)
    if profile:
        data['profile'] = profile

    # Get officers
    time.sleep(DELAY)
    officers = get_officers(company_number)
    if officers and 'items' in officers:
        data['officers'] = officers['items']

    # Get filing history
    time.sleep(DELAY)
    filings = get_filing_history(company_number)
    if filings and 'items' in filings:
        data['filings'] = filings['items']

    # Get charges
    time.sleep(DELAY)
    charges = get_charges(company_number)
    if charges and 'items' in charges:
        data['charges'] = charges['items']

    # Get PSC
    time.sleep(DELAY)
    psc = get_psc(company_number)
    if psc and 'items' in psc:
        data['psc'] = psc['items']

    return data

def extract_flat_data(detailed_data):
    """Extract flat data structure for Excel"""
    profile = detailed_data.get('profile', {})
    officers = detailed_data.get('officers', [])
    filings = detailed_data.get('filings', [])
    charges = detailed_data.get('charges', [])
    psc = detailed_data.get('psc', [])

    # Get active directors
    active_directors = [o for o in officers if o.get('resigned_on') is None and o.get('officer_role') in ['director', 'corporate-director']]
    director_names = ', '.join([o.get('name', '') for o in active_directors[:5]])  # Top 5

    # Get latest filing
    latest_filing = filings[0] if filings else {}

    # Get accounts info
    accounts = profile.get('accounts', {})
    last_accounts = accounts.get('last_accounts', {})
    next_accounts = accounts.get('next_accounts', {})

    # Get confirmation statement
    confirmation = profile.get('confirmation_statement', {})

    # Count charges
    total_charges = len(charges)
    outstanding_charges = len([c for c in charges if c.get('status') == 'outstanding'])

    # Get PSC names
    psc_names = ', '.join([p.get('name', p.get('name_elements', {}).get('surname', '')) for p in psc[:3]])

    return {
        'company_number': detailed_data['company_number'],
        'company_name': profile.get('company_name', detailed_data['company_name']),
        'company_status': profile.get('company_status', ''),
        'company_type': profile.get('type', ''),
        'date_of_creation': profile.get('date_of_creation', ''),
        'registered_address': format_address(profile.get('registered_office_address', {})),
        'sic_codes': ', '.join(profile.get('sic_codes', [])),
        'sic_descriptions': get_sic_descriptions(profile.get('sic_codes', [])),

        # Accounts info
        'last_accounts_date': format_date(last_accounts.get('made_up_to', '')),
        'last_accounts_type': last_accounts.get('type', ''),
        'next_accounts_due': format_date(next_accounts.get('due_on', '')),
        'accounting_reference_date': f"{accounts.get('accounting_reference_date', {}).get('day', '')}/{accounts.get('accounting_reference_date', {}).get('month', '')}",

        # Confirmation statement
        'last_confirmation': format_date(confirmation.get('last_made_up_to', '')),
        'next_confirmation_due': format_date(confirmation.get('next_due', '')),

        # Officers
        'total_officers': len(officers),
        'active_directors': len(active_directors),
        'director_names': director_names,

        # Filings
        'total_filings': len(filings),
        'latest_filing_date': latest_filing.get('date', ''),
        'latest_filing_type': latest_filing.get('type', ''),
        'latest_filing_description': latest_filing.get('description', ''),

        # Charges
        'total_charges': total_charges,
        'outstanding_charges': outstanding_charges,

        # PSC
        'psc_count': len(psc),
        'psc_names': psc_names,

        # Additional
        'jurisdiction': profile.get('jurisdiction', ''),
        'has_insolvency_history': profile.get('has_insolvency_history', False),
        'has_charges': profile.get('has_charges', False),
        'can_file': profile.get('can_file', False),
    }

# SIC Code descriptions for common tyre industry codes
SIC_DESCRIPTIONS = {
    '22110': 'Manufacture of rubber tyres and tubes',
    '22190': 'Manufacture of other rubber products',
    '45200': 'Maintenance and repair of motor vehicles',
    '45310': 'Wholesale trade of motor vehicle parts',
    '45320': 'Retail trade of motor vehicle parts',
    '45400': 'Sale, maintenance and repair of motorcycles',
    '46690': 'Wholesale of other machinery and equipment',
    '46900': 'Non-specialised wholesale trade',
    '47300': 'Retail sale of automotive fuel',
    '49410': 'Freight transport by road',
    '77110': 'Renting and leasing of cars and light motor vehicles',
}

def get_sic_descriptions(sic_codes):
    """Get descriptions for SIC codes"""
    if not sic_codes:
        return ""
    descriptions = []
    for code in sic_codes:
        if code in SIC_DESCRIPTIONS:
            descriptions.append(SIC_DESCRIPTIONS[code])
    return '; '.join(descriptions)

def main():
    print("=" * 60)
    print("UK TRUCK TYRE COMPANIES - DETAILED SCRAPER")
    print("Fetching comprehensive data from Companies House API")
    print("=" * 60)

    all_companies = {}  # Use dict to dedupe by company number

    # Search for companies
    print("\n[1/2] Searching for truck tyre companies...")
    for term in SEARCH_TERMS:
        print(f"  Searching: '{term}'")
        time.sleep(DELAY)

        results = search_companies(term.replace(' ', '+'))
        if results and 'items' in results:
            for company in results['items']:
                name = company.get('title', '')
                number = company.get('company_number', '')
                status = company.get('company_status', '')

                # Only active companies
                if status != 'active':
                    continue

                # Check if truck tyre related
                if is_truck_tyre_company(name):
                    if number not in all_companies:
                        all_companies[number] = {
                            'name': name,
                            'number': number,
                            'address': format_address(company.get('address', {})),
                            'date_created': company.get('date_of_creation', ''),
                            'status': status
                        }
                        print(f"    Found: {name}")

    print(f"\n  Total unique companies found: {len(all_companies)}")

    # Get detailed data for each company
    print("\n[2/2] Fetching detailed data for each company...")
    detailed_companies = []

    for i, (company_number, company_info) in enumerate(all_companies.items(), 1):
        print(f"\n  [{i}/{len(all_companies)}] {company_info['name']}")

        detailed = get_detailed_company_data(company_number, company_info['name'])
        flat_data = extract_flat_data(detailed)
        detailed_companies.append(flat_data)

        # Store raw data too
        company_info['detailed'] = detailed

    # Create Excel workbook
    print("\n[3/3] Creating Excel report...")
    create_excel_report(detailed_companies)

    # Save JSON with all raw data
    with open('UK_TRUCK_TYRE_DETAILED_DATA.json', 'w') as f:
        json.dump(list(all_companies.values()), f, indent=2, default=str)

    print(f"\nDone! Files created:")
    print(f"  - UK_TRUCK_TYRE_DETAILED_REPORT.xlsx")
    print(f"  - UK_TRUCK_TYRE_DETAILED_DATA.json")
    print(f"\nTotal companies with detailed data: {len(detailed_companies)}")

def create_excel_report(companies):
    """Create comprehensive Excel report"""
    wb = Workbook()

    # Style definitions
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ===== SHEET 1: Company Overview =====
    ws1 = wb.active
    ws1.title = "Company Overview"

    headers1 = [
        "Company Number", "Company Name", "Status", "Type", "Date Created",
        "Registered Address", "SIC Codes", "SIC Descriptions", "Jurisdiction"
    ]

    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for row, company in enumerate(companies, 2):
        ws1.cell(row=row, column=1, value=company['company_number']).border = thin_border
        ws1.cell(row=row, column=2, value=company['company_name']).border = thin_border
        ws1.cell(row=row, column=3, value=company['company_status']).border = thin_border
        ws1.cell(row=row, column=4, value=company['company_type']).border = thin_border
        ws1.cell(row=row, column=5, value=company['date_of_creation']).border = thin_border
        ws1.cell(row=row, column=6, value=company['registered_address']).border = thin_border
        ws1.cell(row=row, column=7, value=company['sic_codes']).border = thin_border
        ws1.cell(row=row, column=8, value=company['sic_descriptions']).border = thin_border
        ws1.cell(row=row, column=9, value=company['jurisdiction']).border = thin_border

    # Set column widths
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 40
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 20
    ws1.column_dimensions['E'].width = 15
    ws1.column_dimensions['F'].width = 50
    ws1.column_dimensions['G'].width = 15
    ws1.column_dimensions['H'].width = 40
    ws1.column_dimensions['I'].width = 15

    # ===== SHEET 2: Financial Data =====
    ws2 = wb.create_sheet("Financial & Accounts")

    headers2 = [
        "Company Number", "Company Name", "Last Accounts Date", "Accounts Type",
        "Next Accounts Due", "Accounting Ref Date", "Total Charges", "Outstanding Charges",
        "Has Insolvency History", "Has Charges"
    ]

    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        cell.font = header_font
        cell.border = thin_border

    for row, company in enumerate(companies, 2):
        ws2.cell(row=row, column=1, value=company['company_number']).border = thin_border
        ws2.cell(row=row, column=2, value=company['company_name']).border = thin_border
        ws2.cell(row=row, column=3, value=company['last_accounts_date']).border = thin_border
        ws2.cell(row=row, column=4, value=company['last_accounts_type']).border = thin_border
        ws2.cell(row=row, column=5, value=company['next_accounts_due']).border = thin_border
        ws2.cell(row=row, column=6, value=company['accounting_reference_date']).border = thin_border
        ws2.cell(row=row, column=7, value=company['total_charges']).border = thin_border
        ws2.cell(row=row, column=8, value=company['outstanding_charges']).border = thin_border
        ws2.cell(row=row, column=9, value=str(company['has_insolvency_history'])).border = thin_border
        ws2.cell(row=row, column=10, value=str(company['has_charges'])).border = thin_border

    for i, width in enumerate([15, 40, 18, 15, 18, 18, 15, 18, 20, 15], 1):
        ws2.column_dimensions[chr(64+i)].width = width

    # ===== SHEET 3: Officers & Directors =====
    ws3 = wb.create_sheet("Officers & Directors")

    headers3 = [
        "Company Number", "Company Name", "Total Officers", "Active Directors",
        "Director Names", "PSC Count", "PSC Names"
    ]

    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
        cell.font = header_font
        cell.border = thin_border

    for row, company in enumerate(companies, 2):
        ws3.cell(row=row, column=1, value=company['company_number']).border = thin_border
        ws3.cell(row=row, column=2, value=company['company_name']).border = thin_border
        ws3.cell(row=row, column=3, value=company['total_officers']).border = thin_border
        ws3.cell(row=row, column=4, value=company['active_directors']).border = thin_border
        ws3.cell(row=row, column=5, value=company['director_names']).border = thin_border
        ws3.cell(row=row, column=6, value=company['psc_count']).border = thin_border
        ws3.cell(row=row, column=7, value=company['psc_names']).border = thin_border

    ws3.column_dimensions['A'].width = 15
    ws3.column_dimensions['B'].width = 40
    ws3.column_dimensions['C'].width = 15
    ws3.column_dimensions['D'].width = 18
    ws3.column_dimensions['E'].width = 60
    ws3.column_dimensions['F'].width = 12
    ws3.column_dimensions['G'].width = 50

    # ===== SHEET 4: Filing History =====
    ws4 = wb.create_sheet("Filing History")

    headers4 = [
        "Company Number", "Company Name", "Total Filings", "Latest Filing Date",
        "Latest Filing Type", "Latest Filing Description", "Last Confirmation", "Next Confirmation Due"
    ]

    for col, header in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color="F57C00", end_color="F57C00", fill_type="solid")
        cell.font = header_font
        cell.border = thin_border

    for row, company in enumerate(companies, 2):
        ws4.cell(row=row, column=1, value=company['company_number']).border = thin_border
        ws4.cell(row=row, column=2, value=company['company_name']).border = thin_border
        ws4.cell(row=row, column=3, value=company['total_filings']).border = thin_border
        ws4.cell(row=row, column=4, value=company['latest_filing_date']).border = thin_border
        ws4.cell(row=row, column=5, value=company['latest_filing_type']).border = thin_border
        ws4.cell(row=row, column=6, value=company['latest_filing_description']).border = thin_border
        ws4.cell(row=row, column=7, value=company['last_confirmation']).border = thin_border
        ws4.cell(row=row, column=8, value=company['next_confirmation_due']).border = thin_border

    for i, width in enumerate([15, 40, 15, 18, 20, 50, 18, 20], 1):
        ws4.column_dimensions[chr(64+i)].width = width

    # ===== SHEET 5: Summary =====
    ws5 = wb.create_sheet("Summary")

    ws5.cell(row=1, column=1, value="UK TRUCK TYRE COMPANIES - DETAILED ANALYSIS")
    ws5.cell(row=1, column=1).font = Font(bold=True, size=16)
    ws5.merge_cells('A1:D1')

    ws5.cell(row=3, column=1, value=f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws5.cell(row=4, column=1, value=f"Total Companies Analyzed: {len(companies)}")

    # Statistics
    active_count = len([c for c in companies if c['company_status'] == 'active'])
    with_charges = len([c for c in companies if c['has_charges']])
    with_insolvency = len([c for c in companies if c['has_insolvency_history']])

    ws5.cell(row=6, column=1, value="STATISTICS")
    ws5.cell(row=6, column=1).font = Font(bold=True)

    stats = [
        ("Active Companies", active_count),
        ("Companies with Charges/Loans", with_charges),
        ("Companies with Insolvency History", with_insolvency),
        ("Total Directors Found", sum(c['active_directors'] for c in companies)),
        ("Total Filings Recorded", sum(c['total_filings'] for c in companies)),
    ]

    for i, (label, value) in enumerate(stats, 7):
        ws5.cell(row=i, column=1, value=label)
        ws5.cell(row=i, column=2, value=value)

    ws5.column_dimensions['A'].width = 35
    ws5.column_dimensions['B'].width = 15

    # Save
    wb.save('UK_TRUCK_TYRE_DETAILED_REPORT.xlsx')

if __name__ == "__main__":
    main()
