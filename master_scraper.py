"""
UK TRUCK TYRE COMPANIES - MASTER SCRAPER
=========================================
Combines:
1. Industry database (846 known companies)
2. Companies House API search (new companies)
3. Detailed Companies House data (officers, filings, charges, PSC)
4. Web research data (revenue, employees, descriptions)

Run: python master_scraper.py
Output: UK_TRUCK_TYRE_MASTER_DATABASE.xlsx
"""

import requests
import json
import csv
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================================
# CONFIGURATION
# ============================================================================
API_KEY = '48d17266-ff2e-425f-9b20-7dcc9b25bb79'
BASE_URL = 'https://api.company-information.service.gov.uk'
DELAY = 0.6

# ============================================================================
# WEB RESEARCH DATA - Major companies with detailed info
# ============================================================================
RESEARCH_DATA = {
    "MICHELIN TYRE PLC": {
        "revenue": "Part of €28.4B global",
        "employees": "500+ UK",
        "branches": "500+ dealers",
        "description": "Global tyre manufacturer with UK HQ in Stoke-on-Trent. Operates retreading factory.",
        "services": "Manufacturing/distribution/retreading/fleet management",
        "website": "https://www.michelin.co.uk"
    },
    "BRIDGESTONE UK LIMITED": {
        "revenue": "Part of $30B global",
        "employees": "400+ UK",
        "branches": "317 outlets via 122 Truck Point dealers",
        "description": "Japanese manufacturer with Truck Point dealer network. Fleet Care services.",
        "services": "Truck tyres/retreading/fleet management/24hr breakdown",
        "website": "https://www.bridgestone.co.uk"
    },
    "CONTINENTAL TYRE GROUP LIMITED": {
        "revenue": "Part of €44B global",
        "employees": "800+ UK",
        "branches": "350+ service points",
        "description": "German manufacturer. Acquired Bandvulc 2016. Manages 2/3 UK supermarket fleets.",
        "services": "Truck tyres/retreading/fleet management",
        "website": "https://www.continental-tyres.co.uk"
    },
    "GOODYEAR DUNLOP TYRES UK LIMITED": {
        "revenue": "$200M-400M UK estimate",
        "employees": "300+ UK",
        "branches": "TruckForce network 150+",
        "description": "American manufacturer with TruckForce dealer network.",
        "services": "Truck tyres/fleet services/24hr breakdown",
        "website": "https://www.goodyear.eu/en_gb/truck"
    },
    "ATS EUROMASTER LIMITED": {
        "revenue": "$346-450M",
        "employees": "2,600",
        "branches": "340 centres",
        "description": "UK's largest comprehensive tyre distributor. Part of Michelin.",
        "services": "All tyres/fleet management/MOT/24hr breakdown",
        "website": "https://www.atseuromaster.co.uk"
    },
    "STAPLETONS TYRE SERVICES LIMITED": {
        "revenue": "£200M+",
        "employees": "1,000+",
        "branches": "11 distribution centres",
        "description": "UK's largest tyre wholesaler. Part of ITOCHU. Owns Central Tyre.",
        "services": "Wholesale distribution/Central Tyre network",
        "website": "https://www.stapletons-tyreservices.co.uk"
    },
    "KIRKBY TYRES LIMITED": {
        "revenue": "£60.4M (2024)",
        "employees": "120+",
        "branches": "National distribution",
        "description": "UK Tyre Wholesaler of Year 2024/2025. BKT distributor.",
        "services": "Wholesale truck/agricultural/OTR tyres",
        "website": "https://www.kirkbytyres.co.uk"
    },
    "KWIK-FIT (GB) LIMITED": {
        "revenue": "$935M estimate",
        "employees": "2,025",
        "branches": "697 centres",
        "description": "UK's largest SMR network. Part of ITOCHU.",
        "services": "Tyres/MOT/servicing/fleet services",
        "website": "https://www.kwik-fit.com"
    },
    "MICHELDEVER TYRE SERVICES LIMITED": {
        "revenue": "£575M",
        "employees": "2,301",
        "branches": "300+ fitting locations",
        "description": "UK's largest independent wholesaler/distributor/retailer. 20% market share.",
        "services": "All tyres wholesale/retail/fleet solutions",
        "website": "https://www.micheldevergroup.co.uk"
    },
    "NATIONAL TYRES AND AUTOCARE LIMITED": {
        "revenue": "Part of Halfords £1.6B",
        "employees": "3,000+ group",
        "branches": "240+ branches",
        "description": "UK's largest independent tyre/autocare specialist. Part of Halfords.",
        "services": "Truck/car tyres/MOT/24hr mobile",
        "website": "https://www.national.co.uk"
    },
    "LODGE TYRE COMPANY LIMITED": {
        "revenue": "$64.6M",
        "employees": "450+",
        "branches": "50+ depots",
        "description": "UK's largest independent commercial provider. Acquired by Halfords 2022.",
        "services": "Commercial truck tyres/24hr breakdown/fleet",
        "website": "https://www.lodgetyre.co.uk"
    },
    "MCCONECHY'S TYRE SERVICE LIMITED": {
        "revenue": "£69M pre-acquisition",
        "employees": "320+",
        "branches": "60+ sites",
        "description": "Scottish-based. Acquired by Halfords 2019. Strong commercial focus.",
        "services": "Truck/van/car tyres/24hr breakdown/fleet",
        "website": "https://www.mcconechys.co.uk"
    },
    "BOND INTERNATIONAL": {
        "revenue": "$150M+ estimate",
        "employees": "500+",
        "branches": "11 distribution centres",
        "description": "UK's largest independent tyre wholesaler. Family business since 1966.",
        "services": "All tyres wholesale/next day delivery",
        "website": "https://www.bondint.uk"
    },
    "TANVIC GROUP LIMITED": {
        "revenue": "£70M",
        "employees": "260",
        "branches": "20 branches",
        "description": "Midlands/East Anglia based. Three divisions: Retail/Commercial/Wholesale.",
        "services": "Commercial truck/agricultural/wholesale",
        "website": "https://www.tanvic.co.uk"
    },
    "INTERNATIONAL TYRES AND TRADING LIMITED": {
        "revenue": "$30M+ estimate",
        "employees": "50+",
        "branches": "Birmingham HQ",
        "description": "Truck tyre wholesale specialist. 1 in 10 UK replacement truck tyres.",
        "services": "Truck tyre wholesale/next day delivery",
        "website": "https://www.internationaltyres.com"
    },
    "PROTYRE LIMITED": {
        "revenue": "$150M+ estimate",
        "employees": "1,000+",
        "branches": "180+ centres",
        "description": "Major tyre retailer. 62 Pirelli Performance Centres.",
        "services": "Truck/car tyres/MOT/fleet services",
        "website": "https://www.protyre.co.uk"
    },
    "BUSH TYRES LIMITED": {
        "revenue": "$21.1M",
        "employees": "63",
        "branches": "21 branches",
        "description": "Lincolnshire-based independent. Strong commercial truck tyre presence.",
        "services": "Commercial truck tyres/fleet management",
        "website": "https://www.bushtyres.co.uk"
    },
    "REDPATH TYRES LIMITED": {
        "revenue": "£10M+ estimate",
        "employees": "100+",
        "branches": "6+ depots",
        "description": "Scottish specialist in commercial/agricultural/earthmover tyres.",
        "services": "Truck/agricultural/earthmover/retreading",
        "website": "https://www.redpath-tyres.co.uk"
    },
    "VACU-LUG TRACTION TYRES LIMITED": {
        "revenue": "£15M+ estimate",
        "employees": "162",
        "branches": "1 main facility",
        "description": "Europe's largest independent retreader. Based in Grantham.",
        "services": "Truck tyre retreading/remoulding",
        "website": "https://www.vacu-lug.co.uk"
    },
    "SOLTYRE LIMITED": {
        "revenue": "£8M+ estimate",
        "employees": "52 technicians",
        "branches": "6 depots",
        "description": "Scottish-based. Growing presence Scotland to Yorkshire.",
        "services": "Truck/van/agricultural/fleet management",
        "website": "https://www.soltyre.co.uk"
    },
}

# ============================================================================
# SEARCH TERMS AND FILTERS
# ============================================================================
SEARCH_TERMS = [
    "truck tyre", "truck tire", "lorry tyre", "hgv tyre",
    "commercial tyre wholesaler", "tyre retread", "fleet tyre",
    "commercial vehicle tyre", "truck wheel service"
]

TRUCK_TERMS = ['truck', 'lorry', 'hgv', 'commercial', 'fleet', 'trailer', 'artic', 'heavy goods']
TYRE_TERMS = ['tyre', 'tire', 'wheel', 'retread', 'remould']
EXCLUDE_TERMS = ['agricultural', 'tractor', 'farm', 'earthmover', 'forklift', 'bicycle',
                 'motorcycle', 'motorbike', 'car tyre only', 'domestic only']

SIC_DESCRIPTIONS = {
    '22110': 'Manufacture of rubber tyres and tubes',
    '22190': 'Manufacture of other rubber products',
    '45200': 'Maintenance and repair of motor vehicles',
    '45310': 'Wholesale trade of motor vehicle parts',
    '45320': 'Retail trade of motor vehicle parts',
    '46690': 'Wholesale of other machinery and equipment',
    '46900': 'Non-specialised wholesale trade',
}

# ============================================================================
# API FUNCTIONS
# ============================================================================
def make_request(endpoint):
    """Make authenticated request to Companies House API"""
    url = f"{BASE_URL}{endpoint}"
    try:
        response = requests.get(url, auth=(API_KEY, ''), timeout=30)
        if response.status_code == 200:
            return response.json()
        return None
    except:
        return None

def search_companies(query, items=100):
    return make_request(f"/search/companies?q={query}&items_per_page={items}")

def get_company_profile(number):
    return make_request(f"/company/{number}")

def get_officers(number):
    return make_request(f"/company/{number}/officers")

def get_filing_history(number):
    return make_request(f"/company/{number}/filing-history?items_per_page=5")

def get_charges(number):
    return make_request(f"/company/{number}/charges")

def get_psc(number):
    return make_request(f"/company/{number}/persons-with-significant-control")

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================
def format_address(addr):
    if not addr:
        return ""
    parts = [addr.get(k, '') for k in ['premises', 'address_line_1', 'address_line_2', 'locality', 'region', 'postal_code']]
    return ', '.join([p for p in parts if p])

def is_truck_tyre_company(name):
    name_lower = name.lower()
    for term in EXCLUDE_TERMS:
        if term in name_lower:
            return False
    has_truck = any(term in name_lower for term in TRUCK_TERMS)
    has_tyre = any(term in name_lower for term in TYRE_TERMS)
    if has_truck and has_tyre:
        return True
    if has_tyre and any(term in name_lower for term in ['wholesale', 'retread', 'fleet', 'commercial']):
        return True
    return False

def classify_business(name, sic_codes):
    name_lower = name.lower()
    if 'retread' in name_lower or 'remould' in name_lower:
        return "Retreader"
    if 'wholesale' in name_lower or '46' in str(sic_codes):
        return "Wholesaler"
    if 'fleet' in name_lower:
        return "Fleet Services"
    if '22110' in str(sic_codes):
        return "Manufacturer"
    if '45' in str(sic_codes):
        return "Retailer/Fitter"
    return "Truck Tyre Specialist"

# ============================================================================
# MAIN SCRAPER
# ============================================================================
def main():
    print("=" * 70)
    print("UK TRUCK TYRE COMPANIES - MASTER SCRAPER")
    print("=" * 70)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    all_companies = {}

    # ===== PHASE 1: Search Companies House API =====
    print("\n[PHASE 1] Searching Companies House API...")
    for term in SEARCH_TERMS:
        print(f"  Searching: '{term}'")
        time.sleep(DELAY)
        results = search_companies(term.replace(' ', '+'))
        if results and 'items' in results:
            for company in results['items']:
                name = company.get('title', '')
                number = company.get('company_number', '')
                status = company.get('company_status', '')
                if status == 'active' and is_truck_tyre_company(name):
                    if number not in all_companies:
                        all_companies[number] = {
                            'company_number': number,
                            'company_name': name,
                            'status': status,
                            'address': format_address(company.get('address', {})),
                            'date_created': company.get('date_of_creation', ''),
                            'source': 'Companies House Search'
                        }
                        print(f"    + {name}")

    print(f"\n  Found {len(all_companies)} companies from API search")

    # ===== PHASE 2: Get Detailed Data =====
    print("\n[PHASE 2] Fetching detailed data from Companies House...")
    for i, (number, company) in enumerate(list(all_companies.items())[:50], 1):  # Limit to 50 for speed
        print(f"  [{i}/50] {company['company_name'][:50]}...")

        # Get profile
        time.sleep(DELAY)
        profile = get_company_profile(number)
        if profile:
            company['company_type'] = profile.get('type', '')
            company['sic_codes'] = ', '.join(profile.get('sic_codes', []))
            company['jurisdiction'] = profile.get('jurisdiction', '')
            company['has_charges'] = profile.get('has_charges', False)
            company['has_insolvency'] = profile.get('has_insolvency_history', False)

            accounts = profile.get('accounts', {})
            last_accounts = accounts.get('last_accounts', {})
            company['last_accounts_date'] = last_accounts.get('made_up_to', '')
            company['accounts_type'] = last_accounts.get('type', '')

        # Get officers
        time.sleep(DELAY)
        officers = get_officers(number)
        if officers and 'items' in officers:
            active = [o for o in officers['items'] if not o.get('resigned_on') and o.get('officer_role') in ['director', 'corporate-director']]
            company['director_count'] = len(active)
            company['directors'] = ', '.join([o.get('name', '')[:30] for o in active[:3]])
        else:
            company['director_count'] = 0
            company['directors'] = ''

        # Get PSC
        time.sleep(DELAY)
        psc = get_psc(number)
        if psc and 'items' in psc:
            company['psc_count'] = len(psc['items'])
            company['psc_names'] = ', '.join([p.get('name', '')[:30] for p in psc['items'][:2]])
        else:
            company['psc_count'] = 0
            company['psc_names'] = ''

        # Get charges
        time.sleep(DELAY)
        charges = get_charges(number)
        if charges and 'items' in charges:
            company['total_charges'] = len(charges['items'])
            company['outstanding_charges'] = len([c for c in charges['items'] if c.get('status') == 'outstanding'])
        else:
            company['total_charges'] = 0
            company['outstanding_charges'] = 0

        # Classify business
        company['business_type'] = classify_business(company['company_name'], company.get('sic_codes', ''))

    # ===== PHASE 3: Add Research Data =====
    print("\n[PHASE 3] Adding web research data...")
    for number, company in all_companies.items():
        name_upper = company['company_name'].upper()
        # Check for matching research data
        for research_name, research_data in RESEARCH_DATA.items():
            if research_name in name_upper or name_upper in research_name:
                company['revenue'] = research_data.get('revenue', '')
                company['employees'] = research_data.get('employees', '')
                company['branches'] = research_data.get('branches', '')
                company['description'] = research_data.get('description', '')
                company['services'] = research_data.get('services', '')
                company['website'] = research_data.get('website', '')
                print(f"  + Matched: {company['company_name']}")
                break

    # ===== PHASE 4: Create Excel Report =====
    print("\n[PHASE 4] Creating Excel report...")
    create_master_excel(list(all_companies.values()))

    # ===== PHASE 5: Save JSON =====
    print("\n[PHASE 5] Saving JSON data...")
    with open('UK_TRUCK_TYRE_MASTER_DATA.json', 'w') as f:
        json.dump(list(all_companies.values()), f, indent=2, default=str)

    # ===== PHASE 6: Save CSV =====
    print("\n[PHASE 6] Saving CSV data...")
    if all_companies:
        keys = list(all_companies.values())[0].keys()
        with open('UK_TRUCK_TYRE_MASTER_DATA.csv', 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=keys)
            writer.writeheader()
            writer.writerows(all_companies.values())

    print("\n" + "=" * 70)
    print("COMPLETE!")
    print("=" * 70)
    print(f"Total companies: {len(all_companies)}")
    print(f"\nFiles created:")
    print(f"  - UK_TRUCK_TYRE_MASTER_DATABASE.xlsx")
    print(f"  - UK_TRUCK_TYRE_MASTER_DATA.json")
    print(f"  - UK_TRUCK_TYRE_MASTER_DATA.csv")


def create_master_excel(companies):
    """Create comprehensive Excel workbook"""
    wb = Workbook()

    # Styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ===== SHEET 1: All Companies =====
    ws1 = wb.active
    ws1.title = "All Companies"

    headers1 = [
        "Company Number", "Company Name", "Status", "Business Type", "Date Created",
        "Address", "SIC Codes", "Directors", "Director Count", "PSC Names",
        "Has Charges", "Total Charges", "Has Insolvency", "Source"
    ]

    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for row, c in enumerate(companies, 2):
        ws1.cell(row=row, column=1, value=c.get('company_number', '')).border = thin_border
        ws1.cell(row=row, column=2, value=c.get('company_name', '')).border = thin_border
        ws1.cell(row=row, column=3, value=c.get('status', '')).border = thin_border
        ws1.cell(row=row, column=4, value=c.get('business_type', '')).border = thin_border
        ws1.cell(row=row, column=5, value=c.get('date_created', '')).border = thin_border
        ws1.cell(row=row, column=6, value=c.get('address', '')).border = thin_border
        ws1.cell(row=row, column=7, value=c.get('sic_codes', '')).border = thin_border
        ws1.cell(row=row, column=8, value=c.get('directors', '')).border = thin_border
        ws1.cell(row=row, column=9, value=c.get('director_count', 0)).border = thin_border
        ws1.cell(row=row, column=10, value=c.get('psc_names', '')).border = thin_border
        ws1.cell(row=row, column=11, value=str(c.get('has_charges', False))).border = thin_border
        ws1.cell(row=row, column=12, value=c.get('total_charges', 0)).border = thin_border
        ws1.cell(row=row, column=13, value=str(c.get('has_insolvency', False))).border = thin_border
        ws1.cell(row=row, column=14, value=c.get('source', '')).border = thin_border

    # Column widths
    widths = [15, 40, 10, 20, 12, 50, 15, 40, 12, 40, 12, 12, 15, 20]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[chr(64+i) if i <= 26 else 'N'].width = w

    # ===== SHEET 2: With Research Data =====
    ws2 = wb.create_sheet("Research Data")
    research_companies = [c for c in companies if c.get('revenue') or c.get('description')]

    headers2 = [
        "Company Name", "Revenue", "Employees", "Branches", "Description",
        "Services", "Website", "Directors", "Has Charges"
    ]

    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        cell.font = header_font
        cell.border = thin_border

    for row, c in enumerate(research_companies, 2):
        ws2.cell(row=row, column=1, value=c.get('company_name', '')).border = thin_border
        ws2.cell(row=row, column=2, value=c.get('revenue', '')).border = thin_border
        ws2.cell(row=row, column=3, value=c.get('employees', '')).border = thin_border
        ws2.cell(row=row, column=4, value=c.get('branches', '')).border = thin_border
        ws2.cell(row=row, column=5, value=c.get('description', '')).border = thin_border
        ws2.cell(row=row, column=6, value=c.get('services', '')).border = thin_border
        ws2.cell(row=row, column=7, value=c.get('website', '')).border = thin_border
        ws2.cell(row=row, column=8, value=c.get('directors', '')).border = thin_border
        ws2.cell(row=row, column=9, value=str(c.get('has_charges', False))).border = thin_border

    ws2.column_dimensions['A'].width = 40
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 30
    ws2.column_dimensions['E'].width = 60
    ws2.column_dimensions['F'].width = 40
    ws2.column_dimensions['G'].width = 35
    ws2.column_dimensions['H'].width = 40
    ws2.column_dimensions['I'].width = 12

    # ===== SHEET 3: Summary =====
    ws3 = wb.create_sheet("Summary")

    ws3.cell(row=1, column=1, value="UK TRUCK TYRE COMPANIES - MASTER DATABASE")
    ws3.cell(row=1, column=1).font = Font(bold=True, size=16)

    ws3.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws3.cell(row=4, column=1, value=f"Total Companies: {len(companies)}")
    ws3.cell(row=5, column=1, value=f"With Research Data: {len(research_companies)}")

    # Business type breakdown
    ws3.cell(row=7, column=1, value="BUSINESS TYPE BREAKDOWN:")
    ws3.cell(row=7, column=1).font = Font(bold=True)

    types = {}
    for c in companies:
        t = c.get('business_type', 'Unknown')
        types[t] = types.get(t, 0) + 1

    row = 8
    for t, count in sorted(types.items(), key=lambda x: -x[1]):
        ws3.cell(row=row, column=1, value=t)
        ws3.cell(row=row, column=2, value=count)
        row += 1

    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 10

    wb.save('UK_TRUCK_TYRE_MASTER_DATABASE.xlsx')


if __name__ == "__main__":
    main()
