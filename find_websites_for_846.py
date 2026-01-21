#!/usr/bin/env python3
"""
=============================================================================
FIND REAL WEBSITES FOR THE 846 COMPANIES
=============================================================================
Takes each company name from the existing INDUSTRY_DATABASE (846 companies)
and searches the web to find their REAL website - not invented URLs.

Uses Google Custom Search API to find actual websites.

Date: January 2026
=============================================================================
"""

import requests
from bs4 import BeautifulSoup
import json
import time
import re
import csv
from datetime import datetime
from urllib.parse import urlparse, quote_plus
from scraper import INDUSTRY_DATABASE

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-GB,en;q=0.9',
}

OUTPUT_FILE = 'uk_truck_tyres_846_WITH_REAL_WEBSITES'


def search_website_via_bing(company_name):
    """Search Bing for the company's website"""

    # Clean company name
    search_name = company_name.replace(' LIMITED', '').replace(' LTD', '').replace('(', '').replace(')', '').strip()

    try:
        search_url = f'https://www.bing.com/search?q={quote_plus(search_name + " UK tyres official website")}'

        response = requests.get(search_url, headers=HEADERS, timeout=12)

        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')

            # Find search results
            results = soup.find_all('li', class_='b_algo')

            skip_domains = [
                'facebook.com', 'linkedin.com', 'twitter.com', 'instagram.com',
                'youtube.com', 'pinterest.com', 'tiktok.com',
                'yell.com', 'yelp.com', 'checkatrade.com', 'trustatrader.com',
                'companieshouse.gov.uk', 'find-and-update.company-information',
                'endole.co.uk', 'duedil.com', 'opencorporates.com',
                'amazon.co.uk', 'ebay.co.uk', 'gumtree.com',
                'wikipedia.org', 'bbc.co.uk', 'gov.uk',
                'companieslist.co.uk', 'bizdb.co.uk'
            ]

            for result in results[:5]:
                link = result.find('a')
                if link:
                    href = link.get('href', '')

                    if href.startswith('http'):
                        domain = urlparse(href).netloc.lower()

                        # Skip directories
                        if any(skip in domain for skip in skip_domains):
                            continue

                        return href

    except Exception as e:
        pass

    return None


def search_website_via_duckduckgo(company_name):
    """Search DuckDuckGo for the company's website"""

    search_name = company_name.replace(' LIMITED', '').replace(' LTD', '').strip()

    try:
        search_url = f'https://html.duckduckgo.com/html/?q={quote_plus(search_name + " UK tyres")}'

        response = requests.get(search_url, headers=HEADERS, timeout=12)

        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')

            results = soup.find_all('a', class_='result__a')

            skip_domains = [
                'facebook.com', 'linkedin.com', 'twitter.com',
                'yell.com', 'yelp.com', 'checkatrade.com',
                'companieshouse', 'endole', 'duedil', 'youtube.com'
            ]

            for result in results[:5]:
                href = result.get('href', '')

                if href.startswith('http'):
                    if not any(skip in href.lower() for skip in skip_domains):
                        return href

    except:
        pass

    return None


def verify_website(url):
    """Verify website works and extract contact info"""
    if not url or not url.startswith('http'):
        return None

    try:
        response = requests.get(url, headers=HEADERS, timeout=10, allow_redirects=True)

        if response.status_code == 200:
            text = response.text
            text_lower = text.lower()

            result = {
                'works': True,
                'final_url': response.url,
                'phone': None,
                'email': None,
                'services': [],
                'is_tyre_related': False
            }

            # Check if tyre related
            tyre_words = ['tyre', 'tire', 'wheel', 'fitting']
            result['is_tyre_related'] = any(w in text_lower for w in tyre_words)

            # Extract phone
            phone_patterns = [
                r'0800[\s\-]?\d{3}[\s\-]?\d{3,4}',
                r'0808[\s\-]?\d{3}[\s\-]?\d{3,4}',
                r'01\d{2,3}[\s\-]?\d{3}[\s\-]?\d{3,4}',
                r'02\d[\s\-]?\d{4}[\s\-]?\d{4}',
                r'03\d{2}[\s\-]?\d{3}[\s\-]?\d{4}',
                r'07\d{3}[\s\-]?\d{3}[\s\-]?\d{3}',
            ]
            for pattern in phone_patterns:
                matches = re.findall(pattern, text)
                for m in matches:
                    cleaned = re.sub(r'[\s\-]', '', m)
                    if 10 <= len(cleaned) <= 12 and not cleaned.startswith('00000'):
                        result['phone'] = m.strip()
                        break
                if result['phone']:
                    break

            # Extract email
            emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', text)
            for email in emails:
                if not any(x in email.lower() for x in ['.png', '.jpg', '.css', '.js', 'example.', 'wix.', 'sentry']):
                    result['email'] = email
                    break

            # Services
            if '24' in text_lower and ('hour' in text_lower or '/7' in text_lower):
                result['services'].append('24hr')
            if 'mobile' in text_lower and 'fitting' in text_lower:
                result['services'].append('Mobile')
            if 'fleet' in text_lower:
                result['services'].append('Fleet')
            if 'commercial' in text_lower:
                result['services'].append('Commercial')

            return result

        return None
    except:
        return None


def estimate_revenue(company):
    """Estimate revenue based on characteristics"""
    name = company.get('name', '').lower()
    business_type = company.get('businessType', '').lower()
    services = company.get('services', '')
    verified = company.get('website_verified', False)

    score = 0
    indicators = []

    # Business type
    if 'manufacturer' in business_type:
        score += 50
        indicators.append('Manufacturer')
    elif 'wholesaler' in business_type:
        score += 40
        indicators.append('Wholesaler')
    elif 'national' in business_type:
        score += 35
        indicators.append('National')
    elif 'regional' in business_type:
        score += 25
        indicators.append('Regional')

    # Name indicators
    if any(x in name for x in ['national', 'uk', 'british', 'group']):
        score += 20
        indicators.append('National name')
    if 'network' in name:
        score += 15

    # Services
    if '24hr' in str(services):
        score += 10
        indicators.append('24hr')
    if 'Fleet' in str(services):
        score += 15
        indicators.append('Fleet')
    if 'Commercial' in str(services):
        score += 10

    if verified:
        score += 10

    if score >= 60:
        return {'size': 'Large', 'employees': '50-500+', 'revenue': '£10M - £100M+',
                'rev_low': 10000000, 'rev_high': 100000000, 'indicators': ', '.join(indicators) or 'Major'}
    elif score >= 40:
        return {'size': 'Medium-Large', 'employees': '20-100', 'revenue': '£2M - £15M',
                'rev_low': 2000000, 'rev_high': 15000000, 'indicators': ', '.join(indicators) or 'Regional'}
    elif score >= 25:
        return {'size': 'Medium', 'employees': '10-30', 'revenue': '£500K - £3M',
                'rev_low': 500000, 'rev_high': 3000000, 'indicators': ', '.join(indicators) or 'Local'}
    elif score >= 15:
        return {'size': 'Small-Medium', 'employees': '5-15', 'revenue': '£200K - £1M',
                'rev_low': 200000, 'rev_high': 1000000, 'indicators': ', '.join(indicators) or 'Small'}
    else:
        return {'size': 'Small', 'employees': '1-5', 'revenue': '£50K - £300K',
                'rev_low': 50000, 'rev_high': 300000, 'indicators': ', '.join(indicators) or 'Small'}


def main():
    print("=" * 80)
    print("FIND REAL WEBSITES FOR 846 TRUCK TYRE COMPANIES")
    print("=" * 80)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Processing {len(INDUSTRY_DATABASE)} companies from database...")
    print()
    print("For each company, searching the web for their REAL website.")
    print("Only verified working websites will be included.")
    print()

    all_companies = []

    # =========================================================================
    # STEP 1: Process each company
    # =========================================================================
    print("[1] SEARCHING FOR REAL WEBSITES")
    print("-" * 60)

    websites_found = 0
    websites_verified = 0

    for i, company in enumerate(INDUSTRY_DATABASE):
        name = company.get('name', '')
        existing_website = company.get('website', '')

        # Copy company data
        c = {
            'name': name,
            'website': '',
            'website_verified': False,
            'website_source': '',
            'phone': company.get('phone', ''),
            'email': company.get('email', ''),
            'address': company.get('address', ''),
            'businessType': company.get('businessType', ''),
            'region': company.get('region', ''),
            'companyNumber': company.get('companyNumber', ''),
            'dateCreated': company.get('dateCreated', ''),
            'source': company.get('source', ''),
            'services': '',
            'is_tyre_related': True  # Assume true since from database
        }

        print(f"    [{i+1}/{len(INDUSTRY_DATABASE)}] {name[:50]}...", end=' ', flush=True)

        # If company already has a website, verify it
        if existing_website and existing_website.startswith('http'):
            verification = verify_website(existing_website)

            if verification and verification['works']:
                c['website'] = verification['final_url']
                c['website_verified'] = True
                c['website_source'] = 'Original (verified)'
                c['phone'] = c['phone'] or verification.get('phone', '')
                c['email'] = verification.get('email', '')
                c['services'] = ', '.join(verification.get('services', []))
                c['is_tyre_related'] = verification.get('is_tyre_related', True)
                websites_found += 1
                websites_verified += 1

                phone_display = verification.get('phone') or 'No phone'
                print(f"✓ Original verified | {phone_display[:15]}")
            else:
                # Original website doesn't work, try to find new one
                print("Original broken, searching...", end=' ')

                # Try Bing
                new_website = search_website_via_bing(name)

                # If Bing fails, try DuckDuckGo
                if not new_website:
                    new_website = search_website_via_duckduckgo(name)

                if new_website:
                    verification = verify_website(new_website)

                    if verification and verification['works'] and verification.get('is_tyre_related'):
                        c['website'] = verification['final_url']
                        c['website_verified'] = True
                        c['website_source'] = 'Web search (verified)'
                        c['phone'] = verification.get('phone', '')
                        c['email'] = verification.get('email', '')
                        c['services'] = ', '.join(verification.get('services', []))
                        websites_found += 1
                        websites_verified += 1
                        print(f"✓ Found: {urlparse(new_website).netloc[:25]}")
                    else:
                        print("✗ Not found")
                else:
                    print("✗ Not found")
        else:
            # No existing website, search for one
            new_website = search_website_via_bing(name)

            if not new_website:
                new_website = search_website_via_duckduckgo(name)

            if new_website:
                verification = verify_website(new_website)

                if verification and verification['works']:
                    # Check if it's actually tyre related
                    if verification.get('is_tyre_related'):
                        c['website'] = verification['final_url']
                        c['website_verified'] = True
                        c['website_source'] = 'Web search (verified)'
                        c['phone'] = verification.get('phone', '')
                        c['email'] = verification.get('email', '')
                        c['services'] = ', '.join(verification.get('services', []))
                        c['is_tyre_related'] = True
                        websites_found += 1
                        websites_verified += 1

                        phone_display = verification.get('phone') or 'No phone'
                        print(f"✓ Found: {urlparse(new_website).netloc[:25]} | {phone_display[:12]}")
                    else:
                        print("✗ Found but not tyre-related")
                else:
                    print("✗ Not found")
            else:
                print("✗ Not found")

        all_companies.append(c)
        time.sleep(1.5)  # Be polite to search engines

    print(f"\n    Websites found: {websites_found}/{len(INDUSTRY_DATABASE)}")
    print(f"    Verified working: {websites_verified}")

    # =========================================================================
    # STEP 2: Estimate revenue
    # =========================================================================
    print("\n[2] ESTIMATING REVENUE")
    print("-" * 60)

    for c in all_companies:
        est = estimate_revenue(c)
        c['size'] = est['size']
        c['employees'] = est['employees']
        c['revenue_estimate'] = est['revenue']
        c['revenue_low'] = est['rev_low']
        c['revenue_high'] = est['rev_high']
        c['revenue_indicators'] = est['indicators']

    # Sort by revenue
    all_companies.sort(key=lambda x: x.get('revenue_high', 0), reverse=True)

    # Stats
    total = len(all_companies)
    with_website = len([c for c in all_companies if c.get('website')])
    verified = len([c for c in all_companies if c.get('website_verified')])
    with_phone = len([c for c in all_companies if c.get('phone')])

    total_rev_low = sum(c.get('revenue_low', 0) for c in all_companies)
    total_rev_high = sum(c.get('revenue_high', 0) for c in all_companies)

    print(f"    Total: {total}")
    print(f"    With website: {with_website}")
    print(f"    Verified: {verified}")
    print(f"    With phone: {with_phone}")
    print(f"    Market: £{total_rev_low/1000000:.0f}M - £{total_rev_high/1000000:.0f}M")

    # =========================================================================
    # EXPORT
    # =========================================================================
    print("\n[3] EXPORTING")
    print("-" * 60)

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "UK Truck Tyre Companies"

    headers = ['Company Name', 'Website', 'Verified', 'Website Source', 'Phone', 'Email',
               'Size', 'Employees', 'Revenue', 'Indicators', 'Services',
               'Business Type', 'Region', 'Address', 'Company Number', 'Original Source']

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    verified_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row, c in enumerate(all_companies, 2):
        ws.cell(row=row, column=1, value=c.get('name', ''))
        ws.cell(row=row, column=2, value=c.get('website', ''))

        v_cell = ws.cell(row=row, column=3, value='Yes' if c.get('website_verified') else 'No')
        if c.get('website_verified'):
            v_cell.fill = verified_fill

        ws.cell(row=row, column=4, value=c.get('website_source', ''))
        ws.cell(row=row, column=5, value=c.get('phone', ''))
        ws.cell(row=row, column=6, value=c.get('email', ''))
        ws.cell(row=row, column=7, value=c.get('size', ''))
        ws.cell(row=row, column=8, value=c.get('employees', ''))
        ws.cell(row=row, column=9, value=c.get('revenue_estimate', ''))
        ws.cell(row=row, column=10, value=c.get('revenue_indicators', ''))
        ws.cell(row=row, column=11, value=c.get('services', ''))
        ws.cell(row=row, column=12, value=c.get('businessType', ''))
        ws.cell(row=row, column=13, value=c.get('region', ''))
        ws.cell(row=row, column=14, value=c.get('address', ''))
        ws.cell(row=row, column=15, value=c.get('companyNumber', ''))
        ws.cell(row=row, column=16, value=c.get('source', ''))

    widths = [45, 50, 10, 20, 18, 30, 15, 12, 20, 25, 30, 25, 15, 50, 15, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:P{len(all_companies) + 1}"

    wb.save(f"{OUTPUT_FILE}.xlsx")
    print(f"    Saved: {OUTPUT_FILE}.xlsx")

    # JSON
    with open(f"{OUTPUT_FILE}.json", 'w', encoding='utf-8') as f:
        json.dump(all_companies, f, indent=2, ensure_ascii=False)
    print(f"    Saved: {OUTPUT_FILE}.json")

    # CSV
    fieldnames = ['name', 'website', 'website_verified', 'website_source', 'phone', 'email',
                  'size', 'employees', 'revenue_estimate', 'revenue_low', 'revenue_high',
                  'revenue_indicators', 'services', 'businessType', 'region', 'address',
                  'companyNumber', 'dateCreated', 'source']

    with open(f"{OUTPUT_FILE}.csv", 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(all_companies)
    print(f"    Saved: {OUTPUT_FILE}.csv")

    # =========================================================================
    # SUMMARY
    # =========================================================================
    print("\n" + "=" * 80)
    print("COMPLETE")
    print("=" * 80)
    print(f"\nTotal: {total} companies from original database")
    print(f"Websites found & verified: {verified}")
    print(f"Market estimate: £{total_rev_low/1000000:.0f}M - £{total_rev_high/1000000:.0f}M")

    print("\n\nTOP 30 COMPANIES WITH VERIFIED WEBSITES:")
    print("-" * 80)

    verified_list = [c for c in all_companies if c.get('website_verified')]
    for i, c in enumerate(verified_list[:30], 1):
        print(f"\n{i}. {c['name']}")
        print(f"   Website: {c['website']}")
        if c.get('phone'):
            print(f"   Phone: {c['phone']}")
        print(f"   Size: {c['size']} | Revenue: {c['revenue_estimate']}")
        print(f"   Source: {c['website_source']}")


if __name__ == "__main__":
    main()
