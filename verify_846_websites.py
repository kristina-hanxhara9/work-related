#!/usr/bin/env python3
"""
=============================================================================
VERIFY WEBSITES FOR THE 846 COMPANIES DATABASE
=============================================================================
For each company in the database:
1. If it has a website URL - verify it actually works
2. Scrape phone/email from working websites
3. Estimate revenue

This focuses on verifying existing URLs rather than searching for new ones
(since search engines block automated scraping).

Date: January 2026
=============================================================================
"""

import requests
from bs4 import BeautifulSoup
import json
import time
import re
import csv
from datetime import datetime
from urllib.parse import urlparse
from scraper import INDUSTRY_DATABASE

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
    'Accept': 'text/html,application/xhtml+xml',
}

OUTPUT_FILE = 'uk_truck_tyres_846_WEBSITES_VERIFIED'


def verify_website(url):
    """Verify website works and extract contact info"""
    if not url or not url.startswith('http'):
        return None

    try:
        response = requests.get(url, headers=HEADERS, timeout=12, allow_redirects=True)

        if response.status_code == 200:
            text = response.text
            text_lower = text.lower()

            result = {
                'works': True,
                'final_url': response.url,
                'phone': None,
                'email': None,
                'services': [],
                'is_tyre_related': False
            }

            # Check if tyre related
            tyre_words = ['tyre', 'tire', 'wheel', 'fitting', 'truck', 'hgv', 'commercial', 'fleet']
            result['is_tyre_related'] = sum(1 for w in tyre_words if w in text_lower) >= 2

            # Extract phone
            phone_patterns = [
                r'0800[\s\-]?\d{3}[\s\-]?\d{3,4}',
                r'0808[\s\-]?\d{3}[\s\-]?\d{3,4}',
                r'01\d{2,3}[\s\-]?\d{3}[\s\-]?\d{3,4}',
                r'02\d[\s\-]?\d{4}[\s\-]?\d{4}',
                r'03\d{2}[\s\-]?\d{3}[\s\-]?\d{4}',
                r'07\d{3}[\s\-]?\d{3}[\s\-]?\d{3}',
            ]
            for pattern in phone_patterns:
                matches = re.findall(pattern, text)
                for m in matches:
                    cleaned = re.sub(r'[\s\-]', '', m)
                    if 10 <= len(cleaned) <= 12 and not cleaned.startswith('00000'):
                        result['phone'] = m.strip()
                        break
                if result['phone']:
                    break

            # Extract email
            emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', text)
            for email in emails:
                if not any(x in email.lower() for x in ['.png', '.jpg', '.css', '.js', 'example.', 'wix.', 'sentry']):
                    result['email'] = email
                    break

            # Services
            if '24' in text_lower and ('hour' in text_lower or '/7' in text_lower):
                result['services'].append('24hr')
            if 'mobile' in text_lower and 'fitting' in text_lower:
                result['services'].append('Mobile')
            if 'fleet' in text_lower:
                result['services'].append('Fleet')
            if 'commercial' in text_lower:
                result['services'].append('Commercial')

            return result

        return None
    except:
        return None


def estimate_revenue(company):
    """Estimate revenue based on characteristics"""
    name = company.get('name', '').lower()
    business_type = company.get('businessType', '').lower()
    services = company.get('services', '')
    verified = company.get('website_verified', False)

    score = 0
    indicators = []

    # Business type
    if 'manufacturer' in business_type:
        score += 50
        indicators.append('Manufacturer')
    elif 'wholesaler' in business_type:
        score += 40
        indicators.append('Wholesaler')
    elif 'national' in business_type:
        score += 35
        indicators.append('National')
    elif 'regional' in business_type:
        score += 25
        indicators.append('Regional')

    # Name indicators
    if any(x in name for x in ['national', 'uk', 'british', 'group']):
        score += 20
        indicators.append('National name')
    if 'network' in name:
        score += 15

    # Services
    if '24hr' in str(services):
        score += 10
        indicators.append('24hr')
    if 'Fleet' in str(services):
        score += 15
        indicators.append('Fleet')
    if 'Commercial' in str(services):
        score += 10

    if verified:
        score += 10

    if score >= 60:
        return {'size': 'Large', 'employees': '50-500+', 'revenue': '£10M - £100M+',
                'rev_low': 10000000, 'rev_high': 100000000, 'indicators': ', '.join(indicators) or 'Major'}
    elif score >= 40:
        return {'size': 'Medium-Large', 'employees': '20-100', 'revenue': '£2M - £15M',
                'rev_low': 2000000, 'rev_high': 15000000, 'indicators': ', '.join(indicators) or 'Regional'}
    elif score >= 25:
        return {'size': 'Medium', 'employees': '10-30', 'revenue': '£500K - £3M',
                'rev_low': 500000, 'rev_high': 3000000, 'indicators': ', '.join(indicators) or 'Local'}
    elif score >= 15:
        return {'size': 'Small-Medium', 'employees': '5-15', 'revenue': '£200K - £1M',
                'rev_low': 200000, 'rev_high': 1000000, 'indicators': ', '.join(indicators) or 'Small'}
    else:
        return {'size': 'Small', 'employees': '1-5', 'revenue': '£50K - £300K',
                'rev_low': 50000, 'rev_high': 300000, 'indicators': ', '.join(indicators) or 'Small'}


def main():
    print("=" * 80)
    print("VERIFY WEBSITES FOR 846 TRUCK TYRE COMPANIES")
    print("=" * 80)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Processing {len(INDUSTRY_DATABASE)} companies...")
    print()

    all_companies = []

    # Count companies with existing websites
    has_website = sum(1 for c in INDUSTRY_DATABASE if c.get('website', '').startswith('http'))
    print(f"Companies with website URLs: {has_website}")
    print(f"Companies without website URLs: {len(INDUSTRY_DATABASE) - has_website}")
    print()

    # =========================================================================
    # STEP 1: Process each company
    # =========================================================================
    print("[1] VERIFYING EXISTING WEBSITES")
    print("-" * 60)

    verified_count = 0

    for i, company in enumerate(INDUSTRY_DATABASE):
        name = company.get('name', '')
        existing_website = company.get('website', '')

        # Copy company data
        c = {
            'name': name,
            'website': existing_website if existing_website.startswith('http') else '',
            'website_verified': False,
            'phone': company.get('phone', ''),
            'email': company.get('email', ''),
            'address': company.get('address', ''),
            'businessType': company.get('businessType', ''),
            'region': company.get('region', ''),
            'companyNumber': company.get('companyNumber', ''),
            'dateCreated': company.get('dateCreated', ''),
            'source': company.get('source', ''),
            'services': '',
            'is_tyre_related': True
        }

        # Only verify if there's a website
        if existing_website and existing_website.startswith('http'):
            if (i + 1) % 20 == 0 or i == 0:
                print(f"    Processing {i+1}/{len(INDUSTRY_DATABASE)}...")

            verification = verify_website(existing_website)

            if verification and verification['works']:
                c['website'] = verification['final_url']
                c['website_verified'] = True
                c['phone'] = c['phone'] or verification.get('phone', '')
                c['email'] = verification.get('email', '')
                c['services'] = ', '.join(verification.get('services', []))
                c['is_tyre_related'] = verification.get('is_tyre_related', True)
                verified_count += 1

            time.sleep(0.3)

        all_companies.append(c)

    print(f"\n    Verified working websites: {verified_count}")

    # =========================================================================
    # STEP 2: Estimate revenue
    # =========================================================================
    print("\n[2] ESTIMATING REVENUE")
    print("-" * 60)

    for c in all_companies:
        est = estimate_revenue(c)
        c['size'] = est['size']
        c['employees'] = est['employees']
        c['revenue_estimate'] = est['revenue']
        c['revenue_low'] = est['rev_low']
        c['revenue_high'] = est['rev_high']
        c['revenue_indicators'] = est['indicators']

    # Sort by revenue
    all_companies.sort(key=lambda x: x.get('revenue_high', 0), reverse=True)

    # Stats
    total = len(all_companies)
    with_website = len([c for c in all_companies if c.get('website')])
    verified = len([c for c in all_companies if c.get('website_verified')])
    with_phone = len([c for c in all_companies if c.get('phone')])
    with_email = len([c for c in all_companies if c.get('email')])

    large = len([c for c in all_companies if c.get('size') == 'Large'])
    med_large = len([c for c in all_companies if c.get('size') == 'Medium-Large'])
    medium = len([c for c in all_companies if c.get('size') == 'Medium'])
    small_med = len([c for c in all_companies if c.get('size') == 'Small-Medium'])
    small = len([c for c in all_companies if c.get('size') == 'Small'])

    total_rev_low = sum(c.get('revenue_low', 0) for c in all_companies)
    total_rev_high = sum(c.get('revenue_high', 0) for c in all_companies)

    print(f"    Total: {total}")
    print(f"    With website URL: {with_website}")
    print(f"    Verified working: {verified}")
    print(f"    With phone: {with_phone}")
    print(f"    With email: {with_email}")
    print()
    print(f"    SIZE BREAKDOWN:")
    print(f"    Large: {large}")
    print(f"    Medium-Large: {med_large}")
    print(f"    Medium: {medium}")
    print(f"    Small-Medium: {small_med}")
    print(f"    Small: {small}")
    print()
    print(f"    MARKET ESTIMATE: £{total_rev_low/1000000:.0f}M - £{total_rev_high/1000000:.0f}M")

    # =========================================================================
    # EXPORT
    # =========================================================================
    print("\n[3] EXPORTING")
    print("-" * 60)

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "UK Truck Tyre Companies"

    headers = ['Company Name', 'Website', 'Verified', 'Phone', 'Email',
               'Size', 'Employees', 'Revenue', 'Indicators', 'Services',
               'Business Type', 'Region', 'Address', 'Company Number', 'Source']

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    verified_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    large_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row, c in enumerate(all_companies, 2):
        ws.cell(row=row, column=1, value=c.get('name', ''))
        ws.cell(row=row, column=2, value=c.get('website', ''))

        v_cell = ws.cell(row=row, column=3, value='Yes' if c.get('website_verified') else 'No')
        if c.get('website_verified'):
            v_cell.fill = verified_fill

        ws.cell(row=row, column=4, value=c.get('phone', ''))
        ws.cell(row=row, column=5, value=c.get('email', ''))

        size_cell = ws.cell(row=row, column=6, value=c.get('size', ''))
        if c.get('size') in ['Large', 'Medium-Large']:
            size_cell.fill = large_fill

        ws.cell(row=row, column=7, value=c.get('employees', ''))
        ws.cell(row=row, column=8, value=c.get('revenue_estimate', ''))
        ws.cell(row=row, column=9, value=c.get('revenue_indicators', ''))
        ws.cell(row=row, column=10, value=c.get('services', ''))
        ws.cell(row=row, column=11, value=c.get('businessType', ''))
        ws.cell(row=row, column=12, value=c.get('region', ''))
        ws.cell(row=row, column=13, value=c.get('address', ''))
        ws.cell(row=row, column=14, value=c.get('companyNumber', ''))
        ws.cell(row=row, column=15, value=c.get('source', ''))

    widths = [45, 50, 10, 18, 30, 15, 12, 20, 25, 30, 25, 15, 50, 15, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:O{len(all_companies) + 1}"

    wb.save(f"{OUTPUT_FILE}.xlsx")
    print(f"    Saved: {OUTPUT_FILE}.xlsx")

    # JSON
    with open(f"{OUTPUT_FILE}.json", 'w', encoding='utf-8') as f:
        json.dump(all_companies, f, indent=2, ensure_ascii=False)
    print(f"    Saved: {OUTPUT_FILE}.json")

    # CSV
    fieldnames = ['name', 'website', 'website_verified', 'phone', 'email',
                  'size', 'employees', 'revenue_estimate', 'revenue_low', 'revenue_high',
                  'revenue_indicators', 'services', 'businessType', 'region', 'address',
                  'companyNumber', 'dateCreated', 'source']

    with open(f"{OUTPUT_FILE}.csv", 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(all_companies)
    print(f"    Saved: {OUTPUT_FILE}.csv")

    # =========================================================================
    # SUMMARY
    # =========================================================================
    print("\n" + "=" * 80)
    print("COMPLETE")
    print("=" * 80)
    print(f"\nTotal: {total} companies")
    print(f"Verified working websites: {verified}")
    print(f"Market estimate: £{total_rev_low/1000000:.0f}M - £{total_rev_high/1000000:.0f}M")

    print("\n\nTOP 30 WITH VERIFIED WEBSITES:")
    print("-" * 80)

    verified_list = [c for c in all_companies if c.get('website_verified')]
    for i, c in enumerate(verified_list[:30], 1):
        print(f"\n{i}. {c['name']}")
        print(f"   Website: {c['website']}")
        if c.get('phone'):
            print(f"   Phone: {c['phone']}")
        print(f"   Size: {c['size']} | Revenue: {c['revenue_estimate']}")


if __name__ == "__main__":
    main()
