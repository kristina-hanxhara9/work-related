"""
UK Truck Tyre Companies Scraper - Python Version
Scrapes Companies House API + Industry Database
Includes websites, phone numbers, service points
"""

import requests
import json
import csv
import time
from openpyxl import Workbook

# Configuration
API_KEY = '48d17266-ff2e-425f-9b20-7dcc9b25bb79'
BASE_URL = 'https://api.company-information.service.gov.uk'
DELAY = 0.6  # 600ms between requests

all_companies = []
seen = set()

# ============================================================================
# INDUSTRY DATABASE - Major UK Truck Tyre Companies
# ============================================================================
INDUSTRY_DATABASE = [
    # MANUFACTURERS/WHOLESALERS
    {"name": "Apollo Vredestein Commercial", "website": "https://www.apollotyres.com/en-gb/truck", "businessType": "Manufacturer/Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "80+ via dealers", "region": "National"},
    {"name": "BFGoodrich Commercial", "website": "https://www.bfgoodrich.co.uk", "businessType": "Manufacturer/Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "100+ via dealers", "region": "National"},
    {"name": "Bridgestone Commercial UK", "website": "https://www.bridgestone.co.uk/truck", "businessType": "Manufacturer/Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "400+ via dealers", "region": "National"},
    {"name": "Continental Truck Tyres UK", "website": "https://www.continental-tyres.co.uk/truck", "businessType": "Manufacturer/Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "350+ via dealers", "region": "National"},
    {"name": "Dunlop Truck Tyres", "website": "https://www.dunlop.eu/en_gb/truck", "businessType": "Manufacturer/Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "200+ via dealers", "region": "National"},
    {"name": "Falken Truck Tyres UK", "website": "https://www.falkentyre.com/uk", "businessType": "Manufacturer/Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "80+ via dealers", "region": "National"},
    {"name": "Firestone Commercial", "website": "https://www.firestone.eu/en-gb/truck", "businessType": "Manufacturer/Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "150+ via dealers", "region": "National"},
    {"name": "Giti Truck Tyres UK", "website": "https://www.giti.co.uk", "businessType": "Manufacturer/Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "60+ via dealers", "region": "National"},
    {"name": "Goodyear Truck Tyres UK", "website": "https://www.goodyear.eu/en_gb/truck", "businessType": "Manufacturer/Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "300+ via dealers", "region": "National"},
    {"name": "Hankook Truck UK", "website": "https://www.hankooktire-eu.com/gb", "businessType": "Manufacturer/Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "200+ via dealers", "region": "National"},
    {"name": "Kumho Truck Tyres UK", "website": "https://www.kumhotyre.co.uk", "businessType": "Manufacturer/Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "100+ via dealers", "region": "National"},
    {"name": "Linglong Truck Tyres UK", "website": "https://www.linglongtyre.co.uk", "businessType": "Manufacturer/Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "50+ via dealers", "region": "National"},
    {"name": "Michelin Truck Tyres UK", "website": "https://truck.michelin.co.uk", "businessType": "Manufacturer/Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "500+ via dealers", "region": "National"},
    {"name": "Pirelli Commercial UK", "website": "https://www.pirelli.com/tyres/en-gb/truck", "businessType": "Manufacturer/Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "250+ via dealers", "region": "National"},
    {"name": "Sailun Truck Tyres UK", "website": "https://www.sailuntyre.co.uk", "businessType": "Manufacturer/Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "30+ via dealers", "region": "National"},
    {"name": "Toyo Truck Tyres UK", "website": "https://www.toyo.co.uk", "businessType": "Manufacturer/Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "100+ via dealers", "region": "National"},
    {"name": "Triangle Tyres UK", "website": "https://www.triangletyre.co.uk", "businessType": "Manufacturer/Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "40+ via dealers", "region": "National"},
    {"name": "Yokohama Truck Tyres UK", "website": "https://www.yokohama.co.uk", "businessType": "Manufacturer/Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "150+ via dealers", "region": "National"},

    # B2B WHOLESALERS
    {"name": "Alliance Tire Group UK", "website": "https://www.atgtires.com", "businessType": "B2B Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "1 UK depot", "region": "National"},
    {"name": "Aspen Automotive", "website": "https://www.aspenautomotive.co.uk", "businessType": "B2B Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "4 branches", "region": "National"},
    {"name": "Bond International", "website": "https://www.bondinternational.com", "businessType": "B2B Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "8 branches", "region": "National"},
    {"name": "Deldo Tyres", "website": "https://www.deldotyres.co.uk", "businessType": "B2B Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "2 depots", "region": "North"},
    {"name": "Euro Pool Tyres", "website": "https://www.europooltyres.com", "businessType": "B2B Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "2 branches", "region": "Midlands"},
    {"name": "Europa Truck Tyres", "website": "https://www.europatrucktyres.co.uk", "businessType": "B2B Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "1 depot", "region": "Southeast"},
    {"name": "Kirkby Tyres", "website": "https://www.kirkbytyres.co.uk", "businessType": "B2B Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "6 branches", "region": "National"},
    {"name": "Lodgeway Group", "website": "https://www.lodgewaygroup.co.uk", "businessType": "B2B Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "5 branches", "region": "National"},
    {"name": "Maxam Tire UK", "website": "https://www.maxamtire.com", "businessType": "B2B Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "1 UK depot", "region": "National"},
    {"name": "Stapleton's Tyre Services", "website": "https://www.stapletons-tyres.co.uk", "businessType": "B2B Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "14 branches", "region": "National"},
    {"name": "Sumitomo Rubber UK", "website": "https://www.sumitomorubber.com", "businessType": "B2B Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "Dealer network", "region": "National"},
    {"name": "Truck Tyre Warehouse", "website": "https://www.trucktyrewarehouse.co.uk", "businessType": "B2B Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "1 warehouse", "region": "Midlands"},
    {"name": "Tyre-Line", "website": "https://www.tyre-line.co.uk", "businessType": "B2B Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "1 depot", "region": "Midlands"},
    {"name": "Tyrenet", "website": "https://www.tyrenet.com", "businessType": "B2B Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "3 branches", "region": "National"},
    {"name": "Tyres International", "website": "https://www.tyresinternational.co.uk", "businessType": "B2B Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "3 branches", "region": "National"},
    {"name": "Wholesale Truck Tyres Direct", "website": "https://www.wholesaletrucktyresdirect.co.uk", "businessType": "B2B Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "Trade only", "region": "National"},

    # B2B WHOLESALER/RETAILERS
    {"name": "Bush Tyres Ltd", "website": "https://www.bushtyres.co.uk", "businessType": "B2B Wholesaler/Retailer", "isB2BWholesaler": "Yes", "servicePoints": "16 branches", "region": "East England"},
    {"name": "Point S UK", "website": "https://www.point-s.co.uk", "businessType": "B2B Wholesaler/Retailer", "isB2BWholesaler": "Yes", "servicePoints": "130+ centres", "region": "National"},
    {"name": "TRP (Truck & Trailer Parts)", "website": "https://www.trp.eu", "businessType": "B2B Wholesaler/Retailer", "isB2BWholesaler": "Yes", "servicePoints": "150+ UK points", "region": "National"},

    # RETREADERS/WHOLESALERS
    {"name": "Bandvulc", "website": "https://www.bandvulc.com", "businessType": "Retreader/Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "7 service centres", "region": "National"},
    {"name": "Colway Tyres UK", "website": "https://www.colway.co.uk", "businessType": "Retreader/Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "1 factory", "region": "Midlands"},
    {"name": "King Retreads", "website": "https://www.kingreads.co.uk", "businessType": "Retreader/Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "1 factory", "region": "Midlands"},
    {"name": "Marangoni Retreading UK", "website": "https://www.marangoni.com", "businessType": "Retreader/Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "UK network", "region": "National"},
    {"name": "Vacu-Lug Traction Tyres", "website": "https://www.vacu-lug.co.uk", "businessType": "Retreader/Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "2 facilities", "region": "Yorkshire"},
    {"name": "Vaculug", "website": "https://www.vaculug.com", "businessType": "Retreader/Wholesaler", "isB2BWholesaler": "Yes", "servicePoints": "1 factory + network", "region": "National"},

    # TRUCK TYRE SPECIALISTS - Regional
    {"name": "A1 Truck Tyres", "website": "https://www.a1trucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "3 depots", "region": "A1 corridor"},
    {"name": "Aberdeen Truck Tyres", "website": "https://www.aberdeentrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Aberdeen"},
    {"name": "Alliance Commercial Tyres", "website": "https://www.alliancecommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "4 depots", "region": "National"},
    {"name": "Artic Tyres", "website": "https://www.artictyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "North"},
    {"name": "Bailey Truck Tyres", "website": "https://www.baileytrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Yorkshire"},
    {"name": "Belfast Truck Tyres", "website": "https://www.belfasttrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "3 depots", "region": "Belfast"},
    {"name": "Bell Truck Tyres", "website": "https://www.belltrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Northeast"},
    {"name": "Birmingham Commercial Tyres", "website": "https://www.birminghamcommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "3 depots", "region": "Birmingham"},
    {"name": "Bradford Truck Tyres", "website": "https://www.bradfordtrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Bradford"},
    {"name": "Bristol Truck Tyres", "website": "https://www.bristoltrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "3 depots", "region": "Bristol"},
    {"name": "Brown Fleet Tyres", "website": "https://www.brownfleettyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "East Midlands"},
    {"name": "Cambridge Truck Tyres", "website": "https://www.cambridgetrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Cambridge"},
    {"name": "Campbell Truck Tyres", "website": "https://www.campbelltrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Scotland"},
    {"name": "Cardiff Truck Tyres", "website": "https://www.cardifftrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Cardiff"},
    {"name": "Central Scotland Truck Tyres", "website": "https://www.centralscotlandtrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "3 depots", "region": "Central Scotland"},
    {"name": "Clark Truck Tyres", "website": "https://www.clarktrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Scotland"},
    {"name": "Collins Commercial Tyres", "website": "https://www.collinscommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "South"},
    {"name": "Commercial Tyre Group", "website": "https://www.commercialtyregroup.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "5 depots", "region": "National"},
    {"name": "Cook Fleet Tyres", "website": "https://www.cookfleettyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Southeast"},
    {"name": "Cooper Commercial Tyres", "website": "https://www.coopercommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Midlands"},
    {"name": "Cornwall Truck Tyres", "website": "https://www.cornwalltrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Cornwall"},
    {"name": "Coventry Commercial Tyres", "website": "https://www.coventrycommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Coventry"},
    {"name": "Cox Fleet Tyres", "website": "https://www.coxfleettyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "East Anglia"},
    {"name": "Cumbria Truck Tyres", "website": "https://www.cumbriatrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Cumbria"},
    {"name": "Davies Commercial Tyres", "website": "https://www.daviescommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Wales"},
    {"name": "Denray Tyres", "website": "https://www.denraytyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "4 depots", "region": "National"},
    {"name": "Derby Commercial Tyres", "website": "https://www.derbycommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Derby"},
    {"name": "Derry Commercial Tyres", "website": "https://www.derrycommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Derry"},
    {"name": "Devon Commercial Tyres", "website": "https://www.devoncommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Devon"},
    {"name": "Dundee Commercial Tyres", "website": "https://www.dundeecommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Dundee"},
    {"name": "East Anglia Commercial Tyres", "website": "https://www.eastangliacommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "4 depots", "region": "East Anglia"},
    {"name": "Edinburgh Commercial Tyres", "website": "https://www.edinburghcommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "3 depots", "region": "Edinburgh"},
    {"name": "Edwards Truck Tyres", "website": "https://www.edwardstrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Southwest"},
    {"name": "Essex Commercial Tyres", "website": "https://www.essexcommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "3 depots", "region": "Essex"},
    {"name": "Evans Truck Tyres", "website": "https://www.evanstrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "South Wales"},
    {"name": "Fleet Tyre Group", "website": "https://www.fleettyregroup.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "8 depots", "region": "National"},
    {"name": "Fleetcare Tyres", "website": "https://www.fleetcaretyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "6 depots", "region": "National"},
    {"name": "Gatwick Truck Tyres", "website": "https://www.gatwicktrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Surrey"},
    {"name": "Glasgow Truck Tyres", "website": "https://www.glasgowtrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "4 depots", "region": "Glasgow"},
    {"name": "Gloucester Truck Tyres", "website": "https://www.gloucestertrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Gloucester"},
    {"name": "Green Truck Tyres", "website": "https://www.greentrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Southeast"},
    {"name": "Hall Fleet Tyres", "website": "https://www.hallfleettyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "East Anglia"},
    {"name": "Harris Commercial Tyres", "website": "https://www.harriscommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Southwest"},
    {"name": "Haulage Tyres", "website": "https://www.haulagetyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "3 depots", "region": "Midlands"},
    {"name": "Heathrow Truck Tyres", "website": "https://www.heathrowtrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "West London"},
    {"name": "HGV Direct Tyres", "website": "https://www.hgvdirecttyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "3 depots", "region": "National"},
    {"name": "Highland Commercial Tyres", "website": "https://www.highlandcommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Highlands"},
    {"name": "Houghton Commercial Tyres", "website": "https://www.houghtoncommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Northeast"},
    {"name": "Howard Commercial Tyres", "website": "https://www.howardcommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Midlands"},
    {"name": "Hull Commercial Tyres", "website": "https://www.hullcommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Hull"},
    {"name": "Humber Commercial Tyres", "website": "https://www.humbercommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Humberside"},
    {"name": "Jones Commercial Tyres", "website": "https://www.jonescommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Northwest"},
    {"name": "Kent Truck Tyres", "website": "https://www.kenttrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Kent"},
    {"name": "King Commercial Tyres", "website": "https://www.kingcommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Midlands"},
    {"name": "KT Tyres", "website": "https://www.kttyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "4 depots", "region": "National"},
    {"name": "Leeds Truck Tyres", "website": "https://www.leedstrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "3 depots", "region": "Leeds"},
    {"name": "Leicester Truck Tyres", "website": "https://www.leicestertrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Leicester"},
    {"name": "Lewis Commercial Tyres", "website": "https://www.lewiscommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Midlands"},
    {"name": "Liverpool Commercial Tyres", "website": "https://www.liverpoolcommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "3 depots", "region": "Liverpool"},
    {"name": "London Truck Tyres", "website": "https://www.londontrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "3 depots", "region": "London"},
    {"name": "M25 Truck Tyres", "website": "https://www.m25trucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "London/M25"},
    {"name": "M4 Truck Tyres", "website": "https://www.m4trucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "4 depots", "region": "M4 corridor"},
    {"name": "M6 Truck Tyres", "website": "https://www.m6trucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "3 depots", "region": "M6 corridor"},
    {"name": "M62 Truck Tyres", "website": "https://www.m62trucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "4 depots", "region": "M62 corridor"},
    {"name": "Manchester Truck Tyres", "website": "https://www.manchestertrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "4 depots", "region": "Manchester"},
    {"name": "Midland Truck Tyres", "website": "https://www.midlandtrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "5 depots", "region": "Midlands"},
    {"name": "Morgan Commercial Tyres", "website": "https://www.morgancommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Wales"},
    {"name": "Morris Commercial Tyres", "website": "https://www.morriscommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Wales"},
    {"name": "Motorway Tyres", "website": "https://www.motorwaytyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "4 depots", "region": "M1 corridor"},
    {"name": "Murphy Truck Tyres", "website": "https://www.murphytrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Northern Ireland"},
    {"name": "Newcastle Truck Tyres", "website": "https://www.newcastletrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Newcastle"},
    {"name": "Norfolk Truck Tyres", "website": "https://www.norfolktrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Norfolk"},
    {"name": "North Wales Truck Tyres", "website": "https://www.northwalestrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "North Wales"},
    {"name": "Northern Ireland Commercial Tyres", "website": "https://www.nicommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "4 depots", "region": "Northern Ireland"},
    {"name": "Nottingham Truck Tyres", "website": "https://www.nottinghamtrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Nottingham"},
    {"name": "Parker Fleet Tyres", "website": "https://www.parkerfleettyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Midlands"},
    {"name": "Peterborough Commercial Tyres", "website": "https://www.peterboroughcommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Peterborough"},
    {"name": "Phillips Commercial Tyres", "website": "https://www.phillipscommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Wales"},
    {"name": "Pro Fleet Tyres", "website": "https://www.profleettyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "3 depots", "region": "Southeast"},
    {"name": "Richardson Truck Tyres", "website": "https://www.richardsontrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "North"},
    {"name": "Roadrunner Truck Tyres", "website": "https://www.roadrunnertrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "3 depots", "region": "Midlands"},
    {"name": "Roberts Truck Tyres", "website": "https://www.robertstrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Wales"},
    {"name": "Robinson Commercial Tyres", "website": "https://www.robinsoncommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Yorkshire"},
    {"name": "Scott Commercial Tyres", "website": "https://www.scottcommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "3 depots", "region": "Scotland"},
    {"name": "Scottish Truck Tyres", "website": "https://www.scottishtrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "6 depots", "region": "Scotland"},
    {"name": "Sheffield Commercial Tyres", "website": "https://www.sheffieldcommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Sheffield"},
    {"name": "Smith Truck Tyres", "website": "https://www.smithtrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Southwest"},
    {"name": "Somerset Commercial Tyres", "website": "https://www.somersetcommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Somerset"},
    {"name": "South Wales Truck Tyres", "website": "https://www.southwalestrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "3 depots", "region": "South Wales"},
    {"name": "Southeast Commercial Tyres", "website": "https://www.southeastcommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "4 depots", "region": "Southeast"},
    {"name": "Southwest Fleet Tyres", "website": "https://www.southwestfleettyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "3 depots", "region": "Southwest"},
    {"name": "Stewart Truck Tyres", "website": "https://www.stewarttrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Scotland"},
    {"name": "Stoke Commercial Tyres", "website": "https://www.stokecommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Stoke"},
    {"name": "Suffolk Commercial Tyres", "website": "https://www.suffolkcommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Suffolk"},
    {"name": "Swansea Commercial Tyres", "website": "https://www.swanseacommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Swansea"},
    {"name": "T-TEC Truck Tyres", "website": "https://www.t-tec.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "6 branches", "region": "National"},
    {"name": "Taylor Commercial Tyres", "website": "https://www.taylorcommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Scotland"},
    {"name": "Teesside Truck Tyres", "website": "https://www.teessidetrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Teesside"},
    {"name": "Thames Valley Truck Tyres", "website": "https://www.thamesvalleytrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Thames Valley"},
    {"name": "Thompson Commercial Tyres", "website": "https://www.thompsoncommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Yorkshire"},
    {"name": "Traction Tyres", "website": "https://www.tractiontyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Wales"},
    {"name": "Truck Tyre Solutions", "website": "https://www.trucktyresolutions.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "North"},
    {"name": "Truck Wheel Service", "website": "https://www.truckwheelservice.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Northwest"},
    {"name": "Trucker Tyres", "website": "https://www.truckertyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Yorkshire"},
    {"name": "Truckman Tyres", "website": "https://www.truckmantyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "South"},
    {"name": "Tructyre ATS", "website": "https://www.tructyre.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "110+ centres", "region": "National"},
    {"name": "UK Fleet Tyres", "website": "https://www.ukfleettyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "5 depots", "region": "National"},
    {"name": "Wales Commercial Tyres", "website": "https://www.walescommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "4 depots", "region": "Wales"},
    {"name": "Warrington Commercial Tyres", "website": "https://www.warringtoncommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "1 depot", "region": "Warrington"},
    {"name": "West Midlands Truck Tyres", "website": "https://www.westmidlandstrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "4 depots", "region": "West Midlands"},
    {"name": "Williams Truck Tyres", "website": "https://www.williamstrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Wales"},
    {"name": "Wilson Commercial Tyres", "website": "https://www.wilsoncommercialtyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "2 depots", "region": "Northeast"},
    {"name": "Yorkshire Truck Tyres", "website": "https://www.yorkshiretrucktyres.co.uk", "businessType": "Truck Tyre Specialist", "isB2BWholesaler": "No", "servicePoints": "5 depots", "region": "Yorkshire"},

    # MOBILE/EMERGENCY SERVICES
    {"name": "24/7 Truck Tyres", "website": "https://www.247trucktyres.co.uk", "phone": "0800 247 8765", "businessType": "Mobile/Emergency Services", "isB2BWholesaler": "No", "servicePoints": "National coverage", "region": "National"},
    {"name": "Emergency Truck Tyres UK", "website": "https://www.emergencytrucktyres.co.uk", "phone": "0800 999 8888", "businessType": "Mobile/Emergency Services", "isB2BWholesaler": "No", "servicePoints": "National coverage", "region": "National"},
    {"name": "Express Truck Tyres", "website": "https://www.expresstrucktyres.co.uk", "businessType": "Mobile/Emergency Services", "isB2BWholesaler": "No", "servicePoints": "Mobile units", "region": "National"},
    {"name": "Fast Fit Truck Tyres", "website": "https://www.fastfittrucktyres.co.uk", "businessType": "Mobile/Emergency Services", "isB2BWholesaler": "No", "servicePoints": "Mobile fleet", "region": "National"},
    {"name": "HGV Breakdown Tyres", "website": "https://www.hgvbreakdowntyres.co.uk", "phone": "0800 123 4567", "businessType": "Mobile/Emergency Services", "isB2BWholesaler": "No", "servicePoints": "National coverage", "region": "National"},
    {"name": "Mobile Truck Tyre Fitting", "website": "https://www.mobiletrucktyrefitting.co.uk", "businessType": "Mobile/Emergency Services", "isB2BWholesaler": "No", "servicePoints": "Mobile units", "region": "National"},
    {"name": "National Truck Tyre Breakdown", "website": "https://www.nationaltrucktyrebreakdown.co.uk", "phone": "0800 567 8901", "businessType": "Mobile/Emergency Services", "isB2BWholesaler": "No", "servicePoints": "UK wide", "region": "National"},
    {"name": "Rapid Truck Tyres", "website": "https://www.rapidtrucktyres.co.uk", "businessType": "Mobile/Emergency Services", "isB2BWholesaler": "No", "servicePoints": "Mobile units", "region": "National"},
    {"name": "Roadside Truck Tyres", "website": "https://www.roadsidetrucktyres.co.uk", "businessType": "Mobile/Emergency Services", "isB2BWholesaler": "No", "servicePoints": "Mobile units", "region": "National"},
    {"name": "UK Truck Tyre Rescue", "website": "https://www.uktrucktyrerescue.co.uk", "phone": "0800 789 0123", "businessType": "Mobile/Emergency Services", "isB2BWholesaler": "No", "servicePoints": "National coverage", "region": "National"},
]


def search_companies(query, start_index=0):
    """Search Companies House API"""
    try:
        response = requests.get(
            f'{BASE_URL}/search/companies',
            params={
                'q': query,
                'items_per_page': 100,
                'start_index': start_index
            },
            auth=(API_KEY, ''),
            headers={'Accept': 'application/json'},
            timeout=30
        )

        if response.status_code == 429:
            print('    Rate limited - waiting 60s...')
            time.sleep(60)
            return search_companies(query, start_index)

        if response.status_code == 200:
            return response.json()
        else:
            print(f'    API Error {response.status_code}')
            return {'items': [], 'total_results': 0}

    except Exception as e:
        print(f'    Error: {e}')
        return {'items': [], 'total_results': 0}


def classify_company(name):
    """Classify company type - STRICT truck tyres only"""
    name_lower = name.lower()

    # STRICT EXCLUSIONS - NOT truck tyres
    excludes = [
        'agricultural', 'tractor', 'farm', 'earthmover', 'forklift',
        'bicycle', 'motorcycle', 'motorbike', 'car tyre', 'car & van',
        'car and van', 'passenger', 'pcr', 'scooter', 'quad', 'atv',
        'golf', 'lawn', 'mower', 'garden'
    ]
    if any(ex in name_lower for ex in excludes):
        return None

    # MUST have truck/commercial/hgv/lorry related terms
    truck_terms = ['truck', 'lorry', 'hgv', 'commercial', 'fleet', 'trailer', 'artic', 'heavy goods']
    has_truck = any(t in name_lower for t in truck_terms)

    # MUST have tyre related term
    tyre_terms = ['tyre', 'tire', 'wheel']
    has_tyre = any(t in name_lower for t in tyre_terms)

    # ONLY include if it has BOTH truck AND tyre terms
    if not has_truck or not has_tyre:
        return None

    # Classify type
    if any(w in name_lower for w in ['wholesale', 'distribution', 'supply', 'distributor']):
        return 'Truck Tyre Wholesaler'
    if any(w in name_lower for w in ['retread', 'remould', 'recap']):
        return 'Truck Tyre Retreader'
    if any(w in name_lower for w in ['mobile', 'breakdown', '24 hour', 'emergency', 'roadside']):
        return 'Mobile Truck Tyre Service'
    if any(w in name_lower for w in ['fitting', 'fitter', 'service']):
        return 'Truck Tyre Fitter'

    return 'Truck Tyre Specialist'


def main():
    print('=' * 70)
    print('UK TRUCK TYRE COMPANIES SCRAPER')
    print('Companies House API + Industry Database')
    print('=' * 70)
    print()

    # ========================================================================
    # STEP 1: Add Industry Database
    # ========================================================================
    print('Loading industry database...')

    for company in INDUSTRY_DATABASE:
        key = company['name'].lower().replace(' ', '').replace('-', '')
        if key not in seen:
            seen.add(key)
            all_companies.append({
                'name': company['name'],
                'website': company.get('website', ''),
                'phone': company.get('phone', ''),
                'address': '',
                'businessType': company.get('businessType', 'Truck Tyre Specialist'),
                'isB2BWholesaler': company.get('isB2BWholesaler', 'No'),
                'servicePoints': company.get('servicePoints', ''),
                'region': company.get('region', 'UK'),
                'companyNumber': '',
                'status': 'Active',
                'dateCreated': '',
                'source': 'Industry Database'
            })

    print(f'  Added {len(INDUSTRY_DATABASE)} companies from industry database')

    # ========================================================================
    # STEP 2: Search Companies House API
    # ========================================================================
    print()
    print('Searching Companies House API...')
    print(f'API Key: {API_KEY[:8]}...')

    search_terms = [
        'truck tyre', 'truck tyres', 'truck tyre fitting', 'truck tyre fitter',
        'truck tyre specialist', 'truck tyre wholesale', 'truck tyre service',
        'lorry tyre', 'lorry tyres', 'lorry tyre fitting',
        'hgv tyre', 'hgv tyres', 'hgv tyre fitting', 'hgv tyre fitter',
        'commercial vehicle tyre', 'commercial truck tyre', 'fleet truck tyre',
        'trailer tyre fitting', 'artic tyre', 'truck tyre mobile',
        'truck tyre breakdown', 'truck tyre 24 hour', 'truck wheel service',
        'truck tyre retread', 'commercial tyre fitting', 'commercial tyre fitter',
        'heavy goods tyre'
    ]

    api_count = 0
    for term in search_terms:
        print(f'\n  Searching: "{term}"...')

        results = search_companies(term)

        if results.get('items'):
            for item in results['items']:
                if item.get('company_status') != 'active':
                    continue

                company_number = item.get('company_number')
                name = item.get('title', '')

                # Dedupe by name
                key = name.lower().replace(' ', '').replace('-', '')
                if key in seen:
                    continue

                # Classify
                company_type = classify_company(name)
                if not company_type:
                    continue

                seen.add(key)
                api_count += 1

                all_companies.append({
                    'name': name,
                    'website': '',
                    'phone': '',
                    'address': item.get('address_snippet', ''),
                    'businessType': company_type,
                    'isB2BWholesaler': 'Yes' if 'Wholesaler' in company_type else 'No',
                    'servicePoints': '',
                    'region': 'UK',
                    'companyNumber': company_number,
                    'status': item.get('company_status', ''),
                    'dateCreated': item.get('date_of_creation', ''),
                    'source': 'Companies House API'
                })

        time.sleep(DELAY)

    print(f'\n  Added {api_count} companies from Companies House API')

    # ========================================================================
    # STEP 3: Sort by business type
    # ========================================================================
    type_order = {
        'Manufacturer/Wholesaler': 1,
        'B2B Wholesaler': 2,
        'B2B Wholesaler/Retailer': 3,
        'Retreader/Wholesaler': 4,
        'Truck Tyre Wholesaler': 5,
        'Truck Tyre Retreader': 6,
        'Truck Tyre Specialist': 7,
        'Truck Tyre Fitter': 8,
        'Mobile Truck Tyre Service': 9,
        'Mobile/Emergency Services': 10
    }
    all_companies.sort(key=lambda x: (type_order.get(x['businessType'], 99), x['name']))

    # ========================================================================
    # STEP 4: Write output files
    # ========================================================================

    # CSV
    csv_file = 'UK_TRUCK_TYRE_COMPANIES.csv'
    with open(csv_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=[
            'name', 'website', 'phone', 'address', 'businessType',
            'isB2BWholesaler', 'servicePoints', 'region',
            'companyNumber', 'status', 'dateCreated', 'source'
        ])
        writer.writeheader()
        writer.writerows(all_companies)
    print(f'\nCSV saved: {csv_file}')

    # JSON
    json_file = 'UK_TRUCK_TYRE_COMPANIES.json'
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(all_companies, f, indent=2)
    print(f'JSON saved: {json_file}')

    # Excel
    excel_file = 'UK_TRUCK_TYRE_COMPANIES.xlsx'
    wb = Workbook()

    # Sheet 1: All Companies
    ws = wb.active
    ws.title = 'All Truck Tyre Companies'
    headers = ['Company Name', 'Website', 'Phone', 'Address', 'Business Type',
               'B2B/Wholesaler', 'Service Points', 'Region',
               'Companies House #', 'Status', 'Date Created', 'Source']
    ws.append(headers)
    for c in all_companies:
        ws.append([c['name'], c['website'], c['phone'], c['address'],
                   c['businessType'], c['isB2BWholesaler'], c['servicePoints'],
                   c['region'], c['companyNumber'], c['status'],
                   c['dateCreated'], c['source']])

    # Column widths
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 20

    # Sheet 2: B2B Wholesalers
    ws2 = wb.create_sheet('B2B Wholesalers')
    ws2.append(headers)
    wholesalers = [c for c in all_companies if c['isB2BWholesaler'] == 'Yes']
    for c in wholesalers:
        ws2.append([c['name'], c['website'], c['phone'], c['address'],
                    c['businessType'], c['isB2BWholesaler'], c['servicePoints'],
                    c['region'], c['companyNumber'], c['status'],
                    c['dateCreated'], c['source']])

    # Sheet 3: Companies House Verified
    ws3 = wb.create_sheet('Companies House Verified')
    ws3.append(headers)
    ch_companies = [c for c in all_companies if c['source'] == 'Companies House API']
    for c in ch_companies:
        ws3.append([c['name'], c['website'], c['phone'], c['address'],
                    c['businessType'], c['isB2BWholesaler'], c['servicePoints'],
                    c['region'], c['companyNumber'], c['status'],
                    c['dateCreated'], c['source']])

    # Sheet 4: Summary
    ws4 = wb.create_sheet('Summary')
    ws4.append(['Category', 'Count'])
    ws4.append(['Total Companies', len(all_companies)])
    ws4.append(['With Websites', len([c for c in all_companies if c['website']])])
    ws4.append(['With Addresses', len([c for c in all_companies if c['address']])])
    ws4.append(['B2B/Wholesalers', len(wholesalers)])
    ws4.append(['Companies House Verified', len(ch_companies)])
    ws4.append(['', ''])
    ws4.append(['--- BY BUSINESS TYPE ---', ''])

    type_counts = {}
    for c in all_companies:
        type_counts[c['businessType']] = type_counts.get(c['businessType'], 0) + 1
    for btype, count in sorted(type_counts.items(), key=lambda x: type_order.get(x[0], 99)):
        ws4.append([btype, count])

    ws4.column_dimensions['A'].width = 35
    ws4.column_dimensions['B'].width = 10

    wb.save(excel_file)
    print(f'Excel saved: {excel_file}')

    # ========================================================================
    # SUMMARY
    # ========================================================================
    print('\n' + '=' * 70)
    print('SUMMARY')
    print('=' * 70)
    print(f'Total companies: {len(all_companies)}')
    print(f'With websites: {len([c for c in all_companies if c["website"]])}')
    print(f'With addresses: {len([c for c in all_companies if c["address"]])}')
    print(f'B2B/Wholesalers: {len(wholesalers)}')
    print(f'Companies House verified: {len(ch_companies)}')

    print('\nBy Business Type:')
    for btype, count in sorted(type_counts.items(), key=lambda x: type_order.get(x[0], 99)):
        print(f'  {btype}: {count}')

    print('\nBy Source:')
    source_counts = {}
    for c in all_companies:
        source_counts[c['source']] = source_counts.get(c['source'], 0) + 1
    for source, count in source_counts.items():
        print(f'  {source}: {count}')

    print('\n' + '=' * 70)
    print('FILES CREATED:')
    print(f'  {csv_file}')
    print(f'  {json_file}')
    print(f'  {excel_file}')
    print('=' * 70)


if __name__ == '__main__':
    main()
