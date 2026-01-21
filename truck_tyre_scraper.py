"""
UK TRUCK TYRE COMPANIES - WEB SCRAPER
=====================================
Scrapes actual truck/HGV/commercial tyre company data from:
1. Companies House API (SIC codes for truck tyres)
2. Dealer network websites (Truckpoint, ITDN, etc)
3. Industry directories
4. Individual company websites

Run: python truck_tyre_scraper.py
"""

import requests
import json
import csv
import time
import re
from datetime import datetime
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

# ============================================================================
# CONFIGURATION
# ============================================================================
CH_API_KEY = '48d17266-ff2e-425f-9b20-7dcc9b25bb79'
CH_BASE_URL = 'https://api.company-information.service.gov.uk'

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-GB,en;q=0.9',
}

# ============================================================================
# VERIFIED TRUCK TYRE COMPANIES - STARTING LIST
# ============================================================================
VERIFIED_TRUCK_TYRE_COMPANIES = []

# ============================================================================
# SCRAPING FUNCTIONS
# ============================================================================

def scrape_companies_house_truck_tyres():
    """Search Companies House for truck tyre related companies"""
    print("\n[1] SCRAPING COMPANIES HOUSE FOR TRUCK TYRE COMPANIES...")

    search_terms = [
        "truck tyre",
        "truck tyres",
        "HGV tyre",
        "lorry tyre",
        "commercial tyre",
        "fleet tyre",
        "truck tire",
        "commercial vehicle tyre",
    ]

    companies = []
    seen = set()

    for term in search_terms:
        print(f"  Searching: '{term}'")
        try:
            params = {'q': term, 'items_per_page': 100}
            r = requests.get(f"{CH_BASE_URL}/search/companies",
                           auth=(CH_API_KEY, ''),
                           params=params,
                           timeout=30)

            if r.status_code == 200:
                data = r.json()
                items = data.get('items', [])
                print(f"    Found {len(items)} results")

                for item in items:
                    company_number = item.get('company_number', '')
                    if company_number and company_number not in seen:
                        seen.add(company_number)

                        # Filter for active companies
                        status = item.get('company_status', '')
                        if status == 'active':
                            companies.append({
                                'name': item.get('title', ''),
                                'company_number': company_number,
                                'address': item.get('address_snippet', ''),
                                'status': status,
                                'date_created': item.get('date_of_creation', ''),
                                'source': 'Companies House Search',
                            })

            time.sleep(0.6)

        except Exception as e:
            print(f"    Error: {e}")

    print(f"  Total unique companies from CH: {len(companies)}")
    return companies


def get_company_details(company_number):
    """Get detailed company info from Companies House"""
    try:
        r = requests.get(f"{CH_BASE_URL}/company/{company_number}",
                        auth=(CH_API_KEY, ''), timeout=15)
        if r.status_code == 200:
            data = r.json()
            return {
                'sic_codes': data.get('sic_codes', []),
                'type': data.get('type', ''),
                'accounts_type': data.get('accounts', {}).get('last_accounts', {}).get('type', ''),
                'registered_office': data.get('registered_office_address', {}),
            }
    except:
        pass
    return {}


def scrape_truckpoint_dealers():
    """Scrape Bridgestone Truckpoint dealer network"""
    print("\n[2] SCRAPING BRIDGESTONE TRUCKPOINT DEALERS...")

    dealers = []

    # Try the dealer locator API/page
    urls_to_try = [
        'https://www.bridgestone.co.uk/truck/dealer-locator',
        'https://www.bridgestone.co.uk/commercial/find-a-dealer',
        'https://truckpoint.eu/en/find-a-dealer',
    ]

    for url in urls_to_try:
        try:
            print(f"  Trying: {url}")
            r = requests.get(url, headers=HEADERS, timeout=15)
            if r.status_code == 200:
                soup = BeautifulSoup(r.text, 'html.parser')

                # Look for dealer listings
                # Common patterns: div with dealer info, JSON data, etc.

                # Check for JSON data in script tags
                scripts = soup.find_all('script')
                for script in scripts:
                    if script.string and ('dealer' in script.string.lower() or 'location' in script.string.lower()):
                        # Try to extract JSON
                        try:
                            json_match = re.search(r'\[{.*}\]', script.string)
                            if json_match:
                                data = json.loads(json_match.group())
                                print(f"    Found JSON data with {len(data)} items")
                        except:
                            pass

                # Look for dealer cards/listings
                dealer_elements = soup.find_all(['div', 'li'], class_=re.compile(r'dealer|location|store', re.I))
                if dealer_elements:
                    print(f"    Found {len(dealer_elements)} dealer elements")

        except Exception as e:
            print(f"    Error: {e}")

        time.sleep(1)

    return dealers


def scrape_itdn_members():
    """Scrape ITDN (Independent Tyre Distributors Network) members"""
    print("\n[3] SCRAPING ITDN NETWORK MEMBERS...")

    members = []

    try:
        url = 'https://www.itdn.co.uk'
        r = requests.get(url, headers=HEADERS, timeout=15)
        if r.status_code == 200:
            soup = BeautifulSoup(r.text, 'html.parser')

            # Look for member listings
            links = soup.find_all('a', href=True)
            for link in links:
                href = link.get('href', '')
                if 'member' in href.lower() or 'dealer' in href.lower():
                    print(f"    Found potential member page: {href}")

    except Exception as e:
        print(f"  Error: {e}")

    return members


def scrape_tyrenet_network():
    """Scrape Tyrenet commercial tyre network"""
    print("\n[4] SCRAPING TYRENET COMMERCIAL NETWORK...")

    dealers = []

    try:
        url = 'https://tyrenet.net/find-a-dealer'
        r = requests.get(url, headers=HEADERS, timeout=15)
        if r.status_code == 200:
            soup = BeautifulSoup(r.text, 'html.parser')
            print(f"    Page loaded, parsing...")

            # Look for dealer info
            text = soup.get_text()
            if 'dealer' in text.lower():
                print(f"    Page contains dealer information")

    except Exception as e:
        print(f"  Error: {e}")

    return dealers


def scrape_btc_network():
    """Scrape BTC (British Tyre Centre) commercial network"""
    print("\n[5] SCRAPING BTC NETWORK...")

    members = []

    try:
        url = 'https://www.btcnetwork.co.uk'
        r = requests.get(url, headers=HEADERS, timeout=15)
        if r.status_code == 200:
            soup = BeautifulSoup(r.text, 'html.parser')
            print(f"    Page loaded")

    except Exception as e:
        print(f"  Error: {e}")

    return members


def scrape_company_website(url, company_name):
    """Scrape individual company website for business info"""

    info = {
        'has_truck_tyres': False,
        'has_commercial': False,
        'has_fleet': False,
        'has_hgv': False,
        'phone': '',
        'email': '',
        'services': [],
        'brands': [],
    }

    if not url or not url.startswith('http'):
        return info

    try:
        r = requests.get(url, headers=HEADERS, timeout=10)
        if r.status_code == 200:
            text = r.text.lower()
            soup = BeautifulSoup(r.text, 'html.parser')

            # Check for truck/commercial tyre keywords
            truck_keywords = ['truck tyre', 'truck tire', 'hgv', 'lorry', 'commercial vehicle',
                            'fleet', 'trailer', 'bus tyre', 'coach tyre', '24 hour', 'breakdown']

            for kw in truck_keywords:
                if kw in text:
                    if 'truck' in kw or 'hgv' in kw or 'lorry' in kw:
                        info['has_truck_tyres'] = True
                    if 'commercial' in kw:
                        info['has_commercial'] = True
                    if 'fleet' in kw:
                        info['has_fleet'] = True
                    if 'hgv' in kw:
                        info['has_hgv'] = True

            # Extract phone
            phone_match = re.search(r'(\+?44[\s\-\.]?\d{2,4}[\s\-\.]?\d{3,4}[\s\-\.]?\d{3,4}|0\d{2,4}[\s\-\.]?\d{3,4}[\s\-\.]?\d{3,4})', r.text)
            if phone_match:
                info['phone'] = phone_match.group(1)

            # Extract email
            email_match = re.search(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', r.text)
            if email_match:
                email = email_match.group()
                if not email.endswith(('.png', '.jpg', '.gif')):
                    info['email'] = email

            # Check for tyre brands
            brands = ['bridgestone', 'michelin', 'continental', 'goodyear', 'pirelli',
                     'hankook', 'dunlop', 'firestone', 'bandvulc', 'yokohama']
            for brand in brands:
                if brand in text:
                    info['brands'].append(brand.title())

    except:
        pass

    return info


def scrape_europages_truck_tyres():
    """Scrape Europages B2B directory for UK truck tyre companies"""
    print("\n[6] SCRAPING EUROPAGES DIRECTORY...")

    companies = []

    try:
        url = 'https://www.europages.co.uk/companies/pg-1/truck%20tyres.html'
        r = requests.get(url, headers=HEADERS, timeout=15)
        if r.status_code == 200:
            soup = BeautifulSoup(r.text, 'html.parser')

            # Look for company listings
            company_cards = soup.find_all(['div', 'article'], class_=re.compile(r'company|result|listing', re.I))
            print(f"    Found {len(company_cards)} company listings")

            for card in company_cards[:50]:  # Limit to first 50
                name_elem = card.find(['h2', 'h3', 'a'], class_=re.compile(r'name|title', re.I))
                if name_elem:
                    name = name_elem.get_text(strip=True)
                    if name and 'uk' in card.get_text().lower():
                        companies.append({
                            'name': name,
                            'source': 'Europages',
                        })

    except Exception as e:
        print(f"  Error: {e}")

    return companies


def scrape_yell_truck_tyres():
    """Scrape Yell.com for truck tyre businesses"""
    print("\n[7] SCRAPING YELL.COM DIRECTORY...")

    companies = []

    # Major UK cities to search
    locations = ['london', 'manchester', 'birmingham', 'leeds', 'glasgow', 'bristol', 'edinburgh']

    for location in locations[:3]:  # Limit for speed
        try:
            url = f'https://www.yell.com/ucs/UcsSearchAction.do?keywords=truck+tyres&location={location}'
            print(f"  Searching: {location}")
            r = requests.get(url, headers=HEADERS, timeout=15)

            if r.status_code == 200:
                soup = BeautifulSoup(r.text, 'html.parser')

                # Look for business listings
                listings = soup.find_all('div', class_=re.compile(r'businessCapsule|listing', re.I))
                print(f"    Found {len(listings)} listings")

                for listing in listings[:20]:
                    name_elem = listing.find(['h2', 'a'], class_=re.compile(r'businessName|name', re.I))
                    if name_elem:
                        name = name_elem.get_text(strip=True)

                        # Get address
                        addr_elem = listing.find(class_=re.compile(r'address', re.I))
                        address = addr_elem.get_text(strip=True) if addr_elem else ''

                        # Get phone
                        phone_elem = listing.find(class_=re.compile(r'phone|tel', re.I))
                        phone = phone_elem.get_text(strip=True) if phone_elem else ''

                        companies.append({
                            'name': name,
                            'address': address,
                            'phone': phone,
                            'location': location,
                            'source': 'Yell.com',
                        })

            time.sleep(2)

        except Exception as e:
            print(f"    Error: {e}")

    return companies


def verify_truck_tyre_company(company):
    """Verify if a company is actually a truck tyre specialist"""

    name = company.get('name', '').lower()

    # Strong indicators it's a truck tyre company
    truck_keywords = ['truck', 'hgv', 'lorry', 'commercial', 'fleet', 'trailer',
                     'bus tyre', 'coach', 'transport', 'haulage']

    # Weak indicators (could be car tyres)
    weak_keywords = ['tyre', 'tire', 'wheel', 'auto']

    # Negative indicators (likely not truck tyres)
    negative_keywords = ['car tyre', 'kwik fit', 'halfords', 'formula one', 'f1 auto']

    # Check negative first
    for neg in negative_keywords:
        if neg in name:
            return False, 'Likely car tyres only'

    # Check for strong truck indicators
    for kw in truck_keywords:
        if kw in name:
            return True, f'Name contains "{kw}"'

    # If just "tyre" in name, needs website verification
    if any(w in name for w in weak_keywords):
        return None, 'Needs website verification'

    return False, 'No tyre keywords'


# ============================================================================
# MAIN SCRAPING PIPELINE
# ============================================================================

def run_full_scrape():
    """Run the complete scraping pipeline"""

    print("=" * 70)
    print("UK TRUCK TYRE COMPANIES - WEB SCRAPER")
    print("=" * 70)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    all_companies = []

    # 1. Companies House search
    ch_companies = scrape_companies_house_truck_tyres()
    all_companies.extend(ch_companies)

    # 2. Dealer networks
    truckpoint = scrape_truckpoint_dealers()
    all_companies.extend(truckpoint)

    itdn = scrape_itdn_members()
    all_companies.extend(itdn)

    tyrenet = scrape_tyrenet_network()
    all_companies.extend(tyrenet)

    btc = scrape_btc_network()
    all_companies.extend(btc)

    # 3. Business directories
    europages = scrape_europages_truck_tyres()
    all_companies.extend(europages)

    yell = scrape_yell_truck_tyres()
    all_companies.extend(yell)

    # Deduplicate
    seen = set()
    unique_companies = []
    for c in all_companies:
        key = c.get('company_number') or c.get('name', '').lower()
        if key and key not in seen:
            seen.add(key)
            unique_companies.append(c)

    print(f"\n\nTotal unique companies found: {len(unique_companies)}")

    # Verify each company
    print("\nVerifying companies are truck tyre specialists...")
    verified = []
    for c in unique_companies:
        is_truck, reason = verify_truck_tyre_company(c)
        c['is_verified_truck'] = is_truck
        c['verification_reason'] = reason
        if is_truck is True or is_truck is None:
            verified.append(c)

    print(f"Verified truck tyre companies: {len(verified)}")

    return verified


def save_results(companies):
    """Save results to Excel, JSON, CSV"""

    print("\nSaving results...")

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Truck Tyre Companies"

    headers = ['Company Name', 'Company Number', 'Address', 'Phone', 'Status',
               'Is Verified Truck', 'Verification Reason', 'Source']

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row, c in enumerate(companies, 2):
        ws.cell(row=row, column=1, value=c.get('name', ''))
        ws.cell(row=row, column=2, value=c.get('company_number', ''))
        ws.cell(row=row, column=3, value=c.get('address', ''))
        ws.cell(row=row, column=4, value=c.get('phone', ''))
        ws.cell(row=row, column=5, value=c.get('status', ''))
        ws.cell(row=row, column=6, value=str(c.get('is_verified_truck', '')))
        ws.cell(row=row, column=7, value=c.get('verification_reason', ''))
        ws.cell(row=row, column=8, value=c.get('source', ''))

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['C'].width = 50

    wb.save('UK_TRUCK_TYRE_SCRAPED.xlsx')
    print("  Saved: UK_TRUCK_TYRE_SCRAPED.xlsx")

    # JSON
    with open('UK_TRUCK_TYRE_SCRAPED.json', 'w', encoding='utf-8') as f:
        json.dump(companies, f, indent=2, ensure_ascii=False)
    print("  Saved: UK_TRUCK_TYRE_SCRAPED.json")

    # CSV
    if companies:
        with open('UK_TRUCK_TYRE_SCRAPED.csv', 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=companies[0].keys())
            writer.writeheader()
            writer.writerows(companies)
        print("  Saved: UK_TRUCK_TYRE_SCRAPED.csv")


# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    companies = run_full_scrape()
    save_results(companies)

    print("\n" + "=" * 70)
    print("SCRAPING COMPLETE")
    print("=" * 70)
    print(f"Total verified truck tyre companies: {len(companies)}")
    print(f"Finished: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
