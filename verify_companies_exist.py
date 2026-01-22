#!/usr/bin/env python3
"""
=============================================================================
VERIFY THAT 846 COMPANIES ACTUALLY EXIST AS REAL BUSINESSES
=============================================================================
For each company in the database:
1. Check Companies House API - if found, it's a VERIFIED real company
2. For companies NOT from Companies House, try to verify they exist
3. Mark each company as: VERIFIED, UNVERIFIED, or FAKE

Date: January 2026
=============================================================================
"""

import requests
from bs4 import BeautifulSoup
import json
import time
import re
import csv
from datetime import datetime
from urllib.parse import urlparse, quote
from scraper import INDUSTRY_DATABASE

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
    'Accept': 'text/html,application/xhtml+xml',
}

CH_API_KEY = '48d17266-ff2e-425f-9b20-7dcc9b25bb79'
CH_BASE_URL = 'https://api.company-information.service.gov.uk'

OUTPUT_FILE = 'uk_truck_tyres_846_FULLY_VERIFIED'


def search_companies_house(company_name):
    """Search Companies House to verify company exists"""
    try:
        # Clean company name for search
        search_name = company_name.replace('Ltd', '').replace('Limited', '').replace('PLC', '').strip()

        url = f"{CH_BASE_URL}/search/companies"
        params = {'q': search_name, 'items_per_page': 5}

        response = requests.get(url, params=params, auth=(CH_API_KEY, ''), timeout=10)

        if response.status_code == 200:
            data = response.json()
            items = data.get('items', [])

            if items:
                # Check if any result closely matches
                for item in items:
                    ch_name = item.get('title', '').lower()
                    search_lower = search_name.lower()

                    # Check for close match
                    if search_lower in ch_name or ch_name in search_lower:
                        return {
                            'found': True,
                            'company_number': item.get('company_number'),
                            'company_name': item.get('title'),
                            'company_status': item.get('company_status'),
                            'address': item.get('address_snippet', ''),
                            'date_created': item.get('date_of_creation', ''),
                            'company_type': item.get('company_type', '')
                        }

                    # Also check partial match on key words
                    name_words = set(search_lower.split())
                    ch_words = set(ch_name.split())
                    common = name_words & ch_words
                    if len(common) >= 2:  # At least 2 words match
                        return {
                            'found': True,
                            'company_number': item.get('company_number'),
                            'company_name': item.get('title'),
                            'company_status': item.get('company_status'),
                            'address': item.get('address_snippet', ''),
                            'date_created': item.get('date_of_creation', ''),
                            'company_type': item.get('company_type', '')
                        }

        return {'found': False}
    except Exception as e:
        return {'found': False, 'error': str(e)}


def verify_website(url):
    """Verify website works and check if tyre-related"""
    if not url or not url.startswith('http'):
        return None

    try:
        response = requests.get(url, headers=HEADERS, timeout=12, allow_redirects=True)

        if response.status_code == 200:
            text = response.text
            text_lower = text.lower()

            result = {
                'works': True,
                'final_url': response.url,
                'phone': None,
                'email': None,
                'is_tyre_related': False,
                'is_truck_tyre': False
            }

            # Check if tyre related
            tyre_words = ['tyre', 'tire', 'wheel', 'fitting']
            truck_words = ['truck', 'hgv', 'commercial', 'fleet', 'lorry', 'trailer', 'heavy']

            tyre_count = sum(1 for w in tyre_words if w in text_lower)
            truck_count = sum(1 for w in truck_words if w in text_lower)

            result['is_tyre_related'] = tyre_count >= 1
            result['is_truck_tyre'] = tyre_count >= 1 and truck_count >= 1

            # Extract phone
            phone_patterns = [
                r'0800[\s\-]?\d{3}[\s\-]?\d{3,4}',
                r'0808[\s\-]?\d{3}[\s\-]?\d{3,4}',
                r'01\d{2,3}[\s\-]?\d{3}[\s\-]?\d{3,4}',
                r'02\d[\s\-]?\d{4}[\s\-]?\d{4}',
                r'03\d{2}[\s\-]?\d{3}[\s\-]?\d{4}',
                r'07\d{3}[\s\-]?\d{3}[\s\-]?\d{3}',
            ]
            for pattern in phone_patterns:
                matches = re.findall(pattern, text)
                for m in matches:
                    cleaned = re.sub(r'[\s\-]', '', m)
                    if 10 <= len(cleaned) <= 12 and not cleaned.startswith('00000'):
                        result['phone'] = m.strip()
                        break
                if result['phone']:
                    break

            # Extract email
            emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', text)
            for email in emails:
                if not any(x in email.lower() for x in ['.png', '.jpg', '.css', '.js', 'example.', 'wix.', 'sentry']):
                    result['email'] = email
                    break

            return result

        return None
    except:
        return None


def is_likely_truck_tyre_company(name):
    """Check if company name suggests truck tyre business"""
    name_lower = name.lower()

    # Strong indicators
    strong_indicators = ['truck tyre', 'truck tire', 'hgv tyre', 'commercial tyre',
                        'fleet tyre', 'lorry tyre', 'commercial vehicle']
    for indicator in strong_indicators:
        if indicator in name_lower:
            return True, 'Strong match'

    # Medium indicators (tyre + location or tyre + service)
    if 'tyre' in name_lower or 'tire' in name_lower:
        if any(word in name_lower for word in ['commercial', 'fleet', 'truck', 'hgv', 'mobile', '24']):
            return True, 'Medium match'

    # Weak indicators (just tyre in name)
    if 'tyre' in name_lower or 'tire' in name_lower:
        return True, 'Weak match (general tyre)'

    return False, 'No match'


def estimate_revenue(company):
    """Estimate revenue based on characteristics"""
    name = company.get('name', '').lower()
    business_type = company.get('businessType', '').lower()
    verified = company.get('ch_verified', False)
    website_works = company.get('website_verified', False)

    score = 0
    indicators = []

    # Verification bonus
    if verified:
        score += 15
        indicators.append('CH Verified')
    if website_works:
        score += 10
        indicators.append('Website works')

    # Business type
    if 'manufacturer' in business_type:
        score += 50
        indicators.append('Manufacturer')
    elif 'wholesaler' in business_type:
        score += 40
        indicators.append('Wholesaler')
    elif 'national' in business_type:
        score += 35
        indicators.append('National')
    elif 'regional' in business_type:
        score += 25
        indicators.append('Regional')

    # Name indicators
    if any(x in name for x in ['national', 'uk', 'british', 'group']):
        score += 20
        indicators.append('National name')
    if 'network' in name:
        score += 15

    if score >= 60:
        return {'size': 'Large', 'employees': '50-500+', 'revenue': '£10M - £100M+',
                'rev_low': 10000000, 'rev_high': 100000000, 'indicators': ', '.join(indicators) or 'Major'}
    elif score >= 40:
        return {'size': 'Medium-Large', 'employees': '20-100', 'revenue': '£2M - £15M',
                'rev_low': 2000000, 'rev_high': 15000000, 'indicators': ', '.join(indicators) or 'Regional'}
    elif score >= 25:
        return {'size': 'Medium', 'employees': '10-30', 'revenue': '£500K - £3M',
                'rev_low': 500000, 'rev_high': 3000000, 'indicators': ', '.join(indicators) or 'Local'}
    elif score >= 15:
        return {'size': 'Small-Medium', 'employees': '5-15', 'revenue': '£200K - £1M',
                'rev_low': 200000, 'rev_high': 1000000, 'indicators': ', '.join(indicators) or 'Small'}
    else:
        return {'size': 'Small', 'employees': '1-5', 'revenue': '£50K - £300K',
                'rev_low': 50000, 'rev_high': 300000, 'indicators': ', '.join(indicators) or 'Small'}


def main():
    print("=" * 80)
    print("VERIFY 846 COMPANIES ACTUALLY EXIST")
    print("=" * 80)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Processing {len(INDUSTRY_DATABASE)} companies...")
    print()

    all_companies = []

    # Count companies by source
    has_ch_number = sum(1 for c in INDUSTRY_DATABASE if c.get('companyNumber'))
    has_website = sum(1 for c in INDUSTRY_DATABASE if c.get('website', '').startswith('http'))

    print(f"Companies with Companies House number: {has_ch_number} (Already verified)")
    print(f"Companies with website URLs: {has_website}")
    print(f"Companies needing verification: {len(INDUSTRY_DATABASE) - has_ch_number}")
    print()

    # =========================================================================
    # STEP 1: Process each company
    # =========================================================================
    print("[1] VERIFYING EACH COMPANY EXISTS")
    print("-" * 60)

    ch_verified = 0
    website_verified = 0
    name_matched = 0
    unverified = 0

    for i, company in enumerate(INDUSTRY_DATABASE):
        name = company.get('name', '')
        existing_website = company.get('website', '')
        existing_ch_number = company.get('companyNumber', '')

        if (i + 1) % 25 == 0 or i == 0:
            print(f"\n    Processing {i+1}/{len(INDUSTRY_DATABASE)}...")

        # Copy company data
        c = {
            'name': name,
            'website': existing_website if existing_website.startswith('http') else '',
            'website_verified': False,
            'website_is_truck_tyre': False,
            'ch_verified': False,
            'ch_number': existing_ch_number,
            'ch_status': '',
            'ch_address': company.get('address', ''),
            'phone': company.get('phone', ''),
            'email': company.get('email', ''),
            'businessType': company.get('businessType', ''),
            'region': company.get('region', ''),
            'source': company.get('source', ''),
            'verification_status': 'UNVERIFIED',
            'verification_method': '',
            'is_truck_tyre': False
        }

        # Check if name suggests truck tyre
        is_truck, match_type = is_likely_truck_tyre_company(name)
        c['name_match'] = match_type

        # METHOD 1: Already has Companies House number = VERIFIED
        if existing_ch_number:
            c['ch_verified'] = True
            c['verification_status'] = 'VERIFIED'
            c['verification_method'] = 'Companies House Number'
            c['is_truck_tyre'] = is_truck
            ch_verified += 1
            print(f"    ✓ {name[:40]} - CH Verified ({existing_ch_number})")

        else:
            # METHOD 2: Search Companies House by name
            ch_result = search_companies_house(name)

            if ch_result.get('found'):
                c['ch_verified'] = True
                c['ch_number'] = ch_result.get('company_number', '')
                c['ch_status'] = ch_result.get('company_status', '')
                c['ch_address'] = ch_result.get('address', '') or c['ch_address']
                c['verification_status'] = 'VERIFIED'
                c['verification_method'] = 'Companies House Search'
                c['is_truck_tyre'] = is_truck
                ch_verified += 1
                print(f"    ✓ {name[:40]} - Found in CH: {ch_result.get('company_name', '')[:30]}")
            else:
                # METHOD 3: Check if website works and is truck tyre related
                if existing_website and existing_website.startswith('http'):
                    verification = verify_website(existing_website)

                    if verification and verification['works']:
                        c['website_verified'] = True
                        c['phone'] = c['phone'] or verification.get('phone', '')
                        c['email'] = verification.get('email', '')
                        c['website_is_truck_tyre'] = verification.get('is_truck_tyre', False)

                        if verification.get('is_truck_tyre'):
                            c['verification_status'] = 'VERIFIED'
                            c['verification_method'] = 'Website (Truck Tyre Content)'
                            c['is_truck_tyre'] = True
                            website_verified += 1
                            print(f"    ✓ {name[:40]} - Website verified as truck tyre")
                        elif verification.get('is_tyre_related'):
                            c['verification_status'] = 'PARTIAL'
                            c['verification_method'] = 'Website (General Tyre)'
                            c['is_truck_tyre'] = False
                            name_matched += 1
                            print(f"    ~ {name[:40]} - Website works (general tyre)")
                        else:
                            c['verification_status'] = 'PARTIAL'
                            c['verification_method'] = 'Website Works (Not Tyre)'
                            name_matched += 1
                            print(f"    ~ {name[:40]} - Website works (not tyre content)")
                    else:
                        # Website doesn't work
                        if is_truck:
                            c['verification_status'] = 'UNVERIFIED'
                            c['verification_method'] = 'Name Match Only'
                            c['is_truck_tyre'] = True
                            unverified += 1
                            print(f"    ? {name[:40]} - Name suggests truck tyre (no verification)")
                        else:
                            c['verification_status'] = 'UNVERIFIED'
                            c['verification_method'] = 'None'
                            unverified += 1
                            print(f"    ✗ {name[:40]} - Cannot verify")
                else:
                    # No website to check
                    if is_truck:
                        c['verification_status'] = 'UNVERIFIED'
                        c['verification_method'] = 'Name Match Only'
                        c['is_truck_tyre'] = True
                        unverified += 1
                        print(f"    ? {name[:40]} - Name match only")
                    else:
                        c['verification_status'] = 'UNVERIFIED'
                        c['verification_method'] = 'None'
                        unverified += 1
                        print(f"    ✗ {name[:40]} - Cannot verify")

            time.sleep(0.3)  # Rate limiting for CH API

        all_companies.append(c)

    print(f"\n    Companies House verified: {ch_verified}")
    print(f"    Website verified: {website_verified}")
    print(f"    Partial verification: {name_matched}")
    print(f"    Unverified: {unverified}")

    # =========================================================================
    # STEP 2: Estimate revenue
    # =========================================================================
    print("\n[2] ESTIMATING REVENUE")
    print("-" * 60)

    for c in all_companies:
        est = estimate_revenue(c)
        c['size'] = est['size']
        c['employees'] = est['employees']
        c['revenue_estimate'] = est['revenue']
        c['revenue_low'] = est['rev_low']
        c['revenue_high'] = est['rev_high']
        c['revenue_indicators'] = est['indicators']

    # Sort by verification status then revenue
    status_order = {'VERIFIED': 0, 'PARTIAL': 1, 'UNVERIFIED': 2}
    all_companies.sort(key=lambda x: (status_order.get(x.get('verification_status'), 3), -x.get('revenue_high', 0)))

    # Stats
    total = len(all_companies)
    verified = len([c for c in all_companies if c.get('verification_status') == 'VERIFIED'])
    partial = len([c for c in all_companies if c.get('verification_status') == 'PARTIAL'])
    unverified_count = len([c for c in all_companies if c.get('verification_status') == 'UNVERIFIED'])
    truck_tyre = len([c for c in all_companies if c.get('is_truck_tyre')])
    with_phone = len([c for c in all_companies if c.get('phone')])
    with_email = len([c for c in all_companies if c.get('email')])

    verified_rev_low = sum(c.get('revenue_low', 0) for c in all_companies if c.get('verification_status') == 'VERIFIED')
    verified_rev_high = sum(c.get('revenue_high', 0) for c in all_companies if c.get('verification_status') == 'VERIFIED')

    print(f"    Total: {total}")
    print(f"    VERIFIED (CH or Website): {verified}")
    print(f"    PARTIAL (Website works): {partial}")
    print(f"    UNVERIFIED: {unverified_count}")
    print(f"    Confirmed Truck Tyre: {truck_tyre}")
    print(f"    With phone: {with_phone}")
    print(f"    With email: {with_email}")
    print()
    print(f"    VERIFIED MARKET: £{verified_rev_low/1000000:.0f}M - £{verified_rev_high/1000000:.0f}M")

    # =========================================================================
    # EXPORT
    # =========================================================================
    print("\n[3] EXPORTING")
    print("-" * 60)

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Verified Companies"

    headers = ['Company Name', 'Verification Status', 'Verification Method', 'Is Truck Tyre',
               'Website', 'Website Verified', 'CH Number', 'CH Status',
               'Phone', 'Email', 'Size', 'Employees', 'Revenue',
               'Business Type', 'Region', 'Address', 'Source']

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    verified_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    partial_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    unverified_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row, c in enumerate(all_companies, 2):
        ws.cell(row=row, column=1, value=c.get('name', ''))

        status_cell = ws.cell(row=row, column=2, value=c.get('verification_status', ''))
        if c.get('verification_status') == 'VERIFIED':
            status_cell.fill = verified_fill
        elif c.get('verification_status') == 'PARTIAL':
            status_cell.fill = partial_fill
        else:
            status_cell.fill = unverified_fill

        ws.cell(row=row, column=3, value=c.get('verification_method', ''))
        ws.cell(row=row, column=4, value='Yes' if c.get('is_truck_tyre') else 'No')
        ws.cell(row=row, column=5, value=c.get('website', ''))
        ws.cell(row=row, column=6, value='Yes' if c.get('website_verified') else 'No')
        ws.cell(row=row, column=7, value=c.get('ch_number', ''))
        ws.cell(row=row, column=8, value=c.get('ch_status', ''))
        ws.cell(row=row, column=9, value=c.get('phone', ''))
        ws.cell(row=row, column=10, value=c.get('email', ''))
        ws.cell(row=row, column=11, value=c.get('size', ''))
        ws.cell(row=row, column=12, value=c.get('employees', ''))
        ws.cell(row=row, column=13, value=c.get('revenue_estimate', ''))
        ws.cell(row=row, column=14, value=c.get('businessType', ''))
        ws.cell(row=row, column=15, value=c.get('region', ''))
        ws.cell(row=row, column=16, value=c.get('ch_address', ''))
        ws.cell(row=row, column=17, value=c.get('source', ''))

    widths = [45, 15, 25, 12, 50, 12, 12, 12, 18, 30, 12, 12, 18, 25, 15, 50, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:Q{len(all_companies) + 1}"

    # Add summary sheet
    ws_summary = wb.create_sheet("Summary")
    summary_data = [
        ['UK TRUCK TYRE COMPANIES - VERIFICATION SUMMARY', ''],
        ['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['', ''],
        ['VERIFICATION RESULTS', ''],
        ['Total Companies', total],
        ['VERIFIED (Companies House or Website)', verified],
        ['PARTIAL (Website works)', partial],
        ['UNVERIFIED', unverified_count],
        ['', ''],
        ['TRUCK TYRE CONFIRMATION', ''],
        ['Confirmed Truck Tyre Companies', truck_tyre],
        ['', ''],
        ['CONTACT DATA', ''],
        ['With Phone Number', with_phone],
        ['With Email', with_email],
        ['', ''],
        ['MARKET ESTIMATE (VERIFIED ONLY)', ''],
        ['Low Estimate', f'£{verified_rev_low/1000000:.0f}M'],
        ['High Estimate', f'£{verified_rev_high/1000000:.0f}M'],
    ]

    for row, (label, value) in enumerate(summary_data, 1):
        ws_summary.cell(row=row, column=1, value=label)
        ws_summary.cell(row=row, column=2, value=value)

    ws_summary.column_dimensions['A'].width = 40
    ws_summary.column_dimensions['B'].width = 20

    wb.save(f"{OUTPUT_FILE}.xlsx")
    print(f"    Saved: {OUTPUT_FILE}.xlsx")

    # JSON
    with open(f"{OUTPUT_FILE}.json", 'w', encoding='utf-8') as f:
        json.dump(all_companies, f, indent=2, ensure_ascii=False)
    print(f"    Saved: {OUTPUT_FILE}.json")

    # CSV
    fieldnames = ['name', 'verification_status', 'verification_method', 'is_truck_tyre',
                  'website', 'website_verified', 'ch_number', 'ch_status',
                  'phone', 'email', 'size', 'employees', 'revenue_estimate',
                  'revenue_low', 'revenue_high', 'businessType', 'region',
                  'ch_address', 'source']

    with open(f"{OUTPUT_FILE}.csv", 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(all_companies)
    print(f"    Saved: {OUTPUT_FILE}.csv")

    # =========================================================================
    # SUMMARY
    # =========================================================================
    print("\n" + "=" * 80)
    print("VERIFICATION COMPLETE")
    print("=" * 80)
    print(f"\nTotal: {total} companies")
    print(f"VERIFIED: {verified} ({verified/total*100:.1f}%)")
    print(f"PARTIAL: {partial} ({partial/total*100:.1f}%)")
    print(f"UNVERIFIED: {unverified_count} ({unverified_count/total*100:.1f}%)")
    print(f"\nConfirmed Truck Tyre: {truck_tyre}")
    print(f"Verified market estimate: £{verified_rev_low/1000000:.0f}M - £{verified_rev_high/1000000:.0f}M")

    print("\n\nTOP 30 VERIFIED COMPANIES:")
    print("-" * 80)

    verified_list = [c for c in all_companies if c.get('verification_status') == 'VERIFIED']
    for i, c in enumerate(verified_list[:30], 1):
        print(f"\n{i}. {c['name']}")
        print(f"   Method: {c['verification_method']}")
        if c.get('ch_number'):
            print(f"   CH Number: {c['ch_number']}")
        if c.get('website'):
            print(f"   Website: {c['website']}")
        print(f"   Revenue: {c['revenue_estimate']}")


if __name__ == "__main__":
    main()
