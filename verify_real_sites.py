#!/usr/bin/env python3
"""
=============================================================================
VERIFY REAL UK TRUCK TYRE WEBSITES
=============================================================================
Tests a list of potential truck tyre websites to see which ones actually work.
Only includes websites that pass verification.

Date: January 2026
=============================================================================
"""

import requests
from bs4 import BeautifulSoup
import json
import time
import re
import csv
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
    'Accept': 'text/html,application/xhtml+xml',
}

CH_API_KEY = '48d17266-ff2e-425f-9b20-7dcc9b25bb79'
OUTPUT_FILE = 'uk_truck_tyres_FINAL_VERIFIED'


# =============================================================================
# POTENTIAL WEBSITES TO TEST
# These will be verified - only working ones will be included
# =============================================================================

WEBSITES_TO_TEST = [
    # Major tyre networks
    ('ATS Euromaster', 'https://www.atseuromaster.co.uk'),
    ('National Tyres', 'https://www.national.co.uk'),
    ('Kwik Fit', 'https://www.kwik-fit.com'),
    ('Halfords', 'https://www.halfords.com'),
    ('Protyre', 'https://www.protyre.co.uk'),
    ('Tructyre ATS', 'https://www.tructyre.co.uk'),
    ('Bandvulc', 'https://www.bandvulc.co.uk'),
    ('Bush Tyres', 'https://www.bushtyres.co.uk'),
    ('Watling Tyres', 'https://www.watlingtyres.co.uk'),
    ('Hometyre', 'https://www.hometyre.co.uk'),
    ('Merityre', 'https://www.merityre.co.uk'),
    ('Point S UK', 'https://www.point-s.co.uk'),
    ('Lodge Tyre', 'https://www.lodgetyre.com'),
    ('Kingsway Tyres', 'https://www.kingswaytyres.co.uk'),
    ('McConechy Tyres', 'https://www.mcconechys.co.uk'),

    # Wholesalers
    ('Stapleton Tyres', 'https://www.stapleton-tyres.co.uk'),
    ('Kirkby Tyres', 'https://www.kirkbytyres.co.uk'),
    ('Bond International', 'https://www.bondint.com'),
    ('Micheldever Tyres', 'https://www.micheldever.co.uk'),
    ('Deldo Tyres', 'https://www.deldo.co.uk'),

    # Retreaders
    ('Vacu-Lug', 'https://www.vaculug.com'),
    ('Colway Tyres', 'https://www.colway.co.uk'),

    # Manufacturers UK sites
    ('Michelin UK', 'https://www.michelin.co.uk'),
    ('Bridgestone UK', 'https://www.bridgestone.co.uk'),
    ('Continental UK', 'https://www.continental-tyres.co.uk'),
    ('Goodyear UK', 'https://www.goodyear.eu/en_gb'),
    ('Pirelli UK', 'https://www.pirelli.com/tyres/en-gb'),
    ('Hankook UK', 'https://www.hankooktire.com/uk'),
    ('Yokohama UK', 'https://www.yokohama.co.uk'),
    ('Kumho UK', 'https://www.kumhotyre.co.uk'),
    ('Toyo UK', 'https://www.toyo.co.uk'),
    ('Falken UK', 'https://www.falkentyre.com'),
    ('Dunlop UK', 'https://www.dunlop.eu'),
    ('BFGoodrich UK', 'https://www.bfgoodrich.co.uk'),

    # Mobile/24hr
    ('Tyrenet', 'https://tyrenet.net'),
    ('24hr Mr Tyre', 'https://www.24hrmrtyre.co.uk'),
    ('Tyre Assist 365', 'https://www.tyreassist365.com'),
    ('Mobile Tyre Fitting UK', 'https://www.mobiletyrefittinguk.co.uk'),

    # Truck specific
    ('Big Tyres', 'https://www.bigtyres.co.uk'),
    ('Fleet Tyre Group', 'https://www.fleet-tyres.co.uk'),
    ('Fleetline Tyres', 'https://www.fleetlinetyres.co.uk'),
    ('Venson Automotive', 'https://www.venson.com'),

    # Regional networks
    ('Malvern Tyres', 'https://www.malverntyres.co.uk'),
    ('Just Tyres', 'https://www.justtyres.co.uk'),
    ('Blackcircles', 'https://www.blackcircles.com'),
    ('Tyre Shopper', 'https://www.tyreshopper.co.uk'),
    ('mytyres', 'https://www.mytyres.co.uk'),
    ('Oponeo', 'https://www.oponeo.co.uk'),
    ('Tyres on the Drive', 'https://www.tyresonthedrive.com'),
    ('Event Tyres', 'https://www.eventtyres.co.uk'),
    ('Asda Tyres', 'https://tyres.asda.com'),
    ('Costco Tyres', 'https://www.costco.co.uk/tyres'),

    # Fleet/Commercial specialists
    ('Fleet Operations', 'https://www.fleetoperations.co.uk'),
    ('Zenith Vehicle Contracts', 'https://www.zenith.co.uk'),
    ('Arval', 'https://www.arval.co.uk'),
    ('Lex Autolease', 'https://www.lexautolease.co.uk'),

    # Independent regional
    ('Formula One Autocentres', 'https://www.f1autocentres.co.uk'),
    ('Setyres', 'https://www.setyres.co.uk'),
    ('HiQ Tyres', 'https://www.hiq.co.uk'),
    ('Tyre Pros', 'https://www.tyrepros.co.uk'),
    ('First Stop', 'https://www.firststop.co.uk'),
    ('ETB Autocentres', 'https://www.etbautocentres.com'),

    # More truck/HGV specific
    ('247 Mobile Truck Tyres', 'https://www.247mobiletrucktyres.co.uk'),
    ('HGV Direct', 'https://www.hgvdirect.co.uk'),
    ('Truck Tyres Direct', 'https://www.trucktyresdirect.co.uk'),
    ('Emergency Tyres', 'https://www.emergencytyres.co.uk'),
]


def verify_website(name, url):
    """Test if website works and extract info"""
    try:
        r = requests.get(url, headers=HEADERS, timeout=12, allow_redirects=True)

        if r.status_code == 200:
            text = r.text
            text_lower = text.lower()

            result = {
                'name': name,
                'website': r.url,
                'website_verified': True,
                'phone': None,
                'email': None,
                'services': [],
                'is_truck_tyre': False
            }

            # Check if tyre related
            tyre_words = ['tyre', 'tire', 'wheel', 'fitting']
            truck_words = ['truck', 'hgv', 'lorry', 'commercial', 'fleet', 'trailer']

            has_tyre = any(w in text_lower for w in tyre_words)
            has_truck = any(w in text_lower for w in truck_words)

            result['is_truck_tyre'] = has_tyre and has_truck
            result['is_tyre'] = has_tyre

            # Extract phone
            phone_patterns = [
                r'0800[\s\-]?\d{3}[\s\-]?\d{3,4}',
                r'0808[\s\-]?\d{3}[\s\-]?\d{3,4}',
                r'01\d{2,3}[\s\-]?\d{3}[\s\-]?\d{3,4}',
                r'02\d[\s\-]?\d{4}[\s\-]?\d{4}',
                r'03\d{2}[\s\-]?\d{3}[\s\-]?\d{4}',
            ]
            for pattern in phone_patterns:
                matches = re.findall(pattern, text)
                for m in matches:
                    cleaned = re.sub(r'[\s\-]', '', m)
                    if 10 <= len(cleaned) <= 12 and not cleaned.startswith('00000'):
                        result['phone'] = m.strip()
                        break
                if result['phone']:
                    break

            # Extract email
            emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', text)
            for email in emails:
                if not any(x in email.lower() for x in ['.png', '.jpg', '.css', '.js', 'example.', 'wix.', 'sentry']):
                    result['email'] = email
                    break

            # Services
            if '24' in text_lower and ('hour' in text_lower or '/7' in text_lower):
                result['services'].append('24hr')
            if 'mobile' in text_lower and 'fitting' in text_lower:
                result['services'].append('Mobile')
            if 'fleet' in text_lower:
                result['services'].append('Fleet')
            if 'commercial' in text_lower:
                result['services'].append('Commercial')

            return result

        return None
    except Exception as e:
        return None


def search_companies_house(term):
    """Search Companies House"""
    companies = []
    try:
        url = 'https://api.company-information.service.gov.uk/search/companies'
        params = {'q': term, 'items_per_page': 100}
        r = requests.get(url, auth=(CH_API_KEY, ''), params=params, timeout=15)

        if r.status_code == 200:
            for item in r.json().get('items', []):
                if item.get('company_status') == 'active':
                    companies.append({
                        'name': item.get('title', ''),
                        'company_number': item.get('company_number', ''),
                        'address': item.get('address_snippet', ''),
                        'date_created': item.get('date_of_creation', ''),
                        'source': 'Companies House'
                    })
    except:
        pass
    return companies


def estimate_revenue(company):
    """Estimate revenue"""
    name = company.get('name', '').lower()
    services = company.get('services', '')
    is_truck = company.get('is_truck_tyre', False)
    verified = company.get('website_verified', False)

    score = 0
    indicators = []

    if any(x in name for x in ['national', 'uk', 'british', 'group', 'network']):
        score += 40
        indicators.append('National')
    if any(x in name for x in ['wholesale', 'supply', 'distribution']):
        score += 30
        indicators.append('Wholesale')
    if 'manufacturer' in name or any(x in name.lower() for x in ['michelin', 'bridgestone', 'continental', 'goodyear', 'pirelli']):
        score += 50
        indicators.append('Manufacturer')

    if '24hr' in str(services):
        score += 10
        indicators.append('24hr')
    if 'Fleet' in str(services):
        score += 20
        indicators.append('Fleet')
    if 'Commercial' in str(services):
        score += 15
        indicators.append('Commercial')

    if verified:
        score += 15
    if is_truck:
        score += 10

    if score >= 70:
        return {'size': 'Large', 'employees': '100-1000+', 'revenue': '£20M - £500M+',
                'rev_low': 20000000, 'rev_high': 500000000, 'indicators': ', '.join(indicators) or 'Major'}
    elif score >= 50:
        return {'size': 'Medium-Large', 'employees': '50-200', 'revenue': '£5M - £50M',
                'rev_low': 5000000, 'rev_high': 50000000, 'indicators': ', '.join(indicators) or 'Regional'}
    elif score >= 30:
        return {'size': 'Medium', 'employees': '15-50', 'revenue': '£1M - £10M',
                'rev_low': 1000000, 'rev_high': 10000000, 'indicators': ', '.join(indicators) or 'Regional'}
    elif score >= 15:
        return {'size': 'Small-Medium', 'employees': '5-20', 'revenue': '£300K - £2M',
                'rev_low': 300000, 'rev_high': 2000000, 'indicators': ', '.join(indicators) or 'Local'}
    else:
        return {'size': 'Small', 'employees': '1-10', 'revenue': '£50K - £500K',
                'rev_low': 50000, 'rev_high': 500000, 'indicators': ', '.join(indicators) or 'Small'}


def main():
    print("=" * 80)
    print("VERIFY REAL UK TRUCK TYRE WEBSITES")
    print("=" * 80)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    print(f"Testing {len(WEBSITES_TO_TEST)} potential websites...")
    print()

    verified_companies = []

    # =========================================================================
    # STEP 1: Test each website
    # =========================================================================
    print("[1] VERIFYING WEBSITES")
    print("-" * 60)

    for i, (name, url) in enumerate(WEBSITES_TO_TEST):
        print(f"    [{i+1}/{len(WEBSITES_TO_TEST)}] {name}...", end=' ', flush=True)

        result = verify_website(name, url)

        if result and result['website_verified']:
            verified_companies.append(result)

            truck_mark = '✓' if result.get('is_truck_tyre') else '○'
            phone = result.get('phone') or 'No phone'
            print(f"✓ WORKING [{truck_mark}] | {phone[:15]}")
        else:
            print("✗ Not working")

        time.sleep(0.5)

    print(f"\n    Verified: {len(verified_companies)}/{len(WEBSITES_TO_TEST)}")

    # =========================================================================
    # STEP 2: Add Companies House data
    # =========================================================================
    print("\n[2] ADDING COMPANIES HOUSE DATA")
    print("-" * 60)

    ch_terms = ['truck tyre', 'hgv tyre', 'commercial tyre', 'fleet tyre']
    seen_numbers = set()
    seen_names = set(c['name'].lower() for c in verified_companies)

    for term in ch_terms:
        print(f"    Searching: '{term}'...", end=' ', flush=True)
        results = search_companies_house(term)

        added = 0
        for c in results:
            cn = c['company_number']
            name_lower = c['name'].lower()

            has_truck = any(t in name_lower for t in ['truck', 'hgv', 'commercial', 'lorry', 'fleet'])
            has_tyre = 'tyre' in name_lower

            if has_truck and has_tyre and cn not in seen_numbers:
                if not any(n in name_lower for n in seen_names):
                    seen_numbers.add(cn)
                    seen_names.add(name_lower)
                    c['website'] = ''
                    c['website_verified'] = False
                    c['is_truck_tyre'] = True
                    c['phone'] = ''
                    c['email'] = ''
                    c['services'] = ''
                    verified_companies.append(c)
                    added += 1

        print(f"Added {added}")
        time.sleep(0.5)

    print(f"\n    Total companies: {len(verified_companies)}")

    # =========================================================================
    # STEP 3: Estimate revenue
    # =========================================================================
    print("\n[3] ESTIMATING REVENUE")
    print("-" * 60)

    for c in verified_companies:
        est = estimate_revenue(c)
        c['size'] = est['size']
        c['employees'] = est['employees']
        c['revenue_estimate'] = est['revenue']
        c['revenue_low'] = est['rev_low']
        c['revenue_high'] = est['rev_high']
        c['revenue_indicators'] = est['indicators']
        c['services'] = ', '.join(c.get('services', [])) if isinstance(c.get('services'), list) else c.get('services', '')

    verified_companies.sort(key=lambda x: x.get('revenue_high', 0), reverse=True)

    # Stats
    total = len(verified_companies)
    with_website = len([c for c in verified_companies if c.get('website')])
    verified = len([c for c in verified_companies if c.get('website_verified')])
    truck_related = len([c for c in verified_companies if c.get('is_truck_tyre')])
    with_phone = len([c for c in verified_companies if c.get('phone')])

    total_rev_low = sum(c.get('revenue_low', 0) for c in verified_companies)
    total_rev_high = sum(c.get('revenue_high', 0) for c in verified_companies)

    print(f"    Total: {total}")
    print(f"    With verified website: {verified}")
    print(f"    Truck/Commercial tyre: {truck_related}")
    print(f"    With phone: {with_phone}")
    print(f"    Market: £{total_rev_low/1000000:.0f}M - £{total_rev_high/1000000:.0f}M")

    # =========================================================================
    # EXPORT
    # =========================================================================
    print("\n[4] EXPORTING")
    print("-" * 60)

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "UK Truck Tyre Companies"

    headers = ['Company Name', 'Website', 'Verified', 'Truck/HGV', 'Phone', 'Email',
               'Size', 'Employees', 'Revenue', 'Indicators', 'Services',
               'Address', 'Company Number', 'Source']

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    verified_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    truck_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row, c in enumerate(verified_companies, 2):
        ws.cell(row=row, column=1, value=c.get('name', ''))
        ws.cell(row=row, column=2, value=c.get('website', ''))

        v_cell = ws.cell(row=row, column=3, value='Yes' if c.get('website_verified') else 'No')
        if c.get('website_verified'):
            v_cell.fill = verified_fill

        t_cell = ws.cell(row=row, column=4, value='Yes' if c.get('is_truck_tyre') else 'No')
        if c.get('is_truck_tyre'):
            t_cell.fill = truck_fill

        ws.cell(row=row, column=5, value=c.get('phone', ''))
        ws.cell(row=row, column=6, value=c.get('email', ''))
        ws.cell(row=row, column=7, value=c.get('size', ''))
        ws.cell(row=row, column=8, value=c.get('employees', ''))
        ws.cell(row=row, column=9, value=c.get('revenue_estimate', ''))
        ws.cell(row=row, column=10, value=c.get('revenue_indicators', ''))
        ws.cell(row=row, column=11, value=c.get('services', ''))
        ws.cell(row=row, column=12, value=c.get('address', ''))
        ws.cell(row=row, column=13, value=c.get('company_number', ''))
        ws.cell(row=row, column=14, value=c.get('source', 'Website Verified'))

    widths = [35, 45, 10, 10, 18, 30, 15, 12, 20, 25, 30, 45, 15, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:N{len(verified_companies) + 1}"

    wb.save(f"{OUTPUT_FILE}.xlsx")
    print(f"    Saved: {OUTPUT_FILE}.xlsx")

    # JSON
    with open(f"{OUTPUT_FILE}.json", 'w', encoding='utf-8') as f:
        json.dump(verified_companies, f, indent=2, ensure_ascii=False)
    print(f"    Saved: {OUTPUT_FILE}.json")

    # CSV
    fieldnames = ['name', 'website', 'website_verified', 'is_truck_tyre', 'phone', 'email',
                  'size', 'employees', 'revenue_estimate', 'revenue_low', 'revenue_high',
                  'revenue_indicators', 'services', 'address', 'company_number', 'source']

    with open(f"{OUTPUT_FILE}.csv", 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(verified_companies)
    print(f"    Saved: {OUTPUT_FILE}.csv")

    # =========================================================================
    # SUMMARY
    # =========================================================================
    print("\n" + "=" * 80)
    print("COMPLETE")
    print("=" * 80)
    print(f"\nTotal: {total} companies")
    print(f"Verified websites: {verified}")
    print(f"Truck/HGV related: {truck_related}")
    print(f"Market estimate: £{total_rev_low/1000000:.0f}M - £{total_rev_high/1000000:.0f}M")

    print("\n\nVERIFIED COMPANIES WITH WORKING WEBSITES:")
    print("-" * 80)

    working = [c for c in verified_companies if c.get('website_verified')]
    for i, c in enumerate(working[:30], 1):
        truck_mark = '✓' if c.get('is_truck_tyre') else '○'
        print(f"\n{i}. [{truck_mark}] {c['name']}")
        print(f"   Website: {c['website']}")
        if c.get('phone'):
            print(f"   Phone: {c['phone']}")
        print(f"   Size: {c['size']} | Revenue: {c['revenue_estimate']}")


if __name__ == "__main__":
    main()
