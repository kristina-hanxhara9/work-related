"""
UK TRUCK TYRE COMPANIES - AI BUSINESS INTELLIGENCE SCRAPER
===========================================================
Uses web scraping + Gemini AI to estimate:
- Revenue/turnover estimates
- Market share estimates
- Products and services offered
- Company size indicators
- Web traffic estimates
- Business descriptions

Run: python ai_business_intel_scraper.py
"""

import requests
import json
import csv
import time
import re
import sys
from datetime import datetime
from urllib.parse import urlparse, urljoin
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
import google.generativeai as genai

# ============================================================================
# CONFIGURATION
# ============================================================================
GEMINI_API_KEY = 'AIzaSyCmMw66NyVY9jPiRLTWdqhvNFfv5CSyUGc'
CH_API_KEY = '48d17266-ff2e-425f-9b20-7dcc9b25bb79'
CH_BASE_URL = 'https://api.company-information.service.gov.uk'

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-GB,en;q=0.9',
}

# Configure Gemini
genai.configure(api_key=GEMINI_API_KEY)
model = genai.GenerativeModel('gemini-1.5-flash')

# ============================================================================
# IMPORT THE 846 COMPANIES
# ============================================================================
try:
    from scraper import INDUSTRY_DATABASE
    print(f"✓ Loaded {len(INDUSTRY_DATABASE)} companies from scraper.py")
except ImportError:
    print("ERROR: Could not import INDUSTRY_DATABASE from scraper.py")
    sys.exit(1)

# ============================================================================
# WEB SCRAPING FUNCTIONS
# ============================================================================
def scrape_website_content(url, max_pages=3):
    """Scrape website content for AI analysis"""
    if not url or not url.startswith('http'):
        return None

    content_parts = []
    pages_scraped = 0

    try:
        # Main page
        r = requests.get(url, headers=HEADERS, timeout=15, allow_redirects=True)
        if r.status_code == 200:
            content_parts.append(extract_text_content(r.text))
            pages_scraped += 1

        # Try common pages
        common_pages = ['/about', '/about-us', '/services', '/products', '/contact', '/our-services']
        for page in common_pages:
            if pages_scraped >= max_pages:
                break
            try:
                page_url = urljoin(url, page)
                r = requests.get(page_url, headers=HEADERS, timeout=10)
                if r.status_code == 200:
                    content_parts.append(extract_text_content(r.text))
                    pages_scraped += 1
            except:
                pass
            time.sleep(0.5)

        return '\n\n'.join(content_parts)[:15000]  # Limit content size

    except Exception as e:
        return None

def extract_text_content(html):
    """Extract meaningful text from HTML"""
    # Remove scripts, styles, etc.
    import re

    # Remove script and style content
    html = re.sub(r'<script[^>]*>.*?</script>', '', html, flags=re.DOTALL | re.IGNORECASE)
    html = re.sub(r'<style[^>]*>.*?</style>', '', html, flags=re.DOTALL | re.IGNORECASE)
    html = re.sub(r'<nav[^>]*>.*?</nav>', '', html, flags=re.DOTALL | re.IGNORECASE)
    html = re.sub(r'<footer[^>]*>.*?</footer>', '', html, flags=re.DOTALL | re.IGNORECASE)

    # Remove HTML tags
    text = re.sub(r'<[^>]+>', ' ', html)

    # Clean up whitespace
    text = re.sub(r'\s+', ' ', text)
    text = re.sub(r'\n\s*\n', '\n', text)

    return text.strip()[:5000]

def extract_basic_info(html, url):
    """Extract basic info directly from HTML"""
    info = {
        'phones': [],
        'emails': [],
        'addresses': [],
        'social_media': [],
    }

    # Phone numbers
    phone_pattern = r'(?:tel:|phone:|call us:?)?\s*(\+?44[\s\-\.]?\d{2,4}[\s\-\.]?\d{3,4}[\s\-\.]?\d{3,4}|\b0\d{2,4}[\s\-\.]?\d{3,4}[\s\-\.]?\d{3,4}\b)'
    phones = re.findall(phone_pattern, html, re.IGNORECASE)
    info['phones'] = list(set([p.strip() for p in phones if len(re.sub(r'\D', '', p)) >= 10]))[:3]

    # Emails
    email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
    emails = re.findall(email_pattern, html)
    info['emails'] = list(set([e for e in emails if not e.endswith('.png') and not e.endswith('.jpg')]))[:3]

    # Social media
    social_patterns = [
        r'(facebook\.com/[a-zA-Z0-9._-]+)',
        r'(twitter\.com/[a-zA-Z0-9._-]+)',
        r'(linkedin\.com/company/[a-zA-Z0-9._-]+)',
        r'(instagram\.com/[a-zA-Z0-9._-]+)',
    ]
    for pattern in social_patterns:
        matches = re.findall(pattern, html, re.IGNORECASE)
        if matches:
            info['social_media'].append(matches[0])

    return info

# ============================================================================
# GEMINI AI ANALYSIS
# ============================================================================
def analyze_company_with_ai(company_name, website_content, company_data):
    """Use Gemini AI to analyze company and estimate business metrics"""

    prompt = f"""Analyze this UK truck/commercial tyre company and provide business intelligence estimates.

COMPANY: {company_name}
BUSINESS TYPE: {company_data.get('businessType', 'Unknown')}
REGION: {company_data.get('region', 'Unknown')}
SERVICE POINTS: {company_data.get('servicePoints', 'Unknown')}

WEBSITE CONTENT:
{website_content[:8000] if website_content else 'No website content available'}

Based on this information, provide estimates in the following JSON format ONLY (no other text):
{{
    "estimated_revenue": "e.g., £5M-10M, £50M+, or Unknown if no indicators",
    "revenue_confidence": "Low/Medium/High",
    "revenue_reasoning": "Brief explanation of how you estimated revenue",
    "estimated_employees": "e.g., 10-20, 50-100, 500+, or Unknown",
    "employee_confidence": "Low/Medium/High",
    "market_position": "e.g., Local independent, Regional player, National chain, Market leader",
    "estimated_market_share": "e.g., <1%, 1-5%, 5-10%, 10%+, or Unknown for UK truck tyre market",
    "products_services": ["list", "of", "main", "products", "or", "services"],
    "target_customers": "e.g., Haulage companies, Fleet operators, Owner-drivers, Retailers",
    "business_model": "e.g., B2B Wholesaler, Retail fitter, Mobile service, Manufacturer",
    "competitive_advantages": ["list", "of", "advantages", "if", "apparent"],
    "website_quality": "Poor/Basic/Good/Professional/Enterprise",
    "digital_presence": "Minimal/Basic/Moderate/Strong",
    "growth_indicators": "Declining/Stable/Growing/Rapidly growing/Unknown",
    "key_brands": ["tyre", "brands", "they", "sell", "or", "represent"],
    "geographic_coverage": "Local/Regional/National/International",
    "company_description": "2-3 sentence description of what this company does"
}}

IMPORTANT:
- Base estimates on real indicators from the content (branch counts, fleet size, staff mentioned, etc.)
- For companies with no website content, use industry averages for their type
- Be conservative with estimates when uncertain
- Return ONLY valid JSON, no other text"""

    try:
        response = model.generate_content(prompt)
        response_text = response.text.strip()

        # Clean up response - extract JSON
        if '```json' in response_text:
            response_text = response_text.split('```json')[1].split('```')[0]
        elif '```' in response_text:
            response_text = response_text.split('```')[1].split('```')[0]

        # Parse JSON
        result = json.loads(response_text)
        return result

    except Exception as e:
        print(f" [AI Error: {str(e)[:30]}]", end="")
        return None

# ============================================================================
# COMPANIES HOUSE DATA
# ============================================================================
def get_ch_data(company_number):
    """Get basic Companies House data"""
    if not company_number:
        return {}

    try:
        r = requests.get(f"{CH_BASE_URL}/company/{company_number}",
                        auth=(CH_API_KEY, ''), timeout=15)
        if r.status_code == 200:
            data = r.json()
            return {
                'ch_status': data.get('company_status', ''),
                'ch_type': data.get('type', ''),
                'ch_created': data.get('date_of_creation', ''),
                'ch_sic_codes': ', '.join(data.get('sic_codes', [])),
                'ch_accounts_type': data.get('accounts', {}).get('last_accounts', {}).get('type', ''),
            }
    except:
        pass
    return {}

# ============================================================================
# MAIN PROCESSING
# ============================================================================
def process_all_companies():
    """Process all 846 companies with AI analysis"""

    print("=" * 70)
    print("UK TRUCK TYRE COMPANIES - AI BUSINESS INTELLIGENCE")
    print("=" * 70)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Total companies: {len(INDUSTRY_DATABASE)}")
    print("Using: Web scraping + Gemini AI for analysis")
    print("=" * 70)
    print()

    results = []
    stats = {
        'total': len(INDUSTRY_DATABASE),
        'with_website': 0,
        'with_ai_analysis': 0,
        'with_revenue_estimate': 0,
    }

    for i, company in enumerate(INDUSTRY_DATABASE, 1):
        company_name = company.get('name', '')
        company_number = company.get('companyNumber', '').strip()
        website = company.get('website', '')

        print(f"[{i}/{len(INDUSTRY_DATABASE)}] {company_name[:40]}...", end=" ")

        result = {
            # Basic info
            'company_name': company_name,
            'company_number': company_number,
            'website': website,
            'original_address': company.get('address', ''),
            'business_type': company.get('businessType', ''),
            'is_b2b_wholesaler': company.get('isB2BWholesaler', ''),
            'region': company.get('region', ''),

            # Companies House data
            'ch_status': '',
            'ch_type': '',
            'ch_created': '',
            'ch_sic_codes': '',
            'ch_accounts_type': '',

            # Scraped data
            'phones': '',
            'emails': '',
            'social_media': '',

            # AI estimates
            'estimated_revenue': '',
            'revenue_confidence': '',
            'revenue_reasoning': '',
            'estimated_employees': '',
            'employee_confidence': '',
            'market_position': '',
            'estimated_market_share': '',
            'products_services': '',
            'target_customers': '',
            'business_model': '',
            'competitive_advantages': '',
            'website_quality': '',
            'digital_presence': '',
            'growth_indicators': '',
            'key_brands': '',
            'geographic_coverage': '',
            'company_description': '',
        }

        # 1. Get Companies House data
        if company_number:
            time.sleep(0.5)
            ch_data = get_ch_data(company_number)
            result.update(ch_data)
            if ch_data:
                print("CH", end=" ")

        # 2. Scrape website
        website_content = None
        if website:
            stats['with_website'] += 1
            time.sleep(1)
            website_content = scrape_website_content(website)

            if website_content:
                # Extract basic info
                try:
                    r = requests.get(website, headers=HEADERS, timeout=10)
                    basic_info = extract_basic_info(r.text, website)
                    result['phones'] = ', '.join(basic_info['phones'])
                    result['emails'] = ', '.join(basic_info['emails'])
                    result['social_media'] = ', '.join(basic_info['social_media'])
                except:
                    pass
                print("Web", end=" ")

        # 3. AI Analysis
        time.sleep(1)  # Rate limit for Gemini
        ai_result = analyze_company_with_ai(company_name, website_content, company)

        if ai_result:
            stats['with_ai_analysis'] += 1
            result['estimated_revenue'] = ai_result.get('estimated_revenue', '')
            result['revenue_confidence'] = ai_result.get('revenue_confidence', '')
            result['revenue_reasoning'] = ai_result.get('revenue_reasoning', '')
            result['estimated_employees'] = ai_result.get('estimated_employees', '')
            result['employee_confidence'] = ai_result.get('employee_confidence', '')
            result['market_position'] = ai_result.get('market_position', '')
            result['estimated_market_share'] = ai_result.get('estimated_market_share', '')
            result['products_services'] = ', '.join(ai_result.get('products_services', []))
            result['target_customers'] = ai_result.get('target_customers', '')
            result['business_model'] = ai_result.get('business_model', '')
            result['competitive_advantages'] = ', '.join(ai_result.get('competitive_advantages', []))
            result['website_quality'] = ai_result.get('website_quality', '')
            result['digital_presence'] = ai_result.get('digital_presence', '')
            result['growth_indicators'] = ai_result.get('growth_indicators', '')
            result['key_brands'] = ', '.join(ai_result.get('key_brands', []))
            result['geographic_coverage'] = ai_result.get('geographic_coverage', '')
            result['company_description'] = ai_result.get('company_description', '')

            if result['estimated_revenue'] and result['estimated_revenue'] != 'Unknown':
                stats['with_revenue_estimate'] += 1

            print(f"AI:{result['estimated_revenue'][:15] if result['estimated_revenue'] else 'N/A'}", end=" ")

        print("")
        results.append(result)

        # Progress update
        if i % 25 == 0:
            print(f"\n--- Progress: {i}/{len(INDUSTRY_DATABASE)} ({i*100//len(INDUSTRY_DATABASE)}%) ---")
            print(f"    With AI analysis: {stats['with_ai_analysis']}")
            print(f"    With revenue estimate: {stats['with_revenue_estimate']}\n")

        # Save intermediate results every 100 companies
        if i % 100 == 0:
            save_intermediate_results(results, i)

    print("\n" + "=" * 70)
    print("PROCESSING COMPLETE")
    print("=" * 70)
    print(f"Total companies: {stats['total']}")
    print(f"With websites: {stats['with_website']}")
    print(f"With AI analysis: {stats['with_ai_analysis']}")
    print(f"With revenue estimates: {stats['with_revenue_estimate']}")

    return results, stats

def save_intermediate_results(results, count):
    """Save intermediate results to prevent data loss"""
    with open(f'UK_TRUCK_TYRE_AI_PARTIAL_{count}.json', 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print(f"    [Saved intermediate results: {count} companies]")

# ============================================================================
# EXCEL EXPORT
# ============================================================================
def create_excel_report(results, stats):
    print("\nCreating Excel report...")
    wb = Workbook()

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=9)
    revenue_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ===== SHEET 1: All Companies with AI Analysis =====
    ws1 = wb.active
    ws1.title = "All Companies"

    headers1 = [
        "Company Name", "Estimated Revenue", "Revenue Confidence", "Revenue Reasoning",
        "Estimated Employees", "Market Position", "Market Share",
        "Products & Services", "Target Customers", "Business Model",
        "Key Brands", "Geographic Coverage", "Website Quality",
        "Growth Indicators", "Company Description",
        "Website", "Phones", "Emails", "Region",
        "CH Status", "CH Created", "Company Number"
    ]

    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for row, r in enumerate(results, 2):
        data = [
            r['company_name'], r['estimated_revenue'], r['revenue_confidence'], r['revenue_reasoning'],
            r['estimated_employees'], r['market_position'], r['estimated_market_share'],
            r['products_services'], r['target_customers'], r['business_model'],
            r['key_brands'], r['geographic_coverage'], r['website_quality'],
            r['growth_indicators'], r['company_description'],
            r['website'], r['phones'], r['emails'], r['region'],
            r['ch_status'], r['ch_created'], r['company_number']
        ]
        for col, value in enumerate(data, 1):
            cell = ws1.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if col == 2 and value and value != 'Unknown':  # Revenue column
                cell.fill = revenue_fill

    # Set column widths
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['D'].width = 40
    ws1.column_dimensions['H'].width = 40
    ws1.column_dimensions['O'].width = 50
    ws1.column_dimensions['P'].width = 30

    # ===== SHEET 2: Revenue Rankings =====
    ws2 = wb.create_sheet("Revenue Rankings")

    # Sort by revenue estimate (rough sorting)
    def revenue_sort_key(r):
        rev = r.get('estimated_revenue', '')
        if not rev or rev == 'Unknown':
            return 0
        # Extract numbers
        numbers = re.findall(r'[\d.]+', rev.replace(',', ''))
        if numbers:
            val = float(numbers[-1])  # Get last number
            if 'M' in rev or 'million' in rev.lower():
                val *= 1000000
            elif 'K' in rev or 'k' in rev:
                val *= 1000
            return val
        return 0

    sorted_results = sorted(results, key=revenue_sort_key, reverse=True)

    headers2 = ["Rank", "Company Name", "Estimated Revenue", "Confidence", "Employees",
                "Market Position", "Market Share", "Business Model", "Website"]

    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        cell.font = header_font

    rank = 1
    for r in sorted_results:
        if r['estimated_revenue'] and r['estimated_revenue'] != 'Unknown':
            ws2.cell(row=rank+1, column=1, value=rank)
            ws2.cell(row=rank+1, column=2, value=r['company_name'])
            ws2.cell(row=rank+1, column=3, value=r['estimated_revenue'])
            ws2.cell(row=rank+1, column=4, value=r['revenue_confidence'])
            ws2.cell(row=rank+1, column=5, value=r['estimated_employees'])
            ws2.cell(row=rank+1, column=6, value=r['market_position'])
            ws2.cell(row=rank+1, column=7, value=r['estimated_market_share'])
            ws2.cell(row=rank+1, column=8, value=r['business_model'])
            ws2.cell(row=rank+1, column=9, value=r['website'])
            rank += 1

    ws2.column_dimensions['B'].width = 35
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['I'].width = 35

    # ===== SHEET 3: Market Analysis =====
    ws3 = wb.create_sheet("Market Analysis")

    ws3.cell(row=1, column=1, value="UK TRUCK TYRE MARKET ANALYSIS")
    ws3.cell(row=1, column=1).font = Font(bold=True, size=14)

    # Count by market position
    positions = {}
    for r in results:
        pos = r.get('market_position', 'Unknown')
        positions[pos] = positions.get(pos, 0) + 1

    ws3.cell(row=3, column=1, value="MARKET POSITION DISTRIBUTION")
    ws3.cell(row=3, column=1).font = Font(bold=True)
    row = 4
    for pos, count in sorted(positions.items(), key=lambda x: x[1], reverse=True):
        ws3.cell(row=row, column=1, value=pos)
        ws3.cell(row=row, column=2, value=count)
        row += 1

    # Count by business model
    models = {}
    for r in results:
        model = r.get('business_model', 'Unknown')
        models[model] = models.get(model, 0) + 1

    ws3.cell(row=row+1, column=1, value="BUSINESS MODEL DISTRIBUTION")
    ws3.cell(row=row+1, column=1).font = Font(bold=True)
    row += 2
    for model, count in sorted(models.items(), key=lambda x: x[1], reverse=True):
        ws3.cell(row=row, column=1, value=model)
        ws3.cell(row=row, column=2, value=count)
        row += 1

    ws3.column_dimensions['A'].width = 35
    ws3.column_dimensions['B'].width = 10

    # ===== SHEET 4: Summary =====
    ws4 = wb.create_sheet("Summary")

    ws4.cell(row=1, column=1, value="AI BUSINESS INTELLIGENCE SUMMARY")
    ws4.cell(row=1, column=1).font = Font(bold=True, size=14)

    ws4.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    summary = [
        ("", ""),
        ("DATA COLLECTION", ""),
        ("Total companies", stats['total']),
        ("Companies with websites", stats['with_website']),
        ("Companies with AI analysis", stats['with_ai_analysis']),
        ("Companies with revenue estimates", stats['with_revenue_estimate']),
        ("", ""),
        ("METHODOLOGY", ""),
        ("", "1. Web scraping of company websites (main page + about/services)"),
        ("", "2. Gemini AI analysis of website content"),
        ("", "3. Companies House API for official status"),
        ("", "4. AI estimates based on indicators found"),
        ("", ""),
        ("DATA SOURCES", ""),
        ("Companies House API", "api.company-information.service.gov.uk"),
        ("Company Websites", "Scraped directly"),
        ("AI Analysis", "Google Gemini 1.5 Flash"),
        ("", ""),
        ("NOTES ON ACCURACY", ""),
        ("", "Revenue estimates are AI-generated based on available indicators"),
        ("", "Confidence levels indicate reliability of estimates"),
        ("", "High confidence = multiple clear indicators"),
        ("", "Medium confidence = some indicators present"),
        ("", "Low confidence = limited information available"),
    ]

    row = 5
    for label, value in summary:
        ws4.cell(row=row, column=1, value=label)
        ws4.cell(row=row, column=2, value=value)
        if label and label.isupper():
            ws4.cell(row=row, column=1).font = Font(bold=True)
        row += 1

    ws4.column_dimensions['A'].width = 35
    ws4.column_dimensions['B'].width = 55

    wb.save('UK_TRUCK_TYRE_AI_INTELLIGENCE.xlsx')
    print("✓ Excel saved: UK_TRUCK_TYRE_AI_INTELLIGENCE.xlsx")

def save_json(results):
    with open('UK_TRUCK_TYRE_AI_INTELLIGENCE.json', 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print("✓ JSON saved: UK_TRUCK_TYRE_AI_INTELLIGENCE.json")

def save_csv(results):
    if not results:
        return
    keys = list(results[0].keys())
    with open('UK_TRUCK_TYRE_AI_INTELLIGENCE.csv', 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=keys)
        writer.writeheader()
        writer.writerows(results)
    print("✓ CSV saved: UK_TRUCK_TYRE_AI_INTELLIGENCE.csv")

# ============================================================================
# MAIN
# ============================================================================
if __name__ == "__main__":
    print()
    print("Testing Gemini API connection...")
    try:
        test_response = model.generate_content("Say 'OK' if you can hear me")
        print(f"✓ Gemini API connected: {test_response.text[:20]}...")
    except Exception as e:
        print(f"✗ Gemini API error: {e}")
        print("Continuing without AI analysis...")

    print()
    results, stats = process_all_companies()
    create_excel_report(results, stats)
    save_json(results)
    save_csv(results)

    print("\n" + "=" * 70)
    print("ALL DONE!")
    print("=" * 70)
    print("Files created:")
    print("  - UK_TRUCK_TYRE_AI_INTELLIGENCE.xlsx (4 sheets)")
    print("  - UK_TRUCK_TYRE_AI_INTELLIGENCE.json")
    print("  - UK_TRUCK_TYRE_AI_INTELLIGENCE.csv")
    print(f"\nFinished: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
