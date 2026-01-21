#!/usr/bin/env python3
"""
=============================================================================
REAL UK TRUCK TYRE WEB SCRAPER
=============================================================================
This scraper ONLY uses real websites found through actual web scraping.
NO made-up URLs. Every website is verified to exist and work.

Date: January 2026
=============================================================================
"""

import requests
from bs4 import BeautifulSoup
import json
import time
import re
import csv
from datetime import datetime
from urllib.parse import urljoin, urlparse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'en-GB,en;q=0.9',
}

CH_API_KEY = '48d17266-ff2e-425f-9b20-7dcc9b25bb79'
OUTPUT_FILE = 'uk_truck_tyres_REAL_SCRAPED'


def verify_website(url):
    """Actually visit the website and verify it works and is truck-related"""
    if not url or not url.startswith('http'):
        return None

    try:
        response = requests.get(url, headers=HEADERS, timeout=10, allow_redirects=True)

        if response.status_code == 200:
            html = response.text.lower()

            # Must have truck/HGV/commercial keywords
            truck_keywords = ['truck', 'hgv', 'lorry', 'commercial vehicle', 'fleet', 'trailer']
            tyre_keywords = ['tyre', 'tire', 'wheel', 'fitting']

            has_truck = any(kw in html for kw in truck_keywords)
            has_tyre = any(kw in html for kw in tyre_keywords)

            if not (has_truck and has_tyre):
                return None  # Not a truck tyre site

            result = {
                'status': 'Working',
                'is_truck_tyre': True,
                'phone': None,
                'email': None,
                'services': [],
                'title': ''
            }

            # Get page title
            soup = BeautifulSoup(response.text, 'html.parser')
            title_tag = soup.find('title')
            if title_tag:
                result['title'] = title_tag.get_text(strip=True)[:100]

            # Extract phone - UK formats
            phone_patterns = [
                r'0800[\s\-]?\d{3}[\s\-]?\d{3,4}',
                r'0808[\s\-]?\d{3}[\s\-]?\d{3,4}',
                r'01onal\d{2,3}[\s\-]?\d{3}[\s\-]?\d{3,4}',
                r'02\d[\s\-]?\d{4}[\s\-]?\d{4}',
                r'03\d{2}[\s\-]?\d{3}[\s\-]?\d{4}',
                r'07\d{3}[\s\-]?\d{3}[\s\-]?\d{3}',
            ]
            for pattern in phone_patterns:
                matches = re.findall(pattern, response.text)
                for m in matches:
                    cleaned = re.sub(r'[\s\-]', '', m)
                    if len(cleaned) >= 10 and len(cleaned) <= 12:
                        if not cleaned.startswith('00000'):
                            result['phone'] = m.strip()
                            break
                if result['phone']:
                    break

            # Extract email
            emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', response.text)
            for email in emails:
                email_lower = email.lower()
                # Skip image/css/js files and common fake domains
                if not any(x in email_lower for x in ['.png', '.jpg', '.gif', '.css', '.js',
                                                       'example.com', 'wix', 'sentry', 'cloudflare',
                                                       'wordpress', 'jquery', 'bootstrap']):
                    result['email'] = email
                    break

            # Detect services
            if '24' in html and ('hour' in html or 'hr' in html or '/7' in html):
                result['services'].append('24hr')
            if 'mobile' in html and ('fitting' in html or 'service' in html):
                result['services'].append('Mobile')
            if 'fleet' in html:
                result['services'].append('Fleet')
            if 'breakdown' in html or 'emergency' in html:
                result['services'].append('Emergency')
            if 'retread' in html or 'remould' in html:
                result['services'].append('Retreading')

            return result

        return None

    except Exception as e:
        return None


def scrape_google_results(query):
    """Scrape search results for truck tyre companies"""
    companies = []

    # Use DuckDuckGo HTML (more scraper-friendly)
    try:
        url = f'https://html.duckduckgo.com/html/?q={query.replace(" ", "+")}'
        response = requests.get(url, headers=HEADERS, timeout=15)

        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')

            # Find result links
            results = soup.find_all('a', class_='result__a')

            for result in results[:20]:
                href = result.get('href', '')
                title = result.get_text(strip=True)

                # Skip directories and social media
                skip_domains = ['facebook.com', 'linkedin.com', 'twitter.com', 'instagram.com',
                               'yell.com', 'yelp.com', 'checkatrade.com', 'trustatrader.com',
                               'companieshouse.gov.uk', 'endole.co.uk', 'duedil.com',
                               'youtube.com', 'pinterest.com', 'wikipedia.org',
                               'amazon.co.uk', 'ebay.co.uk', 'gumtree.com']

                if any(skip in href.lower() for skip in skip_domains):
                    continue

                if href.startswith('http') and 'tyre' in (title.lower() + href.lower()):
                    # Extract the actual domain
                    parsed = urlparse(href)
                    domain = f"{parsed.scheme}://{parsed.netloc}/"

                    companies.append({
                        'url': domain,
                        'title': title,
                        'source': 'DuckDuckGo'
                    })

    except Exception as e:
        print(f"      Search error: {e}")

    return companies


def search_companies_house(term):
    """Search Companies House API for active truck tyre companies"""
    companies = []

    try:
        url = 'https://api.company-information.service.gov.uk/search/companies'
        params = {'q': term, 'items_per_page': 100}

        response = requests.get(url, auth=(CH_API_KEY, ''), params=params, timeout=15)

        if response.status_code == 200:
            data = response.json()

            for item in data.get('items', []):
                if item.get('company_status') == 'active':
                    name = item.get('title', '')

                    # Must be truck/HGV related
                    name_lower = name.lower()
                    truck_terms = ['truck tyre', 'truck tyres', 'hgv tyre', 'hgv tyres',
                                  'lorry tyre', 'commercial tyre', 'fleet tyre',
                                  'mobile truck', 'truck & trailer']

                    if any(t in name_lower for t in truck_terms):
                        # Exclude non-tyre businesses
                        exclude = ['truck hire', 'truck rental', 'truck sales', 'trucking',
                                  'truck repair', 'truck parts', 'forklift', 'fork lift']

                        if not any(ex in name_lower for ex in exclude):
                            companies.append({
                                'name': name,
                                'company_number': item.get('company_number', ''),
                                'address': item.get('address_snippet', ''),
                                'date_created': item.get('date_of_creation', ''),
                                'source': 'Companies House'
                            })

    except Exception as e:
        print(f"      CH error: {e}")

    return companies


def try_find_website_for_company(company_name):
    """Try to find the real website for a company"""

    # Clean company name
    search_name = company_name.replace(' LIMITED', '').replace(' LTD', '').strip()

    try:
        url = f'https://html.duckduckgo.com/html/?q={search_name.replace(" ", "+")}+UK+website'
        response = requests.get(url, headers=HEADERS, timeout=10)

        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')

            results = soup.find_all('a', class_='result__a')

            for result in results[:5]:
                href = result.get('href', '')

                # Skip directories
                skip = ['yell.com', 'yelp.com', 'facebook.com', 'linkedin.com',
                       'companieshouse', 'endole', 'checkatrade', 'trustatrader']

                if any(s in href.lower() for s in skip):
                    continue

                if href.startswith('http'):
                    # Verify it's a real truck tyre site
                    verification = verify_website(href)
                    if verification:
                        return href, verification

    except:
        pass

    return None, None


def estimate_revenue(company):
    """Estimate revenue based on company characteristics"""

    name = company.get('name', '').lower()
    services = company.get('services', '')
    has_website = bool(company.get('website'))
    website_verified = company.get('website_verified', False)

    score = 0
    indicators = []

    # Name indicators
    if any(x in name for x in ['national', 'uk', 'british', 'group']):
        score += 30
        indicators.append('National scope')
    if any(x in name for x in ['wholesale', 'distributor', 'supply']):
        score += 25
        indicators.append('Wholesale/Distribution')
    if 'network' in name:
        score += 20
        indicators.append('Network')

    # Services
    if '24hr' in services:
        score += 15
        indicators.append('24hr service')
    if 'Fleet' in services:
        score += 20
        indicators.append('Fleet services')
    if 'Mobile' in services:
        score += 10
        indicators.append('Mobile fitting')
    if 'Retreading' in services:
        score += 15
        indicators.append('Retreading')

    # Website quality
    if website_verified:
        score += 15
        indicators.append('Verified website')

    # Size estimation
    if score >= 60:
        return {
            'size': 'Large',
            'employees': '50-200+',
            'revenue_low': 5000000,
            'revenue_high': 50000000,
            'revenue_display': '£5M - £50M+',
            'indicators': ', '.join(indicators) if indicators else 'Standard'
        }
    elif score >= 35:
        return {
            'size': 'Medium',
            'employees': '15-50',
            'revenue_low': 1000000,
            'revenue_high': 5000000,
            'revenue_display': '£1M - £5M',
            'indicators': ', '.join(indicators) if indicators else 'Standard'
        }
    elif score >= 15:
        return {
            'size': 'Small-Medium',
            'employees': '5-15',
            'revenue_low': 300000,
            'revenue_high': 1000000,
            'revenue_display': '£300K - £1M',
            'indicators': ', '.join(indicators) if indicators else 'Standard'
        }
    else:
        return {
            'size': 'Small',
            'employees': '1-5',
            'revenue_low': 50000,
            'revenue_high': 300000,
            'revenue_display': '£50K - £300K',
            'indicators': ', '.join(indicators) if indicators else 'Standard'
        }


def main():
    print("=" * 80)
    print("REAL UK TRUCK TYRE WEB SCRAPER")
    print("=" * 80)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    print("This scraper ONLY collects REAL, VERIFIED websites.")
    print("No made-up URLs - every website is verified to exist.")
    print()

    all_companies = []
    seen_domains = set()
    seen_names = set()

    # =========================================================================
    # STEP 1: Search for truck tyre websites directly
    # =========================================================================
    print("[1] SEARCHING FOR REAL TRUCK TYRE WEBSITES")
    print("-" * 60)

    search_queries = [
        'UK truck tyre fitters',
        'UK HGV tyre fitting',
        'mobile truck tyre fitting UK',
        'commercial truck tyres UK',
        '24 hour truck tyre service UK',
        'fleet tyre management UK',
        'truck tyre breakdown UK',
        'HGV mobile tyre fitting',
        'lorry tyre fitters UK',
        'truck tyre wholesalers UK',
    ]

    found_sites = []

    for query in search_queries:
        print(f"    Searching: '{query}'...", end=' ', flush=True)
        results = scrape_google_results(query)
        found_sites.extend(results)
        print(f"Found {len(results)} results")
        time.sleep(2)  # Be polite

    print(f"\n    Total search results: {len(found_sites)}")

    # =========================================================================
    # STEP 2: Verify each website is real and truck-tyre related
    # =========================================================================
    print("\n[2] VERIFYING WEBSITES ARE REAL TRUCK TYRE SITES")
    print("-" * 60)

    for site in found_sites:
        url = site['url']
        domain = urlparse(url).netloc

        if domain in seen_domains:
            continue

        print(f"    Checking: {domain[:40]}...", end=' ', flush=True)

        verification = verify_website(url)

        if verification:
            seen_domains.add(domain)

            company = {
                'name': verification.get('title', domain).split('|')[0].split('-')[0].strip()[:60],
                'website': url,
                'website_verified': True,
                'phone': verification.get('phone', ''),
                'email': verification.get('email', ''),
                'services': ', '.join(verification.get('services', [])),
                'source': 'Web Scrape - Verified'
            }

            all_companies.append(company)
            print(f"✓ VERIFIED - {company['name'][:30]}")
        else:
            print("✗ Not a truck tyre site")

        time.sleep(0.5)

    print(f"\n    Verified truck tyre websites: {len(all_companies)}")

    # =========================================================================
    # STEP 3: Search Companies House and find their websites
    # =========================================================================
    print("\n[3] SEARCHING COMPANIES HOUSE")
    print("-" * 60)

    ch_terms = [
        'truck tyre',
        'truck tyres',
        'hgv tyre',
        'hgv tyres',
        'commercial tyre',
        'lorry tyre',
        'mobile truck tyre',
        'fleet tyre',
    ]

    ch_companies = []
    seen_numbers = set()

    for term in ch_terms:
        print(f"    Searching: '{term}'...", end=' ', flush=True)
        results = search_companies_house(term)

        for c in results:
            if c['company_number'] not in seen_numbers:
                seen_numbers.add(c['company_number'])
                ch_companies.append(c)

        print(f"Found {len(results)} (total unique: {len(ch_companies)})")
        time.sleep(0.5)

    print(f"\n    Companies House results: {len(ch_companies)}")

    # Try to find websites for CH companies
    print("\n[4] FINDING WEBSITES FOR COMPANIES HOUSE RESULTS")
    print("-" * 60)

    for i, company in enumerate(ch_companies):
        name = company['name']
        name_lower = name.lower()

        # Skip if we already have this company
        if any(name_lower in c.get('name', '').lower() for c in all_companies):
            continue

        print(f"    [{i+1}/{len(ch_companies)}] {name[:45]}...", end=' ', flush=True)

        website, verification = try_find_website_for_company(name)

        if website and verification:
            domain = urlparse(website).netloc

            if domain not in seen_domains:
                seen_domains.add(domain)

                company['website'] = website
                company['website_verified'] = True
                company['phone'] = verification.get('phone', '')
                company['email'] = verification.get('email', '')
                company['services'] = ', '.join(verification.get('services', []))

                all_companies.append(company)
                print(f"✓ Found: {website[:35]}")
            else:
                print("(duplicate domain)")
        else:
            # Still add company without website
            company['website'] = ''
            company['website_verified'] = False
            company['phone'] = ''
            company['email'] = ''
            company['services'] = ''
            all_companies.append(company)
            print("✗ No website found")

        time.sleep(1)

    # =========================================================================
    # STEP 5: Estimate revenue for all companies
    # =========================================================================
    print("\n[5] ESTIMATING COMPANY SIZE & REVENUE")
    print("-" * 60)

    for company in all_companies:
        estimates = estimate_revenue(company)
        company['size'] = estimates['size']
        company['employees'] = estimates['employees']
        company['revenue_estimate'] = estimates['revenue_display']
        company['revenue_low'] = estimates['revenue_low']
        company['revenue_high'] = estimates['revenue_high']
        company['revenue_indicators'] = estimates['indicators']

    # Sort by revenue
    all_companies.sort(key=lambda x: x.get('revenue_high', 0), reverse=True)

    # =========================================================================
    # STATISTICS
    # =========================================================================
    total = len(all_companies)
    with_website = len([c for c in all_companies if c.get('website')])
    verified = len([c for c in all_companies if c.get('website_verified')])
    with_phone = len([c for c in all_companies if c.get('phone')])
    with_email = len([c for c in all_companies if c.get('email')])

    large = len([c for c in all_companies if c.get('size') == 'Large'])
    medium = len([c for c in all_companies if c.get('size') == 'Medium'])
    small_med = len([c for c in all_companies if c.get('size') == 'Small-Medium'])
    small = len([c for c in all_companies if c.get('size') == 'Small'])

    total_rev_low = sum(c.get('revenue_low', 0) for c in all_companies)
    total_rev_high = sum(c.get('revenue_high', 0) for c in all_companies)

    print(f"\n    RESULTS:")
    print(f"    Total companies: {total}")
    print(f"    With website: {with_website}")
    print(f"    Verified working: {verified}")
    print(f"    With phone: {with_phone}")
    print(f"    With email: {with_email}")
    print()
    print(f"    SIZE BREAKDOWN:")
    print(f"    Large: {large}")
    print(f"    Medium: {medium}")
    print(f"    Small-Medium: {small_med}")
    print(f"    Small: {small}")
    print()
    print(f"    MARKET ESTIMATE: £{total_rev_low/1000000:.0f}M - £{total_rev_high/1000000:.0f}M")

    # =========================================================================
    # EXPORT
    # =========================================================================
    print("\n[6] EXPORTING DATA")
    print("-" * 60)

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "UK Truck Tyre Companies"

    headers = [
        'Company Name', 'Website', 'Verified', 'Phone', 'Email',
        'Size', 'Employees', 'Revenue Estimate', 'Revenue Indicators',
        'Services', 'Address', 'Company Number', 'Source'
    ]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    verified_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    large_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row, c in enumerate(all_companies, 2):
        ws.cell(row=row, column=1, value=c.get('name', ''))
        ws.cell(row=row, column=2, value=c.get('website', ''))
        ws.cell(row=row, column=3, value='Yes' if c.get('website_verified') else 'No')
        ws.cell(row=row, column=4, value=c.get('phone', ''))
        ws.cell(row=row, column=5, value=c.get('email', ''))
        ws.cell(row=row, column=6, value=c.get('size', ''))
        ws.cell(row=row, column=7, value=c.get('employees', ''))
        ws.cell(row=row, column=8, value=c.get('revenue_estimate', ''))
        ws.cell(row=row, column=9, value=c.get('revenue_indicators', ''))
        ws.cell(row=row, column=10, value=c.get('services', ''))
        ws.cell(row=row, column=11, value=c.get('address', ''))
        ws.cell(row=row, column=12, value=c.get('company_number', ''))
        ws.cell(row=row, column=13, value=c.get('source', ''))

        # Highlight verified
        if c.get('website_verified'):
            ws.cell(row=row, column=3).fill = verified_fill

        # Highlight large companies
        if c.get('size') == 'Large':
            ws.cell(row=row, column=6).fill = large_fill

    # Column widths
    widths = [40, 45, 10, 18, 35, 12, 12, 18, 30, 30, 50, 15, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:M{len(all_companies) + 1}"

    wb.save(f"{OUTPUT_FILE}.xlsx")
    print(f"    Saved: {OUTPUT_FILE}.xlsx")

    # JSON
    with open(f"{OUTPUT_FILE}.json", 'w', encoding='utf-8') as f:
        json.dump(all_companies, f, indent=2, ensure_ascii=False)
    print(f"    Saved: {OUTPUT_FILE}.json")

    # CSV
    fieldnames = ['name', 'website', 'website_verified', 'phone', 'email',
                  'size', 'employees', 'revenue_estimate', 'revenue_low', 'revenue_high',
                  'revenue_indicators', 'services', 'address', 'company_number', 'source']

    with open(f"{OUTPUT_FILE}.csv", 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(all_companies)
    print(f"    Saved: {OUTPUT_FILE}.csv")

    # =========================================================================
    # SUMMARY
    # =========================================================================
    print("\n" + "=" * 80)
    print("SCRAPING COMPLETE")
    print("=" * 80)
    print(f"Total: {total} truck tyre companies")
    print(f"Verified websites: {verified}")
    print(f"Market estimate: £{total_rev_low/1000000:.0f}M - £{total_rev_high/1000000:.0f}M")

    print("\nTOP 20 VERIFIED COMPANIES:")
    print("-" * 80)

    verified_companies = [c for c in all_companies if c.get('website_verified')]
    for i, c in enumerate(verified_companies[:20], 1):
        print(f"\n{i}. {c.get('name', '')}")
        print(f"   Website: {c.get('website', '')}")
        if c.get('phone'):
            print(f"   Phone: {c.get('phone')}")
        print(f"   Size: {c.get('size')} | Revenue: {c.get('revenue_estimate')}")


if __name__ == "__main__":
    main()
