"""
UK TRUCK TYRE COMPANIES - FULL DATA SCRAPER
============================================
This scraper:
1. Loads all 846 companies from the INDUSTRY_DATABASE in scraper.py
2. Fetches detailed Companies House data for EVERY company with a company number
3. Includes explicit source citations for ALL data

Data Sources:
- Companies House API: https://api.company-information.service.gov.uk
- Web Research: Explicitly cited per company

Run: python full_scraper.py
Estimated time: ~15-20 minutes (due to API rate limits)

Output:
- UK_TRUCK_TYRE_FULL_DATABASE.xlsx
- UK_TRUCK_TYRE_FULL_DATABASE.json
- UK_TRUCK_TYRE_FULL_DATABASE.csv
"""

import requests
import json
import csv
import time
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================================
# CONFIGURATION
# ============================================================================
API_KEY = '48d17266-ff2e-425f-9b20-7dcc9b25bb79'
BASE_URL = 'https://api.company-information.service.gov.uk'
DELAY = 0.6  # Rate limit: ~100 requests per minute

# ============================================================================
# WEB RESEARCH DATA WITH EXPLICIT SOURCES
# ============================================================================
WEB_RESEARCH = {
    "MICHELIN TYRE PLC": {
        "revenue": "Part of €28.4B global (2024)",
        "revenue_source": "Michelin Annual Report 2024 - michelin.com/en/finance",
        "employees": "500+ UK staff",
        "employees_source": "LinkedIn Company Page - linkedin.com/company/michelin",
        "branches": "500+ dealers UK",
        "branches_source": "Michelin UK Website - michelin.co.uk/find-dealer",
        "description": "Global tyre manufacturer with UK HQ in Stoke-on-Trent. Operates Bulldog retreading factory in Lincolnshire.",
        "description_source": "Michelin UK Website - michelin.co.uk/about-us",
        "website": "https://www.michelin.co.uk"
    },
    "BRIDGESTONE UK LIMITED": {
        "revenue": "Part of $30B global (2024)",
        "revenue_source": "Bridgestone Corporation Annual Report - bridgestone.com/ir",
        "employees": "400+ UK staff",
        "employees_source": "ZoomInfo - zoominfo.com/c/bridgestone-uk",
        "branches": "317 outlets via 122 Truck Point dealers",
        "branches_source": "Tyrepress Industry Report 2024 - tyrepress.com",
        "description": "Japanese manufacturer with Truck Point dealer network. Fleet Care services. Operates Qualitread retreading.",
        "description_source": "Bridgestone UK Website - bridgestone.co.uk/truck",
        "website": "https://www.bridgestone.co.uk"
    },
    "CONTINENTAL TYRE GROUP LIMITED": {
        "revenue": "Part of €44B global (2024)",
        "revenue_source": "Continental AG Annual Report - continental.com/investors",
        "employees": "800+ UK staff",
        "employees_source": "LinkedIn Company Page - linkedin.com/company/continental",
        "branches": "350+ service points",
        "branches_source": "Continental UK Website - continental-tyres.co.uk",
        "description": "German manufacturer. Acquired Bandvulc 2016. Manages tyres for 2/3 of UK supermarket fleets.",
        "description_source": "Tyrepress News Article 2016 - tyrepress.com/continental-bandvulc-acquisition",
        "website": "https://www.continental-tyres.co.uk"
    },
    "GOODYEAR DUNLOP TYRES UK LIMITED": {
        "revenue": "$200M-400M UK estimate",
        "revenue_source": "Owler Company Profile - owler.com/company/goodyear",
        "employees": "300+ UK staff",
        "employees_source": "LinkedIn Company Page - linkedin.com/company/goodyear",
        "branches": "TruckForce network 150+ dealers",
        "branches_source": "Goodyear Commercial Website - goodyear.eu/en_gb/truck",
        "description": "American manufacturer with TruckForce dealer network. Acquired Cooper Tire 2021.",
        "description_source": "Goodyear Press Release 2021 - goodyear.com/news",
        "website": "https://www.goodyear.eu/en_gb/truck"
    },
    "ATS EUROMASTER LIMITED": {
        "revenue": "$346M-450M",
        "revenue_source": "Growjo Business Database - growjo.com/company/ATS_Euromaster",
        "employees": "2,600 employees",
        "employees_source": "ATS Euromaster Website - atseuromaster.co.uk/about-us",
        "branches": "340 centres, 820+ service vans",
        "branches_source": "ATS Euromaster Website - atseuromaster.co.uk/our-network",
        "description": "UK's largest comprehensive tyre distributor. Part of Michelin via Euromaster.",
        "description_source": "Wikipedia - en.wikipedia.org/wiki/ATS_Euromaster",
        "website": "https://www.atseuromaster.co.uk"
    },
    "STAPLETONS TYRE SERVICES LIMITED": {
        "revenue": "£200M+",
        "revenue_source": "Insider Media News - insidermedia.com",
        "employees": "1,000+ employees",
        "employees_source": "Stapleton's Website - stapletons-tyreservices.co.uk",
        "branches": "11 distribution centres, 400+ delivery vehicles",
        "branches_source": "Stapleton's Website - stapletons-tyreservices.co.uk/about",
        "description": "UK's largest tyre wholesaler. Part of ITOCHU Corporation. Owns Central Tyre (29 sites).",
        "description_source": "Fleet News Article - fleetnews.co.uk",
        "website": "https://www.stapletons-tyreservices.co.uk"
    },
    "KIRKBY TYRES LIMITED": {
        "revenue": "£60.4M (2024)",
        "revenue_source": "UK GlobalDatabase - uk.globaldatabase.com/company/kirkby-tyres",
        "employees": "120+ employees",
        "employees_source": "ZoomInfo - zoominfo.com/c/kirkby-tyres",
        "branches": "National distribution from Liverpool",
        "branches_source": "Kirkby Tyres Website - kirkbytyres.co.uk/about",
        "description": "UK Tyre Wholesaler of the Year 2024/2025. BKT UK distributor.",
        "description_source": "Commercial Tyre Business Magazine 2024 - commercialtyrebusiness.com",
        "website": "https://www.kirkbytyres.co.uk"
    },
    "KWIK-FIT (GB) LIMITED": {
        "revenue": "$935M estimate",
        "revenue_source": "Growjo Business Database - growjo.com/company/Kwik-Fit",
        "employees": "2,025 employees",
        "employees_source": "Growjo Business Database - growjo.com/company/Kwik-Fit",
        "branches": "697 centres, 185 mobile vans",
        "branches_source": "Fleet News Awards Article - fleetnews.co.uk/kwik-fit",
        "description": "UK's largest SMR network. Part of ITOCHU Corporation. Fits 4M tyres annually.",
        "description_source": "Wikipedia - en.wikipedia.org/wiki/Kwik_Fit",
        "website": "https://www.kwik-fit.com"
    },
    "MICHELDEVER TYRE SERVICES LIMITED": {
        "revenue": "£575M turnover",
        "revenue_source": "UK GlobalDatabase - uk.globaldatabase.com/company/micheldever-tyre-services",
        "employees": "2,301 employees",
        "employees_source": "UK GlobalDatabase - uk.globaldatabase.com/company/micheldever-tyre-services",
        "branches": "300+ fitting locations, 1,800 reserve locations",
        "branches_source": "Micheldever Fleet Solutions Website - micheldeverfleetsolutions.co.uk",
        "description": "UK's largest independent wholesaler/distributor/retailer. 20% UK market share. 6M tyres annually.",
        "description_source": "Micheldever Group Website - micheldevergroup.co.uk",
        "website": "https://www.micheldevergroup.co.uk"
    },
    "NATIONAL TYRES AND AUTOCARE LIMITED": {
        "revenue": "Part of Halfords £1.6B group",
        "revenue_source": "Halfords Annual Report 2024 - halfordscompany.com/investors",
        "employees": "3,000+ group staff",
        "employees_source": "Halfords Annual Report 2024",
        "branches": "240+ branches, 200+ vans",
        "branches_source": "National Tyres Website - national.co.uk/branches",
        "description": "UK's largest independent tyre/autocare specialist. Part of Halfords.",
        "description_source": "Wikipedia - en.wikipedia.org/wiki/National_Tyres_and_Autocare",
        "website": "https://www.national.co.uk"
    },
    "LODGE TYRE COMPANY LIMITED": {
        "revenue": "$64.6M (2025)",
        "revenue_source": "Growjo Business Database - growjo.com/company/Lodge_Tyre",
        "employees": "450+ employees",
        "employees_source": "CB Insights - cbinsights.com/company/lodge-tyre-company",
        "branches": "50+ depots, 248 mobile vans",
        "branches_source": "Tyrepress News - tyrepress.com/halfords-lodge-tyre-acquisition",
        "description": "UK's largest independent commercial provider. Acquired by Halfords 2022 for £37.2M.",
        "description_source": "Tyrepress News October 2022 - tyrepress.com",
        "website": "https://www.lodgetyre.co.uk"
    },
    "MCCONECHY'S TYRE SERVICE LIMITED": {
        "revenue": "£69M (pre-acquisition 2019)",
        "revenue_source": "Insider Media Scotland - insider.co.uk/news/scottish-tyre-centre-business",
        "employees": "320+ staff",
        "employees_source": "McConechy's Website - mcconechys.co.uk/commercial",
        "branches": "60+ sites, 130+ breakdown vans",
        "branches_source": "Halfords Acquisition Announcement 2019",
        "description": "Scottish-based. Acquired by Halfords 2019 for £8.5M. 40+ years commercial fleet experience.",
        "description_source": "Tyrepress News November 2019 - tyrepress.com",
        "website": "https://www.mcconechys.co.uk"
    },
    "R & R.C.BOND (WHOLESALE)LIMITED": {
        "revenue": "$150M+ estimate",
        "revenue_source": "Fast Track 100 - fasttrack.co.uk/company/bond-international",
        "employees": "500+ employees",
        "employees_source": "Supply Chain World Magazine - scw-mag.com/bond-international",
        "branches": "11 distribution centres, 350 vehicles",
        "branches_source": "Bond International Website - bondint.uk",
        "description": "UK's largest independent tyre wholesaler. Family business since 1966. Sells 1 tyre every 6 seconds.",
        "description_source": "Supply Chain World Magazine 2024 - scw-mag.com",
        "website": "https://www.bondint.uk"
    },
    "TANVIC GROUP LIMITED": {
        "revenue": "£70M turnover",
        "revenue_source": "Tanvic Website - tanvic.co.uk/about",
        "employees": "260 employees",
        "employees_source": "Tanvic Website - tanvic.co.uk/about",
        "branches": "20 branches, 120+ vehicles",
        "branches_source": "Tanvic Website - tanvic.co.uk/about",
        "description": "Midlands/East Anglia based. Three divisions: Retail, Commercial, Wholesale. 200,000 tyres in stock.",
        "description_source": "Tanvic Website - tanvic.co.uk/about",
        "website": "https://www.tanvic.co.uk"
    },
    "INTERNATIONAL TYRES AND TRADING LIMITED": {
        "revenue": "$30M+ estimate",
        "revenue_source": "Estimate based on market position - 1 in 10 UK truck tyres",
        "employees": "50+ staff",
        "employees_source": "LinkedIn Company Page",
        "branches": "Birmingham HQ, national delivery",
        "branches_source": "International Tyres Website - internationaltyres.com/about",
        "description": "Truck tyre wholesale specialist since 1990. Provides 1 in 10 UK replacement truck tyres. 30,000+ tyres in stock.",
        "description_source": "International Tyres Website - internationaltyres.com/about",
        "website": "https://www.internationaltyres.com"
    },
    "BUSH TYRES LIMITED": {
        "revenue": "$21.1M",
        "revenue_source": "Owler Company Profile - owler.com/company/bushtyres",
        "employees": "63 employees",
        "employees_source": "Owler Company Profile - owler.com/company/bushtyres",
        "branches": "21 branches",
        "branches_source": "Bush Tyres Website - bushtyres.co.uk/branches",
        "description": "Lincolnshire-based independent. Strong commercial truck tyre presence. Family business since 1960s.",
        "description_source": "Bush Tyres Website - bushtyres.co.uk/about",
        "website": "https://www.bushtyres.co.uk"
    },
    "REDPATH TYRES LIMITED": {
        "revenue": "£10M+ estimate",
        "revenue_source": "Estimate based on 6 depots and 60 service vans",
        "employees": "100+ staff",
        "employees_source": "Estimate based on fleet size",
        "branches": "6+ depots, 60 service vans",
        "branches_source": "Redpath Tyres Website - redpath-tyres.co.uk/depot-list",
        "description": "Scottish specialist in commercial/agricultural/earthmover tyres. Est 1974. Michelin approved repairer (1 of 3 UK).",
        "description_source": "Redpath Tyres Website - redpath-tyres.co.uk/about-us",
        "website": "https://www.redpath-tyres.co.uk"
    },
    "VACU-LUG TRACTION TYRES LIMITED": {
        "revenue": "£15M+ estimate",
        "revenue_source": "Estimate based on 162 employees",
        "employees": "162 employees",
        "employees_source": "ZoomInfo - zoominfo.com/c/vacu-lug",
        "branches": "Main facility Grantham + service centres",
        "branches_source": "Vacu-Lug Website - vacu-lug.co.uk",
        "description": "Europe's largest independent retreader. Based in Grantham, Lincolnshire. Member of Marangoni Retread Division.",
        "description_source": "Vacu-Lug Website - vacu-lug.co.uk/about",
        "website": "https://www.vacu-lug.co.uk"
    },
    "SOLTYRE LIMITED": {
        "revenue": "£8M+ estimate",
        "revenue_source": "Estimate based on 6 depots, 52 technicians",
        "employees": "52 technicians + admin",
        "employees_source": "Commercial Tyre Business Article - commercialtyrebusiness.com/fitter-force-for-soltyre",
        "branches": "6 depots Scotland/North England",
        "branches_source": "Soltyre Website - soltyre.co.uk/about-soltyre-ltd",
        "description": "Scottish-based independent. Est 2009 Dumfries. Growing presence Scotland to Yorkshire.",
        "description_source": "Soltyre Website - soltyre.co.uk/about-soltyre-ltd",
        "website": "https://www.soltyre.co.uk"
    },
    "PROTYRE LIMITED": {
        "revenue": "$150M+ estimate",
        "revenue_source": "Estimate - Part of Micheldever Group",
        "employees": "1,000+ staff",
        "employees_source": "LinkedIn Company Page",
        "branches": "180+ centres, 150+ vans",
        "branches_source": "Protyre Website - protyre.co.uk/branches",
        "description": "Major tyre retailer. 62 Pirelli Performance Centres - most of any UK retailer. Micheldever Group company.",
        "description_source": "Pirelli UK Website - pirelli.com/tyres/en-gb/car/about-us/performance-centres",
        "website": "https://www.protyre.co.uk"
    },
}

# ============================================================================
# IMPORT THE 846 COMPANIES FROM scraper.py
# ============================================================================
# We'll import the INDUSTRY_DATABASE from scraper.py
try:
    from scraper import INDUSTRY_DATABASE
    print(f"Loaded {len(INDUSTRY_DATABASE)} companies from scraper.py")
except ImportError:
    print("ERROR: Could not import INDUSTRY_DATABASE from scraper.py")
    print("Make sure scraper.py is in the same directory")
    sys.exit(1)

# ============================================================================
# API FUNCTIONS
# ============================================================================
def make_api_request(endpoint):
    """Make authenticated request to Companies House API"""
    url = f"{BASE_URL}{endpoint}"
    try:
        response = requests.get(url, auth=(API_KEY, ''), timeout=30)
        if response.status_code == 200:
            return response.json()
        elif response.status_code == 404:
            return None
        elif response.status_code == 429:
            print("    Rate limited - waiting 60 seconds...")
            time.sleep(60)
            return make_api_request(endpoint)
        else:
            return None
    except Exception as e:
        print(f"    API Error: {e}")
        return None

def get_company_profile(company_number):
    """Get full company profile from Companies House"""
    return make_api_request(f"/company/{company_number}")

def get_officers(company_number):
    """Get company officers/directors"""
    return make_api_request(f"/company/{company_number}/officers")

def get_filing_history(company_number, items=5):
    """Get recent filing history"""
    return make_api_request(f"/company/{company_number}/filing-history?items_per_page={items}")

def get_charges(company_number):
    """Get company charges (mortgages/loans)"""
    return make_api_request(f"/company/{company_number}/charges")

def get_psc(company_number):
    """Get persons with significant control"""
    return make_api_request(f"/company/{company_number}/persons-with-significant-control")

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================
def format_address(addr):
    """Format address dictionary to string"""
    if not addr:
        return ""
    parts = []
    for key in ['premises', 'address_line_1', 'address_line_2', 'locality', 'region', 'postal_code', 'country']:
        if key in addr and addr[key]:
            parts.append(str(addr[key]))
    return ', '.join(parts)

def format_date(date_val):
    """Format date to string"""
    if not date_val:
        return ""
    if isinstance(date_val, dict):
        y = date_val.get('year', '')
        m = date_val.get('month', '')
        d = date_val.get('day', '')
        if y:
            return f"{y}-{m:02d}-{d:02d}" if m else str(y)
    return str(date_val)

# ============================================================================
# MAIN PROCESSING
# ============================================================================
def process_all_companies():
    """Process all 846 companies with detailed Companies House data"""

    print("=" * 70)
    print("UK TRUCK TYRE COMPANIES - FULL DATA SCRAPER")
    print("=" * 70)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Total companies to process: {len(INDUSTRY_DATABASE)}")
    print(f"Estimated time: ~15-20 minutes")
    print("=" * 70)

    results = []
    companies_with_ch_number = 0
    companies_enriched = 0

    for i, company in enumerate(INDUSTRY_DATABASE, 1):
        company_number = company.get('companyNumber', '').strip()
        company_name = company.get('name', '')

        # Progress indicator
        progress = f"[{i}/{len(INDUSTRY_DATABASE)}]"
        print(f"{progress} {company_name[:50]}...", end=" ")

        # Start with existing data
        result = {
            # Original data from scraper.py
            'company_name': company_name,
            'company_number': company_number,
            'original_address': company.get('address', ''),
            'original_phone': company.get('phone', ''),
            'original_website': company.get('website', ''),
            'original_business_type': company.get('businessType', ''),
            'is_b2b_wholesaler': company.get('isB2BWholesaler', ''),
            'service_points': company.get('servicePoints', ''),
            'region': company.get('region', ''),
            'original_status': company.get('status', ''),
            'original_date_created': company.get('dateCreated', ''),
            'original_sic_codes': company.get('sicCodes', ''),
            'original_source': company.get('source', ''),

            # Companies House data (to be filled)
            'ch_company_name': '',
            'ch_status': '',
            'ch_type': '',
            'ch_date_created': '',
            'ch_registered_address': '',
            'ch_sic_codes': '',
            'ch_jurisdiction': '',
            'ch_has_charges': '',
            'ch_has_insolvency': '',
            'ch_last_accounts_date': '',
            'ch_accounts_type': '',
            'ch_next_accounts_due': '',
            'ch_last_confirmation': '',
            'ch_next_confirmation_due': '',
            'ch_director_count': 0,
            'ch_directors': '',
            'ch_psc_count': 0,
            'ch_psc_names': '',
            'ch_total_charges': 0,
            'ch_outstanding_charges': 0,
            'ch_latest_filing_date': '',
            'ch_latest_filing_type': '',

            # Data sources
            'data_source_company_info': 'scraper.py INDUSTRY_DATABASE',
            'data_source_ch': '',

            # Web research data (to be filled)
            'research_revenue': '',
            'research_revenue_source': '',
            'research_employees': '',
            'research_employees_source': '',
            'research_branches': '',
            'research_branches_source': '',
            'research_description': '',
            'research_description_source': '',
            'research_website': '',
        }

        # If has company number, fetch from Companies House
        if company_number:
            companies_with_ch_number += 1
            time.sleep(DELAY)

            # Get company profile
            profile = get_company_profile(company_number)
            if profile:
                companies_enriched += 1
                result['ch_company_name'] = profile.get('company_name', '')
                result['ch_status'] = profile.get('company_status', '')
                result['ch_type'] = profile.get('type', '')
                result['ch_date_created'] = profile.get('date_of_creation', '')
                result['ch_registered_address'] = format_address(profile.get('registered_office_address', {}))
                result['ch_sic_codes'] = ', '.join(profile.get('sic_codes', []))
                result['ch_jurisdiction'] = profile.get('jurisdiction', '')
                result['ch_has_charges'] = str(profile.get('has_charges', False))
                result['ch_has_insolvency'] = str(profile.get('has_insolvency_history', False))

                # Accounts info
                accounts = profile.get('accounts', {})
                last_accounts = accounts.get('last_accounts', {})
                next_accounts = accounts.get('next_accounts', {})
                result['ch_last_accounts_date'] = format_date(last_accounts.get('made_up_to', ''))
                result['ch_accounts_type'] = last_accounts.get('type', '')
                result['ch_next_accounts_due'] = format_date(next_accounts.get('due_on', ''))

                # Confirmation statement
                confirmation = profile.get('confirmation_statement', {})
                result['ch_last_confirmation'] = format_date(confirmation.get('last_made_up_to', ''))
                result['ch_next_confirmation_due'] = format_date(confirmation.get('next_due', ''))

                result['data_source_ch'] = f"Companies House API - api.company-information.service.gov.uk/company/{company_number}"
                print("Profile OK", end=" ")
            else:
                print("No profile", end=" ")

            # Get officers
            time.sleep(DELAY)
            officers = get_officers(company_number)
            if officers and 'items' in officers:
                active_directors = [o for o in officers['items']
                                   if not o.get('resigned_on')
                                   and o.get('officer_role') in ['director', 'corporate-director']]
                result['ch_director_count'] = len(active_directors)
                result['ch_directors'] = '; '.join([o.get('name', '') for o in active_directors[:5]])
                print(f"Directors:{len(active_directors)}", end=" ")

            # Get PSC
            time.sleep(DELAY)
            psc = get_psc(company_number)
            if psc and 'items' in psc:
                result['ch_psc_count'] = len(psc['items'])
                psc_names = []
                for p in psc['items'][:3]:
                    name = p.get('name', '')
                    if not name and 'name_elements' in p:
                        name = p['name_elements'].get('surname', '')
                    psc_names.append(name)
                result['ch_psc_names'] = '; '.join(psc_names)
                print(f"PSC:{len(psc['items'])}", end=" ")

            # Get charges
            time.sleep(DELAY)
            charges = get_charges(company_number)
            if charges and 'items' in charges:
                result['ch_total_charges'] = len(charges['items'])
                result['ch_outstanding_charges'] = len([c for c in charges['items'] if c.get('status') == 'outstanding'])
                print(f"Charges:{len(charges['items'])}", end=" ")

            # Get filing history
            time.sleep(DELAY)
            filings = get_filing_history(company_number)
            if filings and 'items' in filings and len(filings['items']) > 0:
                latest = filings['items'][0]
                result['ch_latest_filing_date'] = latest.get('date', '')
                result['ch_latest_filing_type'] = latest.get('type', '')
        else:
            print("No company number", end=" ")

        # Check for web research data
        name_upper = company_name.upper()
        for research_name, research_data in WEB_RESEARCH.items():
            if research_name in name_upper or name_upper in research_name:
                result['research_revenue'] = research_data.get('revenue', '')
                result['research_revenue_source'] = research_data.get('revenue_source', '')
                result['research_employees'] = research_data.get('employees', '')
                result['research_employees_source'] = research_data.get('employees_source', '')
                result['research_branches'] = research_data.get('branches', '')
                result['research_branches_source'] = research_data.get('branches_source', '')
                result['research_description'] = research_data.get('description', '')
                result['research_description_source'] = research_data.get('description_source', '')
                result['research_website'] = research_data.get('website', '')
                print("+ Research", end=" ")
                break

        print("")  # New line
        results.append(result)

    print("\n" + "=" * 70)
    print("PROCESSING COMPLETE")
    print("=" * 70)
    print(f"Total companies processed: {len(results)}")
    print(f"Companies with CH number: {companies_with_ch_number}")
    print(f"Companies enriched from CH API: {companies_enriched}")
    print(f"Companies with web research: {len([r for r in results if r['research_revenue']])}")

    return results

def create_excel_report(results):
    """Create comprehensive Excel workbook with source citations"""

    print("\nCreating Excel report...")
    wb = Workbook()

    # Styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=9)
    source_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ===== SHEET 1: All Data =====
    ws1 = wb.active
    ws1.title = "All Companies"

    headers1 = [
        "Company Name", "Company Number", "Status (CH)", "Type", "Date Created",
        "Registered Address (CH)", "SIC Codes", "Directors", "Director Count",
        "PSC Names", "PSC Count", "Has Charges", "Total Charges", "Has Insolvency",
        "Last Accounts", "Accounts Type", "Original Source", "CH Data Source"
    ]

    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for row, r in enumerate(results, 2):
        ws1.cell(row=row, column=1, value=r['company_name']).border = thin_border
        ws1.cell(row=row, column=2, value=r['company_number']).border = thin_border
        ws1.cell(row=row, column=3, value=r['ch_status']).border = thin_border
        ws1.cell(row=row, column=4, value=r['ch_type']).border = thin_border
        ws1.cell(row=row, column=5, value=r['ch_date_created']).border = thin_border
        ws1.cell(row=row, column=6, value=r['ch_registered_address']).border = thin_border
        ws1.cell(row=row, column=7, value=r['ch_sic_codes']).border = thin_border
        ws1.cell(row=row, column=8, value=r['ch_directors']).border = thin_border
        ws1.cell(row=row, column=9, value=r['ch_director_count']).border = thin_border
        ws1.cell(row=row, column=10, value=r['ch_psc_names']).border = thin_border
        ws1.cell(row=row, column=11, value=r['ch_psc_count']).border = thin_border
        ws1.cell(row=row, column=12, value=r['ch_has_charges']).border = thin_border
        ws1.cell(row=row, column=13, value=r['ch_total_charges']).border = thin_border
        ws1.cell(row=row, column=14, value=r['ch_has_insolvency']).border = thin_border
        ws1.cell(row=row, column=15, value=r['ch_last_accounts_date']).border = thin_border
        ws1.cell(row=row, column=16, value=r['ch_accounts_type']).border = thin_border
        ws1.cell(row=row, column=17, value=r['original_source']).border = thin_border
        ws1.cell(row=row, column=18, value=r['data_source_ch']).border = thin_border

    # Set widths
    widths1 = [40, 12, 10, 20, 12, 50, 15, 50, 10, 40, 8, 10, 10, 12, 12, 15, 25, 60]
    for i, w in enumerate(widths1, 1):
        ws1.column_dimensions[chr(64+i) if i <= 26 else 'R'].width = w

    # ===== SHEET 2: Research Data with Sources =====
    ws2 = wb.create_sheet("Research Data (with sources)")

    # Only companies with research data
    research_results = [r for r in results if r['research_revenue'] or r['research_description']]

    headers2 = [
        "Company Name", "Revenue", "Revenue Source", "Employees", "Employees Source",
        "Branches", "Branches Source", "Description", "Description Source", "Website"
    ]

    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        cell.font = header_font
        cell.border = thin_border

    for row, r in enumerate(research_results, 2):
        ws2.cell(row=row, column=1, value=r['company_name']).border = thin_border
        ws2.cell(row=row, column=2, value=r['research_revenue']).border = thin_border
        cell = ws2.cell(row=row, column=3, value=r['research_revenue_source'])
        cell.border = thin_border
        cell.fill = source_fill
        ws2.cell(row=row, column=4, value=r['research_employees']).border = thin_border
        cell = ws2.cell(row=row, column=5, value=r['research_employees_source'])
        cell.border = thin_border
        cell.fill = source_fill
        ws2.cell(row=row, column=6, value=r['research_branches']).border = thin_border
        cell = ws2.cell(row=row, column=7, value=r['research_branches_source'])
        cell.border = thin_border
        cell.fill = source_fill
        ws2.cell(row=row, column=8, value=r['research_description']).border = thin_border
        cell = ws2.cell(row=row, column=9, value=r['research_description_source'])
        cell.border = thin_border
        cell.fill = source_fill
        ws2.cell(row=row, column=10, value=r['research_website']).border = thin_border

    ws2.column_dimensions['A'].width = 40
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 50
    ws2.column_dimensions['D'].width = 20
    ws2.column_dimensions['E'].width = 50
    ws2.column_dimensions['F'].width = 30
    ws2.column_dimensions['G'].width = 50
    ws2.column_dimensions['H'].width = 60
    ws2.column_dimensions['I'].width = 50
    ws2.column_dimensions['J'].width = 35

    # ===== SHEET 3: Data Sources Summary =====
    ws3 = wb.create_sheet("Data Sources")

    ws3.cell(row=1, column=1, value="DATA SOURCES AND CITATIONS")
    ws3.cell(row=1, column=1).font = Font(bold=True, size=14)

    ws3.cell(row=3, column=1, value="PRIMARY DATA SOURCE")
    ws3.cell(row=3, column=1).font = Font(bold=True)

    sources_info = [
        ("Companies House API", "https://api.company-information.service.gov.uk",
         "Official UK government company register. Provides: company name, number, status, address, directors, PSC, charges, filings, accounts dates."),
        ("", "", "API Key used: " + API_KEY),
        ("", "", "Rate limit: 600 requests per 5 minutes"),
        ("", "", ""),
        ("SECONDARY DATA SOURCES (Web Research)", "", ""),
        ("Company Websites", "Various - see individual citations", "Revenue, employee counts, branch numbers, service descriptions"),
        ("UK GlobalDatabase", "uk.globaldatabase.com", "Financial data, employee counts"),
        ("ZoomInfo", "zoominfo.com", "Company profiles, employee estimates"),
        ("Growjo", "growjo.com", "Revenue estimates, employee counts"),
        ("Owler", "owler.com", "Company profiles, revenue estimates"),
        ("LinkedIn", "linkedin.com/company/*", "Employee counts, company descriptions"),
        ("Tyrepress", "tyrepress.com", "Industry news, company information"),
        ("Commercial Tyre Business", "commercialtyrebusiness.com", "Industry awards, company news"),
        ("Fleet News", "fleetnews.co.uk", "Fleet industry news, company profiles"),
        ("Insider Media", "insidermedia.com", "Business news, acquisitions"),
        ("Fast Track 100", "fasttrack.co.uk", "UK business growth rankings"),
        ("Wikipedia", "en.wikipedia.org", "Company history, general information"),
        ("", "", ""),
        ("NOTES ON DATA ACCURACY", "", ""),
        ("Companies House Data", "", "100% accurate - official government source"),
        ("Revenue Figures", "", "Estimates unless from official filings - see individual sources"),
        ("Employee Counts", "", "Approximate - from business databases or company websites"),
        ("Branch Counts", "", "From company websites - accurate at time of research"),
    ]

    row = 4
    for source, url, description in sources_info:
        ws3.cell(row=row, column=1, value=source)
        ws3.cell(row=row, column=2, value=url)
        ws3.cell(row=row, column=3, value=description)
        if source and "PRIMARY" not in source and "SECONDARY" not in source and "NOTES" not in source:
            ws3.cell(row=row, column=1).font = Font(bold=True)
        row += 1

    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 45
    ws3.column_dimensions['C'].width = 80

    # ===== SHEET 4: Summary Statistics =====
    ws4 = wb.create_sheet("Summary")

    ws4.cell(row=1, column=1, value="UK TRUCK TYRE COMPANIES - FULL DATABASE")
    ws4.cell(row=1, column=1).font = Font(bold=True, size=16)

    ws4.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws4.cell(row=4, column=1, value=f"Total Companies: {len(results)}")

    # Stats
    ch_enriched = len([r for r in results if r['ch_status']])
    with_research = len([r for r in results if r['research_revenue']])
    with_directors = len([r for r in results if r['ch_director_count'] > 0])
    with_charges = len([r for r in results if r['ch_has_charges'] == 'True'])
    active = len([r for r in results if r['ch_status'] == 'active'])

    stats = [
        ("", ""),
        ("DATA COVERAGE:", ""),
        ("Companies with Companies House data", ch_enriched),
        ("Companies with web research data", with_research),
        ("Companies with director information", with_directors),
        ("Companies with charges/loans", with_charges),
        ("Active companies (CH verified)", active),
        ("", ""),
        ("TOTAL DIRECTORS FOUND", sum(r['ch_director_count'] for r in results)),
        ("TOTAL CHARGES RECORDED", sum(r['ch_total_charges'] for r in results)),
    ]

    row = 6
    for label, value in stats:
        ws4.cell(row=row, column=1, value=label)
        if value != "":
            ws4.cell(row=row, column=2, value=value)
        if label and ":" in label:
            ws4.cell(row=row, column=1).font = Font(bold=True)
        row += 1

    ws4.column_dimensions['A'].width = 40
    ws4.column_dimensions['B'].width = 15

    # Save
    wb.save('UK_TRUCK_TYRE_FULL_DATABASE.xlsx')
    print("Excel report saved: UK_TRUCK_TYRE_FULL_DATABASE.xlsx")

def save_json(results):
    """Save results as JSON"""
    with open('UK_TRUCK_TYRE_FULL_DATABASE.json', 'w', encoding='utf-8') as f:
        json.dump(results, f, indent=2, ensure_ascii=False)
    print("JSON saved: UK_TRUCK_TYRE_FULL_DATABASE.json")

def save_csv(results):
    """Save results as CSV"""
    if not results:
        return

    # Get all keys
    all_keys = list(results[0].keys())

    with open('UK_TRUCK_TYRE_FULL_DATABASE.csv', 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=all_keys)
        writer.writeheader()
        writer.writerows(results)
    print("CSV saved: UK_TRUCK_TYRE_FULL_DATABASE.csv")

# ============================================================================
# MAIN
# ============================================================================
if __name__ == "__main__":
    results = process_all_companies()
    create_excel_report(results)
    save_json(results)
    save_csv(results)

    print("\n" + "=" * 70)
    print("ALL DONE!")
    print("=" * 70)
    print("Files created:")
    print("  - UK_TRUCK_TYRE_FULL_DATABASE.xlsx (with source citations)")
    print("  - UK_TRUCK_TYRE_FULL_DATABASE.json")
    print("  - UK_TRUCK_TYRE_FULL_DATABASE.csv")
    print(f"\nFinished: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
