"""
UK Truck Tyre Companies - Research Report Generator
Creates Excel workbook with detailed company research data
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Research data collected from web searches
RESEARCH_DATA = [
    {
        "name": "Michelin UK",
        "category": "Manufacturer/Wholesaler",
        "revenue": "Part of €28.4B global",
        "employees": "500+ UK staff",
        "branches": "500+ dealers",
        "serviceVans": "300+",
        "description": "Global tyre manufacturer with UK HQ in Stoke-on-Trent. Operates retreading factory in Lincolnshire (Bulldog brand). Major fleet service provider.",
        "services": "Truck tyres manufacturing/distribution/retreading/fleet management",
        "notableInfo": "Owns Euromaster network. Bulldog retreading 75,000 casings annually",
        "website": "https://www.michelin.co.uk"
    },
    {
        "name": "Bridgestone UK",
        "category": "Manufacturer/Wholesaler",
        "revenue": "Part of $30B global",
        "employees": "400+ UK staff",
        "branches": "317 outlets via 122 Truck Point dealers",
        "serviceVans": "250+",
        "description": "Japanese manufacturer with Truck Point dealer network. Fleet Care services. Operates Qualitread retreading.",
        "services": "Truck tyres/retreading/fleet management/24hr breakdown",
        "notableInfo": "122 Truck Point dealers managing 90,000 tyres. LRQA quality accreditation program",
        "website": "https://www.bridgestone.co.uk"
    },
    {
        "name": "Continental UK",
        "category": "Manufacturer/Wholesaler",
        "revenue": "Part of €44B global",
        "employees": "800+ UK staff",
        "branches": "350+ service points",
        "serviceVans": "400+",
        "description": "German manufacturer. Acquired Bandvulc 2016. Manages tyres for 2/3 of UK supermarket fleets including Tesco, Asda, Sainsbury's.",
        "services": "Truck tyres/retreading/fleet management/ContiPressureCheck",
        "notableInfo": "Bandvulc acquisition makes them major player in UK grocery logistics fleet tyres",
        "website": "https://www.continental-tyres.co.uk"
    },
    {
        "name": "Goodyear UK",
        "category": "Manufacturer/Wholesaler",
        "revenue": "$200M-400M UK estimate",
        "employees": "300+ UK staff",
        "branches": "TruckForce network 150+",
        "serviceVans": "200+",
        "description": "American manufacturer with TruckForce dealer network. Part of Goodyear Dunlop.",
        "services": "Truck tyres/fleet services/24hr breakdown",
        "notableInfo": "Acquired Cooper Tire 2021 for $2.8B",
        "website": "https://www.goodyear.eu/en_gb/truck"
    },
    {
        "name": "Bandvulc (Continental)",
        "category": "Retreader/Fleet Manager",
        "revenue": "£80M+ estimate",
        "employees": "500+ staff",
        "branches": "National coverage",
        "serviceVans": "300+",
        "description": "Europe's largest truck tyre retreader. 200,000 retreads per year. Handles 70% UK grocery delivery fleets.",
        "services": "Retreading/fleet management/24hr breakdown/tyre fitting",
        "notableInfo": "Part of Continental since 2016. ISO 9001 certified",
        "website": "https://www.bandvulc.com"
    },
    {
        "name": "ATS Euromaster (Tructyre)",
        "category": "Wholesaler/Retailer",
        "revenue": "$346M-450M revenue",
        "employees": "2,600 employees",
        "branches": "340 centres",
        "serviceVans": "820+ vans",
        "description": "UK's largest comprehensive tyre distributor. Part of Michelin via Euromaster. Covers cars, vans, trucks, buses, agricultural.",
        "services": "All tyres/fleet management/MOT/servicing/24hr breakdown",
        "notableInfo": "Formerly ATS Euromaster. 550+ technicians. 100+ service vans",
        "website": "https://www.atseuromaster.co.uk"
    },
    {
        "name": "Stapleton's Tyre Services",
        "category": "Wholesaler/Distributor",
        "revenue": "£200M+ revenue",
        "employees": "1,000+ employees",
        "branches": "11 distribution centres",
        "serviceVans": "400+ delivery vehicles",
        "description": "UK's largest tyre wholesaler. Part of ITOCHU Corporation. Supplies Kwik Fit network.",
        "services": "Wholesale distribution/Central Tyre network/Tyre Pros network",
        "notableInfo": "Holds 1.5M+ tyres in stock. Owns Central Tyre (29 sites) truck/agri",
        "website": "https://www.stapletons-tyreservices.co.uk"
    },
    {
        "name": "Kirkby Tyres",
        "category": "Wholesaler/Distributor",
        "revenue": "£60.4M revenue 2024",
        "employees": "120+ employees",
        "branches": "National distribution",
        "serviceVans": "80+",
        "description": "Liverpool-based wholesaler. UK Tyre Wholesaler of the Year 2024/2025. BKT UK distributor.",
        "services": "Wholesale truck/agricultural/OTR tyres",
        "notableInfo": "Imports BKT, Sailun, Double Coin brands. Strong truck tyre focus",
        "website": "https://www.kirkbytyres.co.uk"
    },
    {
        "name": "Bush Tyres",
        "category": "Retailer/Fleet Manager",
        "revenue": "$21.1M revenue",
        "employees": "63 employees",
        "branches": "21 branches",
        "serviceVans": "50+",
        "description": "Lincolnshire-based independent. Strong commercial truck tyre presence.",
        "services": "Commercial truck tyres/fleet management/car tyres",
        "notableInfo": "Family business since 1960s. ITDN member",
        "website": "https://www.bushtyres.co.uk"
    },
    {
        "name": "Point S UK",
        "category": "Network/Wholesaler",
        "revenue": "$33M revenue",
        "employees": "200+ via network",
        "branches": "300+ branches",
        "serviceVans": "150+",
        "description": "European buying group. UK members are independent dealers.",
        "services": "Truck tyres/car tyres/fleet services",
        "notableInfo": "Part of Point S International with 6,000+ centres globally",
        "website": "https://www.point-s.co.uk"
    },
    {
        "name": "Vacu-Lug",
        "category": "Retreader",
        "revenue": "£15M+ estimate",
        "employees": "162 employees",
        "branches": "1 main facility + service centres",
        "serviceVans": "40+",
        "description": "Europe's largest independent retreader. Based in Grantham, Lincolnshire.",
        "services": "Truck tyre retreading/remoulding/hot cure process",
        "notableInfo": "Member of Marangoni Retread Division. ISO certified",
        "website": "https://www.vacu-lug.co.uk"
    },
    {
        "name": "TRP Parts",
        "category": "Parts/Tyres Distributor",
        "revenue": "$50M+ UK estimate",
        "employees": "300+ UK staff",
        "branches": "140+ UK locations",
        "serviceVans": "National network",
        "description": "PACCAR parts division. Supplies truck tyres alongside parts.",
        "services": "Truck tyres/truck parts/commercial vehicle parts",
        "notableInfo": "Part of DAF/PACCAR network. 80,000+ parts available",
        "website": "https://www.trp.com"
    },
    {
        "name": "Micheldever Group",
        "category": "Wholesaler/Retailer",
        "revenue": "£575M turnover",
        "employees": "2,301 employees",
        "branches": "300+ fitting locations",
        "serviceVans": "1,800 reserve locations",
        "description": "UK's largest independent wholesaler, distributor, retailer. Part of Sumitomo Rubber.",
        "services": "All tyres wholesale/retail/fleet solutions/MFS fleet services",
        "notableInfo": "Sells 6M tyres annually. 20% UK market share. 1.3M tyres in stock",
        "website": "https://www.micheldevergroup.co.uk"
    },
    {
        "name": "National Tyres (Halfords)",
        "category": "Retailer/Fleet Services",
        "revenue": "Part of Halfords £1.6B",
        "employees": "3,000+ group staff",
        "branches": "240+ branches",
        "serviceVans": "200+ vans",
        "description": "UK's largest independent tyre/autocare specialist. Part of Halfords since merger.",
        "services": "Truck tyres/car tyres/MOT/servicing/24hr mobile",
        "notableInfo": "Combined 1,400+ locations with Halfords. National Fleet division",
        "website": "https://www.national.co.uk"
    },
    {
        "name": "Kwik Fit",
        "category": "Retailer/Fleet Services",
        "revenue": "$935M revenue estimate",
        "employees": "2,025 employees",
        "branches": "697 centres",
        "serviceVans": "185 mobile vans",
        "description": "UK's largest SMR network. Part of ITOCHU. Sister company to Stapleton's.",
        "services": "Tyres/MOT/servicing/fleet services",
        "notableInfo": "Fits 4M tyres annually. 1.2M to fleet. Central Tyre 29 truck sites",
        "website": "https://www.kwik-fit.com"
    },
    {
        "name": "Hankook UK",
        "category": "Manufacturer",
        "revenue": "Part of $6.4B global",
        "employees": "50+ UK staff",
        "branches": "200+ via dealers",
        "serviceVans": "Fleet partners",
        "description": "Korean manufacturer. 5.2% global commercial tyre market share.",
        "services": "Truck/bus tyres via dealer network",
        "notableInfo": "Sponsors UK truck shows. Fleet Management Service partners",
        "website": "https://www.hankooktire.com/uk"
    },
    {
        "name": "McConechy's (Halfords)",
        "category": "Retailer/Commercial Fleet",
        "revenue": "£69M pre-acquisition",
        "employees": "320+ staff",
        "branches": "60+ sites",
        "serviceVans": "130+ breakdown vans",
        "description": "Scottish-based. Acquired by Halfords 2019 for £8.5M. Strong commercial vehicle focus.",
        "services": "Truck/van/car tyres/24hr breakdown/fleet management",
        "notableInfo": "6th largest UK tyre retailer. 40+ years commercial fleet experience",
        "website": "https://www.mcconechys.co.uk"
    },
    {
        "name": "Lodge Tyre (Halfords)",
        "category": "Commercial/Fleet",
        "revenue": "$64.6M revenue",
        "employees": "450+ employees",
        "branches": "50+ depots",
        "serviceVans": "248 mobile vans",
        "description": "Midlands-based. Acquired by Halfords 2022 for £37.2M. UK's largest independent commercial.",
        "services": "Commercial truck tyres/24hr breakdown/fleet management",
        "notableInfo": "Makes Halfords UK's largest commercial tyre provider",
        "website": "https://www.lodgetyre.co.uk"
    },
    {
        "name": "Redpath Tyres",
        "category": "Commercial/Ag Specialist",
        "revenue": "£10M+ estimate",
        "employees": "100+ staff",
        "branches": "6+ depots",
        "serviceVans": "60 service vans",
        "description": "Scottish specialist in commercial, agricultural, earthmover tyres. Est 1974.",
        "services": "Truck/agricultural/earthmover tyres/retreading repairs",
        "notableInfo": "Bridgestone fleet award winner. Michelin approved repairer (1 of 3 UK)",
        "website": "https://www.redpath-tyres.co.uk"
    },
    {
        "name": "International Tyres",
        "category": "Wholesaler",
        "revenue": "$30M+ estimate",
        "employees": "50+ staff",
        "branches": "Birmingham HQ",
        "serviceVans": "National delivery",
        "description": "Truck tyre wholesale specialist since 1990. 1 in 10 UK replacement truck tyres.",
        "services": "Truck tyre wholesale/next day delivery UK wide",
        "notableInfo": "30,000+ tyres in stock. Same day local delivery",
        "website": "https://www.internationaltyres.com"
    },
    {
        "name": "Tyrenet",
        "category": "Network/Breakdown",
        "revenue": "£5M+ estimate",
        "employees": "20+ staff + 1,000 dealers",
        "branches": "1,000+ depots via network",
        "serviceVans": "500+ via partners",
        "description": "Largest independent tyre dealer network UK. Breakdown service specialist.",
        "services": "24hr breakdown/truck tyre fitting/fleet management",
        "notableInfo": "200+ vehicles daily. 90 min response. TEN European network shareholder",
        "website": "https://www.tyrenet.net"
    },
    {
        "name": "ITDN",
        "category": "Network/Breakdown",
        "revenue": "£8M+ estimate",
        "employees": "50+ call centre",
        "branches": "600+ member dealers",
        "serviceVans": "Network vans",
        "description": "Independent Tyre Distributors Network since 1985. Member-owned breakdown service.",
        "services": "24hr truck, van, plant breakdown/national coverage",
        "notableInfo": "122,000 breakdowns 2023. 53 min average response. 49,000 tyres fitted",
        "website": "https://www.itdn.org.uk"
    },
    {
        "name": "Tanvic Group",
        "category": "Retailer/Commercial/Wholesale",
        "revenue": "£70M turnover",
        "employees": "260 employees",
        "branches": "20 branches",
        "serviceVans": "120+ vehicles",
        "description": "Midlands/East Anglia based. Three divisions: Retail, Commercial, Wholesale.",
        "services": "Commercial truck/agricultural tyres/24hr breakdown/wholesale",
        "notableInfo": "200,000 tyres in stock. ITDN member. Founded 1970",
        "website": "https://www.tanvic.co.uk"
    },
    {
        "name": "Central Tyre (Stapleton's)",
        "category": "Commercial/Agricultural",
        "revenue": "Part of Stapleton's",
        "employees": "200+ staff",
        "branches": "29 centres",
        "serviceVans": "125+ mobile vans",
        "description": "Commercial and agricultural tyre specialist. Part of ITOCHU via Stapleton's.",
        "services": "Truck/HGV/agricultural/plant tyres/24hr service",
        "notableInfo": "ISO 9001 certified. National 24hr breakdown",
        "website": "https://www.central-tyre.com"
    },
    {
        "name": "Soltyre",
        "category": "Commercial/Fleet",
        "revenue": "£8M+ estimate",
        "employees": "52 technicians",
        "branches": "6 depots",
        "serviceVans": "40+ vans",
        "description": "Scottish-based independent. Est 2009 Dumfries. Growing presence Scotland to Yorkshire.",
        "services": "Truck/van/agricultural tyres/fleet management/24hr",
        "notableInfo": "Uses Fitter-Force digital systems. NTDA/REACT trained staff",
        "website": "https://www.soltyre.co.uk"
    },
    {
        "name": "County Tyre Group",
        "category": "Commercial/Agricultural",
        "revenue": "£15M+ estimate",
        "employees": "150+ staff",
        "branches": "40+ branches",
        "serviceVans": "100+ vans",
        "description": "South West, South Wales, Midlands specialist. Now part of The Tyre Group.",
        "services": "Commercial truck/plant/agricultural/24hr breakdown",
        "notableInfo": "Independently owned. Strong local fleet relationships",
        "website": "https://www.county-tyres.co.uk"
    },
    {
        "name": "Bond International",
        "category": "Wholesaler",
        "revenue": "$150M+ estimate",
        "employees": "500+ employees",
        "branches": "11 distribution centres",
        "serviceVans": "350 vehicles",
        "description": "UK's largest independent tyre wholesaler. Family business since 1966.",
        "services": "All tyres wholesale/next day 92% UK coverage",
        "notableInfo": "Sells 1 tyre every 6 seconds. 99.5% OTIF delivery rate. Marshal exclusive",
        "website": "https://www.bondint.uk"
    },
    {
        "name": "Universal (Halfords)",
        "category": "Commercial/Fleet",
        "revenue": "£31M pre-acquisition",
        "employees": "200+ staff",
        "branches": "20 garages",
        "serviceVans": "89 commercial vans",
        "description": "Kent-based. Acquired by Halfords 2021 for £15M. South East focus.",
        "services": "Truck/van/car tyres/commercial fleet services",
        "notableInfo": "Founded 1923. Strong fleet customer base",
        "website": "Merged into Halfords"
    },
    {
        "name": "Marangoni UK",
        "category": "Retreading Technology",
        "revenue": "Part of Marangoni Group",
        "employees": "Via UK partners",
        "branches": "Partner network",
        "serviceVans": "Partner vans",
        "description": "Italian retreading technology company. RINGTREAD system. UK partners include Tuf Treads, G&S Tyres.",
        "services": "Retreading technology/equipment/compounds",
        "notableInfo": "25% of European pre-cure retreading uses Marangoni tech",
        "website": "https://www.marangoni.com"
    },
    {
        "name": "TIA Tyres",
        "category": "Wholesaler",
        "revenue": "$20M+ estimate",
        "employees": "50+ staff",
        "branches": "85,000 sq ft warehouse",
        "serviceVans": "National delivery",
        "description": "Part of TIA Group. Major wholesaler, Midlands-based.",
        "services": "Truck/bus/car/van tyres wholesale",
        "notableInfo": "Delivery throughout mainland UK",
        "website": "https://www.tiatyres.co.uk"
    },
    {
        "name": "GB Tyres",
        "category": "Wholesaler",
        "revenue": "$25M+ estimate",
        "employees": "60+ staff",
        "branches": "Large warehouse",
        "serviceVans": "National delivery",
        "description": "One of largest budget, mid-range, premium suppliers UK/Europe.",
        "services": "Wholesale all tyres/commercial specialist",
        "notableInfo": "Large commercial truck tyre stock",
        "website": "https://www.gbtyres.net"
    },
    {
        "name": "PB Tyres",
        "category": "Commercial Wholesaler",
        "revenue": "£5M+ estimate",
        "employees": "20+ staff",
        "branches": "Warehouse facility",
        "serviceVans": "UK and abroad delivery",
        "description": "Family business 40+ years commercial tyre industry.",
        "services": "Commercial truck tyre wholesale",
        "notableInfo": "Strong manufacturer relationships",
        "website": "https://www.pbtyres.uk"
    },
    {
        "name": "Protyre",
        "category": "Retailer/Commercial",
        "revenue": "$150M+ estimate",
        "employees": "1,000+ staff",
        "branches": "180+ centres",
        "serviceVans": "150+ vans",
        "description": "Major tyre retailer. 62 Pirelli Performance Centres - most of any UK retailer.",
        "services": "Truck tyres/car tyres/MOT/fleet services",
        "notableInfo": "Micheldever Group company",
        "website": "https://www.protyre.co.uk"
    },
    {
        "name": "Setyres",
        "category": "Retailer/Commercial",
        "revenue": "£20M+ estimate",
        "employees": "150+ staff",
        "branches": "25 retail sites + commercial",
        "serviceVans": "30+ commercial vans",
        "description": "South England specialist. Commercial division for truck, ag, plant.",
        "services": "Commercial truck/agricultural/plant tyres",
        "notableInfo": "New Hailsham agricultural warehouse",
        "website": "https://www.setyres.com"
    },
    {
        "name": "Pirelli UK (Commercial)",
        "category": "Manufacturer",
        "revenue": "Part of €6.6B global",
        "employees": "100+ UK commercial",
        "branches": "150 Performance Centres",
        "serviceVans": "Via dealers",
        "description": "Italian premium manufacturer. Strong focus on independent dealer support.",
        "services": "Premium truck tyres/fleet management/B2Fleet system",
        "notableInfo": "Pharos brand via International Tyres exclusive",
        "website": "https://www.pirelli.com/tyres/en-gb"
    },
    {
        "name": "Yokohama UK",
        "category": "Manufacturer",
        "revenue": "Part of $6B global",
        "employees": "50+ UK staff",
        "branches": "200+ via dealers",
        "serviceVans": "Via dealers",
        "description": "Japanese manufacturer. BluEarth commercial range expanding.",
        "services": "Truck/bus/van tyres/BluEarth fuel efficient range",
        "notableInfo": "Growing commercial segment focus 2024/25",
        "website": "https://www.yokohama.eu"
    },
    {
        "name": "Cooper Tyres (Goodyear)",
        "category": "Manufacturer",
        "revenue": "Part of Goodyear",
        "employees": "Via UK distribution",
        "branches": "300+ via Goodyear network",
        "serviceVans": "Via dealers",
        "description": "American brand now Goodyear owned (2021 $2.8B acquisition). Owns Avon Tyres UK.",
        "services": "Commercial van/light truck tyres/Evolution Van range",
        "notableInfo": "UK Avon Tyres facility Melksham",
        "website": "https://www.coopertire.co.uk"
    },
    {
        "name": "Tuf Treads (Wales)",
        "category": "Retreader",
        "revenue": "£3M+ estimate",
        "employees": "30+ staff",
        "branches": "22,000 sq ft facility",
        "serviceVans": "Local service",
        "description": "Welsh retreader. 45-year Marangoni partnership. New facility opened 2025.",
        "services": "Truck tyre retreading/RINGTREAD system",
        "notableInfo": "Purpose-built production facility Wales",
        "website": "Via Marangoni"
    },
    {
        "name": "R&J Strang",
        "category": "Commercial/Fleet",
        "revenue": "£15M+ estimate",
        "employees": "80+ staff",
        "branches": "3 locations Scotland",
        "serviceVans": "40+ vans",
        "description": "Continental-owned Scottish business. Hamilton, Irvine, Broxburn.",
        "services": "Commercial truck tyres/fleet management",
        "notableInfo": "Part of Continental dealer network",
        "website": "https://www.rjstrang.co.uk"
    },
    {
        "name": "Halfords Commercial Fleet",
        "category": "Commercial Fleet",
        "revenue": "£384M B2B revenue (24% of group)",
        "employees": "3,000+ commercial staff",
        "branches": "National via acquisitions",
        "serviceVans": "500+ combined",
        "description": "UK's largest commercial tyre service provider. Lodge, McConechy's, Universal combined.",
        "services": "All commercial tyres/24hr breakdown/fleet management",
        "notableInfo": "Formed from multiple acquisitions. Largest UK commercial provider",
        "website": "https://www.halfordscfs.co.uk"
    },
]

def create_research_report():
    wb = Workbook()

    # ===== SHEET 1: All Research Data =====
    ws1 = wb.active
    ws1.title = "All Companies Research"

    # Headers
    headers = ["Company Name", "Category", "Revenue Estimate", "Employees", "Branches/Depots",
               "Service Vans", "Description", "Services", "Notable Info", "Website"]

    # Style headers
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Add data
    for row, company in enumerate(RESEARCH_DATA, 2):
        ws1.cell(row=row, column=1, value=company["name"]).border = thin_border
        ws1.cell(row=row, column=2, value=company["category"]).border = thin_border
        ws1.cell(row=row, column=3, value=company["revenue"]).border = thin_border
        ws1.cell(row=row, column=4, value=company["employees"]).border = thin_border
        ws1.cell(row=row, column=5, value=company["branches"]).border = thin_border
        ws1.cell(row=row, column=6, value=company["serviceVans"]).border = thin_border
        ws1.cell(row=row, column=7, value=company["description"]).border = thin_border
        ws1.cell(row=row, column=8, value=company["services"]).border = thin_border
        ws1.cell(row=row, column=9, value=company["notableInfo"]).border = thin_border
        ws1.cell(row=row, column=10, value=company["website"]).border = thin_border

        # Wrap text for description columns
        ws1.cell(row=row, column=7).alignment = Alignment(wrap_text=True, vertical='top')
        ws1.cell(row=row, column=8).alignment = Alignment(wrap_text=True, vertical='top')
        ws1.cell(row=row, column=9).alignment = Alignment(wrap_text=True, vertical='top')

    # Set column widths
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 22
    ws1.column_dimensions['D'].width = 18
    ws1.column_dimensions['E'].width = 25
    ws1.column_dimensions['F'].width = 15
    ws1.column_dimensions['G'].width = 50
    ws1.column_dimensions['H'].width = 40
    ws1.column_dimensions['I'].width = 45
    ws1.column_dimensions['J'].width = 35

    # ===== SHEET 2: Wholesalers Only =====
    ws2 = wb.create_sheet("Wholesalers & Distributors")

    wholesalers = [c for c in RESEARCH_DATA if 'Wholesaler' in c['category'] or 'Distributor' in c['category']]

    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    for row, company in enumerate(wholesalers, 2):
        ws2.cell(row=row, column=1, value=company["name"]).border = thin_border
        ws2.cell(row=row, column=2, value=company["category"]).border = thin_border
        ws2.cell(row=row, column=3, value=company["revenue"]).border = thin_border
        ws2.cell(row=row, column=4, value=company["employees"]).border = thin_border
        ws2.cell(row=row, column=5, value=company["branches"]).border = thin_border
        ws2.cell(row=row, column=6, value=company["serviceVans"]).border = thin_border
        ws2.cell(row=row, column=7, value=company["description"]).border = thin_border
        ws2.cell(row=row, column=8, value=company["services"]).border = thin_border
        ws2.cell(row=row, column=9, value=company["notableInfo"]).border = thin_border
        ws2.cell(row=row, column=10, value=company["website"]).border = thin_border

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws2.column_dimensions[col].width = ws1.column_dimensions[col].width

    # ===== SHEET 3: Manufacturers =====
    ws3 = wb.create_sheet("Manufacturers")

    manufacturers = [c for c in RESEARCH_DATA if 'Manufacturer' in c['category']]

    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    for row, company in enumerate(manufacturers, 2):
        ws3.cell(row=row, column=1, value=company["name"]).border = thin_border
        ws3.cell(row=row, column=2, value=company["category"]).border = thin_border
        ws3.cell(row=row, column=3, value=company["revenue"]).border = thin_border
        ws3.cell(row=row, column=4, value=company["employees"]).border = thin_border
        ws3.cell(row=row, column=5, value=company["branches"]).border = thin_border
        ws3.cell(row=row, column=6, value=company["serviceVans"]).border = thin_border
        ws3.cell(row=row, column=7, value=company["description"]).border = thin_border
        ws3.cell(row=row, column=8, value=company["services"]).border = thin_border
        ws3.cell(row=row, column=9, value=company["notableInfo"]).border = thin_border
        ws3.cell(row=row, column=10, value=company["website"]).border = thin_border

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws3.column_dimensions[col].width = ws1.column_dimensions[col].width

    # ===== SHEET 4: Retailers & Fleet =====
    ws4 = wb.create_sheet("Retailers & Fleet Services")

    retailers = [c for c in RESEARCH_DATA if 'Retailer' in c['category'] or 'Fleet' in c['category']]

    for col, header in enumerate(headers, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color="F57C00", end_color="F57C00", fill_type="solid")
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    for row, company in enumerate(retailers, 2):
        ws4.cell(row=row, column=1, value=company["name"]).border = thin_border
        ws4.cell(row=row, column=2, value=company["category"]).border = thin_border
        ws4.cell(row=row, column=3, value=company["revenue"]).border = thin_border
        ws4.cell(row=row, column=4, value=company["employees"]).border = thin_border
        ws4.cell(row=row, column=5, value=company["branches"]).border = thin_border
        ws4.cell(row=row, column=6, value=company["serviceVans"]).border = thin_border
        ws4.cell(row=row, column=7, value=company["description"]).border = thin_border
        ws4.cell(row=row, column=8, value=company["services"]).border = thin_border
        ws4.cell(row=row, column=9, value=company["notableInfo"]).border = thin_border
        ws4.cell(row=row, column=10, value=company["website"]).border = thin_border

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws4.column_dimensions[col].width = ws1.column_dimensions[col].width

    # ===== SHEET 5: Networks & Retreaders =====
    ws5 = wb.create_sheet("Networks & Retreaders")

    networks = [c for c in RESEARCH_DATA if 'Network' in c['category'] or 'Retread' in c['category']]

    for col, header in enumerate(headers, 1):
        cell = ws5.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color="7B1FA2", end_color="7B1FA2", fill_type="solid")
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    for row, company in enumerate(networks, 2):
        ws5.cell(row=row, column=1, value=company["name"]).border = thin_border
        ws5.cell(row=row, column=2, value=company["category"]).border = thin_border
        ws5.cell(row=row, column=3, value=company["revenue"]).border = thin_border
        ws5.cell(row=row, column=4, value=company["employees"]).border = thin_border
        ws5.cell(row=row, column=5, value=company["branches"]).border = thin_border
        ws5.cell(row=row, column=6, value=company["serviceVans"]).border = thin_border
        ws5.cell(row=row, column=7, value=company["description"]).border = thin_border
        ws5.cell(row=row, column=8, value=company["services"]).border = thin_border
        ws5.cell(row=row, column=9, value=company["notableInfo"]).border = thin_border
        ws5.cell(row=row, column=10, value=company["website"]).border = thin_border

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws5.column_dimensions[col].width = ws1.column_dimensions[col].width

    # ===== SHEET 6: Summary Statistics =====
    ws6 = wb.create_sheet("Summary")

    ws6.cell(row=1, column=1, value="UK TRUCK TYRE INDUSTRY - RESEARCH SUMMARY")
    ws6.cell(row=1, column=1).font = Font(bold=True, size=16)
    ws6.merge_cells('A1:D1')

    ws6.cell(row=3, column=1, value="Category")
    ws6.cell(row=3, column=2, value="Count")
    ws6.cell(row=3, column=3, value="Key Players")
    ws6.cell(row=3, column=1).font = Font(bold=True)
    ws6.cell(row=3, column=2).font = Font(bold=True)
    ws6.cell(row=3, column=3).font = Font(bold=True)

    ws6.cell(row=4, column=1, value="Manufacturers")
    ws6.cell(row=4, column=2, value=len(manufacturers))
    ws6.cell(row=4, column=3, value="Michelin, Bridgestone, Continental, Goodyear, Hankook, Pirelli, Yokohama")

    ws6.cell(row=5, column=1, value="Wholesalers/Distributors")
    ws6.cell(row=5, column=2, value=len(wholesalers))
    ws6.cell(row=5, column=3, value="Stapleton's, Kirkby, Bond International, Micheldever, International Tyres")

    ws6.cell(row=6, column=1, value="Retailers/Fleet Services")
    ws6.cell(row=6, column=2, value=len(retailers))
    ws6.cell(row=6, column=3, value="Halfords Fleet, Kwik Fit, National Tyres, ATS Euromaster, Lodge Tyre")

    ws6.cell(row=7, column=1, value="Networks/Retreaders")
    ws6.cell(row=7, column=2, value=len(networks))
    ws6.cell(row=7, column=3, value="ITDN, Tyrenet, Bandvulc, Vacu-Lug, Marangoni")

    ws6.cell(row=8, column=1, value="TOTAL RESEARCHED")
    ws6.cell(row=8, column=2, value=len(RESEARCH_DATA))
    ws6.cell(row=8, column=1).font = Font(bold=True)
    ws6.cell(row=8, column=2).font = Font(bold=True)

    ws6.cell(row=10, column=1, value="KEY MARKET INSIGHTS:")
    ws6.cell(row=10, column=1).font = Font(bold=True, size=12)

    insights = [
        "1. Halfords is now UK's largest commercial tyre provider after acquiring Lodge, McConechy's, Universal",
        "2. ITOCHU Corporation (Japan) owns Stapleton's + Kwik Fit - controlling major distribution",
        "3. Continental's Bandvulc acquisition (2016) gave them 70% of UK grocery fleet tyres",
        "4. Micheldever Group sells 6M tyres annually - 20% UK market share",
        "5. ITDN network handled 122,000 breakdowns in 2023 with 53 min average response",
        "6. Bond International is UK's largest independent wholesaler - sells 1 tyre every 6 seconds",
        "7. Bridgestone Truck Point has 122 dealers managing 90,000 tyres through 317 outlets",
        "8. Kirkby Tyres won UK Tyre Wholesaler of Year 2024/2025 - £60.4M revenue",
        "9. ATS Euromaster operates 340 centres with 820+ service vans (Michelin owned)",
        "10. Retreading market led by Bandvulc (200k/year) and Vacu-Lug (Europe's largest independent)"
    ]

    for i, insight in enumerate(insights, 11):
        ws6.cell(row=i, column=1, value=insight)

    ws6.column_dimensions['A'].width = 40
    ws6.column_dimensions['B'].width = 10
    ws6.column_dimensions['C'].width = 70

    # Save workbook
    filename = 'UK_TRUCK_TYRE_RESEARCH_REPORT.xlsx'
    wb.save(filename)
    print(f"Research report saved to: {filename}")
    print(f"Total companies researched: {len(RESEARCH_DATA)}")
    print(f"- Manufacturers: {len(manufacturers)}")
    print(f"- Wholesalers/Distributors: {len(wholesalers)}")
    print(f"- Retailers/Fleet: {len(retailers)}")
    print(f"- Networks/Retreaders: {len(networks)}")

if __name__ == "__main__":
    create_research_report()
