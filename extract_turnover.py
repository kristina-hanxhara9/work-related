#!/usr/bin/env python3
"""
=============================================================================
EXTRACT TURNOVER FROM COMPANIES HOUSE iXBRL FILINGS
=============================================================================

IMPORTANT REALITY CHECK:
- Most UK small/medium companies DON'T report turnover in their filings
- Only companies filing "full accounts" include turnover data
- Many large companies file PDF-only (not parseable automatically)

This script:
1. Identifies which companies file "full" or "small" accounts
2. Extracts turnover where available from iXBRL documents
3. Records accounts type for all companies

Expected results for typical UK SME database:
- ~2-5% will have extractable turnover
- ~20-30% will file micro-entity/dormant (no financial data)
- ~50-60% will file total-exemption (no turnover)

Date: January 2026
=============================================================================
"""

import requests
import json
import time
import re
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION
# =============================================================================

CH_API_KEY = '48d17266-ff2e-425f-9b20-7dcc9b25bb79'
CH_BASE_URL = 'https://api.company-information.service.gov.uk'
CH_DOCUMENT_URL = 'https://document-api.company-information.service.gov.uk'

INPUT_FILE = 'uk_truck_tyres_846_FULLY_VERIFIED.json'
OUTPUT_FILE = 'uk_truck_tyres_WITH_TURNOVER'


def get_filing_history(company_number):
    """Get accounts filing history"""
    try:
        url = f"{CH_BASE_URL}/company/{company_number}/filing-history"
        params = {'category': 'accounts', 'items_per_page': 3}
        response = requests.get(url, params=params, auth=(CH_API_KEY, ''), timeout=15)
        if response.status_code == 200:
            return response.json()
        return None
    except:
        return None


def get_document_and_extract_turnover(filing):
    """Download document and extract turnover if available"""
    try:
        doc_meta_url = filing.get('links', {}).get('document_metadata')
        if not doc_meta_url:
            return None

        # Get document metadata
        meta_response = requests.get(doc_meta_url, auth=(CH_API_KEY, ''), timeout=15)
        if meta_response.status_code != 200:
            return None

        meta_data = meta_response.json()
        resources = meta_data.get('resources', {})
        doc_url = meta_data.get('links', {}).get('document')

        # Only process if xhtml available
        if 'application/xhtml+xml' not in resources:
            return {'format': 'pdf_only', 'turnover': None}

        # Download xhtml
        headers = {'Accept': 'application/xhtml+xml'}
        doc_response = requests.get(doc_url, auth=(CH_API_KEY, ''), headers=headers, timeout=30)

        if doc_response.status_code != 200:
            return None

        content = doc_response.content.decode('utf-8', errors='ignore')

        # Extract turnover from ix:nonFraction elements
        # Pattern: <ix:nonFraction ... name="...Turnover..." scale="X">VALUE</ix:nonFraction>
        pattern = r'<ix:nonFraction[^>]*name="([^"]*)"[^>]*(?:scale="(\d+)")?[^>]*>([^<]+)</ix:nonFraction>'
        matches = re.findall(pattern, content)

        turnovers = []
        for name_attr, scale, value in matches:
            name_lower = name_attr.lower()
            if 'turnover' in name_lower or 'revenue' in name_lower:
                try:
                    # Clean and convert value
                    val = float(value.replace(',', '').replace(' ', '').strip())

                    # Apply scale if present
                    if scale:
                        val = val * (10 ** int(scale))

                    val = int(val)

                    if val > 0:
                        turnovers.append({
                            'value': val,
                            'tag': name_attr
                        })
                except:
                    pass

        if turnovers:
            # Return the largest turnover value found (usually the annual total)
            best = max(turnovers, key=lambda x: x['value'])
            return {
                'format': 'xhtml',
                'turnover': best['value'],
                'turnover_tag': best['tag'],
                'all_turnovers': turnovers
            }

        return {'format': 'xhtml', 'turnover': None}

    except Exception as e:
        return {'error': str(e)}


def classify_accounts_type(description):
    """Classify the accounts type and whether turnover is likely available"""
    desc_lower = description.lower() if description else ''

    if 'dormant' in desc_lower:
        return 'dormant', False, 'Company not trading'
    elif 'micro-entity' in desc_lower:
        return 'micro-entity', False, 'Simplified accounts - no turnover required'
    elif 'total-exemption' in desc_lower:
        return 'total-exemption', False, 'Exempt from reporting turnover'
    elif 'audit-exemption-subsiduary' in desc_lower:
        return 'subsidiary-exempt', False, 'Subsidiary exemption'
    elif 'full' in desc_lower and 'exemption' not in desc_lower:
        return 'full', True, 'Full accounts - turnover included'
    elif 'small' in desc_lower:
        return 'small', True, 'Small company accounts - may include turnover'
    elif 'group' in desc_lower:
        return 'group', True, 'Group accounts - turnover included'
    elif 'abridged' in desc_lower:
        return 'abridged', False, 'Abridged accounts - turnover optional'
    else:
        return 'unknown', False, 'Unknown accounts type'


def format_turnover(value):
    """Format turnover as readable string"""
    if not value:
        return ''
    if value >= 1000000000:
        return f"£{value/1000000000:.1f}B"
    elif value >= 1000000:
        return f"£{value/1000000:.1f}M"
    elif value >= 1000:
        return f"£{value/1000:.0f}K"
    else:
        return f"£{value}"


def main():
    print("=" * 80)
    print("EXTRACT TURNOVER FROM COMPANIES HOUSE FILINGS")
    print("=" * 80)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    # Load companies
    print(f"Loading companies from {INPUT_FILE}...")
    with open(INPUT_FILE, 'r', encoding='utf-8') as f:
        companies = json.load(f)

    companies_with_ch = [c for c in companies if c.get('ch_number')]
    print(f"Total companies: {len(companies)}")
    print(f"With CH numbers: {len(companies_with_ch)}")
    print()

    # Process each company
    print("[1] ANALYZING ACCOUNTS FILINGS")
    print("-" * 60)

    results = []
    stats = {
        'total': 0,
        'has_accounts': 0,
        'turnover_possible': 0,
        'turnover_extracted': 0,
        'by_type': {}
    }

    for i, company in enumerate(companies_with_ch):
        ch_number = company.get('ch_number', '').strip()
        name = company.get('name', '')

        if (i + 1) % 50 == 0:
            print(f"\n  Processing {i + 1}/{len(companies_with_ch)}...")

        stats['total'] += 1

        result = {
            'name': name,
            'ch_number': ch_number,
            'address': company.get('ch_address', ''),
            'website': company.get('website', ''),
            'phone': company.get('phone', ''),
            'email': company.get('email', ''),
            'businessType': company.get('businessType', ''),
            'region': company.get('region', ''),
            'verification_status': company.get('verification_status', ''),

            # Accounts data
            'accounts_type': '',
            'accounts_description': '',
            'turnover_possible': False,
            'turnover_note': '',
            'filing_date': '',

            # Turnover data
            'turnover': None,
            'turnover_formatted': '',
            'turnover_year': '',
            'turnover_tag': '',
            'document_format': ''
        }

        # Get filing history
        filings_data = get_filing_history(ch_number)

        if not filings_data or not filings_data.get('items'):
            result['accounts_type'] = 'no_filings'
            result['turnover_note'] = 'No accounts filings found'
            results.append(result)
            continue

        stats['has_accounts'] += 1

        # Get most recent accounts filing
        filing = filings_data['items'][0]
        description = filing.get('description', '')
        result['accounts_description'] = description
        result['filing_date'] = filing.get('date', '')

        # Get made up date for year
        desc_values = filing.get('description_values', {})
        made_up_date = desc_values.get('made_up_date', '')
        result['turnover_year'] = made_up_date[:4] if made_up_date else ''

        # Classify accounts type
        acc_type, turnover_possible, note = classify_accounts_type(description)
        result['accounts_type'] = acc_type
        result['turnover_possible'] = turnover_possible
        result['turnover_note'] = note

        # Track by type
        stats['by_type'][acc_type] = stats['by_type'].get(acc_type, 0) + 1

        # Only try to extract turnover if it's likely available
        if turnover_possible:
            stats['turnover_possible'] += 1

            # Try to extract from document
            doc_result = get_document_and_extract_turnover(filing)

            if doc_result:
                result['document_format'] = doc_result.get('format', '')

                if doc_result.get('turnover'):
                    turnover = doc_result['turnover']
                    result['turnover'] = turnover
                    result['turnover_formatted'] = format_turnover(turnover)
                    result['turnover_tag'] = doc_result.get('turnover_tag', '')
                    stats['turnover_extracted'] += 1

                    print(f"    ✓ {name[:40]}... {result['turnover_formatted']}")

            time.sleep(0.4)  # Rate limiting
        else:
            time.sleep(0.15)

        results.append(result)

    # Add companies without CH numbers
    for company in companies:
        if not company.get('ch_number'):
            results.append({
                'name': company.get('name', ''),
                'ch_number': '',
                'address': company.get('ch_address', ''),
                'website': company.get('website', ''),
                'phone': company.get('phone', ''),
                'email': company.get('email', ''),
                'businessType': company.get('businessType', ''),
                'region': company.get('region', ''),
                'verification_status': company.get('verification_status', ''),
                'accounts_type': 'no_ch_number',
                'accounts_description': '',
                'turnover_possible': False,
                'turnover_note': 'No Companies House number',
                'filing_date': '',
                'turnover': None,
                'turnover_formatted': '',
                'turnover_year': '',
                'turnover_tag': '',
                'document_format': ''
            })

    # Sort by turnover (highest first), then by name
    results.sort(key=lambda x: (-(x.get('turnover') or 0), x.get('name', '')))

    # Summary
    print("\n" + "-" * 60)
    print("\n  RESULTS SUMMARY:")
    print(f"    Total with CH numbers: {stats['total']}")
    print(f"    With accounts filings: {stats['has_accounts']}")
    print(f"    Turnover potentially available: {stats['turnover_possible']}")
    print(f"    Turnover successfully extracted: {stats['turnover_extracted']}")

    print("\n  ACCOUNTS TYPE BREAKDOWN:")
    for acc_type, count in sorted(stats['by_type'].items(), key=lambda x: -x[1]):
        print(f"    {count:4d} - {acc_type}")

    # Calculate totals
    total_turnover = sum(r.get('turnover', 0) or 0 for r in results)
    print(f"\n  TOTAL TURNOVER FOUND: {format_turnover(total_turnover)}")

    # Export
    print("\n[2] EXPORTING DATA")
    print("-" * 60)

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Companies with Turnover"

    headers = ['Company Name', 'Turnover', 'Turnover (Raw)', 'Year', 'Accounts Type',
               'Turnover Possible', 'Note', 'CH Number', 'Filing Date',
               'Address', 'Website', 'Phone', 'Email', 'Business Type', 'Region']

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    turnover_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    possible_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row, r in enumerate(results, 2):
        ws.cell(row=row, column=1, value=r.get('name', ''))

        turnover_cell = ws.cell(row=row, column=2, value=r.get('turnover_formatted', ''))
        if r.get('turnover'):
            turnover_cell.fill = turnover_fill

        ws.cell(row=row, column=3, value=r.get('turnover'))
        ws.cell(row=row, column=4, value=r.get('turnover_year', ''))

        type_cell = ws.cell(row=row, column=5, value=r.get('accounts_type', ''))
        if r.get('turnover_possible'):
            type_cell.fill = possible_fill

        ws.cell(row=row, column=6, value='Yes' if r.get('turnover_possible') else 'No')
        ws.cell(row=row, column=7, value=r.get('turnover_note', ''))
        ws.cell(row=row, column=8, value=r.get('ch_number', ''))
        ws.cell(row=row, column=9, value=r.get('filing_date', ''))
        ws.cell(row=row, column=10, value=r.get('address', ''))
        ws.cell(row=row, column=11, value=r.get('website', ''))
        ws.cell(row=row, column=12, value=r.get('phone', ''))
        ws.cell(row=row, column=13, value=r.get('email', ''))
        ws.cell(row=row, column=14, value=r.get('businessType', ''))
        ws.cell(row=row, column=15, value=r.get('region', ''))

    # Column widths
    widths = [45, 15, 15, 8, 18, 12, 40, 12, 12, 50, 40, 18, 30, 25, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:O{len(results) + 1}"

    # Summary sheet
    ws_sum = wb.create_sheet("Summary")
    summary_data = [
        ['TURNOVER EXTRACTION SUMMARY', ''],
        ['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['', ''],
        ['OVERVIEW', ''],
        ['Total Companies', len(results)],
        ['With CH Numbers', stats['total']],
        ['With Accounts Filings', stats['has_accounts']],
        ['', ''],
        ['TURNOVER EXTRACTION', ''],
        ['Turnover Potentially Available', stats['turnover_possible']],
        ['Turnover Successfully Extracted', stats['turnover_extracted']],
        ['Total Turnover Found', format_turnover(total_turnover)],
        ['', ''],
        ['ACCOUNTS TYPE BREAKDOWN', ''],
    ]

    for acc_type, count in sorted(stats['by_type'].items(), key=lambda x: -x[1]):
        summary_data.append([acc_type, count])

    summary_data.extend([
        ['', ''],
        ['NOTE', ''],
        ['Most UK SMEs file micro-entity or total-exemption accounts', ''],
        ['These do NOT include turnover figures', ''],
        ['Only "full" and "small" accounts may include turnover', ''],
    ])

    for row, (label, value) in enumerate(summary_data, 1):
        ws_sum.cell(row=row, column=1, value=label)
        ws_sum.cell(row=row, column=2, value=value)

    ws_sum.column_dimensions['A'].width = 45
    ws_sum.column_dimensions['B'].width = 20

    wb.save(f"{OUTPUT_FILE}.xlsx")
    print(f"  Saved: {OUTPUT_FILE}.xlsx")

    # JSON
    output_json = {
        'generated': datetime.now().isoformat(),
        'summary': {
            'total_companies': len(results),
            'with_ch_numbers': stats['total'],
            'with_accounts': stats['has_accounts'],
            'turnover_possible': stats['turnover_possible'],
            'turnover_extracted': stats['turnover_extracted'],
            'total_turnover': total_turnover,
            'total_turnover_formatted': format_turnover(total_turnover),
            'accounts_types': stats['by_type']
        },
        'companies': results
    }

    with open(f"{OUTPUT_FILE}.json", 'w', encoding='utf-8') as f:
        json.dump(output_json, f, indent=2, ensure_ascii=False)
    print(f"  Saved: {OUTPUT_FILE}.json")

    # CSV
    csv_fields = ['name', 'turnover_formatted', 'turnover', 'turnover_year', 'accounts_type',
                  'turnover_possible', 'turnover_note', 'ch_number', 'filing_date',
                  'address', 'website', 'phone', 'email', 'businessType', 'region']

    with open(f"{OUTPUT_FILE}.csv", 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=csv_fields, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(results)
    print(f"  Saved: {OUTPUT_FILE}.csv")

    # Final summary
    print("\n" + "=" * 80)
    print("EXTRACTION COMPLETE")
    print("=" * 80)
    print(f"\nTotal companies: {len(results)}")
    print(f"Turnover extracted: {stats['turnover_extracted']}")
    print(f"Total turnover: {format_turnover(total_turnover)}")

    if stats['turnover_extracted'] > 0:
        print("\n\nCOMPANIES WITH TURNOVER DATA:")
        print("-" * 60)

        for r in results:
            if r.get('turnover'):
                print(f"  {r['name'][:45]}")
                print(f"    Turnover: {r['turnover_formatted']} ({r['turnover_year']})")
                print(f"    CH: {r['ch_number']}")


if __name__ == "__main__":
    main()
