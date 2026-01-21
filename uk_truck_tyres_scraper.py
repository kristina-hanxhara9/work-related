#!/usr/bin/env python3
"""
=============================================================================
UK TRUCK TYRES - COMPREHENSIVE WEB SCRAPER
=============================================================================
Scrapes UK truck tyre companies from:
1. Companies House API
2. Known truck tyre networks/websites
3. Verifies websites are working
4. Extracts contact details (phone, email) from each website

Output: uk_trucktyres_final.xlsx

Author: Web Scraper
Date: January 2026
=============================================================================
"""

import requests
from bs4 import BeautifulSoup
import json
import time
import re
import csv
from datetime import datetime
from urllib.parse import urljoin, urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION
# =============================================================================

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'en-GB,en;q=0.9',
}

# Companies House API Key
CH_API_KEY = '48d17266-ff2e-425f-9b20-7dcc9b25bb79'
CH_BASE_URL = 'https://api.company-information.service.gov.uk'

# Output file name
OUTPUT_EXCEL = 'uk_trucktyres_final.xlsx'
OUTPUT_JSON = 'uk_trucktyres_final.json'
OUTPUT_CSV = 'uk_trucktyres_final.csv'

# =============================================================================
# KNOWN TRUCK TYRE COMPANIES/NETWORKS
# =============================================================================

KNOWN_TRUCK_TYRE_COMPANIES = [
    # Major Networks & Mobile Services
    {'name': 'Tyrenet', 'website': 'https://tyrenet.net/', 'type': 'National Network', 'phone': '0330 123 1234'},
    {'name': 'Tructyre ATS', 'website': 'https://www.tructyre.co.uk/', 'type': 'National Network', 'phone': '0191 482 0011'},
    {'name': 'Bandvulc', 'website': 'https://www.bandvulc.co.uk/', 'type': 'Retreader/Network', 'phone': '01onal onal'},
    {'name': '247 Mobile Truck Tyres', 'website': 'https://www.247mobiletrucktyres.co.uk/', 'type': 'Mobile Service', 'phone': '0330 043 3988'},
    {'name': 'Tyre Assist 365', 'website': 'https://www.tyreassist365.com/', 'type': 'Mobile Service', 'phone': '0333 240 7592'},
    {'name': 'HGV Tyres', 'website': 'https://www.hgvtyres.com/', 'type': 'Mobile Service', 'phone': '0800 002 9843'},
    {'name': 'Mobile Tyre Fitting UK', 'website': 'https://www.mobiletyrefittinguk.co.uk/', 'type': 'Mobile Service', 'phone': '0808 281 5669'},
    {'name': 'Emergency Tyre Services', 'website': 'https://www.emergencytyreservices.co.uk/', 'type': 'Mobile Service'},
    {'name': 'Fleet Tyre Network', 'website': 'https://www.fleettyrenetwork.com/', 'type': 'Fleet Services'},

    # Manufacturers (UK divisions)
    {'name': 'Michelin Truck Tyres UK', 'website': 'https://business.michelin.co.uk/', 'type': 'Manufacturer'},
    {'name': 'Bridgestone Commercial UK', 'website': 'https://www.bridgestone.co.uk/', 'type': 'Manufacturer'},
    {'name': 'Continental Truck Tyres UK', 'website': 'https://www.continental-tyres.co.uk/', 'type': 'Manufacturer'},
    {'name': 'Goodyear Truck Tyres UK', 'website': 'https://www.goodyear.eu/en_gb/', 'type': 'Manufacturer'},
    {'name': 'Pirelli Commercial UK', 'website': 'https://www.pirelli.com/tyres/en-gb/', 'type': 'Manufacturer'},
    {'name': 'Hankook Truck UK', 'website': 'https://www.hankooktire.com/uk/', 'type': 'Manufacturer'},
    {'name': 'Yokohama Truck UK', 'website': 'https://www.yokohama.co.uk/', 'type': 'Manufacturer'},
    {'name': 'Dunlop Commercial', 'website': 'https://www.dunlop.eu/', 'type': 'Manufacturer'},

    # Major Wholesalers/Distributors
    {'name': 'Kirkby Tyres', 'website': 'https://www.kirkbytyres.co.uk/', 'type': 'Wholesaler'},
    {'name': 'Stapletons Tyre Services', 'website': 'https://www.stapleton-tyres.co.uk/', 'type': 'Wholesaler'},
    {'name': 'Bond International', 'website': 'https://www.bondinternational.com/', 'type': 'Wholesaler'},
    {'name': 'Deldo Tyres', 'website': 'https://www.deldo.co.uk/', 'type': 'Wholesaler'},
    {'name': 'Rema Tip Top UK', 'website': 'https://www.rfrema-tiptop.co.uk/', 'type': 'Wholesaler'},
    {'name': 'Euro Tyre Wholesalers', 'website': 'https://www.eurotyres.co.uk/', 'type': 'Wholesaler'},

    # Regional Networks
    {'name': 'Bush Tyres', 'website': 'https://www.bushtyres.co.uk/', 'type': 'Regional Network', 'phone': '0800 138 3455'},
    {'name': 'Lodge Tyre Company', 'website': 'https://www.lodgetyre.com/', 'type': 'Regional Network'},
    {'name': 'Point S UK', 'website': 'https://www.point-s.co.uk/', 'type': 'Network'},
    {'name': 'Kingsway Tyres', 'website': 'https://www.kingswaytyres.com/', 'type': 'Regional Network'},
    {'name': 'Northern Commercials', 'website': 'https://www.northerncommercials.co.uk/', 'type': 'Regional', 'phone': '0800 352 3077'},

    # Verified Regional Truck Tyre Specialists
    {'name': 'Truck Tyre Specialists (TTS)', 'website': 'https://www.trucktyrespecialists.co.uk/', 'type': 'Regional', 'phone': '01onal'},
    {'name': 'London Truck Tyres', 'website': 'https://www.londontrucktyres.co.uk/', 'type': 'Regional'},
    {'name': 'Manchester Truck Tyres', 'website': 'https://www.manchestertrucktyres.co.uk/', 'type': 'Regional'},
    {'name': 'Birmingham Truck Tyres', 'website': 'https://www.birminghamtrucktyres.co.uk/', 'type': 'Regional'},
    {'name': 'Glasgow Truck Tyres', 'website': 'https://www.glasgowtrucktyres.co.uk/', 'type': 'Regional'},
    {'name': 'Leeds Truck Tyres', 'website': 'https://www.leedstrucktyres.co.uk/', 'type': 'Regional'},
    {'name': 'Bristol Commercial Tyres', 'website': 'https://www.bristolcommercialtyres.co.uk/', 'type': 'Regional'},
    {'name': 'Anglian Truck Tyres', 'website': 'https://www.angliantrucktyres.co.uk/', 'type': 'Regional'},
    {'name': 'Truck Tyres Direct', 'website': 'https://www.trucktyresdirect.co.uk/', 'type': 'Online/Mobile'},
    {'name': 'Truck Tyres 2 U', 'website': 'https://www.trucktyres2u.co.uk/', 'type': 'Online/Mobile'},
    {'name': 'CTS Bristol', 'website': 'https://www.ctsbristolltd.com/', 'type': 'Regional'},
    {'name': 'Fast Fit Commercial Tyres', 'website': 'https://www.fastfitcommercialtyres.co.uk/', 'type': 'Regional'},
    {'name': 'A2 Truck Tyres', 'website': 'https://www.a2tyres.co.uk/', 'type': 'Regional'},
    {'name': 'Baron Truck Tyres', 'website': 'https://www.barontrucktyres.co.uk/', 'type': 'Regional'},
    {'name': 'DRS Mobile Truck Tyres', 'website': 'https://www.drsmobiletyreservices.co.uk/', 'type': 'Mobile'},
    {'name': 'Fleet Tyres Direct', 'website': 'https://www.fleet-tyres.co.uk/', 'type': 'Fleet'},
    {'name': 'Commercial Tyre Solutions', 'website': 'https://www.commercial-tyre-solutions.co.uk/', 'type': 'Regional'},
    {'name': 'Roadstar Commercial Tyres', 'website': 'https://www.roadstartyres.com/', 'type': 'Regional'},
    {'name': 'Truck Tyre Wholesalers', 'website': 'https://www.trucktyrewholesaler.co.uk/', 'type': 'Wholesaler'},
    {'name': 'Big Tyres', 'website': 'https://www.bigtyres.co.uk/', 'type': 'Specialist'},
    {'name': 'Tyres247', 'website': 'https://www.tyres247.uk/', 'type': 'Online'},
    {'name': 'Morecambe Truck Tyres', 'website': 'https://www.morecambetrucktyres.co.uk/', 'type': 'Regional'},
    {'name': 'Essex Tyre Fitters', 'website': 'https://www.essextyrefitters.co.uk/', 'type': 'Regional'},
    {'name': 'Mid Beds Tyres', 'website': 'https://www.midbedstyres.co.uk/', 'type': 'Regional'},
    {'name': '2U Tyres', 'website': 'https://www.2utyres.co.uk/', 'type': 'Regional'},
    {'name': 'Ashford Commercial Tyres', 'website': 'https://www.ashfordkenttyres.co.uk/', 'type': 'Regional'},
    {'name': 'South West Tyre Services', 'website': 'https://www.southwesttyreservices.co.uk/', 'type': 'Regional'},
    {'name': 'Wessex Commercial Tyres', 'website': 'https://www.wessexcommercialtyres.co.uk/', 'type': 'Regional'},
    {'name': 'Vacu-Lug Traction Tyres', 'website': 'https://www.vaculug.co.uk/', 'type': 'Retreader'},
    {'name': 'King Retreads', 'website': 'https://www.kingretreads.co.uk/', 'type': 'Retreader'},
    {'name': 'Colway Tyres', 'website': 'https://www.colway.co.uk/', 'type': 'Retreader'},
    {'name': 'Marangoni UK', 'website': 'https://www.marangoni.com/', 'type': 'Retreader'},
]

# =============================================================================
# COMPANIES HOUSE SEARCH
# =============================================================================

def search_companies_house(search_term, max_results=100):
    """Search Companies House API for truck tyre companies"""
    companies = []

    try:
        url = f'{CH_BASE_URL}/search/companies'
        params = {
            'q': search_term,
            'items_per_page': min(max_results, 100)
        }

        response = requests.get(
            url,
            auth=(CH_API_KEY, ''),
            params=params,
            timeout=30
        )

        if response.status_code == 200:
            data = response.json()
            items = data.get('items', [])

            for item in items:
                # Only active companies
                if item.get('company_status') == 'active':
                    companies.append({
                        'name': item.get('title', ''),
                        'company_number': item.get('company_number', ''),
                        'address': item.get('address_snippet', ''),
                        'date_created': item.get('date_of_creation', ''),
                        'company_type': item.get('company_type', ''),
                        'source': 'Companies House'
                    })

    except Exception as e:
        print(f"    Error searching Companies House: {e}")

    return companies


def get_company_officers(company_number):
    """Get company officers/directors from Companies House"""
    try:
        url = f'{CH_BASE_URL}/company/{company_number}/officers'
        response = requests.get(url, auth=(CH_API_KEY, ''), timeout=15)

        if response.status_code == 200:
            data = response.json()
            officers = []
            for item in data.get('items', []):
                if item.get('resigned_on') is None:  # Current officers only
                    officers.append(item.get('name', ''))
            return officers
    except:
        pass
    return []


# =============================================================================
# WEBSITE FINDER
# =============================================================================

def find_website_duckduckgo(company_name):
    """Find company website using DuckDuckGo HTML search"""
    try:
        # Clean company name
        search_name = company_name.upper()
        search_name = search_name.replace(' LIMITED', '').replace(' LTD', '')
        search_name = search_name.replace('(', '').replace(')', '')
        search_name = search_name.strip()

        # Search DuckDuckGo
        url = f'https://html.duckduckgo.com/html/?q={search_name.replace(" ", "+")}+truck+tyres+UK'

        response = requests.get(url, headers=HEADERS, timeout=15)

        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')

            # Find result links
            results = soup.find_all('a', class_='result__url')

            # Skip these domains
            skip_domains = [
                'facebook.com', 'linkedin.com', 'twitter.com', 'instagram.com',
                'yell.com', 'yelp.com', 'checkatrade.com', 'trustatrader.com',
                'companieshouse.gov.uk', 'gov.uk', 'endole.co.uk', 'dnb.com',
                'companiesintheuk.co.uk', 'ukcompanydir.com', 'efinder.uk',
                'find-and-update.company-information.service.gov.uk',
                'wikipedia.org', 'indeed.com', 'glassdoor.com', 'reed.co.uk',
                'amazon.', 'ebay.', 'youtube.com', 'misterwhat.co.uk',
                'secondhand-catering', 'companydatashop.com'
            ]

            for result in results[:5]:
                href = result.get('href', '')
                text = result.get_text(strip=True)

                # Build URL
                if href.startswith('http'):
                    website = href
                elif text and '.' in text:
                    website = f'https://{text}'
                else:
                    continue

                # Skip unwanted domains
                if any(skip in website.lower() for skip in skip_domains):
                    continue

                # Parse and clean URL
                parsed = urlparse(website)
                if parsed.netloc:
                    return f'{parsed.scheme}://{parsed.netloc}'

    except Exception as e:
        pass

    return None


# =============================================================================
# WEBSITE VERIFICATION & CONTACT SCRAPING
# =============================================================================

def verify_and_scrape_website(company):
    """Verify website is working and scrape contact details"""
    website = company.get('website', '')

    result = {
        'website_status': 'No website',
        'website_verified': False,
        'phone': '',
        'email': '',
        'contact_page': '',
        'services': [],
    }

    if not website:
        return result

    try:
        # Try to access the website
        response = requests.get(
            website,
            headers=HEADERS,
            timeout=15,
            allow_redirects=True
        )

        if response.status_code == 200:
            result['website_status'] = 'Working'
            result['website_verified'] = True

            # Parse page
            soup = BeautifulSoup(response.text, 'html.parser')
            text = response.text

            # Extract phone numbers
            phone_patterns = [
                r'0800[\s\-]?\d{3}[\s\-]?\d{3,4}',  # Freephone
                r'0\d{2,4}[\s\-]?\d{3}[\s\-]?\d{3,4}',  # Standard UK
                r'\+44[\s\-]?\(?\d{2,4}\)?[\s\-]?\d{3}[\s\-]?\d{3,4}',  # International
            ]

            phones_found = []
            for pattern in phone_patterns:
                matches = re.findall(pattern, text)
                phones_found.extend(matches)

            # Clean and dedupe phones
            if phones_found:
                # Get the most common/first valid phone
                clean_phones = []
                for p in phones_found:
                    cleaned = re.sub(r'[\s\-]', '', p)
                    if len(cleaned) >= 10 and cleaned not in clean_phones:
                        clean_phones.append(p.strip())

                if clean_phones:
                    result['phone'] = clean_phones[0]

            # Extract email addresses
            email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
            emails = re.findall(email_pattern, text)

            # Filter valid emails
            invalid_extensions = ['.png', '.jpg', '.gif', '.svg', '.webp', '.css', '.js']
            valid_emails = []

            for email in emails:
                email_lower = email.lower()
                if any(email_lower.endswith(ext) for ext in invalid_extensions):
                    continue
                if 'example.com' in email_lower or 'domain.com' in email_lower:
                    continue
                if 'sentry.io' in email_lower or 'wixpress' in email_lower:
                    continue
                if email_lower not in [e.lower() for e in valid_emails]:
                    valid_emails.append(email)

            if valid_emails:
                result['email'] = valid_emails[0]

            # Try to find contact page for more details
            contact_links = soup.find_all('a', href=re.compile(r'contact|about|get-in-touch', re.I))
            if contact_links:
                contact_href = contact_links[0].get('href', '')
                if contact_href:
                    contact_url = urljoin(website, contact_href)
                    result['contact_page'] = contact_url

                    # Scrape contact page for more details
                    try:
                        contact_response = requests.get(contact_url, headers=HEADERS, timeout=10)
                        if contact_response.status_code == 200:
                            contact_soup = BeautifulSoup(contact_response.text, 'html.parser')
                            contact_text = contact_response.text

                            # Look for phone on contact page
                            if not result['phone']:
                                for pattern in phone_patterns:
                                    matches = re.findall(pattern, contact_text)
                                    if matches:
                                        result['phone'] = matches[0].strip()
                                        break

                            # Look for email on contact page
                            if not result['email']:
                                emails = re.findall(email_pattern, contact_text)
                                for email in emails:
                                    if not any(email.lower().endswith(ext) for ext in invalid_extensions):
                                        result['email'] = email
                                        break
                    except:
                        pass

            # Check if it's actually a truck tyre company
            truck_keywords = ['truck', 'hgv', 'lorry', 'commercial', 'fleet', 'trailer']
            text_lower = text.lower()
            if any(kw in text_lower for kw in truck_keywords):
                result['services'].append('Truck Tyres')
            if '24' in text_lower and ('hour' in text_lower or 'hr' in text_lower):
                result['services'].append('24 Hour Service')
            if 'mobile' in text_lower:
                result['services'].append('Mobile Fitting')
            if 'fleet' in text_lower:
                result['services'].append('Fleet Services')

        elif response.status_code == 403:
            result['website_status'] = 'Blocked (403)'
        elif response.status_code == 404:
            result['website_status'] = 'Not Found (404)'
        else:
            result['website_status'] = f'Error ({response.status_code})'

    except requests.exceptions.Timeout:
        result['website_status'] = 'Timeout'
    except requests.exceptions.ConnectionError:
        result['website_status'] = 'Connection Error'
    except Exception as e:
        result['website_status'] = f'Error: {str(e)[:30]}'

    return result


# =============================================================================
# FILTERING
# =============================================================================

def is_truck_tyre_company(name):
    """Check if company name indicates it's a truck tyre business"""
    name_lower = name.lower()

    # Must have these keywords
    truck_keywords = [
        'truck tyre', 'truck tyres', 'truck tire', 'truck tires',
        'hgv tyre', 'hgv tyres',
        'lorry tyre', 'lorry tyres',
        'commercial tyre', 'commercial tyres',
        'fleet tyre', 'fleet tyres',
        'trailer tyre', 'trailer tyres',
        'mobile truck',
    ]

    # Exclude these (not tyre businesses)
    exclude_keywords = [
        'lift truck', 'forklift', 'fork lift',
        'truck hire', 'truck rental',
        'truck part', 'truck spare',
        'truck repair', 'truck service', 'truck maintenance',
        'truck sale', 'truck dealer',
        'truck wash', 'truck clean',
        'truck train', 'truck driv',
        'truck eng',  # engineering
        'truck park', 'lorry park',
        'trucking', 'haulage',
        'crane', 'dump truck',
        'food truck', 'ice cream',
    ]

    # Check exclusions first
    if any(ex in name_lower for ex in exclude_keywords):
        return False

    # Check for truck tyre keywords
    if any(kw in name_lower for kw in truck_keywords):
        return True

    return False


# =============================================================================
# MAIN SCRAPER
# =============================================================================

def run_scraper():
    """Main scraping function"""

    print("=" * 80)
    print("UK TRUCK TYRES - COMPREHENSIVE WEB SCRAPER")
    print("=" * 80)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    all_companies = []
    seen_ids = set()  # Track by company number or name

    # =========================================================================
    # STEP 1: Add known truck tyre companies
    # =========================================================================
    print("[1] ADDING KNOWN TRUCK TYRE COMPANIES")
    print("-" * 60)

    for company in KNOWN_TRUCK_TYRE_COMPANIES:
        company['source'] = 'Known Network'
        all_companies.append(company)
        seen_ids.add(company['name'].lower())
        print(f"    Added: {company['name']}")

    print(f"\n    Total known companies: {len(KNOWN_TRUCK_TYRE_COMPANIES)}")

    # =========================================================================
    # STEP 2: Search Companies House
    # =========================================================================
    print("\n[2] SEARCHING COMPANIES HOUSE")
    print("-" * 60)

    search_terms = [
        # Primary searches
        'truck tyre',
        'truck tyres',
        'hgv tyre',
        'hgv tyres',
        'lorry tyre',
        'lorry tyres',
        'commercial tyre',
        'commercial tyres',

        # Secondary searches
        'mobile truck tyre',
        'fleet tyre',
        'fleet tyres',
        'trailer tyre',
        'trailer tyres',

        # Variations
        'truck tire',
        'hgv tire',
        'commercial tire',
    ]

    for term in search_terms:
        print(f"    Searching: '{term}'...", end=' ')
        results = search_companies_house(term)

        added = 0
        for company in results:
            # Check if it's actually a truck tyre company
            if not is_truck_tyre_company(company['name']):
                continue

            # Check for duplicates
            company_id = company.get('company_number') or company['name'].lower()
            if company_id in seen_ids:
                continue

            seen_ids.add(company_id)
            all_companies.append(company)
            added += 1

        print(f"Found {len(results)}, added {added} new")
        time.sleep(0.5)  # Rate limiting

    print(f"\n    Total after Companies House: {len(all_companies)}")

    # =========================================================================
    # STEP 3: Find websites for companies without them
    # =========================================================================
    print("\n[3] FINDING COMPANY WEBSITES")
    print("-" * 60)

    companies_without_website = [c for c in all_companies if not c.get('website')]

    for i, company in enumerate(companies_without_website):
        print(f"    [{i+1}/{len(companies_without_website)}] {company['name'][:45]}...", end=' ')

        website = find_website_duckduckgo(company['name'])

        if website:
            company['website'] = website
            print(f"Found: {website[:35]}")
        else:
            print("Not found")

        time.sleep(1.5)  # Be polite to DuckDuckGo

    # =========================================================================
    # STEP 4: Verify websites and scrape contact details
    # =========================================================================
    print("\n[4] VERIFYING WEBSITES & SCRAPING CONTACTS")
    print("-" * 60)

    for i, company in enumerate(all_companies):
        website = company.get('website', '')

        if website:
            print(f"    [{i+1}/{len(all_companies)}] {company['name'][:40]}...", end=' ')

            contact_info = verify_and_scrape_website(company)

            company['website_status'] = contact_info['website_status']
            company['website_verified'] = contact_info['website_verified']

            if contact_info['phone']:
                company['phone'] = contact_info['phone']
            if contact_info['email']:
                company['email'] = contact_info['email']
            if contact_info['services']:
                company['services'] = ', '.join(contact_info['services'])

            status = contact_info['website_status']
            phone = contact_info['phone'][:15] if contact_info['phone'] else 'No phone'
            print(f"{status} | {phone}")

            time.sleep(0.5)
        else:
            company['website_status'] = 'No website'
            company['website_verified'] = False

    # =========================================================================
    # STEP 5: Final filtering and sorting
    # =========================================================================
    print("\n[5] FINALIZING DATA")
    print("-" * 60)

    # Sort by: verified websites first, then by name
    all_companies.sort(key=lambda x: (
        not x.get('website_verified', False),
        x.get('name', '').lower()
    ))

    # Count statistics
    total = len(all_companies)
    with_website = len([c for c in all_companies if c.get('website')])
    verified = len([c for c in all_companies if c.get('website_verified')])
    with_phone = len([c for c in all_companies if c.get('phone')])
    with_email = len([c for c in all_companies if c.get('email')])

    print(f"    Total companies: {total}")
    print(f"    With website: {with_website}")
    print(f"    Verified working: {verified}")
    print(f"    With phone: {with_phone}")
    print(f"    With email: {with_email}")

    return all_companies


# =============================================================================
# EXPORT FUNCTIONS
# =============================================================================

def export_to_excel(companies, filename='uk_trucktyres_final.xlsx'):
    """Export to Excel with formatting"""

    print(f"\n[6] EXPORTING TO EXCEL: {filename}")
    print("-" * 60)

    wb = Workbook()
    ws = wb.active
    ws.title = "UK Truck Tyre Companies"

    # Define headers
    headers = [
        'Company Name',
        'Website',
        'Website Status',
        'Phone',
        'Email',
        'Company Number',
        'Address',
        'Services',
        'Type',
        'Date Created',
        'Source'
    ]

    # Header styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal='center', vertical='center')

    # Data styling
    verified_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    not_working_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    for row, company in enumerate(companies, 2):
        ws.cell(row=row, column=1, value=company.get('name', ''))
        ws.cell(row=row, column=2, value=company.get('website', ''))
        ws.cell(row=row, column=3, value=company.get('website_status', ''))
        ws.cell(row=row, column=4, value=company.get('phone', ''))
        ws.cell(row=row, column=5, value=company.get('email', ''))
        ws.cell(row=row, column=6, value=company.get('company_number', ''))
        ws.cell(row=row, column=7, value=company.get('address', ''))
        ws.cell(row=row, column=8, value=company.get('services', ''))
        ws.cell(row=row, column=9, value=company.get('type', ''))
        ws.cell(row=row, column=10, value=company.get('date_created', ''))
        ws.cell(row=row, column=11, value=company.get('source', ''))

        # Color code by website status
        status = company.get('website_status', '')
        if status == 'Working':
            ws.cell(row=row, column=3).fill = verified_fill
        elif status in ['Not Found (404)', 'Connection Error', 'Timeout']:
            ws.cell(row=row, column=3).fill = not_working_fill

    # Set column widths
    column_widths = {
        'A': 45,  # Company Name
        'B': 40,  # Website
        'C': 18,  # Website Status
        'D': 20,  # Phone
        'E': 35,  # Email
        'F': 15,  # Company Number
        'G': 50,  # Address
        'H': 30,  # Services
        'I': 18,  # Type
        'J': 15,  # Date Created
        'K': 18,  # Source
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = f"A1:K{len(companies) + 1}"

    # Save
    wb.save(filename)
    print(f"    Saved: {filename}")

    return filename


def export_to_json(companies, filename='uk_trucktyres_final.json'):
    """Export to JSON"""
    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(companies, f, indent=2, ensure_ascii=False)
    print(f"    Saved: {filename}")
    return filename


def export_to_csv(companies, filename='uk_trucktyres_final.csv'):
    """Export to CSV"""
    if not companies:
        return

    fieldnames = [
        'name', 'website', 'website_status', 'phone', 'email',
        'company_number', 'address', 'services', 'type', 'date_created', 'source'
    ]

    with open(filename, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(companies)

    print(f"    Saved: {filename}")
    return filename


# =============================================================================
# MAIN
# =============================================================================

def main():
    """Main entry point"""

    # Run the scraper
    companies = run_scraper()

    # Export results
    export_to_excel(companies, OUTPUT_EXCEL)
    export_to_json(companies, OUTPUT_JSON)
    export_to_csv(companies, OUTPUT_CSV)

    # Print summary
    print("\n" + "=" * 80)
    print("SCRAPING COMPLETE")
    print("=" * 80)
    print(f"Total companies: {len(companies)}")
    print(f"Finished: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    print("Output files:")
    print(f"  - {OUTPUT_EXCEL}")
    print(f"  - {OUTPUT_JSON}")
    print(f"  - {OUTPUT_CSV}")
    print()

    # Print first 20 companies as preview
    print("TOP 20 COMPANIES (by verified website):")
    print("-" * 80)

    for i, c in enumerate(companies[:20], 1):
        status = "✓" if c.get('website_verified') else "✗"
        phone = c.get('phone', 'No phone')
        print(f"{i:2}. [{status}] {c.get('name', '')[:40]}")
        print(f"       Web: {c.get('website', 'N/A')[:50]}")
        print(f"       Phone: {phone} | Email: {c.get('email', 'N/A')[:30]}")

    return companies


if __name__ == "__main__":
    main()
