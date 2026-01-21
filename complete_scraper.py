"""
UK TRUCK TYRE COMPANIES - COMPLETE DATA SCRAPER
================================================
Processes ALL 846 companies with:
1. Companies House API data (profile, directors, charges, filings)
2. Financial data from iXBRL accounts (where available - ~5% of companies)
3. Website scraping for branches/contact info
4. Pre-researched data for major companies

WHY MOST COMPANIES DON'T HAVE REVENUE DATA:
- UK law allows small/micro companies to file "abbreviated" accounts
- Only large companies (>£10.2M turnover) must file full accounts with revenue
- Most tyre companies are SMEs and file micro/small/dormant accounts
- Financial databases (Duedil, Experian) are PAID services

Data Sources (all cited):
- Companies House API: https://api.company-information.service.gov.uk
- Companies House Document API: For iXBRL accounts
- Company websites: Scraped for contact info
- Pre-researched data: From annual reports, news articles, company websites

Run: python complete_scraper.py
Estimated time: ~30-45 minutes

Output:
- UK_TRUCK_TYRE_COMPLETE.xlsx
- UK_TRUCK_TYRE_COMPLETE.json
- UK_TRUCK_TYRE_COMPLETE.csv
"""

import requests
import json
import csv
import time
import re
import sys
from datetime import datetime
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================================
# CONFIGURATION
# ============================================================================
API_KEY = '48d17266-ff2e-425f-9b20-7dcc9b25bb79'
BASE_URL = 'https://api.company-information.service.gov.uk'
DELAY = 0.6

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Accept': 'text/html,application/xhtml+xml',
}

# ============================================================================
# PRE-RESEARCHED DATA FOR MAJOR COMPANIES (with sources)
# This is what we KNOW from public sources - used to enrich the database
# ============================================================================
KNOWN_COMPANY_DATA = {
    "MICHELIN TYRE PLC": {
        "revenue": "Part of €28.4B Michelin Group (2024)",
        "employees": "500+ UK",
        "branches": "500+ dealers UK-wide",
        "description": "Global tyre manufacturer, UK HQ in Stoke-on-Trent",
        "source": "Michelin Annual Report 2024 - michelin.com/finance"
    },
    "BRIDGESTONE UK LIMITED": {
        "revenue": "Part of $30B Bridgestone Corp (2024)",
        "employees": "400+ UK",
        "branches": "317 outlets via 122 Truck Point dealers",
        "description": "Japanese manufacturer with Truck Point dealer network",
        "source": "Bridgestone Annual Report - bridgestone.com/ir"
    },
    "CONTINENTAL TYRE GROUP LIMITED": {
        "revenue": "Part of €44B Continental AG (2024)",
        "employees": "800+ UK",
        "branches": "350+ service points",
        "description": "German manufacturer, acquired Bandvulc 2016",
        "source": "Continental AG Annual Report - continental.com/investors"
    },
    "GOODYEAR DUNLOP TYRES UK LIMITED": {
        "revenue": "$200M-400M UK estimate",
        "employees": "300+ UK",
        "branches": "TruckForce network 150+ dealers",
        "description": "American manufacturer with TruckForce network",
        "source": "Owler Company Profile - owler.com/company/goodyear"
    },
    "ATS EUROMASTER LIMITED": {
        "revenue": "$346M-450M",
        "employees": "2,600",
        "branches": "340 centres, 820+ service vans",
        "description": "UK's largest comprehensive tyre distributor, Michelin-owned",
        "source": "Growjo - growjo.com/company/ATS_Euromaster"
    },
    "STAPLETONS TYRE SERVICES LIMITED": {
        "revenue": "£200M+",
        "employees": "1,000+",
        "branches": "11 distribution centres, 400+ delivery vehicles",
        "description": "UK's largest tyre wholesaler, Part of ITOCHU",
        "source": "Insider Media - insidermedia.com"
    },
    "KIRKBY(TYRES)LIMITED": {
        "revenue": "£60.4M (2024)",
        "employees": "120+",
        "branches": "National distribution from Liverpool",
        "description": "UK Tyre Wholesaler of the Year 2024/2025, BKT UK distributor",
        "source": "UK GlobalDatabase - uk.globaldatabase.com"
    },
    "KWIK-FIT (GB) LIMITED": {
        "revenue": "$935M estimate",
        "employees": "2,025",
        "branches": "697 centres, 185 mobile vans",
        "description": "UK's largest SMR network, Part of ITOCHU, fits 4M tyres/year",
        "source": "Growjo - growjo.com/company/Kwik-Fit"
    },
    "MICHELDEVER TYRE SERVICES LIMITED": {
        "revenue": "£575M turnover",
        "employees": "2,301",
        "branches": "300+ fitting locations",
        "description": "UK's largest independent, 20% UK market share",
        "source": "UK GlobalDatabase - uk.globaldatabase.com"
    },
    "NATIONAL TYRES AND AUTOCARE LIMITED": {
        "revenue": "Part of Halfords £1.6B group",
        "employees": "3,000+ group",
        "branches": "240+ branches, 200+ vans",
        "description": "UK's largest independent tyre/autocare specialist",
        "source": "Halfords Annual Report 2024"
    },
    "LODGE TYRE COMPANY LIMITED": {
        "revenue": "$64.6M (2025)",
        "employees": "450+",
        "branches": "50+ depots, 248 mobile vans",
        "description": "UK's largest independent commercial provider, acquired by Halfords 2022",
        "source": "Growjo - growjo.com/company/Lodge_Tyre"
    },
    "MCCONECHY'S TYRE SERVICE LIMITED": {
        "revenue": "£69M (pre-acquisition)",
        "employees": "320+",
        "branches": "60+ sites, 130+ breakdown vans",
        "description": "Scottish-based, acquired by Halfords 2019",
        "source": "Tyrepress News - tyrepress.com"
    },
    "R & R.C.BOND (WHOLESALE)LIMITED": {
        "revenue": "$150M+ estimate",
        "employees": "500+",
        "branches": "11 distribution centres, 350 vehicles",
        "description": "UK's largest independent wholesaler, sells 1 tyre every 6 seconds",
        "source": "Fast Track 100 - fasttrack.co.uk"
    },
    "TANVIC GROUP LIMITED": {
        "revenue": "£70M turnover",
        "employees": "260",
        "branches": "20 branches, 120+ vehicles",
        "description": "Midlands/East Anglia based, 200,000 tyres in stock",
        "source": "Tanvic Website - tanvic.co.uk/about"
    },
    "INTERNATIONAL TYRES AND TRADING LIMITED": {
        "revenue": "$30M+ estimate",
        "employees": "50+",
        "branches": "Birmingham HQ, national delivery",
        "description": "Provides 1 in 10 UK replacement truck tyres, 30,000+ tyres in stock",
        "source": "International Tyres Website - internationaltyres.com"
    },
    "BUSH TYRES LIMITED": {
        "revenue": "$21.1M",
        "employees": "63",
        "branches": "21 branches",
        "description": "Lincolnshire-based independent, strong commercial presence",
        "source": "Owler - owler.com/company/bushtyres"
    },
    "PROTYRE LIMITED": {
        "revenue": "$150M+ estimate",
        "employees": "1,000+",
        "branches": "180+ centres, 150+ vans",
        "description": "Major retailer, 62 Pirelli Performance Centres",
        "source": "Pirelli UK Website - pirelli.com"
    },
    "VACU-LUG TRACTION TYRES LIMITED": {
        "revenue": "£15M+ estimate",
        "employees": "162",
        "branches": "Main facility Grantham + service centres",
        "description": "Europe's largest independent retreader",
        "source": "ZoomInfo - zoominfo.com/c/vacu-lug"
    },
    "SOLTYRE LIMITED": {
        "revenue": "£8M+ estimate",
        "employees": "52 technicians + admin",
        "branches": "6 depots Scotland/North England",
        "description": "Scottish-based independent, Est 2009",
        "source": "Commercial Tyre Business - commercialtyrebusiness.com"
    },
    "HANKOOK TYRE UK LIMITED": {
        "revenue": "Part of $6.8B Hankook Global",
        "employees": "50+ UK",
        "branches": "Via dealer network",
        "description": "Korean manufacturer, growing UK presence",
        "source": "Hankook Annual Report - hankooktire.com"
    },
}

# ============================================================================
# IMPORT THE 846 COMPANIES
# ============================================================================
try:
    from scraper import INDUSTRY_DATABASE
    print(f"✓ Loaded {len(INDUSTRY_DATABASE)} companies from scraper.py")
except ImportError:
    print("ERROR: Could not import INDUSTRY_DATABASE from scraper.py")
    sys.exit(1)

# ============================================================================
# COMPANIES HOUSE API FUNCTIONS
# ============================================================================
def make_api_request(endpoint):
    """Make authenticated request to Companies House API"""
    url = f"{BASE_URL}{endpoint}"
    try:
        response = requests.get(url, auth=(API_KEY, ''), timeout=30)
        if response.status_code == 200:
            return response.json()
        elif response.status_code == 429:
            print(" [Rate limited - waiting 60s]", end="")
            time.sleep(60)
            return make_api_request(endpoint)
        return None
    except:
        return None

def get_company_profile(company_number):
    return make_api_request(f"/company/{company_number}")

def get_officers(company_number):
    return make_api_request(f"/company/{company_number}/officers")

def get_charges(company_number):
    return make_api_request(f"/company/{company_number}/charges")

def get_psc(company_number):
    return make_api_request(f"/company/{company_number}/persons-with-significant-control")

def get_filing_history(company_number, category=None):
    endpoint = f"/company/{company_number}/filing-history?items_per_page=10"
    if category:
        endpoint += f"&category={category}"
    return make_api_request(endpoint)

# ============================================================================
# FINANCIAL DATA FROM iXBRL (where available)
# ============================================================================
def get_financial_data_from_accounts(company_number):
    """Try to extract financial data from filed accounts (works for ~5% of companies)"""
    try:
        # Get accounts filings
        filings = get_filing_history(company_number, category='accounts')
        if not filings or 'items' not in filings:
            return {}

        for item in filings['items'][:2]:  # Check last 2 filings
            if 'links' not in item or 'document_metadata' not in item['links']:
                continue

            time.sleep(DELAY)
            r = requests.get(item['links']['document_metadata'], auth=(API_KEY, ''), timeout=30)
            if r.status_code != 200:
                continue

            doc_data = r.json()
            if 'application/xhtml+xml' not in doc_data.get('resources', {}):
                continue

            time.sleep(DELAY)
            doc_url = doc_data['links']['document']
            r = requests.get(doc_url, auth=(API_KEY, ''), headers={'Accept': 'application/xhtml+xml'}, timeout=60)
            if r.status_code != 200:
                continue

            content = r.text
            result = {'accounts_date': item.get('date', '')}

            # Look for turnover
            turnover_patterns = [
                r'<ix:nonFraction[^>]*name="[^"]*[Tt]urnover[^"]*"[^>]*>([0-9,]+)</ix:nonFraction>',
                r'<ix:nonFraction[^>]*name="[^"]*[Rr]evenue[^"]*"[^>]*>([0-9,]+)</ix:nonFraction>',
            ]
            for pattern in turnover_patterns:
                matches = re.findall(pattern, content)
                if matches:
                    values = [int(m.replace(',', '')) for m in matches]
                    result['turnover'] = max(values)
                    break

            # Look for employees
            emp_patterns = [
                r'<ix:nonFraction[^>]*name="[^"]*AverageNumberEmployees[^"]*"[^>]*>([0-9,]+)</ix:nonFraction>',
                r'<ix:nonFraction[^>]*name="[^"]*NumberEmployees[^"]*"[^>]*>([0-9,]+)</ix:nonFraction>',
            ]
            for pattern in emp_patterns:
                matches = re.findall(pattern, content)
                if matches:
                    values = [int(m.replace(',', '')) for m in matches]
                    result['employees'] = max(values)
                    break

            if result.get('turnover') or result.get('employees'):
                return result

        return {}
    except:
        return {}

# ============================================================================
# WEBSITE SCRAPING
# ============================================================================
def scrape_website(url):
    """Scrape company website for contact info and branch data"""
    if not url or not url.startswith('http'):
        return {}

    result = {
        'phones': [],
        'emails': [],
        'branch_count': None,
        'locations': [],
    }

    try:
        r = requests.get(url, headers=HEADERS, timeout=15, allow_redirects=True)
        if r.status_code != 200:
            return result

        content = r.text.lower()
        original = r.text

        # Phone numbers (UK format)
        phone_patterns = [
            r'(?:tel|phone|call)[:\s]*([0-9\s\-\(\)]{10,20})',
            r'(\+44[\s\-]?[0-9\s\-]{9,12})',
            r'(0[0-9]{2,4}[\s\-]?[0-9]{5,8})',
        ]
        for pattern in phone_patterns:
            matches = re.findall(pattern, content)
            for m in matches[:3]:
                cleaned = re.sub(r'[^\d+]', '', m)
                if len(cleaned) >= 10:
                    result['phones'].append(m.strip())

        result['phones'] = list(set(result['phones']))[:3]

        # Emails
        email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
        emails = re.findall(email_pattern, original)
        result['emails'] = list(set(emails))[:3]

        # Branch counts
        branch_patterns = [
            r'(\d+)\s*(?:branches|depots|locations|centres|centers|sites|outlets)',
            r'(?:over|more than)\s*(\d+)\s*(?:branches|depots|locations)',
            r'(\d+)\+?\s*(?:service\s*)?(?:points|vans|vehicles)',
            r'network of\s*(\d+)',
        ]
        for pattern in branch_patterns:
            matches = re.findall(pattern, content)
            if matches:
                numbers = [int(m) for m in matches if m.isdigit() and int(m) < 1000]
                if numbers:
                    result['branch_count'] = max(numbers)
                    break

        # UK cities mentioned
        uk_cities = ['london', 'birmingham', 'manchester', 'leeds', 'glasgow', 'edinburgh',
                     'liverpool', 'bristol', 'sheffield', 'newcastle', 'nottingham', 'cardiff']
        for city in uk_cities:
            if city in content:
                result['locations'].append(city.title())
        result['locations'] = result['locations'][:5]

    except:
        pass

    return result

# ============================================================================
# HELPERS
# ============================================================================
def format_address(addr):
    if not addr:
        return ""
    parts = []
    for key in ['premises', 'address_line_1', 'address_line_2', 'locality', 'region', 'postal_code', 'country']:
        if key in addr and addr[key]:
            parts.append(str(addr[key]))
    return ', '.join(parts)

def format_currency(value):
    if not value:
        return ""
    if value >= 1000000:
        return f"£{value/1000000:.1f}M"
    elif value >= 1000:
        return f"£{value/1000:.0f}K"
    return f"£{value}"

# ============================================================================
# MAIN PROCESSING
# ============================================================================
def process_all_companies():
    """Process all 846 companies"""

    print("=" * 70)
    print("UK TRUCK TYRE COMPANIES - COMPLETE DATA SCRAPER")
    print("=" * 70)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Total companies: {len(INDUSTRY_DATABASE)}")
    print()
    print("Data sources:")
    print("  1. Companies House API (profiles, directors, charges)")
    print("  2. iXBRL accounts (turnover where available)")
    print("  3. Website scraping (contacts, branches)")
    print("  4. Pre-researched data (20 major companies)")
    print("=" * 70)
    print()

    results = []
    stats = {
        'total': len(INDUSTRY_DATABASE),
        'with_ch_data': 0,
        'with_known_data': 0,
        'with_website_data': 0,
        'with_ixbrl_financials': 0,
    }

    for i, company in enumerate(INDUSTRY_DATABASE, 1):
        company_number = company.get('companyNumber', '').strip()
        company_name = company.get('name', '')
        website = company.get('website', '')

        print(f"[{i}/{len(INDUSTRY_DATABASE)}] {company_name[:45]}...", end=" ")

        result = {
            # Basic info
            'company_name': company_name,
            'company_number': company_number,
            'original_address': company.get('address', ''),
            'original_phone': company.get('phone', ''),
            'website': website,
            'business_type': company.get('businessType', ''),
            'is_b2b_wholesaler': company.get('isB2BWholesaler', ''),
            'service_points': company.get('servicePoints', ''),
            'region': company.get('region', ''),
            'original_source': company.get('source', ''),

            # Companies House data
            'ch_status': '',
            'ch_type': '',
            'ch_date_created': '',
            'ch_registered_address': '',
            'ch_sic_codes': '',
            'ch_accounts_type': '',
            'ch_has_charges': '',
            'ch_has_insolvency': '',
            'ch_director_count': 0,
            'ch_directors': '',
            'ch_psc_count': 0,
            'ch_psc_names': '',
            'ch_total_charges': 0,

            # Financial data (from iXBRL or known data)
            'revenue': '',
            'employees': '',
            'branches': '',
            'description': '',
            'financial_source': '',

            # Website scraped data
            'scraped_phones': '',
            'scraped_emails': '',
            'scraped_branch_count': '',
            'scraped_locations': '',

            # Data sources
            'source_ch': '',
            'source_financial': '',
            'source_website': '',
        }

        # ===== 1. CHECK FOR KNOWN/RESEARCHED DATA =====
        name_upper = company_name.upper()
        for known_name, known_data in KNOWN_COMPANY_DATA.items():
            if known_name in name_upper or name_upper in known_name:
                stats['with_known_data'] += 1
                result['revenue'] = known_data.get('revenue', '')
                result['employees'] = known_data.get('employees', '')
                result['branches'] = known_data.get('branches', '')
                result['description'] = known_data.get('description', '')
                result['financial_source'] = known_data.get('source', '')
                print("Known", end=" ")
                break

        # ===== 2. COMPANIES HOUSE PROFILE =====
        if company_number:
            time.sleep(DELAY)
            profile = get_company_profile(company_number)

            if profile:
                stats['with_ch_data'] += 1
                result['ch_status'] = profile.get('company_status', '')
                result['ch_type'] = profile.get('type', '')
                result['ch_date_created'] = profile.get('date_of_creation', '')
                result['ch_registered_address'] = format_address(profile.get('registered_office_address', {}))
                result['ch_sic_codes'] = ', '.join(profile.get('sic_codes', []))
                result['ch_accounts_type'] = profile.get('accounts', {}).get('last_accounts', {}).get('type', '')
                result['ch_has_charges'] = str(profile.get('has_charges', False))
                result['ch_has_insolvency'] = str(profile.get('has_insolvency_history', False))
                result['source_ch'] = f"api.company-information.service.gov.uk/company/{company_number}"
                print("CH", end=" ")

            # Get officers
            time.sleep(DELAY)
            officers = get_officers(company_number)
            if officers and 'items' in officers:
                active = [o for o in officers['items'] if not o.get('resigned_on') and 'director' in o.get('officer_role', '').lower()]
                result['ch_director_count'] = len(active)
                result['ch_directors'] = '; '.join([o.get('name', '') for o in active[:5]])

            # Get PSC
            time.sleep(DELAY)
            psc = get_psc(company_number)
            if psc and 'items' in psc:
                result['ch_psc_count'] = len(psc['items'])
                result['ch_psc_names'] = '; '.join([p.get('name', '') for p in psc['items'][:3]])

            # Get charges
            time.sleep(DELAY)
            charges = get_charges(company_number)
            if charges and 'items' in charges:
                result['ch_total_charges'] = len(charges['items'])

            # ===== 3. TRY iXBRL ACCOUNTS (if no known data) =====
            if not result['revenue']:
                time.sleep(DELAY)
                financial = get_financial_data_from_accounts(company_number)
                if financial:
                    stats['with_ixbrl_financials'] += 1
                    if 'turnover' in financial:
                        result['revenue'] = format_currency(financial['turnover'])
                        result['financial_source'] = f"Companies House iXBRL ({financial.get('accounts_date', '')})"
                        print(f"Rev:{result['revenue']}", end=" ")
                    if 'employees' in financial and not result['employees']:
                        result['employees'] = str(financial['employees'])
                        print(f"Emp:{financial['employees']}", end=" ")

        # ===== 4. WEBSITE SCRAPING =====
        if website:
            time.sleep(1)
            web_data = scrape_website(website)

            if any([web_data.get('phones'), web_data.get('emails'), web_data.get('branch_count')]):
                stats['with_website_data'] += 1
                result['scraped_phones'] = ', '.join(web_data.get('phones', []))
                result['scraped_emails'] = ', '.join(web_data.get('emails', []))
                if web_data.get('branch_count'):
                    result['scraped_branch_count'] = str(web_data['branch_count'])
                result['scraped_locations'] = ', '.join(web_data.get('locations', []))
                result['source_website'] = website
                print("Web", end=" ")

        print("")
        results.append(result)

        if i % 50 == 0:
            print(f"\n--- Progress: {i}/{len(INDUSTRY_DATABASE)} ({i*100//len(INDUSTRY_DATABASE)}%) ---\n")

    print("\n" + "=" * 70)
    print("PROCESSING COMPLETE")
    print("=" * 70)
    print(f"Total companies: {stats['total']}")
    print(f"With Companies House data: {stats['with_ch_data']}")
    print(f"With known/researched financial data: {stats['with_known_data']}")
    print(f"With iXBRL financials: {stats['with_ixbrl_financials']}")
    print(f"With website data: {stats['with_website_data']}")

    return results, stats

# ============================================================================
# EXCEL EXPORT
# ============================================================================
def create_excel_report(results, stats):
    print("\nCreating Excel report...")
    wb = Workbook()

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=9)
    money_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ===== SHEET 1: All Companies =====
    ws1 = wb.active
    ws1.title = "All Companies"

    headers1 = [
        "Company Name", "Company Number", "Status", "Type", "Date Created",
        "Registered Address", "SIC Codes", "Accounts Type",
        "Revenue", "Employees", "Branches", "Description",
        "Directors", "Director Count", "PSC Names",
        "Has Charges", "Total Charges", "Has Insolvency",
        "Website", "Scraped Phones", "Scraped Emails", "Scraped Branch Count",
        "B2B Wholesaler", "Region", "Financial Source", "CH Source"
    ]

    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    for row, r in enumerate(results, 2):
        col = 1
        for key in ['company_name', 'company_number', 'ch_status', 'ch_type', 'ch_date_created',
                    'ch_registered_address', 'ch_sic_codes', 'ch_accounts_type',
                    'revenue', 'employees', 'branches', 'description',
                    'ch_directors', 'ch_director_count', 'ch_psc_names',
                    'ch_has_charges', 'ch_total_charges', 'ch_has_insolvency',
                    'website', 'scraped_phones', 'scraped_emails', 'scraped_branch_count',
                    'is_b2b_wholesaler', 'region', 'financial_source', 'source_ch']:
            cell = ws1.cell(row=row, column=col, value=r.get(key, ''))
            cell.border = thin_border
            if key == 'revenue' and r.get(key):
                cell.fill = money_fill
            col += 1

    # Column widths
    ws1.column_dimensions['A'].width = 40
    ws1.column_dimensions['F'].width = 50
    ws1.column_dimensions['I'].width = 20
    ws1.column_dimensions['L'].width = 50
    ws1.column_dimensions['M'].width = 40
    ws1.column_dimensions['S'].width = 35

    # ===== SHEET 2: Companies with Financial Data =====
    ws2 = wb.create_sheet("With Financial Data")

    financial_companies = [r for r in results if r['revenue'] or r['employees'] or r['branches']]

    headers2 = ["Company Name", "Revenue", "Employees", "Branches", "Description", "Status", "Source"]

    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        cell.font = header_font

    for row, r in enumerate(financial_companies, 2):
        ws2.cell(row=row, column=1, value=r['company_name'])
        ws2.cell(row=row, column=2, value=r['revenue'])
        ws2.cell(row=row, column=3, value=r['employees'])
        ws2.cell(row=row, column=4, value=r['branches'])
        ws2.cell(row=row, column=5, value=r['description'])
        ws2.cell(row=row, column=6, value=r['ch_status'])
        ws2.cell(row=row, column=7, value=r['financial_source'])

    ws2.column_dimensions['A'].width = 40
    ws2.column_dimensions['E'].width = 60
    ws2.column_dimensions['G'].width = 50

    # ===== SHEET 3: Data Sources =====
    ws3 = wb.create_sheet("Data Sources")

    ws3.cell(row=1, column=1, value="DATA SOURCES AND METHODOLOGY")
    ws3.cell(row=1, column=1).font = Font(bold=True, size=14)

    sources = [
        ("", "", ""),
        ("PRIMARY DATA SOURCE", "", ""),
        ("Companies House API", "api.company-information.service.gov.uk",
         "Company profiles, officers, PSC, charges, filing history - 100% accurate government data"),
        ("", "", ""),
        ("FINANCIAL DATA SOURCES", "", ""),
        ("iXBRL Accounts", "document-api.company-information.service.gov.uk",
         "Parsed from filed accounts - only ~5% of SMEs file full accounts with turnover"),
        ("Pre-Researched Data", "Various - see individual citations",
         "Revenue, employees, branches for 20 major companies from annual reports, news, company websites"),
        ("", "", ""),
        ("WHY MOST COMPANIES DON'T HAVE REVENUE DATA:", "", ""),
        ("", "", "UK law allows small/micro companies to file 'abbreviated' accounts without turnover"),
        ("", "", "Only large companies (>£10.2M turnover) must disclose revenue in filings"),
        ("", "", "Financial databases (Duedil, Experian, etc.) with full data are PAID services"),
        ("", "", ""),
        ("WEBSITE SCRAPING", "", ""),
        ("Company Websites", "Various URLs",
         "Phone numbers, emails, branch counts, locations - scraped directly from company sites"),
        ("", "", ""),
        ("KNOWN DATA SOURCES (for major companies):", "", ""),
        ("Company Annual Reports", "Various", "Official revenue and employee figures"),
        ("UK GlobalDatabase", "uk.globaldatabase.com", "Business database with financial estimates"),
        ("Growjo", "growjo.com", "Revenue and employee estimates"),
        ("Owler", "owler.com", "Company profiles and estimates"),
        ("Tyrepress", "tyrepress.com", "Industry news and company information"),
        ("Commercial Tyre Business", "commercialtyrebusiness.com", "Industry awards, news"),
    ]

    row = 3
    for source, url, description in sources:
        ws3.cell(row=row, column=1, value=source)
        ws3.cell(row=row, column=2, value=url)
        ws3.cell(row=row, column=3, value=description)
        if source and source.isupper():
            ws3.cell(row=row, column=1).font = Font(bold=True)
        row += 1

    ws3.column_dimensions['A'].width = 35
    ws3.column_dimensions['B'].width = 45
    ws3.column_dimensions['C'].width = 75

    # ===== SHEET 4: Summary =====
    ws4 = wb.create_sheet("Summary")

    ws4.cell(row=1, column=1, value="UK TRUCK TYRE COMPANIES DATABASE")
    ws4.cell(row=1, column=1).font = Font(bold=True, size=16)

    ws4.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    summary_data = [
        ("", ""),
        ("STATISTICS", ""),
        ("Total companies", stats['total']),
        ("With Companies House data", stats['with_ch_data']),
        ("With known/researched financial data", stats['with_known_data']),
        ("With iXBRL financial data", stats['with_ixbrl_financials']),
        ("With website scraped data", stats['with_website_data']),
    ]

    row = 5
    for label, value in summary_data:
        ws4.cell(row=row, column=1, value=label)
        if value != "":
            ws4.cell(row=row, column=2, value=value)
        if label.isupper():
            ws4.cell(row=row, column=1).font = Font(bold=True)
        row += 1

    ws4.column_dimensions['A'].width = 45
    ws4.column_dimensions['B'].width = 20

    wb.save('UK_TRUCK_TYRE_COMPLETE.xlsx')
    print("✓ Excel saved: UK_TRUCK_TYRE_COMPLETE.xlsx")

def save_json(results):
    json_results = [{k: v for k, v in r.items() if v} for r in results]
    with open('UK_TRUCK_TYRE_COMPLETE.json', 'w', encoding='utf-8') as f:
        json.dump(json_results, f, indent=2, ensure_ascii=False)
    print("✓ JSON saved: UK_TRUCK_TYRE_COMPLETE.json")

def save_csv(results):
    if not results:
        return
    keys = list(results[0].keys())
    with open('UK_TRUCK_TYRE_COMPLETE.csv', 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=keys)
        writer.writeheader()
        writer.writerows(results)
    print("✓ CSV saved: UK_TRUCK_TYRE_COMPLETE.csv")

# ============================================================================
# MAIN
# ============================================================================
if __name__ == "__main__":
    print()
    results, stats = process_all_companies()
    create_excel_report(results, stats)
    save_json(results)
    save_csv(results)

    print("\n" + "=" * 70)
    print("ALL DONE!")
    print("=" * 70)
    print("Files created:")
    print("  - UK_TRUCK_TYRE_COMPLETE.xlsx (4 sheets)")
    print("  - UK_TRUCK_TYRE_COMPLETE.json")
    print("  - UK_TRUCK_TYRE_COMPLETE.csv")
    print(f"\nFinished: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
