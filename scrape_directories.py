#!/usr/bin/env python3
"""
=============================================================================
SCRAPE TYRE DIRECTORIES FOR REAL UK TRUCK TYRE COMPANIES
=============================================================================
Scrapes real tyre industry directories and association websites to find
actual truck tyre company websites.

Date: January 2026
=============================================================================
"""

import requests
from bs4 import BeautifulSoup
import json
import time
import re
import csv
from datetime import datetime
from urllib.parse import urljoin, urlparse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'en-GB,en;q=0.9',
}

CH_API_KEY = '48d17266-ff2e-425f-9b20-7dcc9b25bb79'
OUTPUT_FILE = 'uk_truck_tyres_SCRAPED_REAL'


def scrape_ntda_members():
    """Scrape NTDA (National Tyre Distributors Association) members"""
    print("    Scraping NTDA members...")
    companies = []

    try:
        # NTDA website
        url = 'https://www.ntda.co.uk/find-a-member'
        r = requests.get(url, headers=HEADERS, timeout=15)

        if r.status_code == 200:
            soup = BeautifulSoup(r.text, 'html.parser')

            # Find member listings
            members = soup.find_all(['div', 'article'], class_=re.compile(r'member|listing|card'))

            for member in members:
                name_tag = member.find(['h2', 'h3', 'h4', 'a'])
                if name_tag:
                    name = name_tag.get_text(strip=True)

                    # Find website link
                    website = ''
                    links = member.find_all('a', href=True)
                    for link in links:
                        href = link.get('href', '')
                        if href.startswith('http') and 'ntda.co.uk' not in href:
                            website = href
                            break

                    if name:
                        companies.append({
                            'name': name,
                            'website': website,
                            'source': 'NTDA'
                        })

        print(f"      Found {len(companies)} from NTDA")

    except Exception as e:
        print(f"      NTDA error: {e}")

    return companies


def scrape_tyresafe():
    """Scrape TyreSafe approved retailers"""
    print("    Scraping TyreSafe...")
    companies = []

    try:
        url = 'https://www.tyresafe.org/find-a-retailer/'
        r = requests.get(url, headers=HEADERS, timeout=15)

        if r.status_code == 200:
            soup = BeautifulSoup(r.text, 'html.parser')

            # Find retailer cards
            retailers = soup.find_all(['div', 'li'], class_=re.compile(r'retailer|member|card'))

            for retailer in retailers:
                name_tag = retailer.find(['h2', 'h3', 'h4', 'span', 'a'])
                if name_tag:
                    name = name_tag.get_text(strip=True)

                    website = ''
                    links = retailer.find_all('a', href=True)
                    for link in links:
                        href = link.get('href', '')
                        if href.startswith('http') and 'tyresafe.org' not in href:
                            website = href
                            break

                    if name and len(name) > 3:
                        companies.append({
                            'name': name,
                            'website': website,
                            'source': 'TyreSafe'
                        })

        print(f"      Found {len(companies)} from TyreSafe")

    except Exception as e:
        print(f"      TyreSafe error: {e}")

    return companies


def scrape_yell_commercial_tyres():
    """Scrape Yell.com for commercial tyre fitters"""
    print("    Scraping Yell.com commercial tyres...")
    companies = []

    try:
        # Search for commercial/truck tyres
        url = 'https://www.yell.com/ucs/UcsSearchAction.do?keywords=commercial+truck+tyres&location=UK'
        r = requests.get(url, headers=HEADERS, timeout=15)

        if r.status_code == 200:
            soup = BeautifulSoup(r.text, 'html.parser')

            # Find business listings
            listings = soup.find_all('div', class_=re.compile(r'businessCapsule|listing'))

            for listing in listings:
                name_tag = listing.find(['h2', 'span'], class_=re.compile(r'businessName|name'))
                if name_tag:
                    name = name_tag.get_text(strip=True)

                    # Find website
                    website = ''
                    website_link = listing.find('a', class_=re.compile(r'website'))
                    if website_link:
                        website = website_link.get('href', '')

                    # Find phone
                    phone = ''
                    phone_tag = listing.find(['span', 'a'], class_=re.compile(r'phone|tel'))
                    if phone_tag:
                        phone = phone_tag.get_text(strip=True)

                    if name:
                        companies.append({
                            'name': name,
                            'website': website,
                            'phone': phone,
                            'source': 'Yell.com'
                        })

        print(f"      Found {len(companies)} from Yell.com")

    except Exception as e:
        print(f"      Yell error: {e}")

    return companies


def scrape_checkatrade_truck_tyres():
    """Scrape Checkatrade for truck tyre fitters"""
    print("    Scraping Checkatrade...")
    companies = []

    try:
        url = 'https://www.checkatrade.com/Search?what=truck+tyre+fitting'
        r = requests.get(url, headers=HEADERS, timeout=15)

        if r.status_code == 200:
            soup = BeautifulSoup(r.text, 'html.parser')

            # Find trade listings
            listings = soup.find_all('div', class_=re.compile(r'trade-card|listing'))

            for listing in listings:
                name_tag = listing.find(['h2', 'h3', 'a'])
                if name_tag:
                    name = name_tag.get_text(strip=True)

                    website = ''
                    links = listing.find_all('a', href=True)
                    for link in links:
                        href = link.get('href', '')
                        if href.startswith('http') and 'checkatrade' not in href:
                            website = href
                            break

                    if name:
                        companies.append({
                            'name': name,
                            'website': website,
                            'source': 'Checkatrade'
                        })

        print(f"      Found {len(companies)} from Checkatrade")

    except Exception as e:
        print(f"      Checkatrade error: {e}")

    return companies


def scrape_freeindex_truck_tyres():
    """Scrape FreeIndex for truck tyre companies"""
    print("    Scraping FreeIndex...")
    companies = []

    try:
        url = 'https://www.freeindex.co.uk/categories/motors/tyres/commercial_vehicle_tyres/'
        r = requests.get(url, headers=HEADERS, timeout=15)

        if r.status_code == 200:
            soup = BeautifulSoup(r.text, 'html.parser')

            # Find business listings
            listings = soup.find_all(['div', 'li'], class_=re.compile(r'listing|result|business'))

            for listing in listings:
                name_tag = listing.find(['h2', 'h3', 'a'])
                if name_tag:
                    name = name_tag.get_text(strip=True)

                    website = ''
                    links = listing.find_all('a', href=True)
                    for link in links:
                        href = link.get('href', '')
                        if href.startswith('http') and 'freeindex' not in href:
                            website = href
                            break

                    if name and len(name) > 3:
                        companies.append({
                            'name': name,
                            'website': website,
                            'source': 'FreeIndex'
                        })

        print(f"      Found {len(companies)} from FreeIndex")

    except Exception as e:
        print(f"      FreeIndex error: {e}")

    return companies


def scrape_thomsonlocal_truck_tyres():
    """Scrape Thomson Local for truck tyre companies"""
    print("    Scraping Thomson Local...")
    companies = []

    try:
        url = 'https://www.thomsonlocal.com/search/commercial-tyres/uk/'
        r = requests.get(url, headers=HEADERS, timeout=15)

        if r.status_code == 200:
            soup = BeautifulSoup(r.text, 'html.parser')

            listings = soup.find_all(['div', 'article'], class_=re.compile(r'listing|result|business'))

            for listing in listings:
                name_tag = listing.find(['h2', 'h3', 'a'])
                if name_tag:
                    name = name_tag.get_text(strip=True)

                    website = ''
                    links = listing.find_all('a', href=True)
                    for link in links:
                        href = link.get('href', '')
                        if href.startswith('http') and 'thomsonlocal' not in href:
                            website = href
                            break

                    if name:
                        companies.append({
                            'name': name,
                            'website': website,
                            'source': 'Thomson Local'
                        })

        print(f"      Found {len(companies)} from Thomson Local")

    except Exception as e:
        print(f"      Thomson Local error: {e}")

    return companies


def verify_website(url):
    """Verify website works and extract contact info"""
    if not url or not url.startswith('http'):
        return None

    try:
        r = requests.get(url, headers=HEADERS, timeout=10, allow_redirects=True)

        if r.status_code == 200:
            text = r.text
            text_lower = text.lower()

            result = {
                'works': True,
                'final_url': r.url,
                'phone': None,
                'email': None,
                'services': [],
                'is_truck_tyre': False
            }

            # Check if truck/commercial tyre related
            truck_words = ['truck', 'hgv', 'lorry', 'commercial', 'fleet', 'trailer']
            tyre_words = ['tyre', 'tire', 'wheel', 'fitting']
            result['is_truck_tyre'] = any(t in text_lower for t in truck_words) and any(t in text_lower for t in tyre_words)

            # Extract phone
            phone_patterns = [
                r'0800[\s\-]?\d{3}[\s\-]?\d{3,4}',
                r'01\d{2,3}[\s\-]?\d{3}[\s\-]?\d{3,4}',
                r'02\d[\s\-]?\d{4}[\s\-]?\d{4}',
                r'03\d{2}[\s\-]?\d{3}[\s\-]?\d{4}',
                r'07\d{3}[\s\-]?\d{3}[\s\-]?\d{3}',
            ]
            for pattern in phone_patterns:
                matches = re.findall(pattern, text)
                for m in matches:
                    cleaned = re.sub(r'[\s\-]', '', m)
                    if 10 <= len(cleaned) <= 12:
                        result['phone'] = m.strip()
                        break
                if result['phone']:
                    break

            # Extract email
            emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', text)
            for email in emails:
                if not any(x in email.lower() for x in ['.png', '.jpg', '.css', '.js', 'example.', 'wix.']):
                    result['email'] = email
                    break

            # Services
            if '24' in text_lower and ('hour' in text_lower or '/7' in text_lower):
                result['services'].append('24hr')
            if 'mobile' in text_lower and 'fitting' in text_lower:
                result['services'].append('Mobile')
            if 'fleet' in text_lower:
                result['services'].append('Fleet')

            return result

        return None
    except:
        return None


def search_companies_house(term):
    """Search Companies House"""
    companies = []

    try:
        url = 'https://api.company-information.service.gov.uk/search/companies'
        params = {'q': term, 'items_per_page': 100}
        r = requests.get(url, auth=(CH_API_KEY, ''), params=params, timeout=15)

        if r.status_code == 200:
            for item in r.json().get('items', []):
                if item.get('company_status') == 'active':
                    companies.append({
                        'name': item.get('title', ''),
                        'company_number': item.get('company_number', ''),
                        'address': item.get('address_snippet', ''),
                        'date_created': item.get('date_of_creation', ''),
                        'source': 'Companies House'
                    })
    except Exception as e:
        print(f"    CH Error: {e}")

    return companies


def estimate_revenue(company):
    """Estimate revenue"""
    name = company.get('name', '').lower()
    services = company.get('services', '')
    verified = company.get('website_verified', False)

    score = 0
    indicators = []

    if any(x in name for x in ['national', 'uk', 'group']):
        score += 30
        indicators.append('National')
    if any(x in name for x in ['wholesale', 'supply']):
        score += 25
        indicators.append('Wholesale')

    if '24hr' in services:
        score += 15
        indicators.append('24hr')
    if 'Fleet' in services:
        score += 20
        indicators.append('Fleet')
    if verified:
        score += 15

    if score >= 60:
        return {'size': 'Large', 'employees': '50-200+', 'revenue': '£5M - £50M+',
                'rev_low': 5000000, 'rev_high': 50000000, 'indicators': ', '.join(indicators) or 'Major'}
    elif score >= 35:
        return {'size': 'Medium', 'employees': '15-50', 'revenue': '£1M - £5M',
                'rev_low': 1000000, 'rev_high': 5000000, 'indicators': ', '.join(indicators) or 'Regional'}
    elif score >= 15:
        return {'size': 'Small-Medium', 'employees': '5-15', 'revenue': '£300K - £1M',
                'rev_low': 300000, 'rev_high': 1000000, 'indicators': ', '.join(indicators) or 'Local'}
    else:
        return {'size': 'Small', 'employees': '1-5', 'revenue': '£50K - £300K',
                'rev_low': 50000, 'rev_high': 300000, 'indicators': ', '.join(indicators) or 'Small'}


def main():
    print("=" * 80)
    print("SCRAPE DIRECTORIES FOR UK TRUCK TYRE COMPANIES")
    print("=" * 80)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    all_companies = []
    seen_names = set()

    # =========================================================================
    # STEP 1: Scrape business directories
    # =========================================================================
    print("[1] SCRAPING BUSINESS DIRECTORIES")
    print("-" * 60)

    # Scrape each directory
    directory_results = []

    directory_results.extend(scrape_ntda_members())
    time.sleep(1)

    directory_results.extend(scrape_tyresafe())
    time.sleep(1)

    directory_results.extend(scrape_yell_commercial_tyres())
    time.sleep(1)

    directory_results.extend(scrape_checkatrade_truck_tyres())
    time.sleep(1)

    directory_results.extend(scrape_freeindex_truck_tyres())
    time.sleep(1)

    directory_results.extend(scrape_thomsonlocal_truck_tyres())

    print(f"\n    Total from directories: {len(directory_results)}")

    # Add unique companies
    for c in directory_results:
        name_lower = c['name'].lower().strip()
        if name_lower and name_lower not in seen_names and len(name_lower) > 3:
            seen_names.add(name_lower)
            all_companies.append(c)

    print(f"    Unique companies: {len(all_companies)}")

    # =========================================================================
    # STEP 2: Search Companies House
    # =========================================================================
    print("\n[2] SEARCHING COMPANIES HOUSE")
    print("-" * 60)

    ch_terms = ['truck tyre', 'truck tyres', 'hgv tyre', 'commercial tyre',
                'lorry tyre', 'mobile truck tyre', 'fleet tyre']

    seen_numbers = set()

    for term in ch_terms:
        print(f"    Searching: '{term}'...", end=' ', flush=True)
        results = search_companies_house(term)

        added = 0
        for c in results:
            cn = c['company_number']
            name_lower = c['name'].lower()

            # Must have truck/commercial AND tyre
            has_truck = any(t in name_lower for t in ['truck', 'hgv', 'commercial', 'lorry', 'fleet'])
            has_tyre = 'tyre' in name_lower or 'tyres' in name_lower

            if has_truck and has_tyre and cn not in seen_numbers:
                if not any(ex in name_lower for ex in ['hire', 'rental', 'sales', 'repair', 'wash']):
                    seen_numbers.add(cn)
                    if name_lower not in seen_names:
                        seen_names.add(name_lower)
                        all_companies.append(c)
                        added += 1

        print(f"Added {added}")
        time.sleep(0.5)

    print(f"\n    Total companies: {len(all_companies)}")

    # =========================================================================
    # STEP 3: Verify websites
    # =========================================================================
    print("\n[3] VERIFYING WEBSITES")
    print("-" * 60)

    verified_count = 0
    companies_with_website = [c for c in all_companies if c.get('website')]

    for i, c in enumerate(companies_with_website):
        url = c['website']
        print(f"    [{i+1}/{len(companies_with_website)}] {urlparse(url).netloc[:35]}...", end=' ', flush=True)

        verification = verify_website(url)

        if verification and verification['works']:
            c['website'] = verification['final_url']
            c['website_verified'] = True
            c['phone'] = c.get('phone') or verification.get('phone', '')
            c['email'] = verification.get('email', '')
            c['services'] = ', '.join(verification.get('services', []))
            c['is_truck_tyre'] = verification.get('is_truck_tyre', False)
            verified_count += 1

            phone = verification.get('phone') or 'No phone'
            truck_mark = '✓' if verification.get('is_truck_tyre') else '?'
            print(f"✓ Working [{truck_mark}] | {phone[:15]}")
        else:
            c['website_verified'] = False
            print("✗ Not working")

        time.sleep(0.3)

    # Set defaults for companies without websites
    for c in all_companies:
        if not c.get('website'):
            c['website_verified'] = False
            c['phone'] = c.get('phone', '')
            c['email'] = c.get('email', '')
            c['services'] = ''

    print(f"\n    Verified: {verified_count}/{len(companies_with_website)}")

    # =========================================================================
    # STEP 4: Estimate revenue
    # =========================================================================
    print("\n[4] ESTIMATING REVENUE")
    print("-" * 60)

    for c in all_companies:
        est = estimate_revenue(c)
        c['size'] = est['size']
        c['employees'] = est['employees']
        c['revenue_estimate'] = est['revenue']
        c['revenue_low'] = est['rev_low']
        c['revenue_high'] = est['rev_high']
        c['revenue_indicators'] = est['indicators']

    all_companies.sort(key=lambda x: x.get('revenue_high', 0), reverse=True)

    # Stats
    total = len(all_companies)
    with_website = len([c for c in all_companies if c.get('website')])
    verified = len([c for c in all_companies if c.get('website_verified')])
    with_phone = len([c for c in all_companies if c.get('phone')])
    truck_related = len([c for c in all_companies if c.get('is_truck_tyre')])

    total_rev_low = sum(c.get('revenue_low', 0) for c in all_companies)
    total_rev_high = sum(c.get('revenue_high', 0) for c in all_companies)

    print(f"    Total: {total}")
    print(f"    With website: {with_website}")
    print(f"    Verified: {verified}")
    print(f"    Truck tyre related: {truck_related}")
    print(f"    With phone: {with_phone}")
    print(f"    Market: £{total_rev_low/1000000:.0f}M - £{total_rev_high/1000000:.0f}M")

    # =========================================================================
    # EXPORT
    # =========================================================================
    print("\n[5] EXPORTING")
    print("-" * 60)

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "UK Truck Tyre Companies"

    headers = ['Company Name', 'Website', 'Verified', 'Truck Tyre', 'Phone', 'Email',
               'Size', 'Employees', 'Revenue', 'Indicators', 'Services',
               'Address', 'Company Number', 'Source']

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    verified_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row, c in enumerate(all_companies, 2):
        ws.cell(row=row, column=1, value=c.get('name', ''))
        ws.cell(row=row, column=2, value=c.get('website', ''))
        v_cell = ws.cell(row=row, column=3, value='Yes' if c.get('website_verified') else 'No')
        if c.get('website_verified'):
            v_cell.fill = verified_fill
        ws.cell(row=row, column=4, value='Yes' if c.get('is_truck_tyre') else 'No')
        ws.cell(row=row, column=5, value=c.get('phone', ''))
        ws.cell(row=row, column=6, value=c.get('email', ''))
        ws.cell(row=row, column=7, value=c.get('size', ''))
        ws.cell(row=row, column=8, value=c.get('employees', ''))
        ws.cell(row=row, column=9, value=c.get('revenue_estimate', ''))
        ws.cell(row=row, column=10, value=c.get('revenue_indicators', ''))
        ws.cell(row=row, column=11, value=c.get('services', ''))
        ws.cell(row=row, column=12, value=c.get('address', ''))
        ws.cell(row=row, column=13, value=c.get('company_number', ''))
        ws.cell(row=row, column=14, value=c.get('source', ''))

    widths = [40, 45, 10, 10, 18, 30, 12, 12, 18, 20, 25, 45, 15, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'
    wb.save(f"{OUTPUT_FILE}.xlsx")
    print(f"    Saved: {OUTPUT_FILE}.xlsx")

    # JSON
    with open(f"{OUTPUT_FILE}.json", 'w', encoding='utf-8') as f:
        json.dump(all_companies, f, indent=2, ensure_ascii=False)
    print(f"    Saved: {OUTPUT_FILE}.json")

    # CSV
    fieldnames = ['name', 'website', 'website_verified', 'is_truck_tyre', 'phone', 'email',
                  'size', 'employees', 'revenue_estimate', 'revenue_low', 'revenue_high',
                  'revenue_indicators', 'services', 'address', 'company_number', 'source']

    with open(f"{OUTPUT_FILE}.csv", 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(all_companies)
    print(f"    Saved: {OUTPUT_FILE}.csv")

    # Summary
    print("\n" + "=" * 80)
    print("COMPLETE")
    print("=" * 80)
    print(f"Total: {total} | Verified websites: {verified}")
    print(f"Market: £{total_rev_low/1000000:.0f}M - £{total_rev_high/1000000:.0f}M")

    print("\n\nVERIFIED TRUCK TYRE COMPANIES:")
    print("-" * 80)

    verified_list = [c for c in all_companies if c.get('website_verified') and c.get('is_truck_tyre')]
    for i, c in enumerate(verified_list[:25], 1):
        print(f"\n{i}. {c['name']}")
        print(f"   Website: {c['website']}")
        if c.get('phone'):
            print(f"   Phone: {c['phone']}")
        print(f"   Size: {c['size']} | Revenue: {c['revenue_estimate']}")


if __name__ == "__main__":
    main()
