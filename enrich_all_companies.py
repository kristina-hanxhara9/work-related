#!/usr/bin/env python3
"""
=============================================================================
ENRICH ALL 846 TRUCK TYRE COMPANIES - FULL DATA EXTRACTION
=============================================================================
This script extracts ALL available data from:
1. Companies House API - company details, directors, owners, addresses
2. Website scraping - phone numbers, emails from working websites

NO ASSUMPTIONS - only real data from verified sources

Date: January 2026
=============================================================================
"""

import requests
from bs4 import BeautifulSoup
import json
import time
import re
import csv
from datetime import datetime
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION
# =============================================================================

CH_API_KEY = '48d17266-ff2e-425f-9b20-7dcc9b25bb79'
CH_BASE_URL = 'https://api.company-information.service.gov.uk'

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'en-GB,en;q=0.9',
}

INPUT_FILE = 'uk_truck_tyres_846_FULLY_VERIFIED.json'
OUTPUT_FILE = 'uk_truck_tyres_ENRICHED'

# =============================================================================
# COMPANIES HOUSE API FUNCTIONS
# =============================================================================

def get_company_profile(company_number):
    """Get full company profile from Companies House"""
    try:
        url = f"{CH_BASE_URL}/company/{company_number}"
        response = requests.get(url, auth=(CH_API_KEY, ''), timeout=15)

        if response.status_code == 200:
            return response.json()
        return None
    except Exception as e:
        print(f"      Error getting profile: {e}")
        return None


def get_company_officers(company_number):
    """Get directors/officers from Companies House"""
    try:
        url = f"{CH_BASE_URL}/company/{company_number}/officers"
        response = requests.get(url, auth=(CH_API_KEY, ''), timeout=15)

        if response.status_code == 200:
            return response.json()
        return None
    except Exception as e:
        print(f"      Error getting officers: {e}")
        return None


def get_company_psc(company_number):
    """Get Persons with Significant Control (owners) from Companies House"""
    try:
        url = f"{CH_BASE_URL}/company/{company_number}/persons-with-significant-control"
        response = requests.get(url, auth=(CH_API_KEY, ''), timeout=15)

        if response.status_code == 200:
            return response.json()
        return None
    except Exception as e:
        print(f"      Error getting PSC: {e}")
        return None


def get_company_charges(company_number):
    """Get charges/mortgages from Companies House"""
    try:
        url = f"{CH_BASE_URL}/company/{company_number}/charges"
        response = requests.get(url, auth=(CH_API_KEY, ''), timeout=15)

        if response.status_code == 200:
            return response.json()
        return None
    except Exception as e:
        return None


def get_company_filing_history(company_number):
    """Get recent filing history from Companies House"""
    try:
        url = f"{CH_BASE_URL}/company/{company_number}/filing-history"
        params = {'items_per_page': 10}
        response = requests.get(url, params=params, auth=(CH_API_KEY, ''), timeout=15)

        if response.status_code == 200:
            return response.json()
        return None
    except Exception as e:
        return None


# =============================================================================
# WEBSITE SCRAPING FUNCTIONS
# =============================================================================

def scrape_website(url):
    """Scrape website for contact information"""
    if not url or not url.startswith('http'):
        return None

    result = {
        'website_works': False,
        'final_url': None,
        'phones_found': [],
        'emails_found': [],
        'social_media': {},
        'page_title': None
    }

    try:
        response = requests.get(url, headers=HEADERS, timeout=15, allow_redirects=True)

        if response.status_code == 200:
            result['website_works'] = True
            result['final_url'] = response.url

            soup = BeautifulSoup(response.text, 'html.parser')
            text = response.text

            # Get page title
            title_tag = soup.find('title')
            if title_tag:
                result['page_title'] = title_tag.get_text(strip=True)[:100]

            # Extract ALL phone numbers found
            phone_patterns = [
                r'0800[\s\-\.]?\d{3}[\s\-\.]?\d{3,4}',
                r'0808[\s\-\.]?\d{3}[\s\-\.]?\d{3,4}',
                r'01onal\d{2,3}[\s\-\.]?\d{3}[\s\-\.]?\d{3,4}',
                r'02\d[\s\-\.]?\d{4}[\s\-\.]?\d{4}',
                r'03\d{2}[\s\-\.]?\d{3}[\s\-\.]?\d{4}',
                r'07\d{3}[\s\-\.]?\d{3}[\s\-\.]?\d{3}',
                r'\+44[\s\-\.]?\d{2,4}[\s\-\.]?\d{3,4}[\s\-\.]?\d{3,4}',
                r'01\d{2,3}[\s\-\.]?\d{6}',
                r'01\d{3}[\s\-\.]?\d{5,6}',
            ]

            phones_found = set()
            for pattern in phone_patterns:
                matches = re.findall(pattern, text)
                for match in matches:
                    cleaned = re.sub(r'[\s\-\.]', '', match)
                    # Validate phone number
                    if 10 <= len(cleaned) <= 14:
                        if not cleaned.startswith('00000'):
                            phones_found.add(match.strip())

            result['phones_found'] = list(phones_found)[:5]  # Max 5 numbers

            # Extract ALL email addresses found
            email_pattern = r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
            emails = re.findall(email_pattern, text)

            valid_emails = set()
            invalid_domains = ['.png', '.jpg', '.jpeg', '.gif', '.css', '.js',
                             'example.com', 'wix.com', 'sentry.io', 'wordpress.com',
                             'gravatar.com', 'schema.org']

            for email in emails:
                email_lower = email.lower()
                if not any(invalid in email_lower for invalid in invalid_domains):
                    valid_emails.add(email)

            result['emails_found'] = list(valid_emails)[:5]  # Max 5 emails

            # Extract social media links
            social_patterns = {
                'facebook': r'facebook\.com/[a-zA-Z0-9._-]+',
                'twitter': r'twitter\.com/[a-zA-Z0-9_]+',
                'linkedin': r'linkedin\.com/(?:company|in)/[a-zA-Z0-9_-]+',
                'instagram': r'instagram\.com/[a-zA-Z0-9._]+',
            }

            for platform, pattern in social_patterns.items():
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    result['social_media'][platform] = match.group(0)

            return result

    except Exception as e:
        result['error'] = str(e)
        return result

    return result


# =============================================================================
# DATA PROCESSING FUNCTIONS
# =============================================================================

def extract_address_parts(address_obj):
    """Extract address components from Companies House address object"""
    if not address_obj:
        return {}

    return {
        'premises': address_obj.get('premises', ''),
        'address_line_1': address_obj.get('address_line_1', ''),
        'address_line_2': address_obj.get('address_line_2', ''),
        'locality': address_obj.get('locality', ''),
        'region': address_obj.get('region', ''),
        'postal_code': address_obj.get('postal_code', ''),
        'country': address_obj.get('country', ''),
        'full_address': ', '.join(filter(None, [
            address_obj.get('premises'),
            address_obj.get('address_line_1'),
            address_obj.get('address_line_2'),
            address_obj.get('locality'),
            address_obj.get('region'),
            address_obj.get('postal_code')
        ]))
    }


def extract_ownership_level(natures_of_control):
    """Extract ownership percentage from PSC control types"""
    if not natures_of_control:
        return ''

    for control in natures_of_control:
        if '75-to-100' in control:
            return '75-100%'
        elif '50-to-75' in control:
            return '50-75%'
        elif '25-to-50' in control:
            return '25-50%'

    return 'Over 25%'


def get_sic_description(sic_code):
    """Get description for SIC code (common tyre-related codes)"""
    sic_descriptions = {
        '45310': 'Wholesale of motor vehicle parts and accessories',
        '45320': 'Retail of motor vehicle parts and accessories',
        '45200': 'Maintenance and repair of motor vehicles',
        '45400': 'Sale, maintenance and repair of motorcycles',
        '46900': 'Non-specialised wholesale trade',
        '47300': 'Retail sale of automotive fuel',
        '49410': 'Freight transport by road',
        '52290': 'Other transportation support activities',
        '22110': 'Manufacture of rubber tyres and tubes',
        '22190': 'Manufacture of other rubber products',
    }
    return sic_descriptions.get(sic_code, '')


# =============================================================================
# MAIN ENRICHMENT FUNCTION
# =============================================================================

def enrich_company(company):
    """Enrich a single company with all available data"""

    enriched = {
        # Original data
        'original_name': company.get('name', ''),
        'original_website': company.get('website', ''),
        'original_phone': company.get('phone', ''),
        'original_email': company.get('email', ''),
        'original_address': company.get('ch_address', ''),
        'original_business_type': company.get('businessType', ''),
        'original_region': company.get('region', ''),
        'original_source': company.get('source', ''),
        'original_verification_status': company.get('verification_status', ''),
        'original_verification_method': company.get('verification_method', ''),

        # Companies House data - will be populated
        'ch_company_name': '',
        'ch_company_number': company.get('ch_number', ''),
        'ch_status': '',
        'ch_date_created': '',
        'ch_company_type': '',
        'ch_jurisdiction': '',
        'ch_sic_codes': '',
        'ch_sic_descriptions': '',
        'ch_has_charges': '',
        'ch_has_insolvency': '',
        'ch_previous_names': '',
        'ch_accounts_next_due': '',
        'ch_confirmation_next_due': '',

        # Address from CH
        'ch_premises': '',
        'ch_address_line_1': '',
        'ch_address_line_2': '',
        'ch_locality': '',
        'ch_region': '',
        'ch_postal_code': '',
        'ch_country': '',
        'ch_full_address': '',

        # Directors
        'ch_director_count': 0,
        'ch_directors': '',
        'ch_director_1_name': '',
        'ch_director_1_role': '',
        'ch_director_1_appointed': '',
        'ch_director_1_nationality': '',
        'ch_director_2_name': '',
        'ch_director_2_role': '',

        # Owners (PSC)
        'ch_owner_count': 0,
        'ch_owners': '',
        'ch_owner_1_name': '',
        'ch_owner_1_ownership': '',
        'ch_owner_1_nationality': '',
        'ch_owner_2_name': '',
        'ch_owner_2_ownership': '',

        # Website scraped data
        'website_works': '',
        'website_final_url': '',
        'website_page_title': '',
        'website_phones': '',
        'website_phone_1': '',
        'website_phone_2': '',
        'website_emails': '',
        'website_email_1': '',
        'website_email_2': '',
        'website_facebook': '',
        'website_twitter': '',
        'website_linkedin': '',
        'website_instagram': '',

        # Metadata
        'enriched_date': datetime.now().isoformat(),
        'data_sources': []
    }

    ch_number = company.get('ch_number', '').strip()

    # =========================================================================
    # PART 1: Companies House API Data
    # =========================================================================

    if ch_number:
        enriched['data_sources'].append('Companies House API')

        # Get company profile
        profile = get_company_profile(ch_number)

        if profile:
            enriched['ch_company_name'] = profile.get('company_name', '')
            enriched['ch_status'] = profile.get('company_status', '')
            enriched['ch_date_created'] = profile.get('date_of_creation', '')
            enriched['ch_company_type'] = profile.get('type', '')
            enriched['ch_jurisdiction'] = profile.get('jurisdiction', '')
            enriched['ch_has_charges'] = 'Yes' if profile.get('has_charges') else 'No'
            enriched['ch_has_insolvency'] = 'Yes' if profile.get('has_insolvency_history') else 'No'

            # SIC codes
            sic_codes = profile.get('sic_codes', [])
            enriched['ch_sic_codes'] = ', '.join(sic_codes)
            enriched['ch_sic_descriptions'] = ', '.join([get_sic_description(s) for s in sic_codes if get_sic_description(s)])

            # Previous names
            prev_names = profile.get('previous_company_names', [])
            enriched['ch_previous_names'] = ', '.join([n.get('name', '') for n in prev_names])

            # Accounts info
            accounts = profile.get('accounts', {})
            enriched['ch_accounts_next_due'] = accounts.get('next_due', '')

            confirmation = profile.get('confirmation_statement', {})
            enriched['ch_confirmation_next_due'] = confirmation.get('next_due', '')

            # Address
            address = extract_address_parts(profile.get('registered_office_address', {}))
            enriched['ch_premises'] = address.get('premises', '')
            enriched['ch_address_line_1'] = address.get('address_line_1', '')
            enriched['ch_address_line_2'] = address.get('address_line_2', '')
            enriched['ch_locality'] = address.get('locality', '')
            enriched['ch_region'] = address.get('region', '')
            enriched['ch_postal_code'] = address.get('postal_code', '')
            enriched['ch_country'] = address.get('country', '')
            enriched['ch_full_address'] = address.get('full_address', '')

        time.sleep(0.2)

        # Get officers/directors
        officers_data = get_company_officers(ch_number)

        if officers_data:
            officers = officers_data.get('items', [])
            active_directors = [o for o in officers if not o.get('resigned_on')]

            enriched['ch_director_count'] = len(active_directors)
            enriched['ch_directors'] = '; '.join([o.get('name', '') for o in active_directors[:5]])

            if len(active_directors) >= 1:
                d1 = active_directors[0]
                enriched['ch_director_1_name'] = d1.get('name', '')
                enriched['ch_director_1_role'] = d1.get('officer_role', '')
                enriched['ch_director_1_appointed'] = d1.get('appointed_on', '')
                enriched['ch_director_1_nationality'] = d1.get('nationality', '')

            if len(active_directors) >= 2:
                d2 = active_directors[1]
                enriched['ch_director_2_name'] = d2.get('name', '')
                enriched['ch_director_2_role'] = d2.get('officer_role', '')

        time.sleep(0.2)

        # Get PSC (owners)
        psc_data = get_company_psc(ch_number)

        if psc_data:
            pscs = psc_data.get('items', [])
            active_pscs = [p for p in pscs if not p.get('ceased')]

            enriched['ch_owner_count'] = len(active_pscs)
            enriched['ch_owners'] = '; '.join([p.get('name', '') for p in active_pscs[:5]])

            if len(active_pscs) >= 1:
                p1 = active_pscs[0]
                enriched['ch_owner_1_name'] = p1.get('name', '')
                enriched['ch_owner_1_ownership'] = extract_ownership_level(p1.get('natures_of_control', []))
                enriched['ch_owner_1_nationality'] = p1.get('nationality', '')

            if len(active_pscs) >= 2:
                p2 = active_pscs[1]
                enriched['ch_owner_2_name'] = p2.get('name', '')
                enriched['ch_owner_2_ownership'] = extract_ownership_level(p2.get('natures_of_control', []))

        time.sleep(0.2)

    # =========================================================================
    # PART 2: Website Scraping
    # =========================================================================

    website_url = company.get('website', '')

    if website_url and website_url.startswith('http'):
        website_data = scrape_website(website_url)

        if website_data:
            enriched['data_sources'].append('Website Scraping')

            enriched['website_works'] = 'Yes' if website_data.get('website_works') else 'No'
            enriched['website_final_url'] = website_data.get('final_url', '')
            enriched['website_page_title'] = website_data.get('page_title', '')

            # Phones
            phones = website_data.get('phones_found', [])
            enriched['website_phones'] = ', '.join(phones)
            if len(phones) >= 1:
                enriched['website_phone_1'] = phones[0]
            if len(phones) >= 2:
                enriched['website_phone_2'] = phones[1]

            # Emails
            emails = website_data.get('emails_found', [])
            enriched['website_emails'] = ', '.join(emails)
            if len(emails) >= 1:
                enriched['website_email_1'] = emails[0]
            if len(emails) >= 2:
                enriched['website_email_2'] = emails[1]

            # Social media
            social = website_data.get('social_media', {})
            enriched['website_facebook'] = social.get('facebook', '')
            enriched['website_twitter'] = social.get('twitter', '')
            enriched['website_linkedin'] = social.get('linkedin', '')
            enriched['website_instagram'] = social.get('instagram', '')

    # Convert data_sources to string
    enriched['data_sources'] = ', '.join(enriched['data_sources'])

    return enriched


# =============================================================================
# EXPORT FUNCTIONS
# =============================================================================

def export_to_excel(companies, filename):
    """Export enriched data to Excel with formatting"""

    wb = Workbook()
    ws = wb.active
    ws.title = "Enriched Companies"

    # Define column groups for headers
    columns = [
        # Basic Info
        ('Company Name', 'ch_company_name'),
        ('CH Number', 'ch_company_number'),
        ('Status', 'ch_status'),
        ('Date Created', 'ch_date_created'),
        ('Company Type', 'ch_company_type'),

        # Address
        ('Full Address', 'ch_full_address'),
        ('Postcode', 'ch_postal_code'),
        ('Locality', 'ch_locality'),
        ('Region', 'ch_region'),
        ('Country', 'ch_country'),

        # Business Info
        ('SIC Codes', 'ch_sic_codes'),
        ('SIC Descriptions', 'ch_sic_descriptions'),
        ('Has Charges/Loans', 'ch_has_charges'),
        ('Has Insolvency', 'ch_has_insolvency'),
        ('Jurisdiction', 'ch_jurisdiction'),

        # Directors
        ('Director Count', 'ch_director_count'),
        ('All Directors', 'ch_directors'),
        ('Director 1 Name', 'ch_director_1_name'),
        ('Director 1 Role', 'ch_director_1_role'),
        ('Director 1 Appointed', 'ch_director_1_appointed'),
        ('Director 1 Nationality', 'ch_director_1_nationality'),
        ('Director 2 Name', 'ch_director_2_name'),

        # Owners
        ('Owner Count', 'ch_owner_count'),
        ('All Owners', 'ch_owners'),
        ('Owner 1 Name', 'ch_owner_1_name'),
        ('Owner 1 Ownership %', 'ch_owner_1_ownership'),
        ('Owner 1 Nationality', 'ch_owner_1_nationality'),
        ('Owner 2 Name', 'ch_owner_2_name'),

        # Website Data
        ('Website Works', 'website_works'),
        ('Website URL', 'website_final_url'),
        ('Page Title', 'website_page_title'),
        ('Phone 1 (from website)', 'website_phone_1'),
        ('Phone 2 (from website)', 'website_phone_2'),
        ('All Phones Found', 'website_phones'),
        ('Email 1 (from website)', 'website_email_1'),
        ('Email 2 (from website)', 'website_email_2'),
        ('All Emails Found', 'website_emails'),
        ('Facebook', 'website_facebook'),
        ('LinkedIn', 'website_linkedin'),
        ('Twitter', 'website_twitter'),

        # Filing Info
        ('Accounts Next Due', 'ch_accounts_next_due'),
        ('Confirmation Next Due', 'ch_confirmation_next_due'),
        ('Previous Names', 'ch_previous_names'),

        # Original Data
        ('Original Name', 'original_name'),
        ('Original Website', 'original_website'),
        ('Original Phone', 'original_phone'),
        ('Original Email', 'original_email'),
        ('Original Business Type', 'original_business_type'),
        ('Verification Status', 'original_verification_status'),

        # Metadata
        ('Data Sources', 'data_sources'),
        ('Enriched Date', 'enriched_date'),
    ]

    # Styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    active_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Write headers
    for col, (header, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    # Write data
    for row, company in enumerate(companies, 2):
        for col, (_, field) in enumerate(columns, 1):
            value = company.get(field, '')
            cell = ws.cell(row=row, column=col, value=value)

            # Highlight active companies
            if field == 'ch_status' and value == 'active':
                cell.fill = active_fill

    # Set column widths
    widths = {
        'A': 40,  # Company Name
        'B': 12,  # CH Number
        'C': 10,  # Status
        'D': 12,  # Date Created
        'E': 15,  # Company Type
        'F': 50,  # Full Address
        'G': 10,  # Postcode
        'H': 15,  # Locality
        'I': 15,  # Region
        'J': 12,  # Country
        'K': 15,  # SIC Codes
        'L': 40,  # SIC Descriptions
        'M': 12,  # Has Charges
        'N': 12,  # Has Insolvency
        'O': 15,  # Jurisdiction
        'P': 12,  # Director Count
        'Q': 50,  # All Directors
        'R': 25,  # Director 1 Name
        'S': 12,  # Director 1 Role
        'T': 12,  # Director 1 Appointed
        'U': 15,  # Director 1 Nationality
        'V': 25,  # Director 2 Name
        'W': 12,  # Owner Count
        'X': 50,  # All Owners
        'Y': 25,  # Owner 1 Name
        'Z': 12,  # Owner 1 Ownership
        'AA': 15,  # Owner 1 Nationality
        'AB': 25,  # Owner 2 Name
        'AC': 12,  # Website Works
        'AD': 40,  # Website URL
        'AE': 40,  # Page Title
        'AF': 18,  # Phone 1
        'AG': 18,  # Phone 2
        'AH': 40,  # All Phones
        'AI': 30,  # Email 1
        'AJ': 30,  # Email 2
        'AK': 50,  # All Emails
        'AL': 30,  # Facebook
        'AM': 30,  # LinkedIn
        'AN': 25,  # Twitter
    }

    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Add autofilter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(companies) + 1}"

    # Add summary sheet
    ws_summary = wb.create_sheet("Summary")

    total = len(companies)
    with_ch = len([c for c in companies if c.get('ch_company_number')])
    active = len([c for c in companies if c.get('ch_status') == 'active'])
    with_website = len([c for c in companies if c.get('website_works') == 'Yes'])
    with_phone = len([c for c in companies if c.get('website_phone_1')])
    with_email = len([c for c in companies if c.get('website_email_1')])
    with_directors = len([c for c in companies if c.get('ch_director_count', 0) > 0])
    with_owners = len([c for c in companies if c.get('ch_owner_count', 0) > 0])

    summary_data = [
        ['UK TRUCK TYRE COMPANIES - ENRICHED DATA', ''],
        ['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['', ''],
        ['COMPANIES HOUSE DATA', ''],
        ['Total Companies', total],
        ['With CH Number', with_ch],
        ['Active Status', active],
        ['With Directors Listed', with_directors],
        ['With Owners Listed', with_owners],
        ['', ''],
        ['WEBSITE DATA', ''],
        ['Working Websites', with_website],
        ['With Phone Number', with_phone],
        ['With Email Address', with_email],
        ['', ''],
        ['DATA QUALITY', ''],
        ['CH Data %', f"{with_ch/total*100:.1f}%" if total > 0 else '0%'],
        ['Website Data %', f"{with_website/total*100:.1f}%" if total > 0 else '0%'],
        ['Contact Data %', f"{max(with_phone, with_email)/total*100:.1f}%" if total > 0 else '0%'],
    ]

    for row, (label, value) in enumerate(summary_data, 1):
        ws_summary.cell(row=row, column=1, value=label)
        ws_summary.cell(row=row, column=2, value=value)

    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20

    wb.save(filename)


def export_to_csv(companies, filename):
    """Export enriched data to CSV"""

    if not companies:
        return

    fieldnames = list(companies[0].keys())

    with open(filename, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(companies)


def export_to_json(companies, filename):
    """Export enriched data to JSON"""

    output = {
        'generated': datetime.now().isoformat(),
        'total_companies': len(companies),
        'companies': companies
    }

    with open(filename, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2, ensure_ascii=False)


# =============================================================================
# MAIN FUNCTION
# =============================================================================

def main():
    print("=" * 80)
    print("ENRICH ALL TRUCK TYRE COMPANIES - FULL DATA EXTRACTION")
    print("=" * 80)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    # Load companies
    print(f"Loading companies from {INPUT_FILE}...")
    with open(INPUT_FILE, 'r', encoding='utf-8') as f:
        companies = json.load(f)

    print(f"Loaded {len(companies)} companies")

    # Count what we have
    with_ch = len([c for c in companies if c.get('ch_number')])
    with_website = len([c for c in companies if c.get('website', '').startswith('http')])

    print(f"  - With CH number: {with_ch}")
    print(f"  - With website URL: {with_website}")
    print()

    # Enrich each company
    print("[1] ENRICHING COMPANIES")
    print("-" * 60)

    enriched_companies = []

    for i, company in enumerate(companies):
        name = company.get('name', 'Unknown')[:40]

        if (i + 1) % 25 == 0 or i == 0:
            print(f"\n  Processing {i + 1}/{len(companies)}...")

        # Enrich
        enriched = enrich_company(company)
        enriched_companies.append(enriched)

        # Show progress for companies with CH numbers
        if company.get('ch_number'):
            status = enriched.get('ch_status', 'N/A')
            phones = enriched.get('website_phone_1', '')
            print(f"    {name}... Status: {status}, Phone: {phones if phones else 'N/A'}")

        # Rate limiting
        time.sleep(0.1)

    # Summary stats
    print("\n" + "-" * 60)

    total = len(enriched_companies)
    with_ch_data = len([c for c in enriched_companies if c.get('ch_company_name')])
    active = len([c for c in enriched_companies if c.get('ch_status') == 'active'])
    with_website = len([c for c in enriched_companies if c.get('website_works') == 'Yes'])
    with_phone = len([c for c in enriched_companies if c.get('website_phone_1')])
    with_email = len([c for c in enriched_companies if c.get('website_email_1')])
    with_directors = len([c for c in enriched_companies if c.get('ch_director_count', 0) > 0])
    with_owners = len([c for c in enriched_companies if c.get('ch_owner_count', 0) > 0])

    print(f"\n  ENRICHMENT RESULTS:")
    print(f"    Total companies: {total}")
    print(f"    With CH data: {with_ch_data}")
    print(f"    Active status: {active}")
    print(f"    With directors: {with_directors}")
    print(f"    With owners: {with_owners}")
    print(f"    Working websites: {with_website}")
    print(f"    With phone (scraped): {with_phone}")
    print(f"    With email (scraped): {with_email}")

    # Export
    print("\n[2] EXPORTING DATA")
    print("-" * 60)

    export_to_excel(enriched_companies, f"{OUTPUT_FILE}.xlsx")
    print(f"  Saved: {OUTPUT_FILE}.xlsx")

    export_to_json(enriched_companies, f"{OUTPUT_FILE}.json")
    print(f"  Saved: {OUTPUT_FILE}.json")

    export_to_csv(enriched_companies, f"{OUTPUT_FILE}.csv")
    print(f"  Saved: {OUTPUT_FILE}.csv")

    # Final summary
    print("\n" + "=" * 80)
    print("ENRICHMENT COMPLETE")
    print("=" * 80)
    print(f"\nTotal: {total} companies enriched")
    print(f"CH Data: {with_ch_data} ({with_ch_data/total*100:.1f}%)")
    print(f"Website Data: {with_website} ({with_website/total*100:.1f}%)")
    print(f"Contact Data: {max(with_phone, with_email)} ({max(with_phone, with_email)/total*100:.1f}%)")

    print("\n\nSAMPLE ENRICHED COMPANIES:")
    print("-" * 60)

    # Show sample of well-enriched companies
    sample = [c for c in enriched_companies if c.get('ch_status') == 'active' and c.get('website_phone_1')][:5]

    for c in sample:
        print(f"\n{c.get('ch_company_name', c.get('original_name', 'N/A'))}")
        print(f"  CH: {c.get('ch_company_number', 'N/A')}")
        print(f"  Status: {c.get('ch_status', 'N/A')}")
        print(f"  Created: {c.get('ch_date_created', 'N/A')}")
        print(f"  Address: {c.get('ch_full_address', 'N/A')[:60]}")
        print(f"  Directors: {c.get('ch_directors', 'N/A')[:50]}")
        print(f"  Owners: {c.get('ch_owners', 'N/A')[:50]}")
        print(f"  Phone: {c.get('website_phone_1', 'N/A')}")
        print(f"  Email: {c.get('website_email_1', 'N/A')}")


if __name__ == "__main__":
    main()
