#!/usr/bin/env python3
"""
=============================================================================
FIND REAL WEBSITES FOR UK TRUCK TYRE COMPANIES
=============================================================================
Takes company names from Companies House and searches the web to find
their REAL websites. No guessing URLs - actually search for them.

Date: January 2026
=============================================================================
"""

import requests
from bs4 import BeautifulSoup
import json
import time
import re
import csv
from datetime import datetime
from urllib.parse import urlparse, quote_plus

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-GB,en;q=0.9',
}

CH_API_KEY = '48d17266-ff2e-425f-9b20-7dcc9b25bb79'
OUTPUT_FILE = 'uk_truck_tyres_REAL_WEBSITES'


def search_companies_house(term):
    """Search Companies House for truck tyre companies"""
    companies = []

    try:
        url = 'https://api.company-information.service.gov.uk/search/companies'
        params = {'q': term, 'items_per_page': 100}

        r = requests.get(url, auth=(CH_API_KEY, ''), params=params, timeout=15)

        if r.status_code == 200:
            for item in r.json().get('items', []):
                if item.get('company_status') == 'active':
                    companies.append({
                        'name': item.get('title', ''),
                        'company_number': item.get('company_number', ''),
                        'address': item.get('address_snippet', ''),
                        'date_created': item.get('date_of_creation', ''),
                        'source': 'Companies House'
                    })
    except Exception as e:
        print(f"    Error: {e}")

    return companies


def search_for_website_bing(company_name):
    """Search Bing for the company's website"""

    # Clean company name for search
    search_name = company_name.replace(' LIMITED', '').replace(' LTD', '').replace('(', '').replace(')', '').strip()

    try:
        # Search Bing
        search_url = f'https://www.bing.com/search?q={quote_plus(search_name + " UK official website")}'

        response = requests.get(search_url, headers=HEADERS, timeout=10)

        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')

            # Find search results
            results = soup.find_all('li', class_='b_algo')

            skip_domains = [
                'facebook.com', 'linkedin.com', 'twitter.com', 'instagram.com',
                'youtube.com', 'pinterest.com', 'tiktok.com',
                'yell.com', 'yelp.com', 'checkatrade.com', 'trustatrader.com',
                'companieshouse.gov.uk', 'find-and-update.company-information',
                'endole.co.uk', 'duedil.com', 'opencorporates.com',
                'amazon.co.uk', 'ebay.co.uk', 'gumtree.com',
                'wikipedia.org', 'bbc.co.uk',
                'gov.uk', 'companieslist.co.uk'
            ]

            for result in results[:5]:
                link = result.find('a')
                if link:
                    href = link.get('href', '')

                    if href.startswith('http'):
                        domain = urlparse(href).netloc.lower()

                        # Skip directories and social media
                        if any(skip in domain for skip in skip_domains):
                            continue

                        return href

    except Exception as e:
        pass

    return None


def search_for_website_google(company_name):
    """Try Google search via alternative method"""

    search_name = company_name.replace(' LIMITED', '').replace(' LTD', '').strip()

    try:
        # Try DuckDuckGo HTML version
        search_url = f'https://html.duckduckgo.com/html/?q={quote_plus(search_name + " UK")}'

        response = requests.get(search_url, headers=HEADERS, timeout=10)

        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')

            # Find result links
            results = soup.find_all('a', class_='result__a')

            skip_domains = [
                'facebook.com', 'linkedin.com', 'twitter.com',
                'yell.com', 'yelp.com', 'checkatrade.com',
                'companieshouse', 'endole', 'duedil',
                'youtube.com', 'wikipedia.org'
            ]

            for result in results[:5]:
                href = result.get('href', '')

                if href.startswith('http'):
                    if not any(skip in href.lower() for skip in skip_domains):
                        return href

    except:
        pass

    return None


def verify_website(url):
    """Verify website works and extract contact info"""
    if not url:
        return None

    try:
        response = requests.get(url, headers=HEADERS, timeout=10, allow_redirects=True)

        if response.status_code == 200:
            text = response.text
            text_lower = text.lower()

            result = {
                'works': True,
                'final_url': response.url,
                'phone': None,
                'email': None,
                'services': [],
                'is_tyre_related': False
            }

            # Check if tyre related
            tyre_keywords = ['tyre', 'tire', 'wheel', 'fitting', 'truck', 'hgv', 'commercial', 'fleet', 'lorry']
            result['is_tyre_related'] = sum(1 for kw in tyre_keywords if kw in text_lower) >= 2

            # Extract phone
            phone_patterns = [
                r'0800[\s\-]?\d{3}[\s\-]?\d{3,4}',
                r'0808[\s\-]?\d{3}[\s\-]?\d{3,4}',
                r'01\d{2,3}[\s\-]?\d{3}[\s\-]?\d{3,4}',
                r'02\d[\s\-]?\d{4}[\s\-]?\d{4}',
                r'03\d{2}[\s\-]?\d{3}[\s\-]?\d{4}',
                r'07\d{3}[\s\-]?\d{3}[\s\-]?\d{3}',
            ]
            for pattern in phone_patterns:
                matches = re.findall(pattern, text)
                for m in matches:
                    cleaned = re.sub(r'[\s\-]', '', m)
                    if 10 <= len(cleaned) <= 12 and not cleaned.startswith('00000'):
                        result['phone'] = m.strip()
                        break
                if result['phone']:
                    break

            # Extract email
            emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', text)
            for email in emails:
                if not any(x in email.lower() for x in ['.png', '.jpg', '.css', '.js', 'example.', 'wix.', 'sentry']):
                    result['email'] = email
                    break

            # Services
            if '24' in text_lower and ('hour' in text_lower or '/7' in text_lower):
                result['services'].append('24hr')
            if 'mobile' in text_lower and ('fitting' in text_lower or 'service' in text_lower):
                result['services'].append('Mobile')
            if 'fleet' in text_lower:
                result['services'].append('Fleet')
            if 'emergency' in text_lower or 'breakdown' in text_lower:
                result['services'].append('Emergency')

            return result

        return None

    except:
        return None


def estimate_revenue(company):
    """Estimate revenue based on characteristics"""
    name = company.get('name', '').lower()
    services = company.get('services', '')
    verified = company.get('website_verified', False)

    score = 0
    indicators = []

    if any(x in name for x in ['national', 'uk', 'british', 'group']):
        score += 30
        indicators.append('National')
    if any(x in name for x in ['wholesale', 'supply', 'distribution']):
        score += 25
        indicators.append('Wholesale')
    if 'network' in name:
        score += 20
        indicators.append('Network')

    if '24hr' in services:
        score += 15
        indicators.append('24hr')
    if 'Fleet' in services:
        score += 20
        indicators.append('Fleet')
    if 'Mobile' in services:
        score += 10

    if verified:
        score += 15

    if score >= 60:
        return {'size': 'Large', 'employees': '50-200+', 'revenue': '£5M - £50M+',
                'rev_low': 5000000, 'rev_high': 50000000, 'indicators': ', '.join(indicators) or 'Major'}
    elif score >= 35:
        return {'size': 'Medium', 'employees': '15-50', 'revenue': '£1M - £5M',
                'rev_low': 1000000, 'rev_high': 5000000, 'indicators': ', '.join(indicators) or 'Regional'}
    elif score >= 15:
        return {'size': 'Small-Medium', 'employees': '5-15', 'revenue': '£300K - £1M',
                'rev_low': 300000, 'rev_high': 1000000, 'indicators': ', '.join(indicators) or 'Local'}
    else:
        return {'size': 'Small', 'employees': '1-5', 'revenue': '£50K - £300K',
                'rev_low': 50000, 'rev_high': 300000, 'indicators': ', '.join(indicators) or 'Small'}


def main():
    print("=" * 80)
    print("FIND REAL WEBSITES FOR UK TRUCK TYRE COMPANIES")
    print("=" * 80)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    print("This scraper searches for REAL websites for each company.")
    print("No guessing - actually searching the web for each company.")
    print()

    all_companies = []
    seen_numbers = set()

    # =========================================================================
    # STEP 1: Get all truck tyre companies from Companies House
    # =========================================================================
    print("[1] SEARCHING COMPANIES HOUSE FOR TRUCK TYRE COMPANIES")
    print("-" * 60)

    ch_terms = [
        'truck tyre', 'truck tyres',
        'hgv tyre', 'hgv tyres',
        'commercial tyre', 'commercial tyres',
        'lorry tyre', 'lorry tyres',
        'mobile truck tyre',
        'fleet tyre', 'fleet tyres',
        'trailer tyre', 'trailer tyres',
    ]

    for term in ch_terms:
        print(f"    Searching: '{term}'...", end=' ', flush=True)
        results = search_companies_house(term)

        added = 0
        for c in results:
            cn = c['company_number']
            if cn not in seen_numbers:
                name = c['name'].lower()

                # Must have truck/hgv/commercial/lorry AND tyre
                truck_words = ['truck', 'hgv', 'commercial', 'lorry', 'fleet', 'trailer', 'mobile']
                tyre_words = ['tyre', 'tyres', 'tire', 'tires']

                has_truck = any(t in name for t in truck_words)
                has_tyre = any(t in name for t in tyre_words)

                if has_truck and has_tyre:
                    # Exclude non-tyre businesses
                    exclude = ['hire', 'rental', 'sales', 'repair only', 'parts', 'wash', 'training']
                    if not any(ex in name for ex in exclude):
                        seen_numbers.add(cn)
                        all_companies.append(c)
                        added += 1

        print(f"Added {added} new companies (total: {len(all_companies)})")
        time.sleep(0.5)

    print(f"\n    Total truck tyre companies from Companies House: {len(all_companies)}")

    # =========================================================================
    # STEP 2: Search for real website for each company
    # =========================================================================
    print("\n[2] SEARCHING FOR REAL WEBSITES")
    print("-" * 60)
    print("    Searching Bing for each company's real website...")
    print()

    found_websites = 0

    for i, company in enumerate(all_companies):
        name = company['name']
        print(f"    [{i+1}/{len(all_companies)}] {name[:50]}...", end=' ', flush=True)

        # Try Bing first
        website = search_for_website_bing(name)

        # If no result, try DuckDuckGo
        if not website:
            website = search_for_website_google(name)

        if website:
            company['website'] = website
            found_websites += 1
            print(f"Found: {urlparse(website).netloc[:30]}")
        else:
            company['website'] = ''
            print("Not found")

        time.sleep(1.5)  # Be polite to search engines

    print(f"\n    Websites found: {found_websites}/{len(all_companies)}")

    # =========================================================================
    # STEP 3: Verify found websites and extract contact info
    # =========================================================================
    print("\n[3] VERIFYING WEBSITES & EXTRACTING CONTACTS")
    print("-" * 60)

    verified_count = 0

    for i, company in enumerate(all_companies):
        website = company.get('website', '')

        if not website:
            company['website_verified'] = False
            company['phone'] = ''
            company['email'] = ''
            company['services'] = ''
            continue

        print(f"    [{i+1}/{len(all_companies)}] Verifying {urlparse(website).netloc[:35]}...", end=' ', flush=True)

        verification = verify_website(website)

        if verification and verification['works']:
            company['website'] = verification['final_url']
            company['website_verified'] = True
            company['phone'] = verification.get('phone', '')
            company['email'] = verification.get('email', '')
            company['services'] = ', '.join(verification.get('services', []))
            company['is_tyre_related'] = verification.get('is_tyre_related', False)
            verified_count += 1

            phone_display = verification.get('phone') or 'No phone'
            tyre_mark = '✓' if verification.get('is_tyre_related') else '?'
            print(f"✓ Working [{tyre_mark}] | {phone_display[:15]}")
        else:
            company['website_verified'] = False
            company['phone'] = ''
            company['email'] = ''
            company['services'] = ''
            print("✗ Not working")

        time.sleep(0.3)

    print(f"\n    Verified working: {verified_count}")

    # =========================================================================
    # STEP 4: Estimate revenue
    # =========================================================================
    print("\n[4] ESTIMATING COMPANY SIZE & REVENUE")
    print("-" * 60)

    for c in all_companies:
        est = estimate_revenue(c)
        c['size'] = est['size']
        c['employees'] = est['employees']
        c['revenue_estimate'] = est['revenue']
        c['revenue_low'] = est['rev_low']
        c['revenue_high'] = est['rev_high']
        c['revenue_indicators'] = est['indicators']

    # Sort by revenue
    all_companies.sort(key=lambda x: x.get('revenue_high', 0), reverse=True)

    # Stats
    total = len(all_companies)
    with_website = len([c for c in all_companies if c.get('website')])
    verified = len([c for c in all_companies if c.get('website_verified')])
    with_phone = len([c for c in all_companies if c.get('phone')])
    tyre_related = len([c for c in all_companies if c.get('is_tyre_related')])

    total_rev_low = sum(c.get('revenue_low', 0) for c in all_companies)
    total_rev_high = sum(c.get('revenue_high', 0) for c in all_companies)

    print(f"\n    RESULTS:")
    print(f"    Total companies: {total}")
    print(f"    Websites found: {with_website}")
    print(f"    Verified working: {verified}")
    print(f"    Tyre-related content: {tyre_related}")
    print(f"    With phone: {with_phone}")
    print(f"    Market estimate: £{total_rev_low/1000000:.0f}M - £{total_rev_high/1000000:.0f}M")

    # =========================================================================
    # EXPORT
    # =========================================================================
    print("\n[5] EXPORTING DATA")
    print("-" * 60)

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "UK Truck Tyre Companies"

    headers = ['Company Name', 'Website', 'Verified', 'Tyre Related', 'Phone', 'Email',
               'Size', 'Employees', 'Revenue Estimate', 'Revenue Indicators',
               'Services', 'Address', 'Company Number', 'Date Created', 'Source']

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    verified_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row, c in enumerate(all_companies, 2):
        ws.cell(row=row, column=1, value=c.get('name', ''))
        ws.cell(row=row, column=2, value=c.get('website', ''))

        verified_cell = ws.cell(row=row, column=3, value='Yes' if c.get('website_verified') else 'No')
        if c.get('website_verified'):
            verified_cell.fill = verified_fill

        ws.cell(row=row, column=4, value='Yes' if c.get('is_tyre_related') else 'No')
        ws.cell(row=row, column=5, value=c.get('phone', ''))
        ws.cell(row=row, column=6, value=c.get('email', ''))
        ws.cell(row=row, column=7, value=c.get('size', ''))
        ws.cell(row=row, column=8, value=c.get('employees', ''))
        ws.cell(row=row, column=9, value=c.get('revenue_estimate', ''))
        ws.cell(row=row, column=10, value=c.get('revenue_indicators', ''))
        ws.cell(row=row, column=11, value=c.get('services', ''))
        ws.cell(row=row, column=12, value=c.get('address', ''))
        ws.cell(row=row, column=13, value=c.get('company_number', ''))
        ws.cell(row=row, column=14, value=c.get('date_created', ''))
        ws.cell(row=row, column=15, value=c.get('source', ''))

    widths = [45, 50, 10, 12, 18, 35, 12, 12, 18, 20, 25, 50, 15, 12, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:O{len(all_companies) + 1}"

    wb.save(f"{OUTPUT_FILE}.xlsx")
    print(f"    Saved: {OUTPUT_FILE}.xlsx")

    # JSON
    with open(f"{OUTPUT_FILE}.json", 'w', encoding='utf-8') as f:
        json.dump(all_companies, f, indent=2, ensure_ascii=False)
    print(f"    Saved: {OUTPUT_FILE}.json")

    # CSV
    fieldnames = ['name', 'website', 'website_verified', 'is_tyre_related', 'phone', 'email',
                  'size', 'employees', 'revenue_estimate', 'revenue_low', 'revenue_high',
                  'revenue_indicators', 'services', 'address', 'company_number', 'date_created', 'source']

    with open(f"{OUTPUT_FILE}.csv", 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(all_companies)
    print(f"    Saved: {OUTPUT_FILE}.csv")

    # =========================================================================
    # SUMMARY
    # =========================================================================
    print("\n" + "=" * 80)
    print("COMPLETE")
    print("=" * 80)

    print(f"\nTotal: {total} truck tyre companies")
    print(f"Websites found: {with_website}")
    print(f"Verified working: {verified}")
    print(f"Market estimate: £{total_rev_low/1000000:.0f}M - £{total_rev_high/1000000:.0f}M")

    print("\n\nCOMPANIES WITH VERIFIED REAL WEBSITES:")
    print("-" * 80)

    verified_list = [c for c in all_companies if c.get('website_verified')]
    for i, c in enumerate(verified_list[:30], 1):
        tyre_mark = '✓' if c.get('is_tyre_related') else '?'
        print(f"\n{i}. [{tyre_mark}] {c['name']}")
        print(f"   Website: {c['website']}")
        if c.get('phone'):
            print(f"   Phone: {c['phone']}")
        print(f"   Size: {c['size']} | Revenue: {c['revenue_estimate']}")

    if len(verified_list) > 30:
        print(f"\n... and {len(verified_list) - 30} more (see Excel file)")


if __name__ == "__main__":
    main()
