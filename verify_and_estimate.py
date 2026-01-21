#!/usr/bin/env python3
"""
=============================================================================
VERIFY & ESTIMATE - UK TRUCK TYRE COMPANIES
=============================================================================
Takes the existing 846 companies database, verifies websites,
scrapes contact details, and estimates annual revenue.

Date: January 2026
=============================================================================
"""

import requests
from bs4 import BeautifulSoup
import json
import time
import re
import csv
from datetime import datetime
from scraper import INDUSTRY_DATABASE

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Accept': 'text/html,application/xhtml+xml',
}

OUTPUT_FILE = 'uk_truck_tyres_846_VERIFIED'


def verify_website(url):
    """Check if website works and is truck tyre related"""
    if not url or not url.startswith('http'):
        return None

    try:
        response = requests.get(url, headers=HEADERS, timeout=10, allow_redirects=True)

        if response.status_code == 200:
            html = response.text.lower()

            # Check if it's truck tyre related
            truck_keywords = ['truck', 'hgv', 'lorry', 'commercial', 'fleet', 'trailer', 'tyre', 'tire']
            keyword_count = sum(1 for kw in truck_keywords if kw in html)

            result = {
                'status': 'Working',
                'is_truck_related': keyword_count >= 2,
                'phone': None,
                'email': None,
                'services': []
            }

            # Extract phone
            phone_patterns = [
                r'0800[\s\-]?\d{3}[\s\-]?\d{3,4}',
                r'0\d{2,4}[\s\-]?\d{3}[\s\-]?\d{3,4}',
            ]
            for pattern in phone_patterns:
                matches = re.findall(pattern, response.text)
                for m in matches:
                    cleaned = re.sub(r'[\s\-]', '', m)
                    if len(cleaned) >= 10 and not cleaned.startswith('00000'):
                        result['phone'] = m.strip()
                        break
                if result['phone']:
                    break

            # Extract email
            emails = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', response.text)
            for email in emails:
                if not any(x in email.lower() for x in ['.png', '.jpg', '.css', '.js', 'example.com', 'wix', 'sentry']):
                    result['email'] = email
                    break

            # Services
            if '24' in html and ('hour' in html or 'hr' in html):
                result['services'].append('24hr')
            if 'mobile' in html:
                result['services'].append('Mobile')
            if 'fleet' in html:
                result['services'].append('Fleet')

            return result

        else:
            return {'status': f'Error {response.status_code}', 'is_truck_related': False}

    except requests.exceptions.Timeout:
        return {'status': 'Timeout', 'is_truck_related': False}
    except requests.exceptions.ConnectionError:
        return {'status': 'Connection Error', 'is_truck_related': False}
    except Exception as e:
        return {'status': f'Error', 'is_truck_related': False}


def estimate_revenue(company):
    """Estimate annual revenue based on company characteristics"""

    name = company.get('name', '').lower()
    business_type = company.get('businessType', '').lower()
    website_verified = company.get('website_verified', False)
    services = company.get('services', '')

    # Base score
    score = 0
    indicators = []

    # Business type scoring
    if 'manufacturer' in business_type:
        score += 80
        indicators.append('Manufacturer')
    elif 'wholesaler' in business_type:
        score += 60
        indicators.append('Wholesaler')
    elif 'national' in business_type or 'network' in business_type:
        score += 50
        indicators.append('National/Network')
    elif 'regional' in business_type:
        score += 30
        indicators.append('Regional')
    elif 'mobile' in business_type or 'emergency' in business_type:
        score += 20
        indicators.append('Mobile/Emergency')
    elif 'independent' in business_type:
        score += 15
        indicators.append('Independent')
    else:
        score += 10

    # Name indicators
    if any(x in name for x in ['group', 'national', 'uk', 'british']):
        score += 20
        indicators.append('National name')
    if any(x in name for x in ['wholesale', 'distributor', 'supply']):
        score += 15
        indicators.append('Wholesale/Distribution')

    # Website and services
    if website_verified:
        score += 10
        if '24hr' in services:
            score += 10
            indicators.append('24hr service')
        if 'Fleet' in services:
            score += 15
            indicators.append('Fleet services')
        if 'Mobile' in services:
            score += 5

    # Revenue estimation based on score
    if score >= 80:
        size = 'Large'
        employees = '50-500+'
        revenue_low = 10000000
        revenue_high = 100000000
    elif score >= 50:
        size = 'Medium-Large'
        employees = '20-100'
        revenue_low = 2000000
        revenue_high = 15000000
    elif score >= 30:
        size = 'Medium'
        employees = '10-30'
        revenue_low = 500000
        revenue_high = 3000000
    elif score >= 20:
        size = 'Small-Medium'
        employees = '5-15'
        revenue_low = 200000
        revenue_high = 1000000
    else:
        size = 'Small'
        employees = '1-5'
        revenue_low = 50000
        revenue_high = 300000

    return {
        'size_category': size,
        'estimated_employees': employees,
        'revenue_low': revenue_low,
        'revenue_high': revenue_high,
        'revenue_display': f'£{revenue_low/1000000:.1f}M - £{revenue_high/1000000:.1f}M' if revenue_low >= 1000000 else f'£{revenue_low/1000:.0f}K - £{revenue_high/1000:.0f}K',
        'score': score,
        'indicators': ', '.join(indicators) if indicators else 'Standard'
    }


def main():
    print("=" * 80)
    print("VERIFY & ESTIMATE - UK TRUCK TYRE COMPANIES")
    print("=" * 80)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Processing {len(INDUSTRY_DATABASE)} companies...")
    print()

    all_companies = []

    # Process each company
    print("[1] VERIFYING WEBSITES & SCRAPING CONTACTS")
    print("-" * 60)

    for i, company in enumerate(INDUSTRY_DATABASE):
        # Copy company data
        c = {
            'name': company.get('name', ''),
            'website': company.get('website', ''),
            'phone': company.get('phone', ''),
            'email': company.get('email', ''),
            'address': company.get('address', ''),
            'city': company.get('city', ''),
            'region': company.get('region', ''),
            'postcode': company.get('postcode', ''),
            'businessType': company.get('businessType', ''),
            'source': company.get('source', ''),
        }

        # Verify website
        if c['website']:
            if (i + 1) % 50 == 0 or i == 0:
                print(f"    Processing {i+1}/{len(INDUSTRY_DATABASE)}...")

            verification = verify_website(c['website'])

            if verification:
                c['website_status'] = verification['status']
                c['website_verified'] = verification['status'] == 'Working'
                c['is_truck_related'] = verification.get('is_truck_related', False)

                # Use scraped contact if we don't have one
                if verification.get('phone') and not c['phone']:
                    c['phone'] = verification['phone']
                if verification.get('email') and not c['email']:
                    c['email'] = verification['email']
                if verification.get('services'):
                    c['services'] = ', '.join(verification['services'])
            else:
                c['website_status'] = 'No website'
                c['website_verified'] = False
                c['is_truck_related'] = False

            time.sleep(0.2)  # Be polite
        else:
            c['website_status'] = 'No website'
            c['website_verified'] = False
            c['is_truck_related'] = True  # Assume true if from our database

        all_companies.append(c)

    # Estimate revenue for all
    print("\n[2] ESTIMATING REVENUE")
    print("-" * 60)

    for c in all_companies:
        estimates = estimate_revenue(c)
        c['size_category'] = estimates['size_category']
        c['estimated_employees'] = estimates['estimated_employees']
        c['revenue_estimate'] = estimates['revenue_display']
        c['revenue_low'] = estimates['revenue_low']
        c['revenue_high'] = estimates['revenue_high']
        c['revenue_indicators'] = estimates['indicators']

    # Sort by revenue
    all_companies.sort(key=lambda x: x.get('revenue_high', 0), reverse=True)

    # Statistics
    total = len(all_companies)
    with_website = len([c for c in all_companies if c.get('website')])
    verified = len([c for c in all_companies if c.get('website_verified')])
    with_phone = len([c for c in all_companies if c.get('phone')])
    with_email = len([c for c in all_companies if c.get('email')])

    large = len([c for c in all_companies if c.get('size_category') == 'Large'])
    med_large = len([c for c in all_companies if c.get('size_category') == 'Medium-Large'])
    medium = len([c for c in all_companies if c.get('size_category') == 'Medium'])
    small_med = len([c for c in all_companies if c.get('size_category') == 'Small-Medium'])
    small = len([c for c in all_companies if c.get('size_category') == 'Small'])

    total_rev_low = sum(c.get('revenue_low', 0) for c in all_companies)
    total_rev_high = sum(c.get('revenue_high', 0) for c in all_companies)

    print(f"\n    STATISTICS:")
    print(f"    Total companies: {total}")
    print(f"    With website URL: {with_website}")
    print(f"    Verified working: {verified}")
    print(f"    With phone: {with_phone}")
    print(f"    With email: {with_email}")
    print()
    print(f"    SIZE BREAKDOWN:")
    print(f"    Large: {large}")
    print(f"    Medium-Large: {med_large}")
    print(f"    Medium: {medium}")
    print(f"    Small-Medium: {small_med}")
    print(f"    Small: {small}")
    print()
    print(f"    TOTAL MARKET ESTIMATE:")
    print(f"    £{total_rev_low/1000000:.0f}M - £{total_rev_high/1000000:.0f}M")

    # Export
    print("\n[3] EXPORTING")
    print("-" * 60)

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "UK Truck Tyre Companies"

    headers = [
        'Company Name', 'Website', 'Website Status', 'Phone', 'Email',
        'Size Category', 'Est. Employees', 'Est. Revenue', 'Revenue Indicators',
        'Business Type', 'Address', 'City', 'Region', 'Postcode', 'Services', 'Source'
    ]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    large_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    medium_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row, c in enumerate(all_companies, 2):
        ws.cell(row=row, column=1, value=c.get('name', ''))
        ws.cell(row=row, column=2, value=c.get('website', ''))
        ws.cell(row=row, column=3, value=c.get('website_status', ''))
        ws.cell(row=row, column=4, value=c.get('phone', ''))
        ws.cell(row=row, column=5, value=c.get('email', ''))
        ws.cell(row=row, column=6, value=c.get('size_category', ''))
        ws.cell(row=row, column=7, value=c.get('estimated_employees', ''))
        ws.cell(row=row, column=8, value=c.get('revenue_estimate', ''))
        ws.cell(row=row, column=9, value=c.get('revenue_indicators', ''))
        ws.cell(row=row, column=10, value=c.get('businessType', ''))
        ws.cell(row=row, column=11, value=c.get('address', ''))
        ws.cell(row=row, column=12, value=c.get('city', ''))
        ws.cell(row=row, column=13, value=c.get('region', ''))
        ws.cell(row=row, column=14, value=c.get('postcode', ''))
        ws.cell(row=row, column=15, value=c.get('services', ''))
        ws.cell(row=row, column=16, value=c.get('source', ''))

        # Color by size
        size = c.get('size_category', '')
        if size == 'Large':
            ws.cell(row=row, column=6).fill = large_fill
        elif size in ['Medium-Large', 'Medium']:
            ws.cell(row=row, column=6).fill = medium_fill

    # Column widths
    widths = [45, 40, 15, 18, 35, 15, 15, 18, 30, 25, 40, 20, 20, 12, 25, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:P{len(all_companies) + 1}"

    wb.save(f"{OUTPUT_FILE}.xlsx")
    print(f"    Saved: {OUTPUT_FILE}.xlsx")

    # JSON
    with open(f"{OUTPUT_FILE}.json", 'w', encoding='utf-8') as f:
        json.dump(all_companies, f, indent=2, ensure_ascii=False)
    print(f"    Saved: {OUTPUT_FILE}.json")

    # CSV
    fieldnames = ['name', 'website', 'website_status', 'website_verified', 'phone', 'email',
                  'size_category', 'estimated_employees', 'revenue_estimate', 'revenue_low', 'revenue_high',
                  'revenue_indicators', 'businessType', 'address', 'city', 'region', 'postcode', 'services', 'source']

    with open(f"{OUTPUT_FILE}.csv", 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(all_companies)
    print(f"    Saved: {OUTPUT_FILE}.csv")

    print("\n" + "=" * 80)
    print("COMPLETE")
    print("=" * 80)
    print(f"Processed: {total} companies")
    print(f"Total market estimate: £{total_rev_low/1000000:.0f}M - £{total_rev_high/1000000:.0f}M")

    # Top 20
    print("\nTOP 20 BY ESTIMATED REVENUE:")
    print("-" * 80)
    for i, c in enumerate(all_companies[:20], 1):
        v = "✓" if c.get('website_verified') else "✗"
        print(f"{i:2}. [{v}] {c.get('name', '')[:40]}")
        print(f"       Size: {c.get('size_category')} | Revenue: {c.get('revenue_estimate')}")
        if c.get('phone'):
            print(f"       Phone: {c.get('phone')}")


if __name__ == "__main__":
    main()
